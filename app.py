import os
import io
import json
import sqlite3
import functools
import tempfile
import pandas as pd

# Load .env file for local development (ignored in production where env vars are set directly)
_env_path = os.path.join(os.path.dirname(__file__), '.env')
if os.path.exists(_env_path):
    with open(_env_path) as _f:
        for _line in _f:
            _line = _line.strip()
            if _line and not _line.startswith('#') and '=' in _line:
                _k, _v = _line.split('=', 1)
                os.environ.setdefault(_k.strip(), _v.strip())
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, g, session
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta

try:
    import outlook_connector as _oc
    _OUTLOOK_AVAILABLE = True
except Exception:
    _oc = None
    _OUTLOOK_AVAILABLE = False

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'cpainc2026')

# In production (Railway), DATA_DIR points to the mounted persistent volume.
# Locally it defaults to the project folder.
_DATA_DIR = os.environ.get('DATA_DIR', os.path.dirname(__file__))
os.makedirs(_DATA_DIR, exist_ok=True)
DATABASE = os.path.join(_DATA_DIR, 'CPAinc.sqlite')

# ── Email configuration (set these as Railway environment variables) ──────────
MAIL_SERVER   = os.environ.get('MAIL_SERVER',   'smtp.gmail.com')
MAIL_PORT     = int(os.environ.get('MAIL_PORT', '587'))
MAIL_USERNAME = os.environ.get('MAIL_USERNAME', '')
MAIL_PASSWORD = os.environ.get('MAIL_PASSWORD', '')
MAIL_FROM     = os.environ.get('MAIL_FROM',     MAIL_USERNAME)
MAIL_TO       = os.environ.get('MAIL_TO',       '')   # comma-separated alert recipients
DAILY_TASK_KEY= os.environ.get('DAILY_TASK_KEY','')   # secret key for /admin/run-daily-tasks

SESSION_TIMEOUT_SECONDS = 28800  # 8 hours


def send_email(to_addrs, subject, body_html, attachments=None):
    """Send an email via SMTP. to_addrs can be a string or list. Returns (ok, error)."""
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text      import MIMEText
    from email.mime.base      import MIMEBase
    from email                import encoders as _enc
    if not MAIL_USERNAME or not MAIL_PASSWORD:
        return False, 'Email not configured (MAIL_USERNAME / MAIL_PASSWORD missing)'
    if isinstance(to_addrs, str):
        to_addrs = [a.strip() for a in to_addrs.split(',') if a.strip()]
    if not to_addrs:
        return False, 'No recipients'
    try:
        msg = MIMEMultipart('mixed')
        msg['From']    = MAIL_FROM or MAIL_USERNAME
        msg['To']      = ', '.join(to_addrs)
        msg['Subject'] = subject
        msg.attach(MIMEText(body_html, 'html'))
        if attachments:
            for fname, data in attachments:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(data)
                _enc.encode_base64(part)
                part.add_header('Content-Disposition', f'attachment; filename="{fname}"')
                msg.attach(part)
        with smtplib.SMTP(MAIL_SERVER, MAIL_PORT, timeout=15) as s:
            s.ehlo()
            s.starttls()
            s.login(MAIL_USERNAME, MAIL_PASSWORD)
            s.sendmail(msg['From'], to_addrs, msg.as_string())
        return True, None
    except Exception as e:
        return False, str(e)

def get_db():
    """
    Open a SQLite connection (cached on flask.g) with a small retry on
    transient Dropbox / cloud-storage hiccups. macOS Dropbox occasionally
    denies reads on synced files while it's reconciling; SQLite surfaces
    those as `sqlite3.DatabaseError: authorization denied` or
    `disk I/O error`. We retry briefly so a hiccup doesn't turn into a 500.
    """
    if 'db' not in g:
        import time as _t
        transient_markers = (
            'authorization denied',
            'disk i/o error',
            'database is locked',
            'unable to open database',
        )
        last_err = None
        for attempt in range(4):  # ~0 + 0.2 + 0.4 + 0.6 s = 1.2 s worst case
            try:
                g.db = sqlite3.connect(DATABASE, timeout=10)
                g.db.row_factory = sqlite3.Row
                return g.db
            except sqlite3.Error as e:
                last_err = e
                msg = str(e).lower()
                if not any(m in msg for m in transient_markers):
                    raise
                if attempt < 3:
                    _t.sleep((attempt + 1) * 0.2)
        raise last_err
    return g.db

@app.teardown_appcontext
def close_db(e=None):
    db = g.pop('db', None)
    if db:
        db.close()

# ── Filters ───────────────────────────────────────────────────────────────────

@app.template_filter('fromjson')
def fromjson_filter(value):
    if not value:
        return {}
    try:
        return json.loads(value)
    except Exception:
        return {}

@app.template_filter('us_date')
def us_date_filter(value):
    """Convert YYYY-MM-DD (or datetime) to the configured display format."""
    if not value:
        return '—'
    try:
        s = str(value)[:10]
        return datetime.strptime(s, '%Y-%m-%d').strftime(get_date_format())
    except Exception:
        return value

@app.template_filter('fmtdate')
def _to_iso(raw):
    """Convert any pipeline date value to a YYYY-MM-DD string, or return None.
    Handles ISO format (2026-10-06 or 2026-10-06T00:00:00.000Z)
    and US format (10/6/2026 or 10/23/2026) as stored by some Railway imports."""
    if not raw:
        return None
    s = str(raw).strip()
    if not s:
        return None
    # ISO: starts with 4-digit year
    if len(s) >= 10 and s[4] == '-':
        return s[:10]
    # US format MM/DD/YYYY or M/D/YYYY
    try:
        return datetime.strptime(s.split('T')[0], '%m/%d/%Y').strftime('%Y-%m-%d')
    except Exception:
        pass
    return None


def fmt_date(val):
    if not val:
        return ''
    try:
        s = str(val)[:10]
        return datetime.strptime(s, '%Y-%m-%d').strftime(get_date_format())
    except Exception:
        return str(val)[:10]

@app.template_filter('fmtcurrency')
def fmt_currency(val):
    if val is None or val == '':
        return ''
    try:
        return f'${float(val):,.2f}'
    except Exception:
        return str(val)

@app.template_filter('fmtpct')
def fmt_pct(val):
    if val is None or val == '':
        return ''
    try:
        return f'{float(val)*100:.1f}%'
    except Exception:
        return str(val)

@app.template_filter('format_number')
def format_number(val):
    try:
        return f'{int(val):,}'
    except Exception:
        return str(val) if val is not None else '—'

# ── Settings helpers ──────────────────────────────────────────────────────────

COMMISSION_SPLIT = 0.60

def get_commission_split():
    try:
        row = get_db().execute('SELECT value FROM Settings WHERE key = "commission_split"').fetchone()
        return float(row[0]) if row else COMMISSION_SPLIT
    except Exception:
        return COMMISSION_SPLIT

def get_payment_tolerance():
    try:
        row = get_db().execute('SELECT value FROM Settings WHERE key = "payment_tolerance"').fetchone()
        return float(row[0]) if row else 0.01
    except Exception:
        return 0.01

def get_account_splits():
    try:
        rows = get_db().execute('SELECT account_name, countries, split_rate FROM AccountSplits').fetchall()
        result = []
        for r in rows:
            countries = set(c.strip().upper() for c in (r[1] or '').split(',') if c.strip())
            result.append((r[0].lower(), countries, float(r[2])))
        return result
    except Exception:
        return []

def get_kristin_split():
    try:
        row = get_db().execute('SELECT value FROM Settings WHERE key = "kristin_split"').fetchone()
        return float(row[0]) if row else 0.70
    except Exception:
        return 0.70

def get_kristin_cut():
    try:
        row = get_db().execute('SELECT value FROM Settings WHERE key = "kristin_cut"').fetchone()
        return float(row[0]) if row else 0.10
    except Exception:
        return 0.10

# Supported display formats: (strftime_pattern, label, example)
DATE_FORMAT_OPTIONS = [
    ('%m/%d/%Y',  'MM/DD/YYYY',   'US standard — 05/09/2026'),
    ('%d/%m/%Y',  'DD/MM/YYYY',   'European — 09/05/2026'),
    ('%Y-%m-%d',  'YYYY-MM-DD',   'ISO 8601 — 2026-05-09'),
    ('%d-%m-%Y',  'DD-MM-YYYY',   'European dashes — 09-05-2026'),
    ('%d.%m.%Y',  'DD.MM.YYYY',   'European dots — 09.05.2026'),
    ('%b %d, %Y', 'Mon DD, YYYY', 'Abbreviated — May 09, 2026'),
    ('%d %b %Y',  'DD Mon YYYY',  'International — 09 May 2026'),
    ('%B %d, %Y', 'Month DD, YYYY','Long — May 09, 2026'),
]

def get_date_format():
    """Return the configured strftime date format, cached on flask g per request."""
    if hasattr(g, '_date_fmt'):
        return g._date_fmt
    try:
        row = get_db().execute('SELECT value FROM Settings WHERE key="date_format"').fetchone()
        fmt = row[0] if row else '%m/%d/%Y'
        # Validate it's one of our known patterns
        known = {o[0] for o in DATE_FORMAT_OPTIONS}
        if fmt not in known:
            fmt = '%m/%d/%Y'
    except Exception:
        fmt = '%m/%d/%Y'
    g._date_fmt = fmt
    return fmt

def split_for_account(account, country, account_splits, default_split):
    if not account:
        return default_split
    acct_lower = account.lower()
    country_upper = (country or '').strip().upper()
    for name, countries, rate in account_splits:
        if name == acct_lower:
            if not countries:
                return rate
            if country_upper in countries:
                return rate
    return default_split

# Booking types suppressed from all Team (non-Kristin) views
_TEAM_SUPPRESS_SQL = (
    "LOWER(COALESCE({alias}BookingAssociate,'')) = 'kristin house' OR "
    "LOWER(COALESCE({alias}BookingType,'')) NOT IN "
    "('other services', 'other', 'conference management', 'cm')"
)

def effective_split(associate, account, country, account_splits, default_split, kristin_split, kristin_cut):
    """Net split for a booking row, accounting for Kristin House structure."""
    if associate and associate.strip().lower() == 'kristin house':
        return kristin_split
    base = split_for_account(account, country, account_splits, default_split)
    return max(0.0, base - kristin_cut)

def split_label(associate, split, default_split, kristin_split, kristin_cut):
    """Human-readable label for the split column."""
    if associate and associate.strip().lower() == 'kristin house':
        return f"Kristin {split*100:.0f}%"
    base = split + kristin_cut   # reconstruct gross before Kristin cut
    return f"{base*100:.0f}% − {kristin_cut*100:.0f}% = {split*100:.0f}%"

# ── Dashboard ─────────────────────────────────────────────────────────────────

@app.route('/dashboard')
def dashboard():
    user = get_current_user()
    if not has_permission(user, 'dashboard'):
        flash('You do not have access to the Dashboard.', 'error')
        return redirect(url_for('pipeline'))
    from datetime import timedelta
    db      = get_db()
    today   = datetime.today().strftime('%Y-%m-%d')
    next30  = (datetime.today() + timedelta(days=30)).strftime('%Y-%m-%d')
    year    = datetime.today().strftime('%Y')

    who = request.args.get('who', 'kristin')  # 'kristin' or 'team'
    is_kristin_view = (who == 'kristin')
    title_suffix = 'Kristin House' if is_kristin_view else 'Team Associates'

    default_split  = get_commission_split()
    account_splits = get_account_splits()
    kristin_split  = get_kristin_split()
    kristin_cut    = get_kristin_cut()

    # Bookings that received a payment this year
    paid_rows = db.execute('''
        SELECT r.BookingId AS bid, r.AccountName, r.Country,
               substr(r.StartDate,1,10) AS start_date,
               substr(r.EndDate,1,10)   AS end_date,
               COALESCE(r.USDRevenue, r.Revenue, 0) AS rev,
               r.CommissionPercent AS comm_pct,
               r.RoomRate AS room_rate,
               r.BookingAssociate AS associate,
               SUM(c.FinalPayment) AS total_paid
        FROM ReportPipeline r
        JOIN ChkRegNote c ON c.BookingID = r.BookingId
        WHERE (r.BookingStatus IS NULL OR r.BookingStatus NOT LIKE "%Cancel%")
          AND (c.Cancelled IS NULL OR c.Cancelled = 0)
          AND c.FinalPayment > 0
          AND substr(c.DateOnCheck,1,4) = ?
          AND (LOWER(COALESCE(r.BookingAssociate,'')) = 'kristin house'
               OR LOWER(COALESCE(r.BookingType,'')) NOT IN ('other services', 'other', 'conference management', 'cm'))
        GROUP BY r.BookingId
    ''', (year,)).fetchall()
    paid_bids = {r['bid'] for r in paid_rows}

    # Bookings with this-year start dates and no payment yet
    unpaid_rows = db.execute('''
        SELECT r.BookingId AS bid, r.AccountName, r.Country,
               substr(r.StartDate,1,10) AS start_date,
               substr(r.EndDate,1,10)   AS end_date,
               COALESCE(r.USDRevenue, r.Revenue, 0) AS rev,
               r.CommissionPercent AS comm_pct,
               r.RoomRate AS room_rate,
               r.BookingAssociate AS associate
        FROM ReportPipeline r
        WHERE (r.BookingStatus IS NULL OR r.BookingStatus NOT LIKE "%Cancel%")
          AND substr(r.StartDate,1,4) = ?
          AND (LOWER(COALESCE(r.BookingAssociate,'')) = 'kristin house'
               OR LOWER(COALESCE(r.BookingType,'')) NOT IN ('other services', 'other', 'conference management', 'cm'))
    ''', (year,)).fetchall()

    pickup_map = {r[0]: float(r[1] or 0) for r in db.execute('''
        SELECT BookingID, SUM(ActualPickup) FROM Pickup
        WHERE ActualPickup IS NOT NULL AND ActualPickup > 0
        GROUP BY BookingID
    ''').fetchall()}

    def _is_kristin(associate):
        return (associate or '').strip().lower() == 'kristin house'

    cats = {
        1: {'label': 'Completed & Paid',               'color': '#198754', 'count': 0, 'revenue': 0, 'value': 0},
        2: {'label': 'Future & Paid (Incentive)',       'color': '#20c997', 'count': 0, 'revenue': 0, 'value': 0},
        3: {'label': 'Completed – Has HHR, Unpaid',  'color': '#fd7e14', 'count': 0, 'revenue': 0, 'value': 0},
        4: {'label': 'Completed – No HHR Yet',        'color': '#dc3545', 'count': 0, 'revenue': 0, 'value': 0},
        5: {'label': 'Future / Not Paid (2026-04-10 – 2026-09-30)', 'color': '#1a3a5c', 'count': 0, 'revenue': 0, 'value': 0},
    }
    meetings_next30 = 0

    for b in paid_rows:
        if is_kristin_view != _is_kristin(b['associate']):
            continue
        end_date   = b['end_date'] or ''
        start_date = b['start_date'] or ''
        rev        = float(b['rev'] or 0)
        total_paid_amt = float(b['total_paid'] or 0)
        if start_date and today <= start_date <= next30:
            meetings_next30 += 1
        if end_date and end_date < today:
            cats[1]['count']   += 1
            cats[1]['revenue'] += rev
            cats[1]['value']   += total_paid_amt
        else:
            cats[2]['count']   += 1
            cats[2]['revenue'] += rev
            cats[2]['value']   += total_paid_amt

    for b in unpaid_rows:
        if is_kristin_view != _is_kristin(b['associate']):
            continue
        bid        = b['bid']
        is_prepaid = bid in paid_bids
        end_date   = b['end_date'] or ''
        start_date = b['start_date'] or ''
        rev        = float(b['rev'] or 0)
        comm_pct   = float(b['comm_pct'] or 0)
        room_rate  = float(b['room_rate'] or 0)
        split      = effective_split(b['associate'], b['AccountName'], b['Country'],
                                     account_splits, default_split, kristin_split, kristin_cut)
        est_comm   = rev * comm_pct * split
        if start_date and today <= start_date <= next30:
            meetings_next30 += 1
        is_done = end_date and end_date < today
        if is_prepaid and is_done:
            continue  # completed and paid — already in cat 1
        if is_kristin_view:
            if not is_prepaid and is_done and bid in pickup_map:
                actual_comm = pickup_map[bid] * room_rate * comm_pct * split if room_rate else est_comm
                cats[3]['count']   += 1; cats[3]['revenue'] += rev; cats[3]['value'] += actual_comm
            elif not is_prepaid and is_done:
                cats[4]['count']   += 1; cats[4]['revenue'] += rev; cats[4]['value'] += est_comm
            elif not start_date or start_date < '2026-10-01':
                cats[5]['count']   += 1; cats[5]['revenue'] += rev; cats[5]['value'] += est_comm
        else:
            # Team view: future only — show Kristin's 10% cut of team bookings
            if not is_done and (not start_date or start_date < '2026-10-01'):
                kristin_est = rev * comm_pct * kristin_cut
                cats[5]['count']   += 1; cats[5]['revenue'] += rev; cats[5]['value'] += kristin_est

    total_paid              = cats[1]['value'] + cats[2]['value']
    total_commission        = sum(cats[i]['value'] for i in cats)
    missing_commission_count = cats[3]['count'] + cats[4]['count']

    assoc_filter = '' if not is_kristin_view else ''
    upcoming = db.execute('''
        SELECT BookingId, EventName, Customer, AccountName, StartDate, BookingStatus,
               BookingAssociate AS associate
        FROM ReportPipeline
        WHERE substr(StartDate,1,10) >= ?
          AND (BookingStatus IS NULL OR BookingStatus NOT LIKE "%Cancel%")
          AND (LOWER(COALESCE(BookingAssociate,'')) = 'kristin house'
               OR LOWER(COALESCE(BookingType,'')) NOT IN ('other services', 'other', 'conference management', 'cm'))
        ORDER BY StartDate ASC
    ''', (today,)).fetchall()
    upcoming = [u for u in upcoming if is_kristin_view == _is_kristin(u['associate'])][:10]

    return render_template('dashboard.html',
        cats=cats, total_paid=total_paid, total_commission=total_commission,
        meetings_next30=meetings_next30, missing_commission_count=missing_commission_count,
        upcoming=upcoming, now=datetime.today(), who=who, title_suffix=title_suffix)

# ── Home / Bookings List ──────────────────────────────────────────────────────

@app.route('/')
def index():
    return redirect(url_for('status_board'))

@app.route('/pipeline')
def pipeline():
    user = get_current_user()
    if not has_permission(user, 'bookings_view'):
        return render_template('no_access.html', section='Bookings')
    db = get_db()
    search    = request.args.get('search', '').strip()
    status    = request.args.get('status', '').strip()
    associate = request.args.get('associate', '').strip()

    query = 'SELECT * FROM ReportPipeline WHERE (BookingStatus IS NULL OR BookingStatus NOT LIKE "%Cancel%")'
    params = []

    # Account-level filter for non-admins
    accounts = get_user_account_filter(user)
    if accounts is not None:
        if accounts:
            ph = ','.join('?' * len(accounts))
            query += f' AND AccountName IN ({ph})'
            params.extend(accounts)
        else:
            query += ' AND 1=0'

    if search:
        query += ' AND (EventName LIKE ? OR "Booking Name" LIKE ? OR AccountName LIKE ? OR Customer LIKE ? OR CAST(BookingId AS TEXT) LIKE ?)'
        params += [f'%{search}%', f'%{search}%', f'%{search}%', f'%{search}%', f'%{search}%']
    if status:
        query += ' AND BookingStatus = ?'
        params.append(status)
    if associate:
        query += ' AND BookingAssociate = ?'
        params.append(associate)

    query += ' ORDER BY BookedDate DESC'
    try:
        bookings  = db.execute(query, params).fetchall()
        statuses  = [r[0] for r in db.execute('SELECT DISTINCT BookingStatus FROM ReportPipeline WHERE BookingStatus IS NOT NULL ORDER BY 1').fetchall()]
        associates = [r[0] for r in db.execute('SELECT DISTINCT BookingAssociate FROM ReportPipeline WHERE BookingAssociate IS NOT NULL ORDER BY 1').fetchall()]
    except Exception:
        bookings, statuses, associates = [], [], []

    # Calculate Kristin's commission for each booking
    kristin_split = get_kristin_split()
    kristin_cut   = get_kristin_cut()
    enriched = []
    for b in bookings:
        row = dict(b)
        total_comm = float(row.get('USDCommissionableAmount') or 0)
        assoc = (row.get('BookingAssociate') or '').strip().lower()
        if assoc == 'kristin house':
            row['kristin_comm'] = total_comm * kristin_split
        else:
            row['kristin_comm'] = total_comm * kristin_cut
        enriched.append(row)

    return render_template('index.html', bookings=enriched, search=search, status=status,
                           associate=associate, statuses=statuses, associates=associates)

# ── Booking Detail ────────────────────────────────────────────────────────────

@app.route('/booking/<int:booking_id>')
def booking_detail(booking_id):
    user = get_current_user()
    db = get_db()
    booking = db.execute('SELECT * FROM ReportPipeline WHERE BookingId = ?', (booking_id,)).fetchone()
    if not booking:
        flash('Booking not found.', 'error')
        return redirect(url_for('pipeline'))
    acct_filter = get_user_account_filter(user)
    if acct_filter is not None and booking['AccountName'] not in acct_filter:
        flash('You do not have access to this booking.', 'error')
        return redirect(url_for('pipeline'))
    pickups = db.execute('SELECT *, rowid FROM Pickup WHERE BookingID = ? ORDER BY rowid', (booking_id,)).fetchall()
    checks  = db.execute('SELECT * FROM ChkRegNote WHERE BookingID = ? ORDER BY DateOnCheck DESC', (booking_id,)).fetchall()
    booking_contracts = db.execute(
        'SELECT id, filename, upload_date FROM booking_contract WHERE CAST(booking_id AS TEXT)=CAST(? AS TEXT) ORDER BY upload_date DESC',
        (booking_id,)
    ).fetchall()
    default_split  = get_commission_split()
    account_splits = get_account_splits()
    kristin_split  = get_kristin_split()
    kristin_cut    = get_kristin_cut()
    split = effective_split(booking['BookingAssociate'], booking['AccountName'], booking['Country'],
                            account_splits, default_split, kristin_split, kristin_cut)
    try:
        comm_pct = float(booking['CommissionPercent'] or 0)
    except Exception:
        comm_pct = 0
    pickup_configs = db.execute(
        "SELECT id, hotel, event_name FROM pickup_config WHERE CAST(booking_id AS TEXT)=CAST(? AS TEXT) AND status='active' ORDER BY id",
        (booking_id,)
    ).fetchall()
    return render_template('booking_detail.html', booking=booking, pickups=pickups, checks=checks,
        comm_pct=comm_pct, split=split, booking_contracts=booking_contracts,
        pickup_configs=pickup_configs)

# ── Cancel Booking ────────────────────────────────────────────────────────────

@app.route('/booking/<int:booking_id>/cancel', methods=['POST'])
def booking_cancel(booking_id):
    db = get_db()
    db.execute('UPDATE ReportPipeline SET BookingStatus = ? WHERE BookingId = ?', ('Cancelled', booking_id))
    db.commit()
    flash('Booking has been cancelled.', 'success')
    return redirect(url_for('booking_detail', booking_id=booking_id))

# ── Booking Contracts ─────────────────────────────────────────────────────────

@app.route('/booking/<booking_id>/contract/upload', methods=['POST'])
def booking_contract_upload(booking_id):
    files = request.files.getlist('contract_files')
    db = get_db()
    last_bcid = None
    for f in files:
        if f and f.filename:
            db.execute(
                'INSERT INTO booking_contract (booking_id, filename, file_data, upload_date) VALUES (?,?,?,date("now"))',
                (booking_id, f.filename, f.read())
            )
            last_bcid = db.execute('SELECT last_insert_rowid()').fetchone()[0]
    db.commit()
    # Go straight into AI parsing flow for the last uploaded file
    if last_bcid:
        return redirect(url_for('booking_contract_extract', booking_id=booking_id, bcid=last_bcid))
    flash('No file selected.', 'error')
    return redirect(url_for('booking_detail', booking_id=booking_id))


@app.route('/booking/<booking_id>/contract/<int:bcid>/extract')
def booking_contract_extract(booking_id, bcid):
    """Route to pickup extraction — picks the right hotel or shows a selector."""
    db = get_db()
    bc = db.execute(
        'SELECT id FROM booking_contract WHERE id=? AND CAST(booking_id AS TEXT)=CAST(? AS TEXT)',
        (bcid, booking_id)
    ).fetchone()
    if not bc:
        flash('Contract not found.', 'error')
        return redirect(url_for('booking_detail', booking_id=booking_id))
    configs = db.execute(
        "SELECT * FROM pickup_config WHERE CAST(booking_id AS TEXT)=CAST(? AS TEXT) AND status != 'archived' ORDER BY id",
        (booking_id,)
    ).fetchall()
    if len(configs) == 0:
        flash('No pickup records found for this booking — add one first.', 'warning')
        return redirect(url_for('booking_detail', booking_id=booking_id))
    if len(configs) == 1:
        return redirect(url_for('booking_contract_parse_for_pickup',
                                booking_id=booking_id, bcid=bcid, cid=configs[0]['id']))
    return redirect(url_for('booking_contract_select_pickup',
                            booking_id=booking_id, bcid=bcid))


@app.route('/booking/<booking_id>/contract/<int:bcid>/select-pickup')
def booking_contract_select_pickup(booking_id, bcid):
    """Choose which pickup hotel record a booking-level contract belongs to."""
    db = get_db()
    bc = db.execute(
        'SELECT id, filename FROM booking_contract WHERE id=? AND CAST(booking_id AS TEXT)=CAST(? AS TEXT)',
        (bcid, booking_id)
    ).fetchone()
    if not bc:
        flash('Contract not found.', 'error')
        return redirect(url_for('booking_detail', booking_id=booking_id))
    configs = db.execute(
        "SELECT * FROM pickup_config WHERE CAST(booking_id AS TEXT)=CAST(? AS TEXT) AND status != 'archived' ORDER BY id",
        (booking_id,)
    ).fetchall()
    return render_template('booking_contract_select_pickup.html',
                           booking_id=booking_id, bc=bc, configs=configs)


@app.route('/booking/<booking_id>/contract/<int:bcid>/parse-for-pickup/<int:cid>')
def booking_contract_parse_for_pickup(booking_id, bcid, cid):
    """Parse a stored booking contract and show the pickup review page for a chosen hotel."""
    from pickup_utils import parse_contract_document
    import base64
    db = get_db()
    bc = db.execute(
        'SELECT filename, file_data FROM booking_contract WHERE id=? AND CAST(booking_id AS TEXT)=CAST(? AS TEXT)',
        (bcid, booking_id)
    ).fetchone()
    if not bc or not bc['file_data']:
        flash('Contract file not found.', 'error')
        return redirect(url_for('booking_detail', booking_id=booking_id))
    config = db.execute('SELECT * FROM pickup_config WHERE id=?', (cid,)).fetchone()
    if not config:
        flash('Pickup record not found.', 'error')
        return redirect(url_for('booking_detail', booking_id=booking_id))
    extracted = parse_contract_document(bc['file_data'], filename=bc['filename'])
    if extracted.get('error'):
        flash(f'Could not parse contract: {extracted["error"]}', 'error')
        return redirect(url_for('booking_detail', booking_id=booking_id))
    # file_b64 left empty — file already stored in booking_contract, no need to round-trip it
    return render_template('pickup_contract_review.html',
                           config=config, extracted=extracted,
                           filename=bc['filename'], file_b64='')

@app.route('/booking/<booking_id>/contract/<int:cid>/download')
def booking_contract_download(booking_id, cid):
    db = get_db()
    row = db.execute(
        'SELECT filename, file_data FROM booking_contract WHERE id=? AND CAST(booking_id AS TEXT)=CAST(? AS TEXT)',
        (cid, booking_id)
    ).fetchone()
    if not row:
        flash('Contract file not found.', 'error')
        return redirect(url_for('booking_detail', booking_id=booking_id))
    from flask import send_file
    import io
    return send_file(io.BytesIO(row['file_data']), download_name=row['filename'], as_attachment=True)

@app.route('/booking/<booking_id>/contract/<int:cid>/delete', methods=['POST'])
def booking_contract_delete(booking_id, cid):
    db = get_db()
    db.execute('DELETE FROM booking_contract WHERE id=? AND CAST(booking_id AS TEXT)=CAST(? AS TEXT)', (cid, booking_id))
    db.commit()
    flash('Contract deleted.', 'success')
    return redirect(url_for('booking_detail', booking_id=booking_id))

# ── New Booking ───────────────────────────────────────────────────────────────

@app.route('/booking/new', methods=['GET', 'POST'])
def booking_new():
    user = get_current_user()
    if not has_permission(user, 'bookings_edit'):
        flash('You do not have permission to add bookings.', 'error')
        return redirect(url_for('pipeline'))
    if request.method == 'POST':
        db = get_db()
        fields = ['BookingId','BookingName','BookingAssociate','ShareType','BookingType','BookingStatus',
                  'Addendum','BookedDate','StartDate','EndDate','AccountName','EventName',
                  'Customer','Address','City','State','Country','Brand','Chain','Currency',
                  'ExchangeRate','Advance','ContractedAmount','PeakRooms','TotalRoomNights',
                  'RoomRate','Revenue','OtherRevenue','USDRate','USDRevenue',
                  'CommissionPercent','USDCommissionableAmount']
        values = [request.form.get(f, '').strip() or None for f in fields]
        placeholders = ', '.join(fields)
        qmarks = ', '.join(['?' for _ in fields])
        try:
            db.execute(f'INSERT INTO ReportPipeline ({placeholders}) VALUES ({qmarks})', values)
            db.commit()
            flash('Booking added successfully.', 'success')
            return redirect(url_for('booking_detail', booking_id=request.form.get('BookingId')))
        except Exception as e:
            flash(f'Error saving booking: {e}', 'error')
    return render_template('booking_new.html', prefill=request.args)

# ── Edit Booking ──────────────────────────────────────────────────────────────

@app.route('/booking/<int:booking_id>/edit', methods=['GET', 'POST'])
def booking_edit(booking_id):
    user = get_current_user()
    if not has_permission(user, 'bookings_edit'):
        flash('You do not have permission to edit bookings.', 'error')
        return redirect(url_for('booking_detail', booking_id=booking_id))
    db = get_db()
    booking = db.execute('SELECT * FROM ReportPipeline WHERE BookingId = ?', (booking_id,)).fetchone()
    if not booking:
        flash('Booking not found.', 'error')
        return redirect(url_for('pipeline'))
    if request.method == 'POST':
        fields = ['BookingName','BookingAssociate','ShareType','BookingType','BookingStatus','Addendum',
                  'BookedDate','StartDate','EndDate','AccountName','EventName','Customer',
                  'Address','City','State','Country','Brand','Chain','Currency','ExchangeRate',
                  'Advance','ContractedAmount','PeakRooms','TotalRoomNights','RoomRate',
                  'Revenue','OtherRevenue','USDRate','USDRevenue','CommissionPercent',
                  'USDCommissionableAmount']
        set_clause = ', '.join([f'{f} = ?' for f in fields])
        values = [request.form.get(f, '').strip() or None for f in fields]
        values.append(booking_id)
        try:
            db.execute(f'UPDATE ReportPipeline SET {set_clause} WHERE BookingId = ?', values)
            db.commit()
            flash('Booking updated.', 'success')
            return redirect(url_for('booking_detail', booking_id=booking_id))
        except Exception as e:
            flash(f'Error updating booking: {e}', 'error')
    return render_template('booking_edit.html', booking=booking)

# ── Import Bookings ───────────────────────────────────────────────────────────

@app.route('/import', methods=['GET', 'POST'])
def import_bookings():
    user = get_current_user()
    if not has_permission(user, 'import_bookings'):
        flash('You do not have access to this import tool.', 'error')
        return redirect(url_for('pipeline'))
    if request.method == 'POST':
        file = request.files.get('file')
        if not file or not file.filename:
            flash('No file selected.', 'error')
            return redirect(url_for('import_bookings'))
        ext = file.filename.rsplit('.', 1)[-1].lower()
        tmp_path = f'/tmp/cpainc_import.{ext}'
        file.save(tmp_path)
        try:
            if ext == 'csv':
                df = pd.read_csv(tmp_path, dtype=str, encoding='utf-8-sig')
                # Normalise column names immediately — no header-row scan needed
                df.columns = [str(c).strip() for c in df.columns]
            elif ext == 'xls':
                # ConferenceDirect exports .xls files that are sometimes HTML tables.
                # Try xlrd first; fall back to HTML parsing when xlrd rejects it.
                try:
                    df = pd.read_excel(tmp_path, engine='xlrd', header=None)
                except Exception:
                    tables = pd.read_html(tmp_path)
                    if not tables:
                        raise ValueError('No table found in file.')
                    df = tables[0].astype(str).replace('nan', '')
                    # read_html already parses <th> as column headers — check if
                    # we still need to find the header row or if columns are correct.
                if not any(str(v).strip().lower() == 'booking id' for v in df.columns):
                    header_row = None
                    for i, row in df.iterrows():
                        if any(str(v).strip().lower() == 'booking id' for v in row.values):
                            header_row = i
                            break
                    if header_row is None:
                        full_text = ' '.join(str(v) for v in df.values.flatten() if str(v) != 'nan')
                        if any(kw in full_text.upper() for kw in
                               ('POST EVENT HISTORY', 'FINAL TOTAL PICKUP', 'AUDIT DETAIL',
                                'HOUSING HISTORY', 'TOTAL ESTIMATED ROOMS COMMISSION')):
                            flash(
                                'This looks like a Housing History Report (HHR), not a bookings export. '
                                'Please use "Import Housing History" from the nav menu instead.',
                                'error'
                            )
                        else:
                            flash('Could not find header row with "Booking Id" in the file.', 'error')
                        return redirect(url_for('import_bookings'))
                    df.columns = df.iloc[header_row]
                    df = df.iloc[header_row + 1:].reset_index(drop=True)
            else:
                df = pd.read_excel(tmp_path, engine='openpyxl', header=None)
                header_row = None
                for i, row in df.iterrows():
                    if any(str(v).strip().lower() == 'booking id' for v in row.values):
                        header_row = i
                        break
                if header_row is None:
                    full_text = ' '.join(str(v) for v in df.values.flatten() if str(v) != 'nan')
                    if any(kw in full_text.upper() for kw in
                           ('POST EVENT HISTORY', 'FINAL TOTAL PICKUP', 'AUDIT DETAIL',
                            'HOUSING HISTORY', 'TOTAL ESTIMATED ROOMS COMMISSION')):
                        flash(
                            'This looks like a Housing History Report (HHR), not a bookings export. '
                            'Please use "Import Housing History" from the nav menu instead.',
                            'error'
                        )
                    else:
                        flash('Could not find header row with "Booking Id" in the file.', 'error')
                    return redirect(url_for('import_bookings'))
                df.columns = df.iloc[header_row]
                df = df.iloc[header_row + 1:].reset_index(drop=True)
            # ── Normalise column names from CSV/Excel exports ───────────────
            col_map = {
                'booking id':                               'Booking Id',
                # Event ID (parent event)
                'event: event id':                          'Event ID',
                # Associate
                'booking associate: full name':             'Booking Associate',
                'booking associate':                        'Booking Associate',
                # Event name
                'event: event name':                        'Event Name',
                'event: event  name':                       'Event Name',
                'booking name':                             'Booking Name',
                # Account / Customer
                'account name: account name':               'Account Name',
                'account name':                             'Account Name',
                'customer account name':                    'Account Name',
                'entity to invoice: account name':          'Customer',
                'entity to invoice':                        'Customer',
                # Address
                'entity to invoice: billing street':        'Address',
                'entity to invoice: billing city':          'City',
                'entity to invoice: billing state/province': 'State',
                'entity to invoice: billing country':       'Country',
                # Status & type
                'status':                                   'Booking Status',
                'type':                                     'Booking Type',
                'subtype':                                  'Booking Type',
                'addendum submitted':                       'Addendum',
                # Financial
                'total room revenue':                       'Revenue',
                'total room nights':                        'Total Room Nights',
                'peak rooms':                               'Peak Rooms',
                'contract commission %':                    'Commission Percent',
                'contract rate':                            'Room Rate',
                'gross booking value (calc) (converted)':   'Contracted Amount',
                'gross booking value':                      'Contracted Amount',
                'net contract value (converted)':           'USD Commissionable Amount',
                'total commission amount':                  'USD Commissionable Amount',
            }
            df.columns = [col_map.get(str(c).strip().lower(), str(c).strip()) for c in df.columns]
            # Deduplicate: keep first occurrence if two columns map to same name
            seen = {}
            for i, c in enumerate(df.columns):
                if c not in seen:
                    seen[c] = i
            df = df.iloc[:, sorted(seen.values())]

            bid_col = next((c for c in df.columns if c.lower() == 'booking id' or c == 'Booking Id'), None)
            if bid_col and bid_col != 'Booking Id':
                df = df.rename(columns={bid_col: 'Booking Id'})
            if 'Booking Id' not in df.columns:
                flash('Could not find Booking ID column in the file.', 'error')
                return redirect(url_for('import_bookings'))

            df = df.dropna(subset=['Booking Id'])
            df['Booking Id'] = df['Booking Id'].astype(str).str.strip()
            df = df[df['Booking Id'] != '']

            db = get_db()
            added, skipped, updated = 0, 0, 0
            currency_fields = {'Contracted Amount', 'Room Rate', 'Revenue', 'Other Revenue', 'USD Commissionable Amount'}

            fields = [
                'Booking Id', 'Event ID', 'Booking Name', 'Booking Associate', 'Share Type', 'Booking Type', 'Booking Status',
                'Addendum', 'Booked Date', 'Start Date', 'End Date', 'Account Name', 'Event Name',
                'Customer', 'Address', 'City', 'State', 'Country', 'Brand', 'Chain', 'Currency',
                'Exchange Rate', 'Advance', 'Contracted Amount', 'Peak Rooms', 'Total Room Nights',
                'Room Rate', 'Revenue', 'Other Revenue', 'USD Rate', 'USD Revenue',
                'Commission Percent', 'USD Commissionable Amount'
            ]

            for _, row in df.iterrows():
                bid = str(row.get('Booking Id', '')).strip().split('.')[0]
                if not bid:
                    continue

                values = []
                for f in fields:
                    v = row.get(f)
                    is_na = pd.isna(v) if not isinstance(v, str) else False
                    if is_na or v is None:
                        values.append(None)
                        continue
                    val = str(v).strip()
                    # Treat '-' as blank (CD uses '-' for empty fields like Event ID)
                    if val == '-':
                        values.append(None)
                        continue
                    if f in currency_fields:
                        val = val.replace('USD', '').replace('$', '').replace(',', '').strip()
                    # Normalise Commission Percent: "10.00%" → "0.10"
                    if f == 'Commission Percent' and val.endswith('%'):
                        try:
                            val = str(float(val.rstrip('%')) / 100)
                        except ValueError:
                            val = None
                    values.append(val if val else None)

                # Map from display field names to DB column names
                db_map = {
                    'Booking Id': 'BookingId', 'Event ID': 'EventID',
                    'Booking Name': 'BookingName',
                    'Booking Associate': 'BookingAssociate', 'Share Type': 'ShareType',
                    'Booking Type': 'BookingType', 'Booking Status': 'BookingStatus',
                    'Addendum': 'Addendum', 'Booked Date': 'BookedDate',
                    'Start Date': 'StartDate', 'End Date': 'EndDate',
                    'Account Name': 'AccountName', 'Event Name': 'EventName',
                    'Customer': 'Customer', 'Address': 'Address', 'City': 'City',
                    'State': 'State', 'Country': 'Country', 'Brand': 'Brand',
                    'Chain': 'Chain', 'Currency': 'Currency',
                    'Exchange Rate': 'ExchangeRate', 'Advance': 'Advance',
                    'Contracted Amount': 'ContractedAmount', 'Peak Rooms': 'PeakRooms',
                    'Total Room Nights': 'TotalRoomNights', 'Room Rate': 'RoomRate',
                    'Revenue': 'Revenue', 'Other Revenue': 'OtherRevenue',
                    'USD Rate': 'USDRate', 'USD Revenue': 'USDRevenue',
                    'Commission Percent': 'CommissionPercent',
                    'USD Commissionable Amount': 'USDCommissionableAmount',
                }
                db_fields = [db_map[f] for f in fields]

                exists = db.execute('SELECT 1 FROM ReportPipeline WHERE BookingId = ?', (bid,)).fetchone()
                if exists:
                    # Update Commission Percent and Event ID if previously missing
                    comm_idx    = fields.index('Commission Percent')
                    eventid_idx = fields.index('Event ID')
                    changed = False
                    if values[comm_idx] is not None:
                        db.execute(
                            'UPDATE ReportPipeline SET CommissionPercent = ? WHERE BookingId = ? AND (CommissionPercent IS NULL OR CommissionPercent = "")',
                            (values[comm_idx], bid)
                        )
                        if db.execute('SELECT changes()').fetchone()[0]:
                            changed = True
                    if values[eventid_idx] is not None:
                        db.execute(
                            'UPDATE ReportPipeline SET EventID = ? WHERE BookingId = ? AND (EventID IS NULL OR EventID = "")',
                            (values[eventid_idx], bid)
                        )
                        if db.execute('SELECT changes()').fetchone()[0]:
                            changed = True
                    if changed:
                        updated += 1
                    skipped += 1
                    continue

                placeholders_str = ', '.join(db_fields)
                qmarks_str = ', '.join(['?' for _ in db_fields])
                db.execute(f'INSERT INTO ReportPipeline ({placeholders_str}) VALUES ({qmarks_str})', values)
                added += 1

                # Auto-create pickup tracking entry for this new booking
                val_dict = dict(zip(db_fields, values))
                _create_pickup_config_from_booking(
                    db, bid,
                    account   = val_dict.get('AccountName'),
                    event     = val_dict.get('BookingName') or val_dict.get('EventName'),
                    hotel     = val_dict.get('Customer'),
                    start_str = val_dict.get('StartDate'),
                    end_str   = val_dict.get('EndDate'),
                    peak_rooms= val_dict.get('PeakRooms'),
                    room_rate = val_dict.get('RoomRate'),
                )

            db.commit()
            msg = f'Import complete: {added} added, {skipped} already existed (skipped).'
            if updated:
                msg += f' {updated} existing record{"s" if updated != 1 else ""} updated with missing commission % or Event ID.'
            flash(msg, 'success')
        except Exception as e:
            flash(f'Import error: {e}', 'error')
        return redirect(url_for('pipeline'))
    return render_template('import.html')

# ── Import Cancelled Meetings ─────────────────────────────────────────────────

@app.route('/import/cancelled', methods=['GET', 'POST'])
def import_cancelled():
    user = get_current_user()
    if not has_permission(user, 'import_cancelled'):
        flash('You do not have access to this import tool.', 'error')
        return redirect(url_for('pipeline'))
    from datetime import datetime as _dt, timedelta as _td

    def excel_date(v):
        """Convert an Excel serial date, datetime object, or date string to YYYY-MM-DD."""
        if v is None:
            return None
        # pandas may return datetime/Timestamp objects directly
        if hasattr(v, 'strftime'):
            return v.strftime('%Y-%m-%d')
        try:
            f = float(v)
            if f > 0:
                return (_dt(1899, 12, 30) + _td(days=f)).strftime('%Y-%m-%d')
        except (ValueError, TypeError):
            pass
        s = str(v).strip()
        for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y'):
            try:
                return _dt.strptime(s, fmt).strftime('%Y-%m-%d')
            except ValueError:
                pass
        return s or None

    if request.method == 'POST':
        file = request.files.get('file')
        if not file or not file.filename:
            flash('No file selected.', 'error')
            return redirect(url_for('import_cancelled'))
        ext = file.filename.rsplit('.', 1)[-1].lower()
        tmp_path = f'/tmp/cpainc_cancelled.{ext}'
        file.save(tmp_path)
        try:
            df = pd.read_excel(tmp_path, engine='xlrd' if ext == 'xls' else 'openpyxl', header=None)
            header_row = None
            bid_col = None
            for i, row in df.iterrows():
                for v in row.values:
                    if str(v).strip().lower() == 'booking id':
                        header_row = i
                        bid_col = str(v).strip()
                        break
                if header_row is not None:
                    break
            if header_row is None:
                flash('Could not find header row with "Booking ID" in the file.', 'error')
                return redirect(url_for('import_cancelled'))
            df.columns = df.iloc[header_row]
            df = df.iloc[header_row + 1:].reset_index(drop=True)
            # Normalise column name to 'Booking Id' regardless of capitalisation
            col_map = {c: 'Booking Id' for c in df.columns if str(c).strip().lower() == 'booking id'}
            df = df.rename(columns=col_map)
            df = df.dropna(subset=['Booking Id'])
            df['Booking Id'] = df['Booking Id'].astype(str).str.strip()
            df = df[df['Booking Id'] != '']

            def clean(v, is_currency=False):
                if pd.isna(v) if hasattr(pd, 'isna') else v != v:
                    return None
                s = str(v).strip()
                if is_currency:
                    s = s.replace('$', '').replace(',', '').strip()
                return s or None

            db = get_db()
            updated, inserted, skipped = 0, 0, 0
            pickups_archived = 0
            import datetime as _dt
            cancel_note = f"CANCELLED — imported {_dt.date.today().strftime('%Y-%m-%d')}"

            def archive_pickup_cards(bid):
                nonlocal pickups_archived
                cards = db.execute(
                    "SELECT id, notes FROM pickup_config WHERE CAST(booking_id AS TEXT)=CAST(? AS TEXT) AND status != 'archived'",
                    (bid,)
                ).fetchall()
                for card in cards:
                    existing_notes = (card['notes'] or '').strip()
                    new_notes = cancel_note + ('\n' + existing_notes if existing_notes else '')
                    db.execute(
                        "UPDATE pickup_config SET status='archived', notes=? WHERE id=?",
                        (new_notes, card['id'])
                    )
                    pickups_archived += len(cards)

            for _, row in df.iterrows():
                bid = str(row.get('Booking Id', '')).strip().split('.')[0]
                if not bid:
                    continue

                exists = db.execute(
                    'SELECT BookingStatus FROM ReportPipeline WHERE BookingId = ?', (bid,)
                ).fetchone()

                if exists:
                    current_status = (exists['BookingStatus'] or '').strip()
                    if 'cancel' in current_status.lower():
                        skipped += 1
                        continue
                    db.execute(
                        'UPDATE ReportPipeline SET BookingStatus = ? WHERE BookingId = ?',
                        ('Cancelled', bid)
                    )
                    db.execute(
                        'UPDATE ChkRegNote SET Cancelled = 1 WHERE BookingID = ? AND (Cancelled IS NULL OR Cancelled = 0)',
                        (bid,)
                    )
                    archive_pickup_cards(bid)
                    updated += 1
                else:
                    db.execute('''INSERT INTO ReportPipeline
                        (BookingId,BookingAssociate,ShareType,BookingType,BookingStatus,
                         Addendum,BookedDate,StartDate,EndDate,AccountName,EventName,
                         Customer,Address,City,State,Country,Brand,Chain,Currency,
                         ExchangeRate,Advance,ContractedAmount,PeakRooms,TotalRoomNights,
                         RoomRate,Revenue,OtherRevenue,USDRate,USDRevenue,
                         CommissionPercent,USDCommissionableAmount)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''', (
                        bid,
                        clean(row.get('Booking Associate')),
                        clean(row.get('Share Type')),
                        clean(row.get('Booking Type')),
                        'Cancelled',
                        clean(row.get('Addendum')),
                        excel_date(row.get('Booked Date')),
                        excel_date(row.get('Start Date')),
                        excel_date(row.get('End Date')),
                        clean(row.get('Account Name')),
                        clean(row.get('Event Name')),
                        clean(row.get('Customer')),
                        clean(row.get('Address')),
                        clean(row.get('City')),
                        clean(row.get('State')),
                        clean(row.get('Country')),
                        clean(row.get('Brand')),
                        clean(row.get('Chain')),
                        clean(row.get('Currency')),
                        clean(row.get('Exchange Rate')),
                        clean(row.get('Advance')),
                        clean(row.get('Contracted Amount'), True),
                        clean(row.get('Peak Rooms')),
                        clean(row.get('Total Room Nights')),
                        clean(row.get('Room Rate'), True),
                        clean(row.get('Revenue'), True),
                        clean(row.get('Other Revenue'), True),
                        clean(row.get('USD Rate')),
                        clean(row.get('USD Revenue'), True),
                        clean(row.get('Commission Percent')),
                        clean(row.get('USD Commissionable Amount'), True),
                    ))
                    archive_pickup_cards(bid)
                    inserted += 1

            db.commit()
            pickup_msg = f', {pickups_archived} pickup card{"s" if pickups_archived != 1 else ""} archived' if pickups_archived else ''
            flash(
                f'Cancelled import complete: {updated} updated to Cancelled, '
                f'{inserted} new bookings added, {skipped} already Cancelled (skipped){pickup_msg}.',
                'success'
            )
        except Exception as e:
            flash(f'Import error: {e}', 'error')
        return redirect(url_for('pipeline'))
    return render_template('import_cancelled.html')

# ── Import Payments ───────────────────────────────────────────────────────────

@app.route('/import/payments', methods=['GET', 'POST'])
def import_payments():
    user = get_current_user()
    if not has_permission(user, 'import_payments'):
        flash('You do not have access to this import tool.', 'error')
        return redirect(url_for('pipeline'))
    if request.method == 'POST':
        file = request.files.get('file')
        if not file or not file.filename:
            flash('No file selected.', 'error')
            return redirect(url_for('import_payments'))
        ext = file.filename.rsplit('.', 1)[-1].lower()
        tmp_path = f'/tmp/cpainc_payments.{ext}'
        file.save(tmp_path)
        try:
            if ext == 'csv':
                df = pd.read_csv(tmp_path, dtype=str, encoding='utf-8-sig')
                df.columns = [str(c).strip() for c in df.columns]
                # Normalize common CSV column name variants → expected names
                col_map = {
                    'Booking ID': 'Booking Number',
                    'BookingID': 'Booking Number',
                    'Check Number': 'Reference',
                    'Check #': 'Reference',
                    'Payment Amount': 'Amount',
                    'Payment Date': 'Date',
                }
                df.rename(columns=col_map, inplace=True)
                if 'Booking Number' not in df.columns:
                    flash('CSV must contain a "Booking Number" or "Booking ID" column.', 'error')
                    return redirect(url_for('import_payments'))
                df = df[pd.to_numeric(df['Booking Number'], errors='coerce').notna()]
                df['Booking Number'] = df['Booking Number'].astype(str).str.strip().str.split('.').str[0]
            else:
                # Excel — locate the header row. Two layouts are supported:
                #   • Legacy: a "Booking Number" column
                #   • Associate Payment (1099) Report: header at row B10 with
                #     "Document Number" / "Booking: Booking ID" columns
                df = pd.read_excel(tmp_path, engine='xlrd' if ext == 'xls' else 'openpyxl', header=None)
                header_row, fmt = None, None
                for i, row in df.iterrows():
                    vals = [str(v).strip() for v in row.values]
                    if 'Booking Number' in vals:
                        header_row, fmt = i, 'legacy'
                        break
                    if 'Document Number' in vals or 'Booking: Booking ID' in vals:
                        header_row, fmt = i, '1099'
                        break
                if header_row is None:
                    flash('Could not find a recognised header row '
                          '("Booking Number" or "Document Number") in the file.', 'error')
                    return redirect(url_for('import_payments'))

                df.columns = df.iloc[header_row]
                df = df.iloc[header_row + 1:].reset_index(drop=True)

                if fmt == '1099':
                    # Map the verbose 1099 column names onto the internal names
                    col_map = {
                        'Booking: Booking ID':                                'Booking Number',
                        'Booking: Booking Associate: Associate Name':         'Booking Associate',
                        'Booking: Subtype':                                   'Booking Type',
                        'Booking: Account Name: Entity to Invoice Name':      'Account Name',
                        'Booking: Event: Event  Name':                        'Event Name',
                        'Booking: Event: Event Name':                         'Event Name',
                        'Booking: Start Date':                                'Start Date',
                        'Booking: End Date':                                  'End Date',
                        'Booking: Entity to Invoice: Entity to Invoice Name': 'Hotel',
                        'AP Invoice Number':                                  'Invoice Number',
                        'Check #':                                            'Reference',
                        'Date':                                               'Date',
                        'Currency':                                           'Currency',
                        'Payment':                                            'Payment',
                        'Converted Currency':                                 'Converted Currency',
                        'Payment (converted)':                                'Amount (USD)',
                    }
                    df.columns = [col_map.get(str(c).strip(), str(c).strip()) for c in df.columns]

                    if 'Document Number' in df.columns:
                        # Drop the trailing "Total" summary row (+ record-count row)
                        is_total = df['Document Number'].astype(str).str.strip().str.lower().eq('total')
                        if is_total.any():
                            df = df.iloc[:is_total.idxmax()].reset_index(drop=True)
                        # Housing & Registration lines leave "Booking: Booking ID"
                        # blank but carry the booking in the Document Number, e.g.
                        # "169500-DEC 2025" → 169500. Backfill from there.
                        doc_bid = df['Document Number'].astype(str).str.extract(r'^(\d{4,7})', expand=False)
                        bn = (df['Booking Number'].astype(str).str.strip()
                              .replace({'': None, 'nan': None, 'None': None}))
                        df['Booking Number'] = bn.fillna(doc_bid)

                df = df[pd.to_numeric(df['Booking Number'], errors='coerce').notna()]
                df['Booking Number'] = df['Booking Number'].astype(str).str.strip().str.split('.').str[0]

            db = get_db()
            today = datetime.now().strftime('%Y-%m-%d')
            added, skipped, new_bookings = 0, 0, []

            for _, row in df.iterrows():
                booking_id = str(row.get('Booking Number', '')).strip()
                if not booking_id:
                    continue

                raw_date = row.get('Date')
                if pd.isna(raw_date) if hasattr(pd, 'isna') else raw_date != raw_date:
                    payment_date = None
                else:
                    try:
                        payment_date = pd.to_datetime(raw_date).strftime('%Y-%m-%d')
                    except Exception:
                        payment_date = str(raw_date)[:10]

                check_num = str(row.get('Reference', '') or '').strip() or None

                def _amt(val):
                    if val is None: return None
                    try:
                        if pd.isna(val): return None
                    except Exception: pass
                    try:
                        return float(str(val).replace(',', '').replace('$', '').strip())
                    except Exception:
                        return None

                # USD amount to record. Legacy uses 'Amount'; the 1099 report uses
                # 'Amount (USD)' = Payment (converted), always USD. Per the currency
                # rule a non-USD payment is booked at the converted USD figure,
                # never the foreign amount.
                amount = _amt(row.get('Amount'))
                if amount is None:
                    amount = _amt(row.get('Amount (USD)'))
                if amount is None:
                    amount = _amt(row.get('Payment'))
                if amount is not None:
                    amount = round(amount, 2)

                # For non-USD payments, note the original currency + amount.
                currency = str(row.get('Currency', '') or '').strip()
                special_notes = None
                if currency and currency.upper() != 'USD':
                    orig_pay = _amt(row.get('Payment'))
                    parts = [f'Currency: {currency}']
                    if orig_pay is not None:
                        parts.append(f'Original: {orig_pay:,.2f}')
                    special_notes = ' | '.join(parts)

                if payment_date:
                    dup = db.execute(
                        'SELECT 1 FROM ChkRegNote WHERE BookingID = ? AND DateOnCheck LIKE ? AND FinalPayment = ?',
                        (booking_id, f'{payment_date}%', amount)
                    ).fetchone()
                    if dup:
                        skipped += 1
                        continue

                invoice = str(row.get('Invoice Number', '') or '').upper()
                is_advance = 1 if ('HI INC' in invoice or 'MI INC' in invoice) else 0

                booking_exists = db.execute('SELECT 1 FROM ReportPipeline WHERE BookingId = ?', (booking_id,)).fetchone()
                if not booking_exists:
                    def safe_date(val):
                        try:
                            if pd.isna(val): return None
                            return pd.to_datetime(val).strftime('%Y-%m-%d')
                        except Exception:
                            return None
                    db.execute('''INSERT INTO ReportPipeline
                        (BookingId,BookingAssociate,BookingType,AccountName,EventName,Customer,StartDate,EndDate,BookingStatus)
                        VALUES (?,?,?,?,?,?,?,?,"Definite")''', (
                        booking_id,
                        str(row.get('Booking Associate', '') or '').strip() or None,
                        str(row.get('Booking Type', '') or '').strip() or None,
                        str(row.get('Account Name', '') or '').strip() or None,
                        str(row.get('Event Name', '') or '').strip() or None,
                        str(row.get('Hotel', '') or '').strip() or None,
                        safe_date(row.get('Start Date')),
                        safe_date(row.get('End Date')),
                    ))
                    new_bookings.append({
                        'booking_id': booking_id,
                        'account':    str(row.get('Account Name', '') or '').strip(),
                        'event':      str(row.get('Event Name', '') or '').strip(),
                        'hotel':      str(row.get('Hotel', '') or '').strip(),
                        'amount':     amount,
                    })

                db.execute('''INSERT INTO ChkRegNote
                    (BookingID,FinalPayment,Check_,DateOnCheck,EntryDate,SpecialNotes,Cancelled,AuditFlag,Advance)
                    VALUES (?,?,?,?,?,?,0,0,?)''',
                    (booking_id, amount, check_num, payment_date, today, special_notes, is_advance))
                added += 1

            db.commit()
            return render_template('import_payments_result.html',
                                   added=added, skipped=skipped, new_bookings=new_bookings)
        except Exception as e:
            flash(f'Import error: {e}', 'error')
            return redirect(url_for('import_payments'))
    return render_template('import_payments.html')

# ── Import Payment Voucher (PDF) ──────────────────────────────────────────────

@app.route('/import/voucher', methods=['GET', 'POST'])
def import_voucher():
    user = get_current_user()
    if not has_permission(user, 'import_payments'):
        flash('You do not have access to this import tool.', 'error')
        return redirect(url_for('pipeline'))

    if request.method == 'GET':
        return render_template('import_voucher.html')

    # ── Parse uploaded PDFs ───────────────────────────────────────────────────
    files = request.files.getlist('files')
    files = [f for f in files if f and f.filename]
    if not files:
        flash('No files selected.', 'error')
        return redirect(url_for('import_voucher'))
    bad = [f.filename for f in files if not f.filename.lower().endswith('.pdf')]
    if bad:
        flash(f'Only PDF files are supported: {", ".join(bad)}', 'error')
        return redirect(url_for('import_voucher'))

    import pdfplumber, re, io as _io

    def _parse_voucher_pdf(raw_bytes):
        """Parse one Payment Voucher PDF. Returns a dict with header fields + booking_rows.

        Line-item approach: each invoice number line in the raw text ends with
        an optional end-date and two amounts, e.g.:
            237264-F1-TD 2026 Green Dot Corp ... 3/12/2026 450.66 450.66
            140699-F1 19.80 19.80
        We regex-scan every line of the extracted text for this pattern.
        """
        all_text = ''
        with pdfplumber.open(io.BytesIO(raw_bytes)) as pdf:
            for page in pdf.pages:
                all_text += (page.extract_text() or '') + '\n'

        def _find(pattern, text, group=1, flags=re.IGNORECASE):
            m = re.search(pattern, text, flags)
            return m.group(group).strip() if m else None

        def _parse_amount(s):
            if not s:
                return None
            try:
                return float(str(s).replace(',', '').strip())
            except Exception:
                return None

        raw_date    = _find(r'Date\s+([\d/]+)', all_text)
        amt_usd_str = _find(r'Amount Paid in USD\s+([\d,\.]+)', all_text)
        currency    = _find(r'Currency\s+([^\n]+)', all_text) or 'USD'
        reference   = _find(r'Reference\s+([\S]+)', all_text)

        payment_date = None
        if raw_date:
            try:
                payment_date = datetime.strptime(raw_date.strip(), '%m/%d/%Y').strftime('%Y-%m-%d')
            except Exception:
                payment_date = raw_date[:10] if raw_date else None

        amt_usd = _parse_amount(amt_usd_str)
        currency_lower = currency.strip().lower()
        is_usd = currency_lower in (
            'usd', 'us dollar', 'us dollars', 'united states dollar', 'united states dollars', ''
        )

        # ── Segment-based parser ─────────────────────────────────────────────
        # Strategy: find every invoice number position, then for each one
        # grab the text up to the NEXT invoice number.  Within that segment
        # the last two decimal numbers are the amount / amount-paid, and any
        # MM/DD/YYYY is the end date.  This handles rows whose hotel/account
        # text wraps across multiple lines.
        # Matches sourcing commission invoices (e.g. 237264-F1-TD, 237264-F1, 237264-F2-2)
        # and housing/registration invoices (e.g. 169500-APR 2026)
        # Fallback: any 5-9 digit booking ID followed by a hyphen and anything alphanumeric
        inv_re        = re.compile(r'\d{4,9}-(?:F\d+(?:-[A-Z0-9]+)?|[A-Z]{2,4}\s+\d{4})', re.IGNORECASE)
        inv_re_fallbk = re.compile(r'\b(\d{5,9})-([A-Z0-9][-A-Z0-9]*)\b', re.IGNORECASE)
        date_re  = re.compile(r'\b(\d{1,2}/\d{1,2}/\d{4})\b')
        amt_re   = re.compile(r'[\d,]+\.\d{2}')

        positions = [(m.start(), m.group()) for m in inv_re.finditer(all_text)]

        # If primary regex found nothing, try the broader fallback
        if not positions:
            seen_inv = set()
            for m in inv_re_fallbk.finditer(all_text):
                # Exclude hotel-chain ID numbers (preceded by a digit or inside a hotel string)
                full = m.group(0)
                if full not in seen_inv:
                    seen_inv.add(full)
                    positions.append((m.start(), full))

        booking_rows = []
        seen = set()
        for idx, (pos, invoice_raw) in enumerate(positions):
            if invoice_raw in seen:
                continue
            seen.add(invoice_raw)

            # Segment runs from this invoice number to the next (or end of text)
            next_pos = positions[idx + 1][0] if idx + 1 < len(positions) else len(all_text)
            segment  = all_text[pos:next_pos]

            # Truncate at any page-header boundary so that the repeated
            # "Amount Paid 15674.72" header on continuation pages is never
            # mistaken for the line-item amount.
            for _marker in ('Payment Voucher', 'Amount Paid', 'Amount Paid in USD'):
                _mi = segment.find(_marker)
                if _mi != -1:
                    segment = segment[:_mi]
                    break

            # Last two decimal amounts in the segment = amount, amt_paid
            amounts = amt_re.findall(segment)
            if len(amounts) < 2:
                continue
            amt_raw      = amounts[-2]
            amt_paid_raw = amounts[-1]

            # First date found in segment (if any)
            end_date = None
            dm = date_re.search(segment)
            if dm:
                try:
                    end_date = datetime.strptime(dm.group(1), '%m/%d/%Y').strftime('%Y-%m-%d')
                except Exception:
                    pass

            num_match  = re.match(r'^(\d+)', invoice_raw)
            booking_id = int(num_match.group(1)) if num_match else None

            booking_rows.append({
                'booking_id':  booking_id,
                'invoice_raw': invoice_raw,
                'bname':       '',
                'account':     '',
                'hotel':       '',
                'btype':       '',
                'end_date':    end_date,
                'amount':      _parse_amount(amt_raw),
                'amt_paid':    _parse_amount(amt_paid_raw),
            })

        return {
            'payment_date': payment_date,
            'reference':    reference,
            'currency':     currency,
            'amt_usd':      amt_usd,
            'is_usd':       is_usd,
            'booking_rows': booking_rows,
        }

    # ── Process every uploaded file ───────────────────────────────────────────
    try:
        db    = get_db()
        today = datetime.now().strftime('%Y-%m-%d')
        total_added, total_skipped = 0, 0
        all_errors, all_new_bookings, file_summaries = [], [], []

        for f in files:
            raw_bytes = f.read()
            try:
                v = _parse_voucher_pdf(raw_bytes)
            except Exception as e:
                all_errors.append(f'{f.filename}: parse error — {e}')
                file_summaries.append({'filename': f.filename, 'added': 0, 'skipped': 0, 'error': str(e)})
                continue

            if not v['booking_rows']:
                all_errors.append(f'{f.filename}: no booking table found')
                file_summaries.append({'filename': f.filename, 'added': 0, 'skipped': 0, 'error': 'No booking table found'})
                continue

            payment_date = v['payment_date']
            reference    = v['reference']
            currency     = v['currency']
            amt_usd      = v['amt_usd']
            is_usd       = v['is_usd']
            booking_rows = v['booking_rows']

            total_orig = sum(r['amt_paid'] or r['amount'] or 0 for r in booking_rows) or 1

            def row_usd(row):
                if is_usd:
                    return row['amt_paid'] or row['amount']
                row_orig = row['amt_paid'] or row['amount'] or 0
                if amt_usd and total_orig:
                    return round(amt_usd * (row_orig / total_orig), 2)
                return row['amount']

            file_added, file_skipped = 0, 0

            for row in booking_rows:
                bid = row['booking_id']
                if not bid:
                    all_errors.append(f"{f.filename}: could not parse BookingID from '{row['invoice_raw']}'")
                    continue

                usd_amount  = row_usd(row)
                invoice_num = row['invoice_raw']   # e.g. "237264-F1-TD"
                check_num   = reference            # voucher reference e.g. "00001176/106"

                # Duplicate: same BookingID + same check/reference + same invoice line,
                # OR same BookingID + same date + same amount from a DIFFERENT voucher (catch re-imports).
                # The date+amount check is scoped to OTHER vouchers only — the same voucher can
                # legitimately carry two lines for the same booking at the same amount (e.g. F1 + F1-TD).
                # Use 'Invoice: {num} |' (with delimiter) to prevent '177316-F1' matching '177316-F1-TD'.
                dup = db.execute(
                    '''SELECT 1 FROM ChkRegNote
                       WHERE BookingID = ?
                         AND (
                           (Check_ = ? AND SpecialNotes LIKE ?)
                           OR (Check_ != ? AND DateOnCheck LIKE ? AND ABS(COALESCE(FinalPayment,0) - ?) < 0.02)
                         )''',
                    (bid, check_num, f'%Invoice: {invoice_num} |%', check_num, f'{payment_date}%', usd_amount or 0)
                ).fetchone()
                if dup:
                    file_skipped += 1
                    continue

                notes_parts = []
                notes_parts.append(f'Invoice: {invoice_num}')
                if not is_usd:
                    notes_parts.append(f'Currency: {currency}')
                if not is_usd and row['amt_paid']:
                    notes_parts.append(f'Original: {row["amt_paid"]:,.2f}')
                notes_parts.append('Voucher Import')
                special_notes = ' | '.join(notes_parts)

                booking_exists = db.execute('SELECT 1 FROM ReportPipeline WHERE BookingId=?', (bid,)).fetchone()
                if not booking_exists:
                    db.execute('''INSERT INTO ReportPipeline
                        (BookingId, BookingType, AccountName, EventName, Customer, EndDate, BookingStatus)
                        VALUES (?,?,?,?,?,?,"Definite")''',
                        (bid, row['btype'] or None, row['account'] or None,
                         row['bname'] or None, row['hotel'] or None, row['end_date']))
                    all_new_bookings.append({
                        'booking_id': bid,
                        'account':    row['account'],
                        'event':      row['bname'],
                        'hotel':      row['hotel'],
                    })

                db.execute('''INSERT INTO ChkRegNote
                    (BookingID, FinalPayment, Check_, DateOnCheck, EntryDate, SpecialNotes, Cancelled, AuditFlag, Advance)
                    VALUES (?,?,?,?,?,?,0,0,0)''',
                    (bid, usd_amount, check_num, payment_date, today, special_notes))
                file_added += 1

            total_added   += file_added
            total_skipped += file_skipped
            file_summaries.append({
                'filename':     f.filename,
                'added':        file_added,
                'skipped':      file_skipped,
                'payment_date': payment_date,
                'reference':    reference,
                'currency':     currency,
                'amt_usd':      amt_usd,
                'error':        None,
            })

        db.commit()

        return render_template('import_voucher_result.html',
            added=total_added, skipped=total_skipped,
            errors=all_errors, new_bookings=all_new_bookings,
            file_summaries=file_summaries,
            file_count=len(files))

    except Exception as e:
        flash(f'Import error: {e}', 'error')
        return redirect(url_for('import_voucher'))


# ── Pickup New ────────────────────────────────────────────────────────────────

@app.route('/booking/<int:booking_id>/pickup/new', methods=['GET', 'POST'])
def pickup_new(booking_id):
    db = get_db()
    booking = db.execute('SELECT * FROM ReportPipeline WHERE BookingId = ?', (booking_id,)).fetchone()
    if not booking:
        flash('Booking not found.', 'error')
        return redirect(url_for('pipeline'))
    if request.method == 'POST':
        actual_pickup   = request.form.get('actual_pickup', '').strip() or None
        attrition       = request.form.get('attrition', '').strip() or None
        cutoff          = request.form.get('cutoff', '').strip() or None
        brand           = request.form.get('brand', '').strip() or None
        total_revenue   = request.form.get('total_revenue', '').strip() or None
        try:
            avg_rate = None
            if actual_pickup and total_revenue:
                try:
                    avg_rate = float(total_revenue) / float(actual_pickup)
                except Exception:
                    pass
            db.execute(
                'INSERT INTO Pickup (BookingID, ActualPickup, AttritionPercentage, CutOffDate, Brand, TotalRevenue) VALUES (?,?,?,?,?,?)',
                (booking_id, actual_pickup, attrition, cutoff, brand, total_revenue)
            )
            if avg_rate is not None:
                db.execute('UPDATE ReportPipeline SET RoomRate = ? WHERE BookingId = ?', (avg_rate, booking_id))
            db.commit()
            flash('Pickup record added.', 'success')
            return redirect(url_for('booking_detail', booking_id=booking_id))
        except Exception as e:
            flash(f'Error saving pickup: {e}', 'error')
    default_split  = get_commission_split()
    account_splits = get_account_splits()
    kristin_split  = get_kristin_split()
    kristin_cut    = get_kristin_cut()
    split = effective_split(booking['BookingAssociate'], booking['AccountName'], booking['Country'],
                            account_splits, default_split, kristin_split, kristin_cut)
    try:
        comm_pct = float(booking['CommissionPercent'] or 0)
    except Exception:
        comm_pct = 0
    return render_template('pickup_new.html', booking=booking,
        now=datetime.now().strftime('%Y-%m-%d'), comm_pct=comm_pct, split=split)

# ── Pickup Edit ───────────────────────────────────────────────────────────────

@app.route('/pickup/<int:pickup_id>/edit', methods=['GET', 'POST'])
def pickup_edit(pickup_id):
    db = get_db()
    pickup = db.execute('SELECT *, rowid FROM Pickup WHERE rowid = ?', (pickup_id,)).fetchone()
    if not pickup:
        flash('Pickup record not found.', 'error')
        return redirect(url_for('pipeline'))
    booking_id = pickup['BookingID']
    if request.method == 'POST':
        actual_pickup   = request.form.get('actual_pickup', '').strip() or None
        attrition       = request.form.get('attrition', '').strip() or None
        cutoff          = request.form.get('cutoff', '').strip() or None
        brand           = request.form.get('brand', '').strip() or None
        total_revenue   = request.form.get('total_revenue', '').strip() or None
        try:
            avg_rate = None
            if actual_pickup and total_revenue:
                try:
                    avg_rate = float(total_revenue) / float(actual_pickup)
                except Exception:
                    pass
            db.execute(
                'UPDATE Pickup SET ActualPickup=?, AttritionPercentage=?, CutOffDate=?, Brand=?, TotalRevenue=? WHERE rowid=?',
                (actual_pickup, attrition, cutoff, brand, total_revenue, pickup_id)
            )
            if avg_rate is not None:
                db.execute('UPDATE ReportPipeline SET RoomRate = ? WHERE BookingId = ?', (avg_rate, booking_id))
            db.commit()
            flash('Pickup updated.', 'success')
            return redirect(url_for('booking_detail', booking_id=booking_id))
        except Exception as e:
            flash(f'Error updating pickup: {e}', 'error')
    booking = db.execute('SELECT * FROM ReportPipeline WHERE BookingId = ?', (pickup['BookingID'],)).fetchone()
    default_split  = get_commission_split()
    account_splits = get_account_splits()
    kristin_split  = get_kristin_split()
    kristin_cut    = get_kristin_cut()
    split = effective_split(
        booking['BookingAssociate'] if booking else None,
        booking['AccountName'] if booking else None,
        booking['Country'] if booking else None,
        account_splits, default_split, kristin_split, kristin_cut)
    try:
        comm_pct = float((booking['CommissionPercent'] if booking else None) or 0)
    except Exception:
        comm_pct = 0
    return render_template('pickup_edit.html', pickup=pickup, comm_pct=comm_pct, split=split)

# ── Pickup Delete ─────────────────────────────────────────────────────────────

@app.route('/pickup/<int:pickup_id>/delete', methods=['POST'])
def pickup_delete(pickup_id):
    db = get_db()
    pickup = db.execute('SELECT BookingID FROM Pickup WHERE rowid = ?', (pickup_id,)).fetchone()
    if pickup:
        booking_id = pickup['BookingID']
        db.execute('DELETE FROM Pickup WHERE rowid = ?', (pickup_id,))
        db.commit()
        flash('Pickup record deleted.', 'success')
    else:
        booking_id = None
        flash('Pickup record not found.', 'error')
    if booking_id:
        return redirect(url_for('booking_detail', booking_id=booking_id))
    return redirect(url_for('pipeline'))

# ── Check New ─────────────────────────────────────────────────────────────────

@app.route('/booking/<int:booking_id>/check/new', methods=['GET', 'POST'])
def check_new(booking_id):
    db = get_db()
    booking = db.execute('SELECT * FROM ReportPipeline WHERE BookingId = ?', (booking_id,)).fetchone()
    if not booking:
        flash('Booking not found.', 'error')
        return redirect(url_for('pipeline'))
    if request.method == 'POST':
        try:
            db.execute('''INSERT INTO ChkRegNote
                (BookingID, FinalPayment, Check_, DateOnCheck, SpecialNotes,
                 Cancelled, AuditFlag, EntryDate, DepositDate)
                VALUES (?,?,?,?,?,?,?,?,?)''', (
                booking_id,
                request.form.get('final_payment', '').strip() or None,
                request.form.get('check_number', '').strip() or None,
                request.form.get('date_on_check', '').strip() or None,
                request.form.get('special_notes', '').strip() or None,
                1 if request.form.get('cancelled') else 0,
                1 if request.form.get('audit_flag') else 0,
                datetime.now().strftime('%Y-%m-%d'),
                request.form.get('deposit_date', '').strip() or None,
            ))
            db.commit()
            flash('Payment record added.', 'success')
            return redirect(url_for('booking_detail', booking_id=booking_id))
        except Exception as e:
            flash(f'Error saving payment: {e}', 'error')
    return render_template('check_new.html', booking=booking)

# ── Check Edit ────────────────────────────────────────────────────────────────

@app.route('/check/<int:check_id>/edit', methods=['GET', 'POST'])
def check_edit(check_id):
    db = get_db()
    check = db.execute('SELECT * FROM ChkRegNote WHERE ChkRegID = ?', (check_id,)).fetchone()
    if not check:
        flash('Payment record not found.', 'error')
        return redirect(url_for('pipeline'))
    booking_id = check['BookingID']
    if request.method == 'POST':
        try:
            db.execute('''UPDATE ChkRegNote
                SET FinalPayment=?, Check_=?, DateOnCheck=?, SpecialNotes=?,
                    Cancelled=?, AuditFlag=?, DepositDate=?
                WHERE ChkRegID=?''', (
                request.form.get('final_payment', '').strip() or None,
                request.form.get('check_number', '').strip() or None,
                request.form.get('date_on_check', '').strip() or None,
                request.form.get('special_notes', '').strip() or None,
                1 if request.form.get('cancelled') else 0,
                1 if request.form.get('audit_flag') else 0,
                request.form.get('deposit_date', '').strip() or None,
                check_id,
            ))
            db.commit()
            flash('Payment updated.', 'success')
            return redirect(url_for('booking_detail', booking_id=booking_id))
        except Exception as e:
            flash(f'Error updating payment: {e}', 'error')
    return render_template('check_edit.html', check=check)

@app.route('/check/<int:check_id>/delete', methods=['POST'])
def check_delete(check_id):
    db = get_db()
    check = db.execute('SELECT * FROM ChkRegNote WHERE ChkRegID = ?', (check_id,)).fetchone()
    if not check:
        flash('Payment record not found.', 'error')
        return redirect(url_for('pipeline'))
    booking_id = check['BookingID']
    db.execute('DELETE FROM ChkRegNote WHERE ChkRegID = ?', (check_id,))
    db.commit()
    flash('Payment record deleted.', 'success')
    return redirect(url_for('booking_detail', booking_id=booking_id))


# ── Commission Report ─────────────────────────────────────────────────────────

@app.route('/reports/commission')
def report_commission():
    user = get_current_user()
    who_param = request.args.get('who', 'team')
    if who_param == 'kristin' and not has_permission(user, 'reports_commission_kristin'):
        flash('You do not have access to that report.', 'error')
        return redirect(url_for('pipeline'))
    if who_param == 'team' and not has_permission(user, 'reports_commission_team'):
        flash('You do not have access to that report.', 'error')
        return redirect(url_for('pipeline'))
    date_from = request.args.get('date_from', '').strip()
    date_to   = request.args.get('date_to', '').strip()
    who       = who_param
    today     = datetime.now().strftime('%Y-%m-%d')
    rows, totals = [], {}

    if date_from and date_to:
        db = get_db()
        query = '''
            SELECT
                r.BookingId          AS booking_id,
                r.AccountName        AS account,
                r.Customer           AS hotel,
                r.BookingAssociate   AS associate,
                r.BookedDate         AS booked_date,
                r.StartDate          AS start_date,
                r.EndDate            AS end_date,
                r.RoomRate           AS room_rate,
                r.CommissionPercent  AS comm_pct,
                r.Revenue            AS revenue,
                r.USDRevenue         AS usd_revenue,
                r.TotalRoomNights    AS total_room_nights,
                r.Country            AS country,
                COALESCE(p.total_pickup, 0) AS actual_pickup,
                CASE WHEN pay.pay_count > 0 THEN 1 ELSE 0 END AS is_paid,
                COALESCE(pay.total_paid, 0) AS prepaid
            FROM ReportPipeline r
            LEFT JOIN (
                SELECT BookingID, SUM(ActualPickup) AS total_pickup
                FROM Pickup GROUP BY BookingID
            ) p ON p.BookingID = r.BookingId
            LEFT JOIN (
                SELECT BookingID, COUNT(*) AS pay_count,
                       SUM(COALESCE(FinalPayment, 0)) AS total_paid
                FROM ChkRegNote
                WHERE (Cancelled IS NULL OR Cancelled = 0)
                GROUP BY BookingID
            ) pay ON pay.BookingID = r.BookingId
            WHERE DATE(r.StartDate) BETWEEN ? AND ?
            AND (pay.pay_count IS NULL OR pay.pay_count = 0 OR DATE(r.StartDate) > ?)
            AND (r.BookingStatus IS NULL OR r.BookingStatus != 'Cancelled')
            AND NOT EXISTS (
                SELECT 1 FROM ChkRegNote c2
                WHERE c2.BookingID = r.BookingId AND c2.Cancelled = 1
            )
            AND (LOWER(COALESCE(r.BookingAssociate,'')) = 'kristin house'
                 OR LOWER(COALESCE(r.BookingType,'')) NOT IN ('other services', 'other', 'conference management', 'cm'))
            ORDER BY r.StartDate
        '''
        rows = [dict(r) for r in db.execute(query, (date_from, date_to, today)).fetchall()]

        default_split  = get_commission_split()
        account_splits = get_account_splits()
        kristin_split  = get_kristin_split()
        kristin_cut    = get_kristin_cut()
        for r in rows:
            split = effective_split(r.get('associate'), r.get('account'), r.get('country'),
                                    account_splits, default_split, kristin_split, kristin_cut)
            r['split_pct'] = split_label(r.get('associate'), split, default_split, kristin_split, kristin_cut)
            try:
                rev = float(r['usd_revenue'] or 0) or float(r['revenue'] or 0)
                r['est_commission'] = rev * float(r['comm_pct'] or 0) * split
            except Exception:
                r['est_commission'] = 0
            try:
                r['actual_commission'] = float(r['actual_pickup'] or 0) * float(r['room_rate'] or 0) * float(r['comm_pct'] or 0) * split
            except Exception:
                r['actual_commission'] = 0

        if who == 'kristin':
            rows = [r for r in rows if (r.get('associate') or '').strip().lower() == 'kristin house']
        else:
            rows = [r for r in rows if (r.get('associate') or '').strip().lower() != 'kristin house']

        totals = {
            'est_commission':    sum(float(r['est_commission'] or 0) for r in rows),
            'actual_commission': sum(r['actual_commission'] for r in rows),
            'prepaid':           sum(float(r.get('prepaid') or 0) for r in rows),
        }

    title = 'Kristin House — Missing Commission' if who == 'kristin' else 'Team — Missing Commission'
    return render_template('report_commission.html',
                           rows=rows, totals=totals, title=title, who=who,
                           date_from=date_from, date_to=date_to, today=today)

# ── Payment Report ────────────────────────────────────────────────────────────

@app.route('/reports/payments')
def report_payments():
    user = get_current_user()
    if not has_permission(user, 'reports_payments'):
        flash('You do not have access to the Payment Report.', 'error')
        return redirect(url_for('pipeline'))
    date_from = request.args.get('date_from', '').strip()
    date_to   = request.args.get('date_to', '').strip()
    today     = datetime.now().strftime('%Y-%m-%d')
    rows, totals = [], {}

    if date_from and date_to:
        db = get_db()
        query = '''
            SELECT
                r.BookingId          AS booking_id,
                r.AccountName        AS account,
                r.EventName          AS event_name,
                r.Customer           AS hotel,
                r.StartDate          AS start_date,
                r.RoomRate           AS room_rate,
                r.CommissionPercent  AS comm_pct,
                r.Revenue            AS revenue,
                r.USDRevenue         AS usd_revenue,
                r.Country            AS country,
                r.BookingAssociate   AS associate,
                c.DateOnCheck        AS date_on_check,
                c.Check_             AS check_number,
                c.FinalPayment       AS final_payment,
                c.ChkRegID           AS chk_id,
                c.Advance            AS is_advance,
                COALESCE(p.total_pickup, 0) AS actual_pickup,
                COALESCE(all_pay.pay_count, 0) AS total_pay_count,
                COALESCE(all_pay.pay_total, 0) AS total_pay_amount
            FROM ChkRegNote c
            JOIN ReportPipeline r ON r.BookingId = c.BookingID
            LEFT JOIN (
                SELECT BookingID, SUM(ActualPickup) AS total_pickup
                FROM Pickup GROUP BY BookingID
            ) p ON p.BookingID = r.BookingId
            LEFT JOIN (
                SELECT BookingID, COUNT(*) AS pay_count, SUM(COALESCE(FinalPayment,0)) AS pay_total
                FROM ChkRegNote WHERE (Cancelled IS NULL OR Cancelled = 0)
                GROUP BY BookingID
            ) all_pay ON all_pay.BookingID = c.BookingID
            WHERE DATE(c.DateOnCheck) BETWEEN ? AND ?
            AND (c.Cancelled IS NULL OR c.Cancelled = 0)
            AND (LOWER(COALESCE(r.BookingAssociate,'')) = 'kristin house'
                 OR LOWER(COALESCE(r.BookingType,'')) NOT IN ('other services', 'other', 'conference management', 'cm'))
            ORDER BY c.DateOnCheck DESC
        '''
        raw = db.execute(query, (date_from, date_to)).fetchall()

        default_split  = get_commission_split()
        account_splits = get_account_splits()
        tolerance      = get_payment_tolerance()
        kristin_split  = get_kristin_split()
        kristin_cut    = get_kristin_cut()

        for r in raw:
            r = dict(r)
            is_kristin = (r.get('associate') or '').strip().lower() == 'kristin house'
            r['is_kristin'] = is_kristin
            final = float(r['final_payment'] or 0)
            r['final_payment_f'] = final
            r['is_future'] = r.get('start_date', '') and str(r['start_date'])[:10] > today
            total_paid  = float(r.get('total_pay_amount') or 0)
            total_count = int(r.get('total_pay_count') or 0)
            prior_total = total_paid - final
            prior_count = total_count - 1
            r['prior_count']      = prior_count if prior_count > 0 else 0
            r['prior_total']      = prior_total if prior_count > 0 else 0
            r['grand_total_paid'] = total_paid
            if is_kristin:
                split = effective_split(r.get('associate'), r.get('account'), r.get('country'),
                                        account_splits, default_split, kristin_split, kristin_cut)
                r['split_pct'] = split_label(r.get('associate'), split, default_split, kristin_split, kristin_cut)
                try:
                    actual = float(r['actual_pickup'] or 0) * float(r['room_rate'] or 0) * float(r['comm_pct'] or 0) * split
                    if actual == 0:
                        rev = float(r['usd_revenue'] or 0) or float(r['revenue'] or 0)
                        actual = rev * float(r['comm_pct'] or 0) * split
                except Exception:
                    actual = 0
                r['actual_commission'] = actual
                if actual > 0 and final < actual:
                    r['out_of_tolerance'] = (actual - final) / actual > tolerance
                else:
                    r['out_of_tolerance'] = False
            else:
                r['split_pct']        = None
                r['actual_commission'] = None
                r['out_of_tolerance'] = False
            rows.append(r)

        totals = {
            'final_payment':     sum(r['final_payment_f'] for r in rows),
            'actual_commission': sum(r['actual_commission'] or 0 for r in rows),
            'future_count':      sum(1 for r in rows if r['is_future']),
            'tolerance_count':   sum(1 for r in rows if r['out_of_tolerance']),
        }

    return render_template('report_payments.html',
                           rows=rows, totals=totals,
                           date_from=date_from, date_to=date_to, today=today,
                           tolerance=get_payment_tolerance())

# ── Payment Report Export ─────────────────────────────────────────────────────

@app.route('/reports/payments/export')
def report_payments_export():
    date_from = request.args.get('date_from', '').strip()
    date_to   = request.args.get('date_to', '').strip()
    today     = datetime.now().strftime('%Y-%m-%d')
    if not date_from or not date_to:
        flash('Please run the report with a date range before exporting.', 'error')
        return redirect(url_for('report_payments'))

    db = get_db()
    query = '''
        SELECT r.BookingId AS booking_id, r.AccountName AS account, r.EventName AS event_name,
               r.Customer AS hotel, r.StartDate AS start_date, r.RoomRate AS room_rate,
               r.CommissionPercent AS comm_pct, r.Revenue AS revenue, r.USDRevenue AS usd_revenue,
               r.Country AS country, r.BookingAssociate AS associate,
               c.DateOnCheck AS date_on_check, c.Check_ AS check_number,
               c.FinalPayment AS final_payment, c.Advance AS is_advance,
               COALESCE(p.total_pickup,0) AS actual_pickup,
               COALESCE(all_pay.pay_total,0) AS total_pay_amount,
               COALESCE(all_pay.pay_count,0) AS total_pay_count
        FROM ChkRegNote c
        JOIN ReportPipeline r ON r.BookingId = c.BookingID
        LEFT JOIN (SELECT BookingID, SUM(ActualPickup) AS total_pickup FROM Pickup GROUP BY BookingID) p
            ON p.BookingID = r.BookingId
        LEFT JOIN (SELECT BookingID, COUNT(*) AS pay_count, SUM(COALESCE(FinalPayment,0)) AS pay_total
                   FROM ChkRegNote WHERE (Cancelled IS NULL OR Cancelled = 0) GROUP BY BookingID) all_pay
            ON all_pay.BookingID = c.BookingID
        WHERE DATE(c.DateOnCheck) BETWEEN ? AND ?
        AND (c.Cancelled IS NULL OR c.Cancelled = 0)
        AND (LOWER(COALESCE(r.BookingAssociate,'')) = 'kristin house'
             OR LOWER(COALESCE(r.BookingType,'')) NOT IN ('other services', 'other', 'conference management', 'cm'))
        ORDER BY c.DateOnCheck DESC
    '''
    raw = db.execute(query, (date_from, date_to)).fetchall()

    default_split  = get_commission_split()
    account_splits = get_account_splits()
    kristin_split  = get_kristin_split()
    kristin_cut    = get_kristin_cut()
    tolerance      = get_payment_tolerance()

    by_check_rows = []   # Tab 1 – all payments sorted by check #
    kristin_rows  = []   # Tab 2 – Kristin House only
    team_rows     = []   # Tab 3 – everyone else

    def _us(iso):
        try:
            return datetime.strptime(str(iso)[:10], '%Y-%m-%d').strftime('%m/%d/%Y')
        except Exception:
            return str(iso or '')[:10]

    for r in raw:
        r = dict(r)
        is_kristin = (r.get('associate') or '').strip().lower() == 'kristin house'
        final      = float(r['final_payment'] or 0)
        date_str   = _us(r.get('date_on_check')) if r.get('date_on_check') else ''

        by_check_rows.append({
            'Check #':        r['check_number'],
            'Date on Check':  date_str,
            'Associate':      r['associate'] or '',
            'Booking ID':     r['booking_id'],
            'Account':        r['account'] or '',
            'Event Name':     r['event_name'] or '',
            'Hotel':          r['hotel'] or '',
            'Incentive':      'Yes' if r['is_advance'] else '',
            'Final Payment':  final,
        })

        if is_kristin:
            split  = effective_split(r.get('associate'), r.get('account'), r.get('country'),
                                     account_splits, default_split, kristin_split, kristin_cut)
            pickup = float(r['actual_pickup'] or 0)
            try:
                actual = pickup * float(r['room_rate'] or 0) * float(r['comm_pct'] or 0) * split
                if actual == 0:
                    rev = float(r['usd_revenue'] or 0) or float(r['revenue'] or 0)
                    actual = rev * float(r['comm_pct'] or 0) * split
            except Exception:
                actual = 0
            total_paid  = float(r.get('total_pay_amount') or 0)
            prior_count = int(r.get('total_pay_count') or 0) - 1
            prior_total = total_paid - final if prior_count > 0 else 0
            kristin_rows.append({
                'Date on Check':     date_str,
                'Check #':           r['check_number'],
                'Booking ID':        r['booking_id'],
                'Account':           r['account'] or '',
                'Event Name':        r['event_name'] or '',
                'Hotel':             r['hotel'] or '',
                'Start Date':        _us(r['start_date']) if r['start_date'] else '',
                'Split %':           split_label(r.get('associate'), split, default_split, kristin_split, kristin_cut),
                'Final Payment':     final,
                'Actual Commission': round(actual, 2),
                'Difference':        round(final - actual, 2),
                'Prior Payments':    round(prior_total, 2) if prior_count > 0 else 0,
                'Grand Total Paid':  round(total_paid, 2),
                'Diff %':            round((total_paid - actual) / actual, 4) if actual else None,
                'Incentive':         'Yes' if r['is_advance'] else '',
            })
        else:
            team_rows.append({
                'Date on Check':  date_str,
                'Check #':        r['check_number'],
                'Associate':      r['associate'] or '',
                'Booking ID':     r['booking_id'],
                'Account':        r['account'] or '',
                'Event Name':     r['event_name'] or '',
                'Hotel':          r['hotel'] or '',
                'Start Date':     _us(r['start_date']) if r['start_date'] else '',
                'Incentive':      'Yes' if r['is_advance'] else '',
                'Final Payment':  final,
            })

    # Sort Tab 1 by check number then date; Tabs 2 & 3 by date paid
    by_check_rows.sort(key=lambda x: (str(x['Check #'] or ''), x['Date on Check']))
    kristin_rows.sort(key=lambda x: x['Date on Check'])
    team_rows.sort(key=lambda x: x['Date on Check'])

    from itertools import groupby as _groupby
    from openpyxl.styles import PatternFill, Font, Alignment

    HEADER_FILL = PatternFill('solid', start_color='1A3A5C')
    HEADER_FONT = Font(bold=True, color='FFFFFF')
    SUBTOT_FILL = PatternFill('solid', start_color='D9E1F2')
    GRAND_FILL  = PatternFill('solid', start_color='1A3A5C')
    GRAND_FONT  = Font(bold=True, color='FFFFFF')
    BOLD        = Font(bold=True)
    RIGHT       = Alignment(horizontal='right')
    CENTER      = Alignment(horizontal='center', wrap_text=True)

    def style_sheet(ws, currency_cols, col_widths):
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = w
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                if cell.column_letter in currency_cols:
                    cell.number_format = '$#,##0.00'
        # Grand totals row
        last = ws.max_row
        tr = last + 1
        first_curr_idx = min(ord(c) - ord('A') + 1 for c in currency_cols)
        lbl = ws.cell(tr, first_curr_idx - 1)
        lbl.value = 'Grand Total:'
        lbl.font = GRAND_FONT
        lbl.fill = GRAND_FILL
        lbl.alignment = RIGHT
        for cl in currency_cols:
            ci = ord(cl) - ord('A') + 1
            c = ws.cell(tr, ci)
            c.value = f'=SUM({cl}2:{cl}{last})'
            c.number_format = '$#,##0.00'
            c.font = GRAND_FONT
            c.fill = GRAND_FILL

    ORANGE_FILL = PatternFill('solid', start_color='FFD580')

    def write_grouped_sheet(ws, rows, group_key, label_prefix):
        """Write a sheet grouped by group_key with subtotals and 2 blank rows between groups."""
        headers    = ['Date on Check', 'Check #', 'Associate', 'Booking ID',
                      'Account', 'Event Name', 'Hotel', 'Start Date', 'Incentive', 'Final Payment']
        col_widths = [15, 14, 28, 14, 28, 35, 30, 13, 10, 15]
        pay_col    = 10

        for ci, h in enumerate(headers, 1):
            cell = ws.cell(1, ci, h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = w
        ws.freeze_panes = 'A2'

        cur_row     = 2
        grand_total = 0.0

        for grp_val, grp in _groupby(rows, key=lambda x: x[group_key]):
            grp_rows  = list(grp)
            grp_total = 0.0

            for r in grp_rows:
                is_incentive = r['Incentive'] == 'Yes'
                ws.cell(cur_row, 1, r['Date on Check'])
                ws.cell(cur_row, 2, r['Check #'])
                ws.cell(cur_row, 3, r['Associate'])
                ws.cell(cur_row, 4, r['Booking ID'])
                ws.cell(cur_row, 5, r['Account'])
                ws.cell(cur_row, 6, r['Event Name'])
                ws.cell(cur_row, 7, r['Hotel'])
                ws.cell(cur_row, 8, r['Start Date'])
                ws.cell(cur_row, 9, r['Incentive'])
                pay = ws.cell(cur_row, pay_col, r['Final Payment'])
                pay.number_format = '$#,##0.00'
                if is_incentive:
                    for ci in range(1, pay_col + 1):
                        ws.cell(cur_row, ci).fill = ORANGE_FILL
                grp_total += r['Final Payment']
                cur_row += 1

            lbl = ws.cell(cur_row, pay_col - 1, f'{label_prefix} {grp_val} Total:')
            lbl.font = BOLD
            lbl.fill = SUBTOT_FILL
            lbl.alignment = RIGHT
            tot = ws.cell(cur_row, pay_col, grp_total)
            tot.number_format = '$#,##0.00'
            tot.font = BOLD
            tot.fill = SUBTOT_FILL
            grand_total += grp_total
            cur_row += 3

        lbl = ws.cell(cur_row, pay_col - 1, 'Grand Total:')
        lbl.font = GRAND_FONT
        lbl.fill = GRAND_FILL
        lbl.alignment = RIGHT
        tot = ws.cell(cur_row, pay_col, grand_total)
        tot.number_format = '$#,##0.00'
        tot.font = GRAND_FONT
        tot.fill = GRAND_FILL

    def write_by_check_sheet(ws, rows):
        """Write Tab 1 with per-check subtotals, 2 blank rows between groups, frozen header."""
        headers    = ['Check #', 'Date on Check', 'Associate', 'Booking ID',
                      'Account', 'Event Name', 'Hotel', 'Incentive', 'Final Payment']
        col_widths = [14, 15, 28, 14, 28, 35, 30, 10, 15]

        for ci, h in enumerate(headers, 1):
            cell = ws.cell(1, ci, h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = w
        ws.freeze_panes = 'A2'

        cur_row     = 2
        grand_total = 0.0

        for check_num, grp in _groupby(rows, key=lambda x: x['Check #']):
            grp_rows    = list(grp)
            check_total = 0.0

            for r in grp_rows:
                is_incentive = r['Incentive'] == 'Yes'
                ws.cell(cur_row, 1, r['Check #'])
                ws.cell(cur_row, 2, r['Date on Check'])
                ws.cell(cur_row, 3, r['Associate'])
                ws.cell(cur_row, 4, r['Booking ID'])
                ws.cell(cur_row, 5, r['Account'])
                ws.cell(cur_row, 6, r['Event Name'])
                ws.cell(cur_row, 7, r['Hotel'])
                ws.cell(cur_row, 8, r['Incentive'])
                pay = ws.cell(cur_row, 9, r['Final Payment'])
                pay.number_format = '$#,##0.00'
                if is_incentive:
                    for ci in range(1, 10):
                        ws.cell(cur_row, ci).fill = ORANGE_FILL
                check_total += r['Final Payment']
                cur_row += 1

            lbl = ws.cell(cur_row, 8, f'Check {check_num} Total:')
            lbl.font = BOLD
            lbl.fill = SUBTOT_FILL
            lbl.alignment = RIGHT
            tot = ws.cell(cur_row, 9, check_total)
            tot.number_format = '$#,##0.00'
            tot.font = BOLD
            tot.fill = SUBTOT_FILL
            grand_total += check_total
            cur_row += 3

        lbl = ws.cell(cur_row, 8, 'Grand Total:')
        lbl.font = GRAND_FONT
        lbl.fill = GRAND_FILL
        lbl.alignment = RIGHT
        tot = ws.cell(cur_row, 9, grand_total)
        tot.number_format = '$#,##0.00'
        tot.font = GRAND_FONT
        tot.fill = GRAND_FILL

    # Tab 4 sort: by associate then date
    team_by_assoc = sorted(team_rows, key=lambda x: (x['Associate'], x['Date on Check']))

    # ── Summary data ──────────────────────────────────────────────────────────
    # Totals by check number (from all rows)
    check_totals = {}
    for r in by_check_rows:
        k = str(r['Check #'] or '')
        check_totals[k] = check_totals.get(k, 0.0) + r['Final Payment']
    check_totals = sorted(check_totals.items(), key=lambda x: x[0])

    # Totals by associate (Kristin + team)
    assoc_totals = {}
    for r in kristin_rows:
        assoc_totals['Kristin House'] = assoc_totals.get('Kristin House', 0.0) + r['Final Payment']
    for r in team_rows:
        a = r['Associate'] or 'Unknown'
        assoc_totals[a] = assoc_totals.get(a, 0.0) + r['Final Payment']
    assoc_totals = sorted(assoc_totals.items(), key=lambda x: x[0])

    def write_summary_sheet(ws):
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        TITLE_FONT   = Font(bold=True, size=14, color='1A3A5C')
        SECTION_FONT = Font(bold=True, size=11, color='FFFFFF')
        SECTION_FILL = PatternFill('solid', start_color='1A3A5C')
        SUBTOT_FONT  = Font(bold=True)
        CURRENCY_FMT = '$#,##0.00'
        CENTER       = Alignment(horizontal='center')
        RIGHT        = Alignment(horizontal='right')

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 18

        # Title
        ws.merge_cells('A1:B1')
        title = ws.cell(1, 1, 'Conference Planning Associates Inc.')
        title.font = TITLE_FONT
        title.alignment = CENTER

        ws.merge_cells('A2:B2')
        sub = ws.cell(2, 1, 'Check Register Report')
        sub.font = Font(bold=True, size=12, color='1A3A5C')
        sub.alignment = CENTER

        ws.merge_cells('A3:B3')
        dr = ws.cell(3, 1, f'{date_from}  —  {date_to}')
        dr.font = Font(size=10, color='555555')
        dr.alignment = CENTER

        cur = 5   # start data after blank row 4

        # ── By Check Number ────────────────────────────────────────────────
        ws.merge_cells(f'A{cur}:B{cur}')
        hdr = ws.cell(cur, 1, 'Total by Check Number')
        hdr.font = SECTION_FONT
        hdr.fill = SECTION_FILL
        hdr.alignment = CENTER
        cur += 1

        col_hdr_fill = PatternFill('solid', start_color='D9E1F2')
        for col, label in enumerate(['Check #', 'Total'], 1):
            c = ws.cell(cur, col, label)
            c.font = Font(bold=True)
            c.fill = col_hdr_fill
            c.alignment = CENTER
        cur += 1

        check_grand = 0.0
        for check_num, total in check_totals:
            ws.cell(cur, 1, check_num)
            amt = ws.cell(cur, 2, total)
            amt.number_format = CURRENCY_FMT
            check_grand += total
            cur += 1

        ws.cell(cur, 1, 'Grand Total:').font = SUBTOT_FONT
        ws.cell(cur, 1).alignment = RIGHT
        gtc = ws.cell(cur, 2, check_grand)
        gtc.number_format = CURRENCY_FMT
        gtc.font = SUBTOT_FONT
        gtc.fill = SECTION_FILL
        gtc.font = Font(bold=True, color='FFFFFF')
        ws.cell(cur, 1).fill = SECTION_FILL
        ws.cell(cur, 1).font = Font(bold=True, color='FFFFFF')
        cur += 2   # blank separator

        # ── By Associate ───────────────────────────────────────────────────
        ws.merge_cells(f'A{cur}:B{cur}')
        hdr = ws.cell(cur, 1, 'Total by Associate')
        hdr.font = SECTION_FONT
        hdr.fill = SECTION_FILL
        hdr.alignment = CENTER
        cur += 1

        for col, label in enumerate(['Associate', 'Total'], 1):
            c = ws.cell(cur, col, label)
            c.font = Font(bold=True)
            c.fill = col_hdr_fill
            c.alignment = CENTER
        cur += 1

        assoc_grand = 0.0
        for assoc, total in assoc_totals:
            ws.cell(cur, 1, assoc)
            amt = ws.cell(cur, 2, total)
            amt.number_format = CURRENCY_FMT
            assoc_grand += total
            cur += 1

        ws.cell(cur, 1, 'Grand Total:').fill = SECTION_FILL
        ws.cell(cur, 1).font = Font(bold=True, color='FFFFFF')
        ws.cell(cur, 1).alignment = RIGHT
        gta = ws.cell(cur, 2, assoc_grand)
        gta.number_format = CURRENCY_FMT
        gta.fill = SECTION_FILL
        gta.font = Font(bold=True, color='FFFFFF')

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Placeholders to control sheet order; Tabs 1 & 4 written manually
        pd.DataFrame(columns=['Summary']).to_excel(writer, index=False, sheet_name='Summary')
        pd.DataFrame(columns=['Check #','Date on Check','Associate','Booking ID',
                               'Account','Event Name','Hotel','Incentive','Final Payment']
                     ).to_excel(writer, index=False, sheet_name='By Check')
        pd.DataFrame(kristin_rows).to_excel(writer, index=False, sheet_name='Kristin House')
        pd.DataFrame(team_rows).to_excel(writer, index=False, sheet_name='Team')
        pd.DataFrame(columns=['Date on Check','Check #','Associate','Booking ID',
                               'Account','Event Name','Hotel','Start Date','Incentive','Final Payment']
                     ).to_excel(writer, index=False, sheet_name='Total by Associate')

        # ── Summary ───────────────────────────────────────────────────────────
        write_summary_sheet(writer.sheets['Summary'])

        # ── Tab 1: By Check ───────────────────────────────────────────────────
        write_by_check_sheet(writer.sheets['By Check'], by_check_rows)

        # ── Tab 2: Kristin House ──────────────────────────────────────────────
        # Cols A-N: Date Check# BookingID Account Event Hotel Start Split% Payment
        #           ActComm Diff Prior Grand Incentive
        # Cols: A=Date B=Check# C=BookingID D=Account E=Event F=Hotel G=Start H=Split%
        #       I=Payment J=ActComm K=Diff L=Prior M=Grand N=Diff% O=Incentive
        style_sheet(writer.sheets['Kristin House'],
                    ('I', 'J', 'K', 'L', 'M'),
                    [15, 14, 14, 28, 35, 30, 13, 10, 15, 17, 12, 14, 16, 10, 10])
        ws_k = writer.sheets['Kristin House']
        ws_k.freeze_panes = 'A2'
        # Format Diff % column (N) as percentage
        for row in ws_k.iter_rows(min_row=2, max_row=ws_k.max_row):
            if row[13].value is not None:
                row[13].number_format = '+0.0%;-0.0%;0.0%'
        red_fill    = PatternFill('solid', start_color='F8D7DA')
        yellow_fill = PatternFill('solid', start_color='FFF3CD')
        last_data   = ws_k.max_row - 1
        for row in ws_k.iter_rows(min_row=2, max_row=last_data):
            incentive   = row[14].value == 'Yes'   # col O
            fin         = row[8].value              # col I
            act         = row[9].value              # col J
            grand       = row[12].value or 0        # col M – Grand Total Paid
            start_date  = str(row[6].value or '')[:10]
            is_future   = start_date > today
            try:
                covered   = act and act > 0 and grand >= act * (1 - tolerance)
                underpaid = (act and act > 0 and fin is not None and fin < act
                             and (act - fin) / act > tolerance
                             and not covered)
            except Exception:
                underpaid = False
            if incentive:
                fill = ORANGE_FILL
            elif underpaid:
                fill = red_fill
            elif is_future:
                fill = yellow_fill
            else:
                fill = None
            if fill:
                for cell in row:
                    cell.fill = fill

        # ── Tab 3: Team (sorted by date) ──────────────────────────────────────
        # Cols A-J: Date Check# Associate BookingID Account Event Hotel Start Incentive Payment
        style_sheet(writer.sheets['Team'],
                    ('J',),
                    [15, 14, 28, 14, 28, 35, 30, 13, 10, 15])
        ws_t = writer.sheets['Team']
        ws_t.freeze_panes = 'A2'
        last_team = ws_t.max_row - 1
        for row in ws_t.iter_rows(min_row=2, max_row=last_team):
            if row[8].value == 'Yes':   # col I – Incentive
                for cell in row:
                    cell.fill = ORANGE_FILL

        # ── Tab 4: Total by Associate ──────────────────────────────────────────
        write_grouped_sheet(writer.sheets['Total by Associate'],
                            team_by_assoc, 'Associate', 'Associate')

    output.seek(0)
    filename = f'Payment_Report_{date_from}_to_{date_to}.xlsx'
    return send_file(output, download_name=filename, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ── Booking Report (admin only) ───────────────────────────────────────────────
# Modeled on the Payment Report, but reports on the booking date range with
# booking amounts and commission amounts (instead of check dates and payments).

def _booking_report_rows(date_from, date_to):
    """Fetch non-cancelled bookings whose BookedDate (M/D/YYYY or ISO text) falls
    within [date_from, date_to]. Returns dict rows with normalized USD amounts.
    Filtering is done in Python because BookedDate is stored as free-form text."""
    db = get_db()
    raw = db.execute('''
        SELECT BookingId         AS booking_id,
               BookedDate        AS booked_date,
               StartDate         AS start_date,
               AccountName       AS account,
               EventName         AS event_name,
               Customer          AS hotel,
               BookingAssociate  AS associate,
               BookingType       AS booking_type,
               BookingStatus     AS status,
               Currency          AS currency,
               CommissionPercent AS comm_pct,
               Revenue           AS revenue,
               USDRevenue        AS usd_revenue,
               USDCommissionableAmount AS comm_amount
        FROM ReportPipeline
        WHERE LOWER(COALESCE(BookingStatus,'')) != 'cancelled'
          AND BookedDate IS NOT NULL AND TRIM(BookedDate) NOT IN ('', 'nan')
    ''').fetchall()

    rows = []
    for r in raw:
        r = dict(r)
        iso = _to_iso(r['booked_date'])
        if not iso or iso < date_from or iso > date_to:
            continue
        r['booked_iso'] = iso
        amount = float(r['usd_revenue'] or 0) or float(r['revenue'] or 0)
        commission = float(r['comm_amount'] or 0)
        if not commission and amount and r['comm_pct']:
            commission = amount * float(r['comm_pct'] or 0)
        r['amount']     = amount
        r['commission'] = commission
        rows.append(r)

    rows.sort(key=lambda x: x['booked_iso'], reverse=True)
    return rows


@app.route('/reports/bookings')
def report_bookings():
    user = get_current_user()
    if not user or user['role'] != 'admin':
        flash('This report is restricted to administrators.', 'error')
        return redirect(url_for('pipeline'))

    date_from = request.args.get('date_from', '').strip()
    date_to   = request.args.get('date_to', '').strip()
    today     = datetime.now().strftime('%Y-%m-%d')
    rows, totals = [], {}

    if date_from and date_to:
        rows = _booking_report_rows(date_from, date_to)
        totals = {
            'count':      len(rows),
            'amount':     sum(r['amount'] for r in rows),
            'commission': sum(r['commission'] for r in rows),
        }

    return render_template('report_bookings.html',
                           rows=rows, totals=totals,
                           date_from=date_from, date_to=date_to, today=today)


@app.route('/reports/bookings/export')
def report_bookings_export():
    user = get_current_user()
    if not user or user['role'] != 'admin':
        flash('This report is restricted to administrators.', 'error')
        return redirect(url_for('pipeline'))

    date_from = request.args.get('date_from', '').strip()
    date_to   = request.args.get('date_to', '').strip()
    if not date_from or not date_to:
        flash('Please run the report with a date range before exporting.', 'error')
        return redirect(url_for('report_bookings'))

    rows = _booking_report_rows(date_from, date_to)

    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment

    HEADER_FILL = PatternFill('solid', start_color='1A3A5C')
    HEADER_FONT = Font(bold=True, color='FFFFFF')
    SUBTOT_FILL = PatternFill('solid', start_color='D9E1F2')
    GRAND_FILL  = PatternFill('solid', start_color='1A3A5C')
    GRAND_FONT  = Font(bold=True, color='FFFFFF')
    BOLD        = Font(bold=True)
    CENTER      = Alignment(horizontal='center', wrap_text=True)
    RIGHT       = Alignment(horizontal='right')

    def _us(iso):
        try:
            return datetime.strptime(str(iso)[:10], '%Y-%m-%d').strftime('%m/%d/%Y')
        except Exception:
            return str(iso or '')[:10]

    import re as _re
    def _safe_sheet_name(name, used):
        """Excel sheet names: max 31 chars, no []:*?/\\, must be unique."""
        s = _re.sub(r'[\[\]:\*\?/\\]', ' ', str(name)).strip()[:31] or 'Unknown'
        base, i = s, 2
        while s.lower() in used:
            suffix = f' ({i})'
            s = base[:31 - len(suffix)] + suffix
            i += 1
        used.add(s.lower())
        return s

    def _write_detail_sheet(ws, det_rows, include_associate=True):
        """Write a booking-detail sheet (optionally with the Associate column)."""
        if include_associate:
            headers    = ['Booked Date', 'Booking ID', 'Associate', 'Account', 'Event Name',
                          'Hotel', 'Start Date', 'Booking Amount', 'Comm %', 'Commission']
            col_widths = [13, 12, 26, 28, 35, 30, 13, 16, 9, 16]
        else:
            headers    = ['Booked Date', 'Booking ID', 'Account', 'Event Name',
                          'Hotel', 'Start Date', 'Booking Amount', 'Comm %', 'Commission']
            col_widths = [13, 12, 28, 35, 30, 13, 16, 9, 16]
        ncols   = len(headers)
        amt_col = ncols - 2   # Booking Amount
        pct_col = ncols - 1   # Comm %
        com_col = ncols       # Commission

        for ci, h in enumerate(headers, 1):
            c = ws.cell(1, ci, h)
            c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = w
        ws.freeze_panes = 'A2'

        rownum = 2
        for r in det_rows:
            vals = [_us(r['booked_iso']), r['booking_id']]
            if include_associate:
                vals.append(r['associate'] or '')
            vals += [r['account'] or '', r['event_name'] or '', r['hotel'] or '',
                     _us(_to_iso(r['start_date'])) if r['start_date'] else '']
            for ci, v in enumerate(vals, 1):
                ws.cell(rownum, ci, v)
            a = ws.cell(rownum, amt_col, round(r['amount'], 2));      a.number_format = '$#,##0.00'
            p = ws.cell(rownum, pct_col, float(r['comm_pct'] or 0));  p.number_format = '0.0%'
            cm = ws.cell(rownum, com_col, round(r['commission'], 2)); cm.number_format = '$#,##0.00'
            rownum += 1

        from openpyxl.utils import get_column_letter as _gcl
        amt_l, com_l = _gcl(amt_col), _gcl(com_col)
        lbl = ws.cell(rownum, amt_col - 1, 'Grand Total:')
        lbl.font = GRAND_FONT; lbl.fill = GRAND_FILL; lbl.alignment = RIGHT
        ta = ws.cell(rownum, amt_col, f'=SUM({amt_l}2:{amt_l}{rownum-1})')
        ta.number_format = '$#,##0.00'; ta.font = GRAND_FONT; ta.fill = GRAND_FILL
        ws.cell(rownum, pct_col).fill = GRAND_FILL
        tc = ws.cell(rownum, com_col, f'=SUM({com_l}2:{com_l}{rownum-1})')
        tc.number_format = '$#,##0.00'; tc.font = GRAND_FONT; tc.fill = GRAND_FILL

    wb = Workbook()

    # ── Tab 1: Summary by Associate ───────────────────────────────────────────
    ws2 = wb.active
    ws2.title = 'Summary by Associate'
    by_assoc = {}
    for r in rows:
        a = (r['associate'] or 'Unknown').strip() or 'Unknown'
        d = by_assoc.setdefault(a, {'count': 0, 'amount': 0.0, 'commission': 0.0})
        d['count'] += 1
        d['amount'] += r['amount']
        d['commission'] += r['commission']

    sheaders = ['Associate', 'Bookings', 'Booking Amount', 'Commission', "Kristin's Team Cut"]
    for ci, h in enumerate(sheaders, 1):
        c = ws2.cell(1, ci, h)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = CENTER
    for i, w in enumerate([28, 12, 18, 18, 20], 1):
        ws2.column_dimensions[ws2.cell(1, i).column_letter].width = w
    ws2.freeze_panes = 'A2'

    rn = 2
    for a in sorted(by_assoc):
        d = by_assoc[a]
        ws2.cell(rn, 1, a)
        ws2.cell(rn, 2, d['count'])
        am = ws2.cell(rn, 3, round(d['amount'], 2));     am.number_format = '$#,##0.00'
        cc = ws2.cell(rn, 4, round(d['commission'], 2)); cc.number_format = '$#,##0.00'
        kc = ws2.cell(rn, 5, f'=D{rn}*0.1');             kc.number_format = '$#,##0.00'
        rn += 1
    glbl = ws2.cell(rn, 1, 'Grand Total:'); glbl.font = GRAND_FONT; glbl.fill = GRAND_FILL
    gcnt = ws2.cell(rn, 2, sum(d['count'] for d in by_assoc.values())); gcnt.font = GRAND_FONT; gcnt.fill = GRAND_FILL
    ga = ws2.cell(rn, 3, f'=SUM(C2:C{rn-1})'); ga.number_format = '$#,##0.00'; ga.font = GRAND_FONT; ga.fill = GRAND_FILL
    gc = ws2.cell(rn, 4, f'=SUM(D2:D{rn-1})'); gc.number_format = '$#,##0.00'; gc.font = GRAND_FONT; gc.fill = GRAND_FILL
    gk = ws2.cell(rn, 5, f'=SUM(E2:E{rn-1})'); gk.number_format = '$#,##0.00'; gk.font = GRAND_FONT; gk.fill = GRAND_FILL

    # ── Tab 2: All Bookings ───────────────────────────────────────────────────
    _write_detail_sheet(wb.create_sheet('All Bookings'), rows, include_associate=True)

    # ── Per-person detail tabs (one per associate) ────────────────────────────
    used_names = {'summary by associate', 'all bookings'}
    for a in sorted(by_assoc):
        person_rows = [r for r in rows if ((r['associate'] or 'Unknown').strip() or 'Unknown') == a]
        ws_p = wb.create_sheet(_safe_sheet_name(a, used_names))
        _write_detail_sheet(ws_p, person_rows, include_associate=False)

    output = _io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'Booking_Report_{date_from}_to_{date_to}.xlsx'
    return send_file(output, download_name=filename, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ── Commission Report Export ──────────────────────────────────────────────────

@app.route('/reports/commission/export')
def report_commission_export():
    date_from = request.args.get('date_from', '').strip()
    date_to   = request.args.get('date_to', '').strip()
    who       = request.args.get('who', 'team')
    today     = datetime.now().strftime('%Y-%m-%d')
    if not date_from or not date_to:
        flash('Please run the report with a date range before exporting.', 'error')
        return redirect(url_for('report_commission'))

    db = get_db()
    query = '''
        SELECT r.BookingId AS booking_id, r.AccountName AS account, r.Customer AS hotel,
               r.BookedDate AS booked_date, r.StartDate AS start_date, r.EndDate AS end_date,
               r.RoomRate AS room_rate, r.TotalRoomNights AS total_room_nights,
               r.CommissionPercent AS comm_pct, r.Revenue AS revenue, r.USDRevenue AS usd_revenue,
               r.Country AS country, r.BookingAssociate AS associate,
               COALESCE(p.total_pickup,0) AS actual_pickup,
               CASE WHEN pay.pay_count > 0 THEN 'Paid / Future' ELSE 'Unpaid' END AS payment_status
        FROM ReportPipeline r
        LEFT JOIN (SELECT BookingID, SUM(ActualPickup) AS total_pickup FROM Pickup GROUP BY BookingID) p
            ON p.BookingID = r.BookingId
        LEFT JOIN (SELECT BookingID, COUNT(*) AS pay_count FROM ChkRegNote
                   WHERE (Cancelled IS NULL OR Cancelled=0) GROUP BY BookingID) pay
            ON pay.BookingID = r.BookingId
        WHERE DATE(r.StartDate) BETWEEN ? AND ?
        AND (pay.pay_count IS NULL OR pay.pay_count=0 OR DATE(r.StartDate) > ?)
        AND (r.BookingStatus IS NULL OR r.BookingStatus != 'Cancelled')
        AND NOT EXISTS (SELECT 1 FROM ChkRegNote c2 WHERE c2.BookingID = r.BookingId AND c2.Cancelled=1)
        AND (LOWER(COALESCE(r.BookingAssociate,'')) = 'kristin house'
             OR LOWER(COALESCE(r.BookingType,'')) NOT IN ('other services', 'other', 'conference management', 'cm'))
        ORDER BY r.StartDate
    '''
    all_rows = db.execute(query, (date_from, date_to, today)).fetchall()
    if who == 'kristin':
        rows = [r for r in all_rows if (r['associate'] or '').strip().lower() == 'kristin house']
    else:
        rows = [r for r in all_rows if (r['associate'] or '').strip().lower() != 'kristin house']

    default_split  = get_commission_split()
    account_splits = get_account_splits()
    kristin_split  = get_kristin_split()
    kristin_cut    = get_kristin_cut()
    export_rows     = []
    missing_pickups = set()   # 0-based data row indices with no pickup
    for r in rows:
        split = effective_split(r['associate'], r['account'], r['country'],
                                account_splits, default_split, kristin_split, kristin_cut)
        try:
            rev = float(r['usd_revenue'] or 0) or float(r['revenue'] or 0)
            est = rev * float(r['comm_pct'] or 0) * split
        except Exception:
            est = 0
        pickup = float(r['actual_pickup'] or 0)
        try:
            actual = pickup * float(r['room_rate'] or 0) * float(r['comm_pct'] or 0) * split
        except Exception:
            actual = 0
        if pickup == 0:
            missing_pickups.add(len(export_rows))
        def _us2(iso):
            try:
                return datetime.strptime(str(iso)[:10], '%Y-%m-%d').strftime('%m/%d/%Y')
            except Exception:
                return str(iso or '')
        export_rows.append({
            'Booking ID':        r['booking_id'],
            'Account':           r['account'],
            'Hotel':             r['hotel'],
            'Booked Date':       _us2(r['booked_date']) if r['booked_date'] else '',
            'Start Date':        _us2(r['start_date']) if r['start_date'] else '',
            'End Date':          _us2(r['end_date']) if r['end_date'] else '',
            'Room Rate':         r['room_rate'],
            'Contracted Nights': int(r['total_room_nights']) if r['total_room_nights'] else 0,
            'Pickup':            int(r['actual_pickup']) if r['actual_pickup'] else 0,
            'Est. Commission':   round(est, 2),
            'Actual Commission': round(actual, 2),
            'Status':            r['payment_status'],
        })

    df = pd.DataFrame(export_rows)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Missing Commission')
        ws = writer.sheets['Missing Commission']
        ws.freeze_panes = 'A2'
        col_widths = [14,30,35,14,14,14,12,14,10,16,18,14]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(1,i).column_letter].width = width
        from openpyxl.styles import PatternFill, Font, Alignment
        header_fill = PatternFill('solid', start_color='1A3A5C')
        header_font = Font(bold=True, color='FFFFFF')
        red_fill    = PatternFill('solid', start_color='F8D7DA')
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row)):
            for cell in row:
                if cell.column_letter in ('G','J','K'):
                    cell.number_format = '$#,##0.00'
                if cell.column_letter in ('H','I'):
                    cell.number_format = '#,##0'
            if idx in missing_pickups:
                for cell in row:
                    cell.fill = red_fill
        total_row = ws.max_row + 1
        last_data  = total_row - 1
        ws.cell(total_row, 9).value = 'Totals:'
        ws.cell(total_row, 9).font = Font(bold=True)
        ws.cell(total_row, 9).alignment = Alignment(horizontal='right')
        ws.cell(total_row,10).value = f'=SUM(J2:J{last_data})'
        ws.cell(total_row,10).number_format = '$#,##0.00'
        ws.cell(total_row,10).font = Font(bold=True)
        ws.cell(total_row,11).value = f'=SUM(K2:K{last_data})'
        ws.cell(total_row,11).number_format = '$#,##0.00'
        ws.cell(total_row,11).font = Font(bold=True)

    output.seek(0)
    who_label = 'Kristin_House' if who == 'kristin' else 'Team'
    filename = f'Missing_Commission_{who_label}_{date_from}_to_{date_to}.xlsx'
    return send_file(output, download_name=filename, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ── Days Outstanding Report (admin only) ─────────────────────────────────────

@app.route('/reports/days-outstanding')
def report_days_outstanding():
    user = get_current_user()
    if not user or user['role'] != 'admin':
        flash('This report is restricted to administrators.', 'error')
        return redirect(url_for('pipeline'))

    from datetime import date as _date
    default_from = str(_date.today().replace(year=_date.today().year - 3))
    default_to   = str(_date.today())
    date_from = request.args.get('date_from', default_from).strip() or default_from
    date_to   = request.args.get('date_to',   default_to).strip()   or default_to

    db = get_db()

    def _run(group_col, group_label, who_filter):
        sql = f'''
            SELECT
                COALESCE(NULLIF(r.{group_col},''), 'Unknown / Blank') AS grp,
                COUNT(*)                                               AS num_payments,
                ROUND(AVG(julianday(c.DateOnCheck) - julianday(r.EndDate)), 1) AS avg_days,
                ROUND(MIN(julianday(c.DateOnCheck) - julianday(r.EndDate)), 0) AS min_days,
                ROUND(MAX(julianday(c.DateOnCheck) - julianday(r.EndDate)), 0) AS max_days,
                ROUND(SUM(c.FinalPayment), 2)                          AS total_commission
            FROM ChkRegNote c
            JOIN ReportPipeline r ON c.BookingID = r.BookingId
            WHERE c.Cancelled = 0
              AND c.DateOnCheck IS NOT NULL AND c.DateOnCheck != ''
              AND r.EndDate IS NOT NULL AND r.EndDate != ''
              AND r.EndDate >= ? AND r.EndDate <= ?
              AND (julianday(c.DateOnCheck) - julianday(r.EndDate)) > 0
              AND {who_filter}
            GROUP BY r.{group_col}
            ORDER BY avg_days DESC
        '''
        return db.execute(sql, (date_from, date_to)).fetchall()

    def totals(rows):
        n = sum(r['num_payments'] for r in rows)
        total_comm = sum(r['total_commission'] for r in rows)
        weighted = (sum(r['avg_days'] * r['num_payments'] for r in rows) / n) if n else 0
        mn = min((r['min_days'] for r in rows), default=0)
        mx = max((r['max_days'] for r in rows), default=0)
        return {'num_payments': n, 'avg_days': round(weighted, 1),
                'min_days': mn, 'max_days': mx, 'total_commission': round(total_comm, 2)}

    kristin_brand = _run('Brand', 'Brand', "r.BookingAssociate LIKE '%Kristin%'")
    kristin_chain = _run('Chain', 'Chain', "r.BookingAssociate LIKE '%Kristin%'")
    team_brand    = _run('Brand', 'Brand', "r.BookingAssociate NOT LIKE '%Kristin%'")
    team_chain    = _run('Chain', 'Chain', "r.BookingAssociate NOT LIKE '%Kristin%'")

    if request.args.get('export') == 'xlsx':
        import io, openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb   = openpyxl.Workbook()
        thin = Side(style='thin', color='CCCCCC')
        bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
        subtitle_text = f'Period: {date_from} to {date_to}  |  Payments received after event end date only'

        def _build_sheet(ws, title, col_label, rows, totals_row):
            col_headers = [col_label, '# Payments', 'Avg Days Outstanding', 'Min Days', 'Max Days', 'Total Commission ($)']
            col_widths  = [38, 13, 22, 12, 12, 22]

            ws.merge_cells('A1:F1')
            tc = ws['A1']
            tc.value     = title
            tc.font      = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
            tc.fill      = PatternFill('solid', fgColor='1F3864')
            tc.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 22

            ws.merge_cells('A2:F2')
            sc = ws['A2']
            sc.value     = subtitle_text
            sc.font      = Font(name='Calibri', italic=True, color='666666', size=10)
            sc.alignment = Alignment(horizontal='center')
            ws.row_dimensions[2].height = 14
            ws.row_dimensions[3].height = 5

            for ci, h in enumerate(col_headers, 1):
                cell = ws.cell(row=4, column=ci, value=h)
                cell.font      = Font(name='Calibri', bold=True, color='FFFFFF')
                cell.fill      = PatternFill('solid', fgColor='2E75B6')
                cell.alignment = Alignment(horizontal='center')
                cell.border    = bdr

            for ri, r in enumerate(rows):
                rn = ri + 5
                vals = [r['grp'], r['num_payments'], r['avg_days'],
                        int(r['min_days']), int(r['max_days']), r['total_commission']]
                bg = 'DEEAF1' if ri % 2 == 0 else 'FFFFFF'
                for ci, v in enumerate(vals, 1):
                    cell = ws.cell(row=rn, column=ci, value=v)
                    cell.fill      = PatternFill('solid', fgColor=bg)
                    cell.border    = bdr
                    cell.font      = Font(name='Calibri')
                    cell.alignment = Alignment(horizontal='left' if ci == 1 else ('right' if ci == 6 else 'center'))
                ws.cell(row=rn, column=3).number_format = '0.0'
                ws.cell(row=rn, column=6).number_format = '$#,##0.00'

            tr = len(rows) + 5
            tvals = ['TOTAL / AVERAGE', totals_row['num_payments'],
                     round(totals_row['avg_days'], 1),
                     int(totals_row['min_days']), int(totals_row['max_days']),
                     round(totals_row['total_commission'], 2)]
            for ci, v in enumerate(tvals, 1):
                cell = ws.cell(row=tr, column=ci, value=v)
                cell.fill   = PatternFill('solid', fgColor='D9D9D9')
                cell.font   = Font(name='Calibri', bold=True)
                cell.border = bdr
                cell.alignment = Alignment(horizontal='left' if ci == 1 else ('right' if ci == 6 else 'center'))
            ws.cell(row=tr, column=3).number_format = '0.0'
            ws.cell(row=tr, column=6).number_format = '$#,##0.00'

            for ci, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(ci)].width = w
            ws.freeze_panes = 'A5'

        sheets = [
            (wb.active, 'Kristin — Brand', 'Kristin House — Avg Days by Brand', 'Brand', kristin_brand, totals(kristin_brand)),
        ]
        wb.active.title = 'Kristin — Brand'
        for ws_title, sheet_title, col_label, rows, tot in [
            ('Kristin — Brand', 'Kristin House — Avg Days to Payment by Brand', 'Brand', kristin_brand, totals(kristin_brand)),
            ('Kristin — Chain', 'Kristin House — Avg Days to Payment by Chain', 'Chain', kristin_chain, totals(kristin_chain)),
            ('Team — Brand',    'Team — Avg Days to Payment by Brand',           'Brand', team_brand,    totals(team_brand)),
            ('Team — Chain',    'Team — Avg Days to Payment by Chain',           'Chain', team_chain,    totals(team_chain)),
        ]:
            ws = wb.active if ws_title == 'Kristin — Brand' else wb.create_sheet(ws_title)
            ws.title = ws_title
            _build_sheet(ws, sheet_title, col_label, [dict(r) for r in rows], tot)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, download_name='Commission_Days_Outstanding.xlsx',
                         as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    return render_template('report_days_outstanding.html',
                           kristin_brand=kristin_brand, kristin_chain=kristin_chain,
                           team_brand=team_brand,       team_chain=team_chain,
                           kristin_brand_totals=totals(kristin_brand),
                           kristin_chain_totals=totals(kristin_chain),
                           team_brand_totals=totals(team_brand),
                           team_chain_totals=totals(team_chain),
                           date_from=date_from, date_to=date_to)


# ── Customer Summary Report ────────────────────────────────────────────────────

from cs_report_utils import (
    _chain_color, _query_customer_summary, _query_meeting_summary,
    _build_word_doc, _build_pptx, build_city_map_data
)

@app.route('/reports/proforma')
def report_proforma():
    user = get_current_user()
    if not user or user['role'] != 'admin':
        flash('This report is restricted to administrators.', 'error')
        return redirect(url_for('pipeline'))

    from datetime import date as _date, datetime as _dt, timedelta as _td
    from collections import defaultdict
    import json as _json, io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.chart.label import DataLabel

    db    = get_db()
    today = _date.today()
    six_months_ago = str(today - _td(days=182))
    one_year_ago   = str(today - _td(days=365))

    # ── 12-month variance: contracted vs actual (excl. Hilton/Marriott advances) ─
    variance_raw = db.execute('''
        SELECT r.BookingId, r.Brand, r.Chain,
               r.USDCommissionableAmount, r.CommissionPercent,
               c.FinalPayment, c.Advance
        FROM ReportPipeline r
        JOIN ChkRegNote c ON c.BookingID = r.BookingId
        WHERE c.Cancelled=0
          AND c.DateOnCheck IS NOT NULL AND c.DateOnCheck != ''
          AND c.DateOnCheck >= ?
          AND r.USDCommissionableAmount IS NOT NULL AND r.USDCommissionableAmount > 0
          AND r.CommissionPercent IS NOT NULL AND r.CommissionPercent > 0
          AND c.FinalPayment > 0
    ''', (one_year_ago,)).fetchall()

    # Better contracted commission via pickup_config block × rate
    pc_contracted = {}
    for pc in db.execute('''
        SELECT booking_id, contracted_block, contracted_rate FROM pickup_config
        WHERE contracted_block IS NOT NULL AND contracted_block NOT IN ('{}','')
          AND contracted_rate IS NOT NULL AND contracted_rate > 0
    ''').fetchall():
        try:
            block = _json.loads(pc['contracted_block'] or '{}')
            rooms = sum(v for v in block.values() if v)
            if rooms > 0:
                pc_contracted[str(pc['booking_id'])] = rooms * float(pc['contracted_rate'])
        except Exception:
            pass

    booking_actuals = {}
    for row in variance_raw:
        bid   = str(row['BookingId'])
        brand = (row['Brand'] or '').strip()
        chain = (row['Chain'] or '').strip()
        is_hm = ('hilton' in brand.lower() or 'marriott' in brand.lower() or
                 'hilton' in chain.lower() or 'marriott' in chain.lower())
        if bid not in booking_actuals:
            if bid in pc_contracted and row['CommissionPercent']:
                contracted = pc_contracted[bid] * float(row['CommissionPercent'])
            else:
                contracted = float(row['USDCommissionableAmount'] or 0) * float(row['CommissionPercent'] or 0)
            booking_actuals[bid] = {'contracted': contracted, 'actual': 0.0}
        if is_hm and row['Advance'] == 1:
            continue                      # skip Hilton/Marriott advance payments
        booking_actuals[bid]['actual'] += float(row['FinalPayment'] or 0)

    ratios = []
    for d in booking_actuals.values():
        if d['contracted'] > 10 and d['actual'] > 0:
            r_ = d['actual'] / d['contracted']
            if 0.3 <= r_ <= 1.5:          # exclude extreme outliers
                ratios.append(r_)
    avg_ratio   = sum(ratios) / len(ratios) if ratios else 1.0
    avg_pct_off = round((avg_ratio - 1.0) * 100, 1)   # negative = below contracted
    n_variance  = len(ratios)

    # ── Parse default days (fallback when no brand/chain timing data) ─────────
    try:
        default_days = max(1, min(365, int(request.args.get('default_days', 90) or 90)))
    except (ValueError, TypeError):
        default_days = 90

    # ── Landing page (no export) ──────────────────────────────────────────────
    if request.args.get('export') != 'xlsx':
        return render_template('report_proforma.html',
                               avg_ratio_pct=round(avg_ratio * 100, 1),
                               avg_pct_off=avg_pct_off,
                               n_variance=n_variance,
                               default_adj=int(round(avg_pct_off, 0)),
                               default_days=default_days)

    # ── Adjustment % from form (default = historical variance) ────────────────
    try:
        adj_pct = float(request.args.get('adj_pct', avg_pct_off) or avg_pct_off)
    except ValueError:
        adj_pct = avg_pct_off
    adj_mult = 1.0 + adj_pct / 100.0    # e.g., -12% → 0.88×

    # ── 6-month moving average by brand ──────────────────────────────────────
    brand_avg = {}
    for row in db.execute('''
        SELECT COALESCE(NULLIF(r.Brand,''),'') AS b,
               AVG(julianday(c.DateOnCheck) - julianday(r.EndDate)) AS avg_days
        FROM ChkRegNote c JOIN ReportPipeline r ON c.BookingID = r.BookingId
        WHERE c.Cancelled=0 AND c.DateOnCheck >= ?
          AND c.DateOnCheck IS NOT NULL AND c.DateOnCheck != ''
          AND r.EndDate   IS NOT NULL AND r.EndDate   != ''
          AND (julianday(c.DateOnCheck) - julianday(r.EndDate)) > 0
        GROUP BY r.Brand
    ''', (six_months_ago,)).fetchall():
        if row[0]:
            brand_avg[row[0]] = row[1]

    # ── 6-month moving average by chain ──────────────────────────────────────
    chain_avg = {}
    for row in db.execute('''
        SELECT COALESCE(NULLIF(r.Chain,''),'') AS ch,
               AVG(julianday(c.DateOnCheck) - julianday(r.EndDate)) AS avg_days
        FROM ChkRegNote c JOIN ReportPipeline r ON c.BookingID = r.BookingId
        WHERE c.Cancelled=0 AND c.DateOnCheck >= ?
          AND c.DateOnCheck IS NOT NULL AND c.DateOnCheck != ''
          AND r.EndDate   IS NOT NULL AND r.EndDate   != ''
          AND (julianday(c.DateOnCheck) - julianday(r.EndDate)) > 0
        GROUP BY r.Chain
    ''', (six_months_ago,)).fetchall():
        if row[0]:
            chain_avg[row[0]] = row[1]

    def proj_days(brand, chain):
        ch = (chain or '').strip()
        if ch and chain_avg.get(ch):
            return int(round(chain_avg[ch])), ch   # actual chain name
        return default_days, 'Default'

    # ── Max meeting year ──────────────────────────────────────────────────────
    row = db.execute(
        "SELECT MAX(CAST(substr(EndDate,1,4) AS INTEGER)) FROM ReportPipeline "
        "WHERE EndDate IS NOT NULL AND EndDate!='' AND length(EndDate)>=4"
    ).fetchone()
    max_meeting_year = int(row[0]) if row and row[0] else today.year + 2

    # ── Paid rows: payments received in 2026+ (any meeting end date) ──────────
    paid_rows = db.execute('''
        SELECT r.BookingId,
               r.BookingAssociate,
               COALESCE(r.AccountName,'') AS customer_name,
               COALESCE(NULLIF(r.EventName,''), r.AccountName, '') AS meeting_name,
               COALESCE(r.Customer,'')   AS hotel,
               COALESCE(r.StartDate,'') AS start_date,
               COALESCE(r.EndDate,'')   AS end_date,
               r.Brand, r.Chain,
               r.USDCommissionableAmount, r.Revenue,
               c.DateOnCheck, c.FinalPayment, c.Advance
        FROM ChkRegNote c JOIN ReportPipeline r ON c.BookingID = r.BookingId
        WHERE c.Cancelled=0
          AND c.DateOnCheck IS NOT NULL AND c.DateOnCheck != ''
          AND c.FinalPayment > 0
          AND c.DateOnCheck >= '2026-01-01'
        ORDER BY c.DateOnCheck
    ''').fetchall()

    # ── HHR commissions: actual room revenue from hhr_data blob or Pickup table
    hhr_commissions = {}   # booking_id (str) → commission amount
    hhr_revenues    = {}   # booking_id (str) → room revenue (base for % calc)

    # Priority 1: parse uploaded HHR Excel files in pickup_config
    for pc in db.execute('''
        SELECT pc.booking_id, pc.hhr_data, r.CommissionPercent
        FROM pickup_config pc
        JOIN ReportPipeline r ON r.BookingId = pc.booking_id
        WHERE pc.hhr_data IS NOT NULL
          AND r.CommissionPercent IS NOT NULL AND r.CommissionPercent > 0
    ''').fetchall():
        try:
            from pickup_utils import parse_hhr_excel as _parse_hhr
            stats    = _parse_hhr(bytes(pc['hhr_data']))
            room_rev = stats.get('room_revenue') or stats.get('room_revenue_calc')
            if room_rev and room_rev > 0:
                bid = str(pc['booking_id'])
                hhr_commissions[bid] = round(
                    float(room_rev) * float(pc['CommissionPercent']), 2)
                hhr_revenues[bid]    = round(float(room_rev), 2)
        except Exception:
            pass

    # Priority 2: fall back to old Pickup table (actual room nights × rate × pct)
    for row in db.execute('''
        SELECT p.BookingID,
               SUM(p.ActualPickup) AS total_rn,
               r.RoomRate, r.CommissionPercent
        FROM Pickup p
        JOIN ReportPipeline r ON r.BookingId = p.BookingID
        WHERE p.ActualPickup IS NOT NULL AND p.ActualPickup > 0
          AND r.RoomRate IS NOT NULL AND r.RoomRate > 0
          AND r.CommissionPercent IS NOT NULL AND r.CommissionPercent > 0
        GROUP BY p.BookingID
    ''').fetchall():
        bid = str(row['BookingID'])
        if bid not in hhr_commissions:
            rev = round(float(row['total_rn'] or 0) * float(row['RoomRate'] or 0), 2)
            amt = round(rev * float(row['CommissionPercent'] or 0), 2)
            if amt > 0:
                hhr_commissions[bid] = amt
                hhr_revenues[bid]    = rev

    # ── Unpaid/projected rows: meetings ending 2025-07-01+ with no payment ───
    unpaid_rows = db.execute('''
        SELECT r.BookingId,
               r.BookingAssociate,
               COALESCE(r.AccountName,'') AS customer_name,
               COALESCE(NULLIF(r.EventName,''), r.AccountName, '') AS meeting_name,
               COALESCE(r.Customer,'')            AS hotel,
               COALESCE(r.StartDate,'')           AS start_date,
               COALESCE(r.EndDate,'')             AS end_date,
               r.Brand, r.Chain,
               r.USDCommissionableAmount, r.CommissionPercent, r.Revenue
        FROM ReportPipeline r
        WHERE r.EndDate IS NOT NULL AND r.EndDate != ''
          AND r.EndDate >= '2025-07-01'
          AND COALESCE(r.BookingStatus,'') NOT IN ('Cancelled','Lost','Tentative')
          AND r.CommissionPercent IS NOT NULL AND r.CommissionPercent > 0
          AND r.BookingId NOT IN (
              SELECT DISTINCT BookingID FROM ChkRegNote
              WHERE Cancelled=0
                AND DateOnCheck IS NOT NULL AND DateOnCheck != ''
                AND FinalPayment > 0
          )
        ORDER BY r.EndDate
    ''').fetchall()

    # ── Helpers ───────────────────────────────────────────────────────────────
    def parse_dt(s):
        if not s: return None
        try:    return _dt.strptime(s[:10], '%Y-%m-%d').date()
        except: return None

    def fmt_date(s):
        d = parse_dt(s)
        return d.strftime('%m/%d/%Y') if d else ''

    def is_kristin(associate):
        return 'kristin' in (associate or '').lower()

    # ── Assemble year→who→rows ────────────────────────────────────────────────
    kristin_split = get_kristin_split()   # Kristin's personal cut (e.g. 0.70)
    data = defaultdict(lambda: {'kristin': [], 'team': []})

    for r in paid_rows:
        pd = parse_dt(r['DateOnCheck'])
        if not pd or pd.year < 2026:
            continue
        year = pd.year
        who  = 'kristin' if is_kristin(r['BookingAssociate']) else 'team'
        bid  = str(r['BookingId'])
        if who == 'kristin':
            # Kristin: HHR room revenue if available, else contracted commissionable
            rev = hhr_revenues.get(bid) or float(r['USDCommissionableAmount'] or r['Revenue'] or 0)
        else:
            # Team: always Pipeline Revenue
            rev = float(r['Revenue'] or 0)
        data[year][who].append({
            'bid':            bid,
            'customer':       r['customer_name'],
            'associate':      r['BookingAssociate'] or '',
            'meeting':        r['meeting_name'],
            'hotel':          r['hotel'],
            'start':          fmt_date(r['start_date']),
            'end':            fmt_date(r['end_date']),
            'paid_date':      pd,
            'revenue':        rev,
            'amount':         float(r['FinalPayment'] or 0),
            'month':          pd.month,
            'projected':      False,
            'src':            'paid',
            'days_src':       None,
            'proj_days_used': None,
        })

    for r in unpaid_rows:
        ed = parse_dt(r['end_date'])
        if not ed: continue
        days, days_src = proj_days(r['Brand'], r['Chain'])
        pd   = ed + _td(days=days)
        year = min(pd.year, max_meeting_year + 2)
        if year < 2026:
            year = 2026   # if projected into the past, place in 2026

        bid = str(r['BookingId'])
        who = 'kristin' if is_kristin(r['BookingAssociate']) else 'team'

        if who == 'kristin':
            # Kristin: HHR commission → green; else CommissionPercent × revenue → yellow
            # Apply her personal split (e.g. 70%) to all unpaid projections
            if bid in hhr_commissions:
                amt = round(hhr_commissions[bid] * kristin_split, 2)
                rev = hhr_revenues.get(bid, 0.0)
                src = 'hhr'
            else:
                rev = float(r['USDCommissionableAmount'] or r['Revenue'] or 0)
                amt = round(rev * float(r['CommissionPercent'] or 0) * adj_mult * kristin_split, 2)
                src = 'calc'
        else:
            # Team: Pipeline Revenue, always 1% × adj
            rev = float(r['Revenue'] or 0)
            amt = round(rev * 0.01 * adj_mult, 2)
            # Overdue: projected date already passed — flag red and move to today
            if pd < today:
                pd   = today
                year = today.year
                src  = 'overdue'
            else:
                src  = 'calc'

        if amt <= 0:
            continue

        data[year][who].append({
            'bid':            bid,
            'customer':       r['customer_name'],
            'associate':      r['BookingAssociate'] or '',
            'meeting':        r['meeting_name'],
            'hotel':          r['hotel'],
            'start':          fmt_date(r['start_date']),
            'end':            fmt_date(r['end_date']),
            'paid_date':      pd,
            'revenue':        rev,
            'amount':         amt,
            'month':          pd.month,
            'projected':      True,
            'src':            src,
            'days_src':       days_src,
            'proj_days_used': days,
        })

    # Sort each bucket by payment date
    for yr in data:
        for who in ('kristin', 'team'):
            data[yr][who].sort(key=lambda x: x['paid_date'])

    # ── Pre-compute monthly summary totals ────────────────────────────────────
    # summary[year][who] = list of 12 monthly totals
    summary = {}
    for yr in sorted(data.keys()):
        summary[yr] = {'kristin': [0.0]*12, 'team': [0.0]*12}
        for who in ('kristin', 'team'):
            for rd in data[yr][who]:
                mo = rd['month']
                if 1 <= mo <= 12:
                    summary[yr][who][mo-1] += rd['amount']

    # ── Build Excel ───────────────────────────────────────────────────────────
    MONTH_ABBR = ['Jan','Feb','Mar','Apr','May','Jun',
                  'Jul','Aug','Sep','Oct','Nov','Dec']

    thin = Side(style='thin', color='CCCCCC')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hfill(hex_str):
        return PatternFill('solid', fgColor=hex_str)

    C = {
        'title':    '1F3864',   # dark navy
        'header':   '2E75B6',   # medium blue
        'actual_a': 'DEEAF1',   # light blue (alternating even rows)
        'actual_b': 'FFFFFF',   # white (odd rows)
        'hhr':      'E2EFDA',   # light green (HHR-sourced projection)
        'calc':     'FFF9C4',   # light yellow (calculated projection)
        'overdue':  'FFCCCC',   # light red (team unpaid past due)
        'subtotal': 'D9D9D9',   # gray totals row
    }

    wb = openpyxl.Workbook()
    wb.remove(wb.active)    # drop default Sheet

    years = sorted(data.keys())
    if not years:
        years = [today.year]

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet('Summary')
    SUM_HDRS  = ['Year', 'Kristin / Team'] + MONTH_ABBR + ['Year Total']
    SUM_WIDTHS = [8, 14] + [11]*12 + [13]
    n_sum_cols = len(SUM_HDRS)
    sum_last_col = get_column_letter(n_sum_cols)

    adj_sign_str = ('+' if adj_pct >= 0 else '') + f'{adj_pct:.1f}%'

    # Title
    ws_sum.merge_cells(f'A1:{sum_last_col}1')
    t = ws_sum['A1']
    t.value     = 'Commission Proforma — Monthly Summary by Year'
    t.font      = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
    t.fill      = hfill(C['title'])
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws_sum.row_dimensions[1].height = 26

    # Subtitle
    ws_sum.merge_cells(f'A2:{sum_last_col}2')
    s = ws_sum['A2']
    s.value = (f'Generated {today.strftime("%B %d, %Y")}  │  '
               f'Projected adjustment: {adj_sign_str}  │  '
               f'Includes actual payments + HHR-sourced + contracted estimates')
    s.font      = Font(name='Calibri', italic=True, color='555555', size=9)
    s.alignment = Alignment(horizontal='center')
    ws_sum.row_dimensions[2].height = 14
    ws_sum.row_dimensions[3].height = 5

    # Column headers
    for ci, h in enumerate(SUM_HDRS, 1):
        cell = ws_sum.cell(row=4, column=ci, value=h)
        cell.font      = Font(name='Calibri', bold=True, color='FFFFFF', size=9)
        cell.fill      = hfill(C['header'])
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border    = bdr
    ws_sum.row_dimensions[4].height = 30

    grand_total = [0.0] * 12

    data_row = 5
    for yr in years:
        k_mo  = summary[yr]['kristin']
        t_mo  = summary[yr]['team']
        yr_k  = sum(k_mo)
        yr_t  = sum(t_mo)
        yr_tot = yr_k + yr_t

        for mo_i in range(12):
            grand_total[mo_i] += k_mo[mo_i] + t_mo[mo_i]

        ROWS = [
            (str(yr), 'Kristin',    k_mo,                            yr_k,   'DEEAF1', False),
            ('',      'Team',       t_mo,                            yr_t,   'FFFFFF', False),
            ('',      f'{yr} TOTAL',[k_mo[i]+t_mo[i] for i in range(12)], yr_tot, 'D9D9D9', True),
        ]

        for label_yr, label_who, monthly, yr_sum, bg, is_total in ROWS:
            fnt = Font(name='Calibri', size=9, bold=is_total)
            fill = hfill(bg)

            ws_sum.cell(row=data_row, column=1, value=label_yr).font = fnt
            ws_sum.cell(row=data_row, column=1).fill   = fill
            ws_sum.cell(row=data_row, column=1).border = bdr
            ws_sum.cell(row=data_row, column=1).alignment = Alignment(horizontal='center')

            ws_sum.cell(row=data_row, column=2, value=label_who).font  = fnt
            ws_sum.cell(row=data_row, column=2).fill   = fill
            ws_sum.cell(row=data_row, column=2).border = bdr
            ws_sum.cell(row=data_row, column=2).alignment = Alignment(horizontal='left')

            for mo_i, mo_val in enumerate(monthly):
                cell = ws_sum.cell(row=data_row, column=3 + mo_i,
                                   value=mo_val if mo_val else None)
                cell.font          = fnt
                cell.fill          = fill
                cell.border        = bdr
                cell.number_format = '$#,##0'
                cell.alignment     = Alignment(horizontal='right')

            tot_cell = ws_sum.cell(row=data_row, column=n_sum_cols, value=yr_sum if yr_sum else None)
            tot_cell.font          = Font(name='Calibri', size=9, bold=True)
            tot_cell.fill          = fill
            tot_cell.border        = bdr
            tot_cell.number_format = '$#,##0'
            tot_cell.alignment     = Alignment(horizontal='right')

            ws_sum.row_dimensions[data_row].height = 14
            data_row += 1

        # Spacer between years
        for ci in range(1, n_sum_cols + 1):
            ws_sum.cell(row=data_row, column=ci).fill   = hfill('F2F2F2')
            ws_sum.cell(row=data_row, column=ci).border = Border(
                bottom=Side(style='thin', color='AAAAAA'))
        ws_sum.row_dimensions[data_row].height = 4
        data_row += 1

    # Grand Total row
    gt_total = sum(grand_total)
    ws_sum.cell(row=data_row, column=1, value='ALL').font   = Font(name='Calibri', bold=True, size=9, color='FFFFFF')
    ws_sum.cell(row=data_row, column=1).fill   = hfill(C['title'])
    ws_sum.cell(row=data_row, column=1).border = bdr
    ws_sum.cell(row=data_row, column=1).alignment = Alignment(horizontal='center')

    ws_sum.cell(row=data_row, column=2, value='GRAND TOTAL').font  = Font(name='Calibri', bold=True, size=9, color='FFFFFF')
    ws_sum.cell(row=data_row, column=2).fill   = hfill(C['title'])
    ws_sum.cell(row=data_row, column=2).border = bdr
    ws_sum.cell(row=data_row, column=2).alignment = Alignment(horizontal='left')

    for mo_i, mo_val in enumerate(grand_total):
        cell = ws_sum.cell(row=data_row, column=3 + mo_i, value=mo_val if mo_val else None)
        cell.font          = Font(name='Calibri', bold=True, size=9, color='FFFFFF')
        cell.fill          = hfill(C['title'])
        cell.border        = bdr
        cell.number_format = '$#,##0'
        cell.alignment     = Alignment(horizontal='right')

    gt_cell = ws_sum.cell(row=data_row, column=n_sum_cols, value=gt_total if gt_total else None)
    gt_cell.font          = Font(name='Calibri', bold=True, size=9, color='FFFFFF')
    gt_cell.fill          = hfill(C['title'])
    gt_cell.border        = bdr
    gt_cell.number_format = '$#,##0'
    gt_cell.alignment     = Alignment(horizontal='right')
    ws_sum.row_dimensions[data_row].height = 18

    # Column widths & freeze
    for ci, w in enumerate(SUM_WIDTHS, 1):
        ws_sum.column_dimensions[get_column_letter(ci)].width = w
    ws_sum.freeze_panes = 'A5'

    # ── Summary charts ────────────────────────────────────────────────────────
    grand_total_row = data_row   # row number of the Grand Total row

    # Chart 1: Monthly grand totals (column chart) — placed below the table
    cht_monthly = BarChart()
    cht_monthly.type    = 'col'
    cht_monthly.title   = 'Monthly Commission — All Years Combined'
    cht_monthly.y_axis.title = 'Commission ($)'
    cht_monthly.x_axis.title = 'Month'
    cht_monthly.style   = 10
    cht_monthly.width   = 22
    cht_monthly.height  = 12
    cht_monthly.y_axis.numFmt = '$#,##0'
    cht_monthly.y_axis.majorGridlines = None

    data_ref_mo = Reference(ws_sum, min_col=3, max_col=14,
                            min_row=grand_total_row, max_row=grand_total_row)
    cats_ref_mo = Reference(ws_sum, min_col=3, max_col=14, min_row=4)
    cht_monthly.add_data(data_ref_mo, from_rows=True)
    cht_monthly.set_categories(cats_ref_mo)
    cht_monthly.series[0].graphicalProperties.solidFill = '2E75B6'
    cht_monthly.series[0].graphicalProperties.line.solidFill = '1F3864'
    ws_sum.add_chart(cht_monthly, f'A{grand_total_row + 3}')

    # Chart 2: Year totals — Kristin vs Team (clustered column)
    # Build a mini data table below the main table for year-level chart data
    tbl_start = grand_total_row + 3 + 22   # below chart 1 (~22 rows tall at default height)
    ws_sum.cell(row=tbl_start, column=1, value='Year').font = Font(name='Calibri', bold=True, size=8)
    ws_sum.cell(row=tbl_start, column=2, value='Kristin').font = Font(name='Calibri', bold=True, size=8)
    ws_sum.cell(row=tbl_start, column=3, value='Team').font = Font(name='Calibri', bold=True, size=8)
    for yi, yr in enumerate(years):
        r = tbl_start + 1 + yi
        ws_sum.cell(row=r, column=1, value=str(yr)).font = Font(name='Calibri', size=8)
        ws_sum.cell(row=r, column=2, value=sum(summary[yr]['kristin'])).number_format = '$#,##0'
        ws_sum.cell(row=r, column=2).font = Font(name='Calibri', size=8)
        ws_sum.cell(row=r, column=3, value=sum(summary[yr]['team'])).number_format = '$#,##0'
        ws_sum.cell(row=r, column=3).font = Font(name='Calibri', size=8)

    cht_yr = BarChart()
    cht_yr.type    = 'col'
    cht_yr.title   = 'Annual Commission — Kristin vs Team'
    cht_yr.y_axis.title = 'Commission ($)'
    cht_yr.x_axis.title = 'Year'
    cht_yr.style   = 10
    cht_yr.width   = 16
    cht_yr.height  = 12
    cht_yr.y_axis.numFmt = '$#,##0'
    cht_yr.y_axis.majorGridlines = None
    cht_yr.grouping = 'clustered'

    yr_data_rows = tbl_start + 1 + len(years)
    k_ref = Reference(ws_sum, min_col=2, max_col=2,
                      min_row=tbl_start, max_row=yr_data_rows - 1)
    t_ref = Reference(ws_sum, min_col=3, max_col=3,
                      min_row=tbl_start, max_row=yr_data_rows - 1)
    yr_cats = Reference(ws_sum, min_col=1, max_col=1,
                        min_row=tbl_start + 1, max_row=yr_data_rows - 1)
    cht_yr.add_data(k_ref, titles_from_data=True)
    cht_yr.add_data(t_ref, titles_from_data=True)
    cht_yr.set_categories(yr_cats)
    cht_yr.series[0].graphicalProperties.solidFill = '2E75B6'   # Kristin — blue
    cht_yr.series[0].graphicalProperties.line.solidFill = '1F3864'
    cht_yr.series[1].graphicalProperties.solidFill = 'ED7D31'   # Team — orange
    cht_yr.series[1].graphicalProperties.line.solidFill = 'C55A11'
    ws_sum.add_chart(cht_yr, f'N{grand_total_row + 3}')

    # ── Avg Days sheet ───────────────────────────────────────────────────────
    ws_avg = wb.create_sheet('Avg Days — Payment Timing')

    # Title
    ws_avg.merge_cells('A1:E1')
    t = ws_avg['A1']
    t.value     = '6-Month Rolling Average — Days from Meeting End to Commission Received'
    t.font      = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
    t.fill      = hfill(C['title'])
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws_avg.row_dimensions[1].height = 26

    ws_avg.merge_cells('A2:E2')
    s = ws_avg['A2']
    s.value = (f'Based on payments received {six_months_ago} – {today}  │  '
               f'Chain avg used first → {default_days}-day default')
    s.font      = Font(name='Calibri', italic=True, color='555555', size=9)
    s.alignment = Alignment(horizontal='center')
    ws_avg.row_dimensions[2].height = 14
    ws_avg.row_dimensions[3].height = 5

    # Section: Brand averages
    def _avg_section_header(ws, row, label, col_a, col_b, col_c):
        for ci, (col, val, algn) in enumerate([(col_a, label, 'left'),
                                                (col_b, 'Avg Days', 'center'),
                                                (col_c, 'Used By (unpaid meetings)', 'left')], 1):
            cell = ws.cell(row=row, column=[col_a, col_b, col_c][ci-1])
            cell.value     = val
            cell.font      = Font(name='Calibri', bold=True, color='FFFFFF', size=9)
            cell.fill      = hfill(C['header'])
            cell.alignment = Alignment(horizontal=algn, wrap_text=True)
            cell.border    = bdr
        ws.row_dimensions[row].height = 24

    # Build a lookup: which unpaid bookings use each brand/chain?
    brand_users = {}   # brand → list of meeting names
    chain_users = {}   # chain → list of meeting names
    default_users = []
    for yr_data in data.values():
        for who_rows in yr_data.values():
            for rd in who_rows:
                if rd['projected']:
                    pass  # we don't store brand/chain in rd — skip user lists for now

    # Brand section header row
    br_hdr_row = 4
    ws_avg.cell(row=br_hdr_row, column=1, value='Brand').font      = Font(name='Calibri', bold=True, color='FFFFFF', size=9)
    ws_avg.cell(row=br_hdr_row, column=1).fill      = hfill(C['header'])
    ws_avg.cell(row=br_hdr_row, column=1).alignment = Alignment(horizontal='left')
    ws_avg.cell(row=br_hdr_row, column=1).border    = bdr
    ws_avg.cell(row=br_hdr_row, column=2, value='Avg Days (6 mo)').font      = Font(name='Calibri', bold=True, color='FFFFFF', size=9)
    ws_avg.cell(row=br_hdr_row, column=2).fill      = hfill(C['header'])
    ws_avg.cell(row=br_hdr_row, column=2).alignment = Alignment(horizontal='center', wrap_text=True)
    ws_avg.cell(row=br_hdr_row, column=2).border    = bdr
    ws_avg.cell(row=br_hdr_row, column=3, value='# Payments').font      = Font(name='Calibri', bold=True, color='FFFFFF', size=9)
    ws_avg.cell(row=br_hdr_row, column=3).fill      = hfill(C['header'])
    ws_avg.cell(row=br_hdr_row, column=3).alignment = Alignment(horizontal='center')
    ws_avg.cell(row=br_hdr_row, column=3).border    = bdr
    ws_avg.row_dimensions[br_hdr_row].height = 28

    # Brand data rows — fetch with count
    brand_detail = db.execute('''
        SELECT COALESCE(NULLIF(r.Brand,''), 'Unknown') AS b,
               ROUND(AVG(julianday(c.DateOnCheck) - julianday(r.EndDate)), 1) AS avg_days,
               COUNT(*) AS n
        FROM ChkRegNote c JOIN ReportPipeline r ON c.BookingID = r.BookingId
        WHERE c.Cancelled=0 AND c.DateOnCheck >= ?
          AND c.DateOnCheck IS NOT NULL AND c.DateOnCheck != ''
          AND r.EndDate IS NOT NULL AND r.EndDate != ''
          AND (julianday(c.DateOnCheck) - julianday(r.EndDate)) > 0
        GROUP BY r.Brand
        ORDER BY avg_days DESC
    ''', (six_months_ago,)).fetchall()

    br_row = br_hdr_row + 1
    for i, row in enumerate(brand_detail):
        bg = 'F0F4FA' if i % 2 == 0 else 'FFFFFF'
        days = float(row[1] or 90)
        # Color-code by speed
        if days >= 120:   day_color = 'FFC7CE'   # red tint
        elif days >= 90:  day_color = 'FFEB9C'   # yellow tint
        elif days >= 60:  day_color = 'C6EFCE'   # green tint
        else:             day_color = 'DDFFDD'   # bright green

        c1 = ws_avg.cell(row=br_row, column=1, value=row[0])
        c1.fill = hfill(bg); c1.border = bdr
        c1.font = Font(name='Calibri', size=9)
        c1.alignment = Alignment(horizontal='left')

        c2 = ws_avg.cell(row=br_row, column=2, value=days)
        c2.fill = hfill(day_color); c2.border = bdr
        c2.font = Font(name='Calibri', size=9, bold=True)
        c2.number_format = '0.0'
        c2.alignment = Alignment(horizontal='center')

        c3 = ws_avg.cell(row=br_row, column=3, value=int(row[2]))
        c3.fill = hfill(bg); c3.border = bdr
        c3.font = Font(name='Calibri', size=9)
        c3.alignment = Alignment(horizontal='center')

        ws_avg.row_dimensions[br_row].height = 14
        br_row += 1

    # Spacer
    ws_avg.row_dimensions[br_row].height = 8
    br_row += 1

    # Chain section header
    ch_hdr_row = br_row
    for ci, (val, algn) in enumerate([('Chain', 'left'),
                                       ('Avg Days (6 mo)', 'center'),
                                       ('# Payments', 'center')], 1):
        cell = ws_avg.cell(row=ch_hdr_row, column=ci, value=val)
        cell.font      = Font(name='Calibri', bold=True, color='FFFFFF', size=9)
        cell.fill      = hfill('4472C4')   # slightly different blue for chain
        cell.alignment = Alignment(horizontal=algn, wrap_text=True)
        cell.border    = bdr
    ws_avg.row_dimensions[ch_hdr_row].height = 28

    chain_detail = db.execute('''
        SELECT COALESCE(NULLIF(r.Chain,''), 'Unknown') AS ch,
               ROUND(AVG(julianday(c.DateOnCheck) - julianday(r.EndDate)), 1) AS avg_days,
               COUNT(*) AS n
        FROM ChkRegNote c JOIN ReportPipeline r ON c.BookingID = r.BookingId
        WHERE c.Cancelled=0 AND c.DateOnCheck >= ?
          AND c.DateOnCheck IS NOT NULL AND c.DateOnCheck != ''
          AND r.EndDate IS NOT NULL AND r.EndDate != ''
          AND (julianday(c.DateOnCheck) - julianday(r.EndDate)) > 0
        GROUP BY r.Chain
        ORDER BY avg_days DESC
    ''', (six_months_ago,)).fetchall()

    ch_row = ch_hdr_row + 1
    for i, row in enumerate(chain_detail):
        bg = 'F5F0FA' if i % 2 == 0 else 'FFFFFF'
        days = float(row[1] or 90)
        if days >= 120:   day_color = 'FFC7CE'
        elif days >= 90:  day_color = 'FFEB9C'
        elif days >= 60:  day_color = 'C6EFCE'
        else:             day_color = 'DDFFDD'

        c1 = ws_avg.cell(row=ch_row, column=1, value=row[0])
        c1.fill = hfill(bg); c1.border = bdr
        c1.font = Font(name='Calibri', size=9)
        c1.alignment = Alignment(horizontal='left')

        c2 = ws_avg.cell(row=ch_row, column=2, value=days)
        c2.fill = hfill(day_color); c2.border = bdr
        c2.font = Font(name='Calibri', size=9, bold=True)
        c2.number_format = '0.0'
        c2.alignment = Alignment(horizontal='center')

        c3 = ws_avg.cell(row=ch_row, column=3, value=int(row[2]))
        c3.fill = hfill(bg); c3.border = bdr
        c3.font = Font(name='Calibri', size=9)
        c3.alignment = Alignment(horizontal='center')

        ws_avg.row_dimensions[ch_row].height = 14
        ch_row += 1

    # Spacer + default row
    ws_avg.row_dimensions[ch_row].height = 8
    ch_row += 1
    def_cell = ws_avg.cell(row=ch_row, column=1, value='DEFAULT (no brand or chain match)')
    def_cell.font      = Font(name='Calibri', bold=True, size=9, color='FFFFFF')
    def_cell.fill      = hfill('808080')
    def_cell.border    = bdr
    def_cell.alignment = Alignment(horizontal='left')
    def2 = ws_avg.cell(row=ch_row, column=2, value=default_days)
    def2.font          = Font(name='Calibri', bold=True, size=9, color='FFFFFF')
    def2.fill          = hfill('808080')
    def2.border        = bdr
    def2.number_format = '0'
    def2.alignment     = Alignment(horizontal='center')
    def3 = ws_avg.cell(row=ch_row, column=3, value='—')
    def3.font          = Font(name='Calibri', bold=True, size=9, color='FFFFFF')
    def3.fill          = hfill('808080')
    def3.border        = bdr
    def3.alignment     = Alignment(horizontal='center')
    ws_avg.row_dimensions[ch_row].height = 16

    # Legend row
    ch_row += 2
    legend = [('< 60 days', 'DDFFDD'), ('60–89 days', 'C6EFCE'),
              ('90–119 days', 'FFEB9C'), ('120+ days', 'FFC7CE')]
    ws_avg.cell(row=ch_row, column=1, value='Legend:').font = Font(name='Calibri', bold=True, size=8)
    ws_avg.cell(row=ch_row, column=1).alignment = Alignment(horizontal='right')
    for li, (lbl, lclr) in enumerate(legend, 2):
        lc = ws_avg.cell(row=ch_row, column=li, value=lbl)
        lc.fill      = hfill(lclr)
        lc.font      = Font(name='Calibri', size=8)
        lc.alignment = Alignment(horizontal='center')
        lc.border    = Border(left=Side(style='thin', color='AAAAAA'),
                              right=Side(style='thin', color='AAAAAA'),
                              top=Side(style='thin', color='AAAAAA'),
                              bottom=Side(style='thin', color='AAAAAA'))
    ws_avg.row_dimensions[ch_row].height = 14

    # ── Avg Days charts ───────────────────────────────────────────────────────
    n_brands = len(brand_detail)
    n_chains = len(chain_detail)

    def _avg_bar_chart(ws, title, hdr_row, n_rows, anchor, color_hex, width=18, height=None):
        """Horizontal bar chart for avg days by brand or chain."""
        chart = BarChart()
        chart.type    = 'bar'   # horizontal
        chart.barDir  = 'bar'
        chart.title   = title
        chart.x_axis.title = 'Avg Days'
        chart.style   = 10
        chart.width   = width
        chart.height  = height or max(8, min(n_rows * 0.38, 22))
        chart.x_axis.numFmt = '0'
        chart.y_axis.majorGridlines = None
        chart.x_axis.majorGridlines = None
        chart.grouping = 'clustered'

        vals = Reference(ws, min_col=2, max_col=2,
                         min_row=hdr_row, max_row=hdr_row + n_rows)
        cats = Reference(ws, min_col=1, max_col=1,
                         min_row=hdr_row + 1, max_row=hdr_row + n_rows)
        chart.add_data(vals, titles_from_data=True)
        chart.set_categories(cats)
        chart.series[0].graphicalProperties.solidFill    = color_hex
        chart.series[0].graphicalProperties.line.solidFill = color_hex
        # Invert axis so slowest payers appear at top (already sorted desc)
        chart.y_axis.delete = False
        ws.add_chart(chart, anchor)
        return chart

    if n_brands > 0:
        _avg_bar_chart(ws_avg,
                       title  = 'Avg Days to Commission — by Brand  (6-month window)',
                       hdr_row= br_hdr_row,
                       n_rows = n_brands,
                       anchor = 'E4',
                       color_hex = '2E75B6')

    if n_chains > 0:
        # position chain chart to the right of (or below) brand chart
        chain_chart_anchor = f'E{ch_hdr_row}'
        _avg_bar_chart(ws_avg,
                       title  = 'Avg Days to Commission — by Chain  (6-month window)',
                       hdr_row= ch_hdr_row,
                       n_rows = n_chains,
                       anchor = chain_chart_anchor,
                       color_hex = '4472C4')

    # Column widths
    ws_avg.column_dimensions['A'].width = 42
    ws_avg.column_dimensions['B'].width = 16
    ws_avg.column_dimensions['C'].width = 12
    ws_avg.column_dimensions['D'].width = 14
    ws_avg.column_dimensions['E'].width = 14
    ws_avg.freeze_panes = 'A5'

    # ── Year detail sheets ────────────────────────────────────────────────────
    for year in years:
        for who_key, who_label, who_full in [
            ('kristin', 'Kristin', 'Kristin House'),
            ('team',    'Team',    'Team'),
        ]:
            rows = data[year][who_key]
            ws   = wb.create_sheet(f'{year} — {who_label}')
            is_team = (who_key == 'team')

            # Column definitions
            # fixed cols: Booking ID | ...text... | Timing Basis | Days | Revenue | Commission | % Comm | Jan...Dec | Year Total
            if is_team:
                fixed_hdr   = ['Booking ID', 'Customer', 'Team Member', 'Meeting Name', 'Hotel',
                                'Start Date', 'End Date', 'Paid / Proj Date',
                                'Chain', 'Days',
                                'Revenue', 'Commission', '% Comm']
                fixed_w     = [11, 30, 20, 35, 28, 12, 12, 14, 12, 7, 16, 14, 9]
            else:
                fixed_hdr   = ['Booking ID', 'Customer', 'Meeting Name', 'Hotel',
                                'Start Date', 'End Date', 'Paid / Proj Date',
                                'Chain', 'Days',
                                'Revenue', 'Commission', '% Comm']
                fixed_w     = [11, 30, 38, 28, 12, 12, 14, 12, 7, 16, 14, 9]

            nf                 = len(fixed_hdr)
            timing_basis_col_i = nf - 4     # 1-based index of Timing Basis column
            timing_days_col_i  = nf - 3     # 1-based index of Days column
            rev_col_i          = nf - 2     # 1-based index of Revenue column
            comm_col_i         = nf - 1     # 1-based index of Commission column
            pct_col_i          = nf         # 1-based index of % Comm column
            all_hdr     = fixed_hdr + MONTH_ABBR + ['Year Total']
            all_w       = fixed_w + [10]*12 + [14]
            ncols       = len(all_hdr)
            last_col    = get_column_letter(ncols)
            rev_col     = get_column_letter(rev_col_i)
            comm_col    = get_column_letter(comm_col_i)
            pct_col     = get_column_letter(pct_col_i)
            jan_col     = get_column_letter(nf + 1)
            dec_col     = get_column_letter(nf + 12)
            total_col   = get_column_letter(nf + 13)

            # ── Row 1: Title ──────────────────────────────────────────────────
            ws.merge_cells(f'A1:{last_col}1')
            tc = ws['A1']
            tc.value     = f'{who_full} — {year} Commission Proforma'
            tc.font      = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
            tc.fill      = hfill(C['title'])
            tc.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 26

            # ── Row 2: Subtitle ───────────────────────────────────────────────
            ws.merge_cells(f'A2:{last_col}2')
            sc = ws['A2']
            adj_sign = '+' if adj_pct >= 0 else ''
            sc.value = (
                f'Generated {today.strftime("%B %d, %Y")}  │  '
                f'Commission adjustment applied to projected: {adj_sign}{adj_pct:.1f}%  │  '
                f'Blue/White = paid · Green = HHR (Kristin) · Yellow = projected  │  '
                f'Kristin: HHR/contracted rev × comm% · Team: Pipeline rev × 1%  │  Adj: {adj_sign}{adj_pct:.1f}%'
            )
            sc.font      = Font(name='Calibri', italic=True, color='555555', size=9)
            sc.alignment = Alignment(horizontal='center')
            ws.row_dimensions[2].height = 14
            ws.row_dimensions[3].height = 5

            # ── Row 4: Column headers ─────────────────────────────────────────
            for ci, h in enumerate(all_hdr, 1):
                cell = ws.cell(row=4, column=ci, value=h)
                cell.font      = Font(name='Calibri', bold=True, color='FFFFFF', size=9)
                cell.fill      = hfill(C['header'])
                cell.alignment = Alignment(horizontal='center', wrap_text=True)
                cell.border    = bdr
            ws.row_dimensions[4].height = 32

            # ── Data rows ─────────────────────────────────────────────────────
            for ri, rd in enumerate(rows):
                rn      = ri + 5
                is_proj = rd['projected']
                src     = rd.get('src', 'paid')
                if src == 'overdue':
                    bg = C['overdue']
                elif src == 'hhr':
                    bg = C['hhr']
                elif src == 'calc':
                    bg = C['calc']
                else:
                    bg = C['actual_a'] if ri % 2 == 0 else C['actual_b']

                rev = rd.get('revenue', 0.0) or 0.0
                amt = rd['amount']
                t_basis = rd.get('days_src') or ''           # e.g. 'Brand', 'Chain', 'Default', or ''
                t_days  = rd.get('proj_days_used')           # int or None for paid rows
                if is_team:
                    fvals = [rd.get('bid',''), rd.get('customer',''), rd['associate'], rd['meeting'], rd['hotel'],
                             rd['start'], rd['end'],
                             rd['paid_date'].strftime('%m/%d/%Y'),
                             t_basis, t_days,
                             rev if rev else None, amt, None]   # None = % formula below
                else:
                    fvals = [rd.get('bid',''), rd.get('customer',''), rd['meeting'], rd['hotel'],
                             rd['start'], rd['end'],
                             rd['paid_date'].strftime('%m/%d/%Y'),
                             t_basis, t_days,
                             rev if rev else None, amt, None]   # None = % formula below

                monthly = [None] * 12
                mo = rd['month']
                if 1 <= mo <= 12:
                    monthly[mo - 1] = rd['amount']

                all_vals = fvals + monthly + [None]   # Year Total = formula

                for ci, v in enumerate(all_vals, 1):
                    cell = ws.cell(row=rn, column=ci, value=v)
                    cell.fill   = hfill(bg)
                    cell.border = bdr
                    cell.font   = Font(name='Calibri', size=9, italic=is_proj,
                                       color='555555' if is_proj else '000000')
                    if ci <= nf:
                        if ci == timing_basis_col_i:   # Timing Basis
                            cell.alignment = Alignment(horizontal='center')
                        elif ci == timing_days_col_i:  # Days
                            cell.number_format = '0'
                            cell.alignment = Alignment(horizontal='center')
                        elif ci < rev_col_i:   # text columns
                            # wrap Customer, Meeting Name, Hotel; not Booking ID or date/name cols
                            wrap = (2 <= ci <= (5 if is_team else 4))
                            cell.alignment = Alignment(horizontal='left', wrap_text=wrap)
                        elif ci == rev_col_i:   # Revenue
                            cell.number_format = '$#,##0'
                            cell.alignment = Alignment(horizontal='right')
                        elif ci == comm_col_i:   # Commission
                            cell.number_format = '$#,##0.00'
                            cell.alignment = Alignment(horizontal='right')
                        elif ci == pct_col_i:   # % Comm — insert formula
                            if rev and rev > 0:
                                cell.value         = f'={comm_col}{rn}/{rev_col}{rn}'
                                cell.number_format = '0.0%'
                            cell.alignment = Alignment(horizontal='right')
                    else:
                        cell.alignment    = Alignment(horizontal='right')
                        cell.number_format = '$#,##0.00'

                # Year Total formula
                yt = ws.cell(row=rn, column=ncols,
                             value=f'=SUM({jan_col}{rn}:{dec_col}{rn})')
                yt.fill          = hfill(bg)
                yt.border        = bdr
                yt.font          = Font(name='Calibri', size=9, italic=is_proj, bold=True,
                                        color='555555' if is_proj else '000000')
                yt.number_format = '$#,##0.00'
                yt.alignment     = Alignment(horizontal='right')

                ws.row_dimensions[rn].height = 14

            # ── Totals row ────────────────────────────────────────────────────
            tr = len(rows) + 5
            ws.cell(row=tr, column=1, value='TOTAL')
            for ci in range(1, ncols + 1):
                cell = ws.cell(row=tr, column=ci)
                cell.fill   = hfill(C['subtotal'])
                cell.border = bdr
                cell.font   = Font(name='Calibri', bold=True, size=9)
            ws.cell(row=tr, column=1).alignment = Alignment(horizontal='left')

            if rows:
                d_start = 5
                d_end   = tr - 1
                for ci in range(rev_col_i, ncols + 1):
                    col_ltr = get_column_letter(ci)
                    cell = ws.cell(row=tr, column=ci)
                    cell.fill      = hfill(C['subtotal'])
                    cell.border    = bdr
                    cell.font      = Font(name='Calibri', bold=True, size=9)
                    cell.alignment = Alignment(horizontal='right')
                    if ci == pct_col_i:   # % Comm total — blank
                        pass
                    else:
                        cell.value         = f'=SUM({col_ltr}{d_start}:{col_ltr}{d_end})'
                        cell.number_format = '$#,##0.00'
            ws.row_dimensions[tr].height = 16

            # ── No-data placeholder ───────────────────────────────────────────
            if not rows:
                ws.merge_cells(f'A5:{last_col}5')
                nc = ws['A5']
                nc.value     = 'No commission activity for this period.'
                nc.font      = Font(name='Calibri', italic=True, color='888888')
                nc.alignment = Alignment(horizontal='center')

            # ── Column widths & freeze ────────────────────────────────────────
            for ci, w in enumerate(all_w, 1):
                ws.column_dimensions[get_column_letter(ci)].width = w
            ws.freeze_panes = 'A5'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    adj_sign = 'p' if adj_pct >= 0 else 'm'
    fname = f'Commission_Proforma_{today.strftime("%Y%m%d")}_{adj_sign}{abs(int(adj_pct))}pct.xlsx'
    return send_file(buf, download_name=fname, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/reports/customer-summary', methods=['GET', 'POST'])
def report_customer_summary():
    user = get_current_user()
    if not has_permission(user, 'reports_customer_summary'):
        flash('You do not have access to the Customer Summary report.', 'error')
        return redirect(url_for('pipeline'))
    export = request.args.get('export', '').lower()
    conn   = get_db()

    customers = [r[0] for r in conn.execute(
        '''SELECT DISTINCT AccountName FROM ReportPipeline
           WHERE AccountName IS NOT NULL
             AND substr(StartDate,1,10) >= '2025-01-01'
           ORDER BY AccountName'''
    ).fetchall()]

    # Team member dropdown — filtered to 2025+, Kristin House pinned first
    raw_team_members = [r[0] for r in conn.execute(
        '''SELECT DISTINCT BookingAssociate FROM ReportPipeline
           WHERE BookingAssociate IS NOT NULL AND BookingAssociate != ''
             AND BookingAssociate NOT IN ('Unknown','Unkown','Housing Registration')
             AND substr(StartDate,1,10) >= '2025-01-01'
           ORDER BY BookingAssociate'''
    ).fetchall()]
    team_members = []
    if 'Kristin House' in raw_team_members:
        team_members.append('Kristin House')
    team_members += [tm for tm in raw_team_members if tm != 'Kristin House']

    # Build team-member → customers map for dynamic JS filtering
    tm_cust_rows = conn.execute(
        '''SELECT DISTINCT BookingAssociate, AccountName FROM ReportPipeline
           WHERE AccountName IS NOT NULL AND BookingAssociate IS NOT NULL
             AND BookingAssociate != ''
             AND BookingAssociate NOT IN ('Unknown','Unkown','Housing Registration')
             AND substr(StartDate,1,10) >= '2025-01-01'
           ORDER BY BookingAssociate, AccountName'''
    ).fetchall()
    tm_customers = {}
    for tm, acct in tm_cust_rows:
        tm_customers.setdefault(tm, []).append(acct)

    today = datetime.today()
    default_start = f'{today.year}-01-01'
    default_end   = f'{today.year}-12-31'

    if request.method == 'POST':
        selected_customers = request.form.getlist('customer')
    else:
        selected_customers = request.args.getlist('customer')
    selected_customers = [c.strip() for c in selected_customers if c.strip()]

    date_from   = (request.form.get('date_from')   or request.args.get('date_from',   default_start)).strip()
    date_to     = (request.form.get('date_to')     or request.args.get('date_to',     default_end)).strip()
    team_member = (request.form.get('team_member') or request.args.get('team_member', '')).strip()
    submitted   = request.method == 'POST' or bool(export)

    brand_rows, chain_groups, grand_total, grand_bookings = [], {}, 0.0, 0
    city_meetings = {}

    if request.method == 'POST' or export:
        brand_rows, chain_groups, grand_total, grand_bookings = _query_customer_summary(
            conn, selected_customers, date_from, date_to,
            customer_col='AccountName', status_col='BookingStatus',
            date_col='StartDate', contract_col='ContractedAmount',
            revenue_col='Revenue', chain_col='Chain', brand_col='Brand',
            city_col='City', state_col='State', cast_contract=False,
            team_member=team_member, team_member_col='BookingAssociate')
        city_meetings = _query_meeting_summary(
            conn, selected_customers, date_from, date_to,
            customer_col='AccountName', status_col='BookingStatus',
            date_col='StartDate', contract_col='ContractedAmount',
            revenue_col='Revenue', event_col='EventName',
            booking_col='BookingId', hotel_col='Customer',
            city_col='City', state_col='State', cast_contract=False,
            team_member=team_member, team_member_col='BookingAssociate')
        submitted = True

    # Build export filename/subtitle including team member when filtered
    tm_slug = f'_{team_member.replace(" ","_")}' if team_member else ''
    if len(selected_customers) == 1:
        cust_slug = selected_customers[0].replace(' ', '_')
    elif selected_customers:
        cust_slug = f'{len(selected_customers)}_Customers'
    else:
        cust_slug = 'All'

    city_map_data = build_city_map_data(city_meetings) if city_meetings else []

    if export == 'word' and brand_rows:
        buf = _build_word_doc(selected_customers, date_from, date_to, brand_rows, chain_groups,
                              grand_total, grand_bookings, city_meetings,
                              team_member=team_member)
        fname = f'Customer_Summary_{cust_slug}{tm_slug}_{date_from}_{date_to}.docx'
        return send_file(buf, download_name=fname, as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document')

    if export == 'pptx' and brand_rows:
        buf = _build_pptx(selected_customers, date_from, date_to, brand_rows, chain_groups,
                          grand_total, grand_bookings, city_meetings,
                          team_member=team_member, city_map_data=city_map_data)
        fname = f'Customer_Summary_{cust_slug}{tm_slug}_{date_from}_{date_to}.pptx'
        return send_file(buf, download_name=fname, as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.presentationml.presentation')

    brand_labels = [r['chain'] for r in brand_rows]
    brand_values = [r['revenue'] or 0 for r in brand_rows]
    chain_colors = [_chain_color(l, i) for i, l in enumerate(brand_labels)]

    city_rev = {}
    for data in chain_groups.values():
        for c in data['cities']:
            key = f"{c['city']}, {c['state']}" if c['state'] else c['city']
            city_rev[key] = city_rev.get(key, 0) + c['revenue']
    top_cities = sorted(city_rev.items(), key=lambda x: x[1], reverse=True)[:14]
    city_labels = [x[0] for x in top_cities]
    city_values = [x[1] for x in top_cities]
    city_count  = len(city_rev)

    return render_template('report_customer_summary.html',
        customers=customers, selected_customers=selected_customers,
        date_from=date_from, date_to=date_to, submitted=submitted,
        brand_rows=brand_rows, chain_groups=chain_groups,
        grand_total=grand_total, grand_bookings=grand_bookings,
        brand_labels=brand_labels, brand_values=brand_values,
        city_labels=city_labels, city_values=city_values,
        chain_colors=chain_colors, city_count=city_count,
        city_meetings=city_meetings, city_map_data=city_map_data,
        team_members=team_members, team_member=team_member,
        tm_customers=tm_customers,
    )

# ── Settings ──────────────────────────────────────────────────────────────────

@app.route('/settings', methods=['GET', 'POST'])
def settings():
    db = get_db()
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'save_default':
            split     = request.form.get('commission_split', '').strip()
            tolerance = request.form.get('payment_tolerance', '').strip()
            try:
                val = float(split.replace('%',''))
                if val > 1: val = val / 100
                tval = float(tolerance.replace('%',''))
                if tval > 1: tval = tval / 100
                db.execute('UPDATE Settings SET value=? WHERE key="commission_split"', (str(val),))
                db.execute('UPDATE Settings SET value=? WHERE key="payment_tolerance"', (str(tval),))
                db.commit()
                flash(f'Settings saved — split: {val*100:.1f}%, tolerance: {tval*100:.1f}%.', 'success')
            except Exception:
                flash('Invalid value — please enter numbers like 60 or 1.', 'error')
        elif action == 'save_kristin':
            ksplit = request.form.get('kristin_split', '').strip()
            kcut   = request.form.get('kristin_cut', '').strip()
            try:
                kval = float(ksplit.replace('%',''))
                if kval > 1: kval = kval / 100
                cval = float(kcut.replace('%',''))
                if cval > 1: cval = cval / 100
                db.execute('INSERT OR REPLACE INTO Settings (key, value) VALUES ("kristin_split", ?)', (str(kval),))
                db.execute('INSERT OR REPLACE INTO Settings (key, value) VALUES ("kristin_cut", ?)', (str(cval),))
                db.commit()
                flash(f'Kristin House settings saved — Kristin split: {kval*100:.0f}%, team cut: {cval*100:.0f}%.', 'success')
            except Exception:
                flash('Invalid value — please enter numbers like 70 or 10.', 'error')
        elif action == 'add_account':
            account   = request.form.get('account_name', '').strip()
            split     = request.form.get('account_split', '').strip()
            countries = request.form.get('account_countries', '').strip()
            try:
                val = float(split.replace('%',''))
                if val > 1: val = val / 100
                db.execute('INSERT INTO AccountSplits (account_name, split_rate, countries) VALUES (?,?,?)',
                           (account, val, countries.upper()))
                db.commit()
                note = f' ({countries.upper()})' if countries else ' (all countries)'
                flash(f'Account override added: {account}{note} → {val*100:.1f}%.', 'success')
            except Exception:
                flash('Invalid entry — check account name and split value.', 'error')
        elif action == 'delete_account':
            acct_id = request.form.get('acct_id')
            db.execute('DELETE FROM AccountSplits WHERE id = ?', (acct_id,))
            db.commit()
            flash('Account override removed.', 'success')
        elif action == 'save_date_format':
            fmt = request.form.get('date_format', '%m/%d/%Y')
            known = {o[0] for o in DATE_FORMAT_OPTIONS}
            if fmt not in known:
                flash('Unknown date format selected.', 'error')
            else:
                db.execute('INSERT OR REPLACE INTO Settings (key, value) VALUES ("date_format", ?)', (fmt,))
                db.commit()
                label = next((o[1] for o in DATE_FORMAT_OPTIONS if o[0] == fmt), fmt)
                flash(f'Date format updated to {label}.', 'success')
        elif action == 'save_user_profile':
            for key in ('user_full_name', 'user_email', 'user_phone'):
                val = request.form.get(key, '').strip()
                db.execute('INSERT OR REPLACE INTO Settings (key, value) VALUES (?, ?)', (key, val))
            db.commit()
            flash('User profile saved — used on Hotel Points forms.', 'success')
        elif action == 'save_my_loyalty_profile':
            current = get_current_user()
            if not current:
                flash('Please log in to save your loyalty profile.', 'error')
            else:
                uid = current['id']
                fields = {}
                for k in ('full_name', 'email', 'phone',
                          'marriott_number', 'hyatt_number', 'hilton_number',
                          'ihg_number', 'omni_number', 'choice_number', 'sonesta_number'):
                    fields[k] = request.form.get(k, '').strip() or None
                # Ensure a row exists
                db.execute('''INSERT OR IGNORE INTO user_loyalty_profile (user_id, full_name, email)
                              VALUES (?, ?, ?)''',
                           (uid, current['name'] or '', current['email'] or ''))
                sets = ', '.join(f'{k}=?' for k in fields)
                db.execute(f'''UPDATE user_loyalty_profile
                               SET {sets}, updated_at=datetime('now')
                               WHERE user_id=?''',
                           list(fields.values()) + [uid])
                db.commit()
                flash('Your loyalty info saved.', 'success')
        elif action == 'save_default_recipient':
            current = get_current_user()
            if not current or current['role'] != 'admin':
                flash('Only an administrator can change the default Hotel Points recipient.', 'error')
            else:
                new_uid = request.form.get('default_recipient_user_id', '').strip()
                try:
                    int(new_uid)  # validate
                except Exception:
                    flash('Invalid user selection.', 'error')
                else:
                    db.execute('INSERT OR REPLACE INTO Settings (key, value) VALUES '
                               '("hotel_points_default_recipient_user_id", ?)',
                               (new_uid,))
                    db.commit()
                    target = db.execute('SELECT name FROM Users WHERE id=?', (new_uid,)).fetchone()
                    flash(f'Default Hotel Points recipient set to {target["name"] if target else new_uid}.',
                          'success')
        return redirect(url_for('settings'))

    current       = get_commission_split()
    tolerance     = get_payment_tolerance()
    kristin_split = get_kristin_split()
    kristin_cut   = get_kristin_cut()
    current_date_fmt = get_date_format()
    today_preview    = datetime.today().strftime(current_date_fmt)
    overrides = db.execute('SELECT id, account_name, split_rate, countries FROM AccountSplits ORDER BY account_name').fetchall()
    accounts  = [r[0] for r in db.execute('SELECT DISTINCT AccountName FROM ReportPipeline WHERE AccountName IS NOT NULL ORDER BY AccountName').fetchall()]
    user_profile = {}
    for key in ('user_full_name', 'user_email', 'user_phone'):
        row = db.execute('SELECT value FROM Settings WHERE key=?', (key,)).fetchone()
        user_profile[key] = row[0] if row else ''

    # Hotel Points multi-user data
    current_user = get_current_user()
    my_loyalty   = None
    if current_user:
        my_loyalty = _get_user_loyalty_profile(db, current_user['id'])
    default_recipient_user_id = _get_default_recipient_user_id(db)
    default_recipient_user    = None
    if default_recipient_user_id:
        default_recipient_user = db.execute(
            'SELECT id, name, email FROM Users WHERE id=?',
            (default_recipient_user_id,)).fetchone()
    all_users = db.execute(
        'SELECT id, name, email FROM Users WHERE active=1 ORDER BY name'
    ).fetchall()

    return render_template('settings.html', commission_split=current, tolerance=tolerance,
                           kristin_split=kristin_split, kristin_cut=kristin_cut,
                           overrides=overrides, accounts=accounts,
                           date_format_options=DATE_FORMAT_OPTIONS,
                           current_date_fmt=current_date_fmt,
                           today_preview=today_preview,
                           user_profile=user_profile,
                           current_user=current_user,
                           my_loyalty=my_loyalty,
                           default_recipient_user=default_recipient_user,
                           default_recipient_user_id=default_recipient_user_id,
                           all_users=all_users)

# ── Summary ───────────────────────────────────────────────────────────────────

@app.route('/summary')
def summary():
    db = get_db()
    by_associate = db.execute('''
        SELECT BookingAssociate,
               COUNT(*) as Bookings,
               SUM(CASE WHEN BookingStatus='Definite' THEN 1 ELSE 0 END) as Definite,
               SUM(USDCommissionableAmount) as TotalCommission,
               SUM(USDRevenue) as TotalRevenue
        FROM ReportPipeline WHERE BookingAssociate IS NOT NULL
        GROUP BY BookingAssociate ORDER BY TotalCommission DESC
    ''').fetchall()
    by_status = db.execute('''
        SELECT BookingStatus, COUNT(*) as Count,
               SUM(USDCommissionableAmount) as TotalCommission,
               SUM(USDRevenue) as TotalRevenue
        FROM ReportPipeline GROUP BY BookingStatus ORDER BY Count DESC
    ''').fetchall()
    by_brand = db.execute('''
        SELECT Brand, COUNT(*) as Count, SUM(USDCommissionableAmount) as TotalCommission
        FROM ReportPipeline WHERE Brand IS NOT NULL
        GROUP BY Brand ORDER BY Count DESC LIMIT 20
    ''').fetchall()
    totals = db.execute('''
        SELECT COUNT(*) as Total,
               SUM(CASE WHEN BookingStatus='Definite' THEN 1 ELSE 0 END) as Definite,
               SUM(USDRevenue) as TotalRevenue,
               SUM(USDCommissionableAmount) as TotalCommission
        FROM ReportPipeline
    ''').fetchone()
    return render_template('summary.html', by_associate=by_associate, by_status=by_status,
                           by_brand=by_brand, totals=totals)

# ── Import Pickup Reports (HHR) ───────────────────────────────────────────────

def get_usd_exchange_rate(currency_code):
    """Fetch live USD exchange rate for a given 3-letter currency code."""
    import urllib.request, json as _json2
    currency_code = currency_code.upper()
    if currency_code == 'USD':
        return 1.0, 'USD'
    try:
        url = f'https://api.exchangerate-api.com/v4/latest/{currency_code}'
        with urllib.request.urlopen(url, timeout=5) as resp:
            data = _json2.loads(resp.read())
        rate = data['rates'].get('USD')
        if rate:
            return float(rate), 'live'
    except Exception:
        pass
    return None, None


CURRENCY_LABEL_MAP = {
    'usd': 'USD', 'us dollar': 'USD', 'dollar': 'USD',
    'euro': 'EUR', 'eur': 'EUR',
    'pound': 'GBP', 'gbp': 'GBP', 'british pound': 'GBP', 'sterling': 'GBP',
    'canadian': 'CAD', 'cad': 'CAD', 'canadian dollar': 'CAD',
    'australian': 'AUD', 'aud': 'AUD', 'australian dollar': 'AUD',
    'mexican': 'MXN', 'mxn': 'MXN', 'peso': 'MXN',
    'swiss': 'CHF', 'chf': 'CHF', 'franc': 'CHF',
    'japanese': 'JPY', 'jpy': 'JPY', 'yen': 'JPY',
}

REVENUE_ROW_KEYWORDS = [
    'TOTAL ROOMS COMMISSION DUE',
    'TOTAL ESTIMATED ROOMS COMMISSION DUE',   # Marriott variant with "ESTIMATED"
    'TOTAL ACCOMMODATION',
]


def parse_pickup_report(filepath):
    """Parse a TeamHouse Housing History Report xlsx and return extracted fields."""
    df = pd.read_excel(filepath, header=None)

    result = {
        'booking_id':       None,
        'hotel':            None,
        'organization':     None,
        'event_name':       None,
        'comm_pct':         None,
        'actual_pickup':    None,
        'pickup_by_night':  None,   # dict {"YYYY-MM-DD": rooms}
        'total_revenue':    None,
        'avg_rate':         None,
        'currency_code':    'USD',
        'currency_label':   'USD',
        'original_revenue': None,
        'exchange_rate':    None,
        'filename':         os.path.basename(filepath),
        'error':            None,
    }

    try:
        import re as _re2
        # Scan first 10 rows for Booking ID, Comm %, and Currency
        for row_idx in range(min(10, len(df))):
            row_raw = list(df.iloc[row_idx].values)
            row_str = [str(v) for v in row_raw]
            for i, cell in enumerate(row_str):
                if 'Booking' in cell and not result['booking_id']:
                    m = _re2.search(r'\d{5,7}', cell)
                    if m:
                        result['booking_id'] = m.group(0)
                    else:
                        for offset in (1, 2, 3):
                            if i + offset < len(row_raw):
                                val = row_raw[i + offset]
                                if val is None or (hasattr(val, '__float__') and str(val) == 'nan'):
                                    continue
                                try:
                                    bid = int(float(str(val)))
                                    if 10000 <= bid <= 9999999:
                                        result['booking_id'] = str(bid)
                                        break
                                except Exception:
                                    break
                if 'Comm %' in cell and not result['comm_pct']:
                    for offset in (1, 2, 3):
                        if i + offset < len(row_raw):
                            try:
                                v = float(row_raw[i + offset])
                                if not pd.isna(v) and v > 0:
                                    result['comm_pct'] = v
                                    break
                            except Exception:
                                pass
                if 'Currency:' in cell:
                    m = _re2.search(r'Currency:\s*(.+)', cell, _re2.IGNORECASE)
                    if m:
                        label = m.group(1).strip()
                        result['currency_label'] = label
                        iso = CURRENCY_LABEL_MAP.get(label.lower())
                        if iso:
                            result['currency_code'] = iso
            if result['booking_id'] and result['comm_pct']:
                break

        # Fallback: scan first 10 rows for a bare 6-7 digit integer
        if not result['booking_id']:
            for row_idx in range(min(10, len(df))):
                for val in df.iloc[row_idx].values:
                    if val is None or (hasattr(val, '__float__') and pd.isna(float(val)) if hasattr(val, '__float__') else False):
                        continue
                    try:
                        candidate = int(float(str(val)))
                        if 100000 <= candidate <= 9999999:
                            result['booking_id'] = str(candidate)
                            break
                    except Exception:
                        pass
                if result['booking_id']:
                    break

        # Row 1: Organization, Row 2: Hotel, Row 3: Event name
        for row_idx, key in [(1, 'organization'), (2, 'hotel'), (3, 'event_name')]:
            row_vals = df.iloc[row_idx].dropna().tolist()
            if len(row_vals) >= 2:
                result[key] = str(row_vals[1]).strip()

        # Find DATE header row → map column index to ISO date
        date_col_map = {}
        for idx, row in df.iterrows():
            row_vals = list(row.values)
            if str(row_vals[0]).strip().upper() == 'DATE':
                for ci, val in enumerate(row_vals):
                    if hasattr(val, 'strftime'):
                        date_col_map[ci] = val.strftime('%Y-%m-%d')
                break

        # Find FINAL TOTAL PICKUP row → total rooms + night-by-night breakdown
        for idx, row in df.iterrows():
            row_str_joined = ' '.join(str(v) for v in row.values if pd.notna(v))
            if 'FINAL TOTAL PICKUP' in row_str_joined.upper():
                nums = []
                pickup_by_night = {}
                for ci, v in enumerate(row.values):
                    try:
                        f = float(v)
                        if not pd.isna(f):
                            nums.append(f)
                            if ci in date_col_map and f > 0:
                                pickup_by_night[date_col_map[ci]] = int(f)
                    except Exception:
                        pass
                if nums:
                    result['actual_pickup'] = float(nums[-1])
                if pickup_by_night:
                    result['pickup_by_night'] = pickup_by_night
                break

        # Find TOTAL ROOMS / TOTAL ACCOMMODATION COMMISSION DUE row → revenue and avg rate
        for idx, row in df.iterrows():
            row_str_joined = ' '.join(str(v) for v in row.values if pd.notna(v)).upper()
            if any(kw in row_str_joined for kw in REVENUE_ROW_KEYWORDS) and 'COMMISSION DUE' in row_str_joined:
                nums = []
                found_label = False
                for v in row.values:
                    v_up = str(v).upper() if pd.notna(v) else ''
                    if pd.notna(v) and any(kw in v_up for kw in REVENUE_ROW_KEYWORDS):
                        found_label = True
                        continue
                    if found_label and pd.notna(v):
                        try:
                            nums.append(float(v))
                        except Exception:
                            pass
                if len(nums) >= 2:
                    result['total_revenue'] = nums[1]
                elif len(nums) == 1 and nums[0] > 10000:
                    # Some formats put only the commission total here; skip — use fallback below
                    pass
                if len(nums) >= 3:
                    result['avg_rate'] = nums[2]
                break

        # Fallback: scan for "Total Actualized Pickup Revenue" (Renaissance-style HHR format)
        # This row contains the actual room revenue when the commission row doesn't have it.
        if not result['total_revenue']:
            for idx, row in df.iterrows():
                row_str_joined = ' '.join(str(v) for v in row.values if pd.notna(v)).upper()
                if 'TOTAL ACTUALIZED PICKUP REVENUE' in row_str_joined or \
                   ('ACTUALIZED' in row_str_joined and 'REVENUE' in row_str_joined and 'PICKUP' in row_str_joined):
                    rev_nums = []
                    for v in row.values:
                        if pd.notna(v):
                            try:
                                f = float(v)
                                if not pd.isna(f) and f > 0:
                                    rev_nums.append(f)
                            except Exception:
                                pass
                    if rev_nums:
                        # Take the largest number as the revenue figure
                        result['total_revenue'] = max(rev_nums)
                    break

        # Fallback: derive avg_rate from revenue ÷ actual_pickup when not directly found
        if result['total_revenue'] and result['actual_pickup'] and not result['avg_rate']:
            try:
                result['avg_rate'] = round(float(result['total_revenue']) / float(result['actual_pickup']), 2)
            except Exception:
                pass

        # Currency conversion: if not USD, fetch live rate and convert revenue to USD
        if result['currency_code'] != 'USD' and result['total_revenue']:
            rate, source = get_usd_exchange_rate(result['currency_code'])
            if rate:
                result['original_revenue'] = result['total_revenue']
                result['exchange_rate']    = rate
                result['total_revenue']    = round(result['total_revenue'] * rate, 2)
                if result['avg_rate']:
                    result['avg_rate'] = round(result['avg_rate'] * rate, 2)
            else:
                result['error'] = (
                    f"Revenue is in {result['currency_label']} but live exchange rate "
                    f"could not be fetched. Please convert manually."
                )

        if not result['booking_id']:
            result['error'] = 'Could not find Booking ID in file'

    except Exception as e:
        result['error'] = str(e)

    return result


@app.route('/import/hhr', methods=['GET', 'POST'])
def import_hhr():
    user = get_current_user()
    if not has_permission(user, 'import_hhr'):
        flash('You do not have access to this import tool.', 'error')
        return redirect(url_for('pipeline'))
    if request.method == 'POST':
        files = request.files.getlist('files')
        if not files or not any(f.filename for f in files):
            flash('No files selected.', 'error')
            return redirect(url_for('import_hhr'))

        db = get_db()
        results = []

        for file in files:
            if not file.filename:
                continue
            ext = file.filename.rsplit('.', 1)[-1].lower()
            if ext not in ('xls', 'xlsx'):
                results.append({'filename': file.filename, 'status': 'skipped',
                                'reason': 'Not an Excel file'})
                continue

            tmp_path = f'/tmp/cpainc_pickup_{file.filename}'
            file.save(tmp_path)
            file_bytes = open(tmp_path, 'rb').read()

            parsed = parse_pickup_report(tmp_path)

            if parsed['error']:
                results.append({'filename': file.filename, 'status': 'error',
                                'reason': parsed['error']})
                continue

            # Reject files with no pickup or revenue data
            has_pickup  = parsed['actual_pickup'] and float(parsed['actual_pickup']) > 0
            has_revenue = parsed['total_revenue'] and float(parsed['total_revenue']) > 0
            if not has_pickup and not has_revenue:
                results.append({'filename': file.filename, 'status': 'error',
                                'booking_id': str(parsed['booking_id']).strip().split('.')[0] if parsed['booking_id'] else None,
                                'reason': 'File contains no pickup rooms or revenue — nothing to import'})
                continue
            if not has_pickup:
                results.append({'filename': file.filename, 'status': 'error',
                                'booking_id': str(parsed['booking_id']).strip().split('.')[0] if parsed['booking_id'] else None,
                                'reason': 'File contains no actual pickup room nights — nothing to import'})
                continue
            if not has_revenue:
                results.append({'filename': file.filename, 'status': 'error',
                                'booking_id': str(parsed['booking_id']).strip().split('.')[0] if parsed['booking_id'] else None,
                                'reason': 'File contains no revenue total — nothing to import'})
                continue

            bid = str(parsed['booking_id']).strip().split('.')[0]

            # Check if booking exists
            booking = db.execute(
                'SELECT BookingId, EventName, Customer FROM ReportPipeline WHERE BookingId = ?', (bid,)
            ).fetchone()

            if not booking:
                from urllib.parse import urlencode
                qs = urlencode({
                    'BookingId':         bid,
                    'EventName':         parsed.get('event_name') or '',
                    'Customer':          parsed.get('hotel') or '',
                    'AccountName':       parsed.get('organization') or '',
                    'CommissionPercent': parsed.get('comm_pct') or '',
                })
                results.append({'filename': file.filename, 'status': 'no_booking',
                                'reason': f'Booking ID {bid} not found in database — click Create Booking to add it, then re-import.',
                                'booking_id': bid, 'create_url': f'/booking/new?{qs}'})
                continue

            # Create/update Pickup record
            existing = db.execute('SELECT ID FROM Pickup WHERE BookingID = ?', (bid,)).fetchone()
            if existing:
                db.execute('''UPDATE Pickup SET ActualPickup=?, TotalRevenue=?, Brand=?
                    WHERE ID=?''',
                    (parsed['actual_pickup'], parsed['total_revenue'],
                     parsed['hotel'] or '', existing['ID']))
                action = 'updated'
            else:
                db.execute('''INSERT INTO Pickup (BookingID, ActualPickup, TotalRevenue, Brand)
                    VALUES (?,?,?,?)''',
                    (bid, parsed['actual_pickup'], parsed['total_revenue'], parsed['hotel'] or ''))
                action = 'imported'

            # Update RoomRate on the booking if avg_rate available
            if parsed['avg_rate']:
                db.execute('UPDATE ReportPipeline SET RoomRate=? WHERE BookingId=?',
                           (parsed['avg_rate'], bid))

            # Archive the raw file
            db.execute(
                'INSERT INTO housing_history_files (booking_id, filename, file_data) VALUES (?,?,?)',
                (bid, file.filename, file_bytes)
            )

            # Create a "Final History" pickup_weekly entry if pickup_config exists
            final_history_result = None
            if parsed.get('pickup_by_night'):
                pc = db.execute(
                    "SELECT * FROM pickup_config WHERE booking_id=? AND status='active'", (bid,)
                ).fetchone()
                if pc:
                    existing_fh = db.execute(
                        "SELECT id FROM pickup_weekly WHERE config_id=? AND label='Final History'",
                        (pc['id'],)
                    ).fetchone()
                    if not existing_fh:
                        pbn       = parsed['pickup_by_night']
                        total_rms = sum(pbn.values())
                        block     = json.loads(pc['contracted_block'] or '{}')
                        total_blk = sum(block.values())
                        pct_blk   = round(total_rms / total_blk * 100, 1) if total_blk else None
                        atr_pct   = pc['attrition_pct']
                        atr_floor = round(total_blk * float(atr_pct), 1) if atr_pct and total_blk else None
                        pct_atr   = round(total_rms / atr_floor * 100, 1) if atr_floor else None
                        prev = db.execute(
                            "SELECT total_rooms FROM pickup_weekly "
                            "WHERE config_id=? AND label IS NULL "
                            "ORDER BY report_date DESC, id DESC LIMIT 1",
                            (pc['id'],)
                        ).fetchone()
                        change = (total_rms - prev['total_rooms']) if prev and prev['total_rooms'] else None
                        db.execute(
                            '''INSERT INTO pickup_weekly
                               (config_id, report_date, pickup_by_night, total_rooms,
                                change_from_last, pct_of_block, pct_of_attrition, label)
                               VALUES (?,?,?,?,?,?,?,?)''',
                            (pc['id'], datetime.now().strftime('%Y-%m-%d'),
                             json.dumps(pbn), total_rms, change, pct_blk, pct_atr, 'Final History')
                        )
                        final_history_result = {
                            'config_id': pc['id'],
                            'event':     pc['event_name'] or pc['organization'],
                            'total':     total_rms,
                            'pct_blk':   pct_blk,
                        }
                    else:
                        final_history_result = {'skipped': True}

            results.append({
                'filename':         file.filename,
                'status':           action,
                'booking_id':       bid,
                'event':            booking['EventName'],
                'hotel':            booking['Customer'],
                'pickup':           parsed['actual_pickup'],
                'revenue':          parsed['total_revenue'],
                'avg_rate':         parsed['avg_rate'],
                'currency_code':    parsed['currency_code'],
                'currency_label':   parsed['currency_label'],
                'original_revenue': parsed['original_revenue'],
                'exchange_rate':    parsed['exchange_rate'],
                'final_history':    final_history_result,
            })

        db.commit()
        return render_template('import_hhr_result.html', results=results)

    return render_template('import_hhr.html')


# ── Pickup Tracking Module ─────────────────────────────────────────────────────

def ensure_pickup_tables():
    db = get_db()
    db.executescript('''
        CREATE TABLE IF NOT EXISTS pickup_config (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            booking_id TEXT,
            tab_name TEXT,
            organization TEXT NOT NULL,
            event_name TEXT,
            hotel TEXT,
            hotel_contact TEXT,
            hotel_contact_email TEXT,
            group_contact TEXT,
            group_contact_email TEXT,
            cutoff_date TEXT,
            attrition_pct REAL,
            contracted_block TEXT NOT NULL DEFAULT '{}',
            contracted_rate REAL,
            shoulder_pre INTEGER DEFAULT 3,
            shoulder_post INTEGER DEFAULT 3,
            hotel_booking_link TEXT,
            status TEXT DEFAULT 'active',
            notes TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS pickup_weekly (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            config_id INTEGER NOT NULL REFERENCES pickup_config(id) ON DELETE CASCADE,
            report_date TEXT NOT NULL,
            pickup_by_night TEXT NOT NULL DEFAULT '{}',
            total_rooms INTEGER,
            change_from_last INTEGER,
            pct_of_block REAL,
            pct_of_attrition REAL,
            ota_rate REAL,
            label TEXT,
            notes TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS pickup_rooming_list (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            config_id INTEGER NOT NULL REFERENCES pickup_config(id) ON DELETE CASCADE,
            weekly_id INTEGER REFERENCES pickup_weekly(id),
            upload_date TEXT DEFAULT (datetime('now')),
            filename TEXT,
            file_data BLOB,
            total_guests INTEGER,
            nights_by_date TEXT,
            reconciliation_status TEXT DEFAULT 'pending',
            discrepancy_notes TEXT,
            guests_json TEXT
        );
        CREATE TABLE IF NOT EXISTS pickup_contact_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            config_id INTEGER NOT NULL REFERENCES pickup_config(id) ON DELETE CASCADE,
            contact_date TEXT NOT NULL,
            contact_type TEXT NOT NULL DEFAULT 'email_sent',
            notes TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS booking_contract (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            booking_id TEXT NOT NULL,
            filename TEXT NOT NULL,
            file_data BLOB NOT NULL,
            upload_date TEXT DEFAULT (date('now'))
        );
        CREATE TABLE IF NOT EXISTS housing_history_files (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            booking_id TEXT NOT NULL,
            filename TEXT NOT NULL,
            file_data BLOB NOT NULL,
            upload_date TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS client_contacts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            email TEXT NOT NULL,
            organization TEXT,
            created_at TEXT,
            updated_at TEXT,
            UNIQUE(email)
        );
        CREATE TABLE IF NOT EXISTS hotel_contact_list (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            email TEXT NOT NULL,
            hotel_name TEXT,
            created_at TEXT,
            updated_at TEXT,
            UNIQUE(email)
        );
        CREATE TABLE IF NOT EXISTS contract_template (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            chain TEXT NOT NULL,
            template_name TEXT NOT NULL,
            description TEXT,
            filename TEXT,
            file_data BLOB,
            sections TEXT DEFAULT '[]',
            merge_fields TEXT DEFAULT '[]',
            is_active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT (datetime('now')),
            updated_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS rfp (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            rfp_code TEXT,
            rfp_name TEXT,
            client_org TEXT NOT NULL,
            event_name TEXT,
            start_date TEXT,
            end_date TEXT,
            alt_start_date TEXT,
            alt_end_date TEXT,
            peak_rooms INTEGER,
            total_room_nights INTEGER,
            f_and_b_budget REAL,
            total_attendees INTEGER,
            response_due_date TEXT,
            decision_due_date TEXT,
            status TEXT DEFAULT 'sourcing',
            booking_id TEXT,
            notes TEXT,
            rfp_filename TEXT,
            rfp_data BLOB,
            contract_filename TEXT,
            contract_data BLOB,
            archived INTEGER DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now')),
            updated_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS rfp_hotel (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            rfp_id INTEGER NOT NULL REFERENCES rfp(id) ON DELETE CASCADE,
            hotel_name TEXT NOT NULL,
            brand TEXT, city TEXT, state TEXT,
            contact_name TEXT, contact_email TEXT,
            contact_phone TEXT, contact_title TEXT,
            status TEXT DEFAULT 'pending',
            proposed_rate REAL,
            commission_pct REAL,
            f_and_b_minimum REAL,
            meeting_room_rental REAL,
            attrition_pct REAL,
            cutoff_days INTEGER,
            concessions TEXT,
            proposal_filename TEXT,
            proposal_data BLOB,
            notes TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            updated_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS rfp_note (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            rfp_id INTEGER NOT NULL REFERENCES rfp(id) ON DELETE CASCADE,
            rfp_hotel_id INTEGER REFERENCES rfp_hotel(id) ON DELETE SET NULL,
            note_date TEXT NOT NULL,
            note_type TEXT DEFAULT 'internal',
            note_text TEXT NOT NULL,
            created_at TEXT DEFAULT (datetime('now'))
        );
    ''')
    db.commit()
    for _col, _typ in [('ota_url', 'TEXT'), ('cc_emails', "TEXT DEFAULT '[]'"),
                        ('force_current', 'INTEGER DEFAULT 0'),
                        ('hotel_contacts', "TEXT DEFAULT '[]'"),
                        ('force_past', 'INTEGER DEFAULT 0')]:
        try:
            db.execute(f'ALTER TABLE pickup_config ADD COLUMN {_col} {_typ}')
            db.commit()
        except Exception:
            pass
    # Add HHR (Housing History Report) columns to pickup_config
    for _col, _typ in [('hhr_filename', 'TEXT'), ('hhr_data', 'BLOB'), ('hhr_stats', 'TEXT')]:
        try:
            db.execute(f'ALTER TABLE pickup_config ADD COLUMN {_col} {_typ}')
            db.commit()
        except Exception:
            pass
    # Add contract storage + block_is_estimated columns to pickup_config
    for _col, _typ in [('contract_filename', 'TEXT'), ('contract_data', 'BLOB'),
                       ('block_is_estimated', 'INTEGER DEFAULT 0')]:
        try:
            db.execute(f'ALTER TABLE pickup_config ADD COLUMN {_col} {_typ}')
            db.commit()
        except Exception:
            pass
    # Add event_start / event_end date columns to pickup_config
    for _col in ('event_start', 'event_end'):
        try:
            db.execute(f'ALTER TABLE pickup_config ADD COLUMN {_col} TEXT')
            db.commit()
        except Exception:
            pass
    # Add rooming_list_required flag to pickup_config
    try:
        db.execute('ALTER TABLE pickup_config ADD COLUMN rooming_list_required INTEGER DEFAULT 0')
        db.commit()
    except Exception:
        pass
    # Add secondary hotel contact columns to pickup_config
    for _col in ('hotel_contact2', 'hotel_contact2_email'):
        try:
            db.execute(f'ALTER TABLE pickup_config ADD COLUMN {_col} TEXT')
            db.commit()
        except Exception:
            pass
    # Add per-week OTA rate + URL to pickup_weekly
    for _col, _typ in [('ota_rate', 'REAL'), ('ota_url', 'TEXT')]:
        try:
            db.execute(f'ALTER TABLE pickup_weekly ADD COLUMN {_col} {_typ}')
            db.commit()
        except Exception:
            pass
    # Add CRF columns and alt date columns to rfp/rfp_hotel if not present
    for _tbl, _col, _typ in [
        ('rfp',       'crf_filename',    'TEXT'),
        ('rfp',       'crf_data',        'BLOB'),
        ('rfp_hotel', 'crf_row_data',    'TEXT'),
        ('rfp_hotel', 'crf_version',     'INTEGER DEFAULT 0'),
        ('rfp',       'alt_start_date_2','TEXT'),
        ('rfp',       'alt_end_date_2',  'TEXT'),
        ('rfp',       'alt_start_date_3','TEXT'),
        ('rfp',       'alt_end_date_3',  'TEXT'),
    ]:
        try:
            db.execute(f'ALTER TABLE {_tbl} ADD COLUMN {_col} {_typ}')
            db.commit()
        except Exception:
            pass
    # Add rebate_per_room to pickup_config if not present
    try:
        db.execute('ALTER TABLE pickup_config ADD COLUMN rebate_per_room REAL')
        db.commit()
    except Exception:
        pass
    # Add BookingName to ReportPipeline if not present
    try:
        db.execute('ALTER TABLE ReportPipeline ADD COLUMN BookingName TEXT')
        db.commit()
    except Exception:
        pass
    # Add EventID to ReportPipeline if not present
    try:
        db.execute('ALTER TABLE ReportPipeline ADD COLUMN EventID TEXT')
        db.commit()
    except Exception:
        pass
    # Add block_review_date to pickup_config if not present
    try:
        db.execute('ALTER TABLE pickup_config ADD COLUMN block_review_date TEXT')
        db.commit()
    except Exception:
        pass
    # Add critical dates columns to rfp if not present
    for _col, _typ in [('critical_dates_json', 'TEXT'), ('critical_dates_sent_at', 'TEXT'),
                       ('checklist', "TEXT DEFAULT '{}'")]:
        try:
            db.execute(f'ALTER TABLE rfp ADD COLUMN {_col} {_typ}')
            db.commit()
        except Exception:
            pass

    _ensure_hotel_points_tables(db)
    _ensure_cost_savings_tables(db)


def _ensure_cost_savings_tables(db):
    """Create cost_savings_report + cost_savings_item tables."""
    db.executescript('''
        CREATE TABLE IF NOT EXISTS cost_savings_report (
            id                      INTEGER PRIMARY KEY AUTOINCREMENT,
            rfp_hotel_id            INTEGER NOT NULL,
            meeting_name            TEXT,
            hotel_name              TEXT,
            meeting_dates           TEXT,
            lead_date_requested     TEXT,
            booked_date             TEXT,
            hours_worked            REAL DEFAULT 0,
            gr_rack_rate            REAL,
            gr_contracted_rate      REAL,
            gr_total_nights         INTEGER,
            gr_notes                TEXT,
            sr_group_rate           REAL,
            sr_contracted_rate      REAL,
            sr_total_nights         INTEGER,
            comp_industry_standard  INTEGER DEFAULT 50,
            comp_negotiated_policy  INTEGER,
            mr_initial              REAL,
            mr_negotiated           REAL,
            mr_notes                TEXT,
            fb_initial              REAL,
            fb_negotiated           REAL,
            fb_notes                TEXT,
            hotel_brand             TEXT,
            status                  TEXT DEFAULT 'draft',
            proposal_extracted_at   TEXT,
            contract_extracted_at   TEXT,
            created_at              TEXT DEFAULT (datetime('now')),
            updated_at              TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (rfp_hotel_id) REFERENCES rfp_hotel(id)
        );

        CREATE TABLE IF NOT EXISTS cost_savings_item (
            id                INTEGER PRIMARY KEY AUTOINCREMENT,
            report_id         INTEGER NOT NULL,
            sort_order        INTEGER DEFAULT 0,
            item_name         TEXT NOT NULL,
            calc_type         TEXT NOT NULL DEFAULT 'simple',
            standard_price    REAL,
            negotiated_price  REAL,
            quantity          REAL,
            cost_savings      REAL,
            notes             TEXT,
            FOREIGN KEY (report_id) REFERENCES cost_savings_report(id) ON DELETE CASCADE
        );

        CREATE INDEX IF NOT EXISTS idx_csr_rfp_hotel ON cost_savings_report(rfp_hotel_id);
        CREATE INDEX IF NOT EXISTS idx_csi_report    ON cost_savings_item(report_id);
    ''')
    db.commit()

    # pickup_config.contract_extracted_data — cache for the richer extractor.
    try:
        db.execute('ALTER TABLE pickup_config ADD COLUMN contract_extracted_data TEXT')
        db.commit()
    except Exception:
        pass

    # Backfill: add Travel Expense rows to any existing report that lacks them.
    travel_items = [
        ('Travel Expenses – Site Visits',
         'Enter total travel cost as Standard, 0 in Negotiated if hotel covered'),
        ('Travel Expenses – Attending Meeting Dates',
         'Enter total travel cost as Standard, 0 in Negotiated if hotel covered'),
    ]
    reports = db.execute('SELECT id FROM cost_savings_report').fetchall()
    for r in reports:
        existing_names = {row[0] for row in db.execute(
            'SELECT item_name FROM cost_savings_item WHERE report_id=?', (r[0],)
        ).fetchall()}
        max_order = db.execute(
            'SELECT COALESCE(MAX(sort_order), -1) FROM cost_savings_item WHERE report_id=?',
            (r[0],)
        ).fetchone()[0]
        for name, note in travel_items:
            if name in existing_names:
                continue
            max_order += 1
            db.execute('''INSERT INTO cost_savings_item
                (report_id, sort_order, item_name, calc_type,
                 standard_price, negotiated_price, quantity, notes)
                VALUES (?, ?, ?, 'simple', 0, 0, 1, ?)''',
                (r[0], max_order, name, note))
    db.commit()


# Default seed items inserted when a Cost Savings Report is first created.
COST_SAVINGS_SEED_ITEMS = [
    {'item_name': 'Attrition Savings vs 90% Baseline',               'calc_type': 'attrition',  'quantity': 0.70},
    {'item_name': 'Complimentary Internet in Guest Rooms',           'calc_type': 'simple',     'standard_price': 12.95, 'negotiated_price': 0},
    {'item_name': 'Complimentary Internet in Meeting Space',         'calc_type': 'simple',     'standard_price': 15.00, 'negotiated_price': 0},
    {'item_name': 'Suite upgrades per night at group rate',          'calc_type': 'simple',     'standard_price': 150.00, 'negotiated_price': 0},
    {'item_name': 'VIP Amenities',                                   'calc_type': 'simple',     'standard_price': 50.00, 'negotiated_price': 0},
    {'item_name': 'Complimentary Overnight Valet Vouchers',          'calc_type': 'simple',     'standard_price': 59.00, 'negotiated_price': 0},
    {'item_name': '15% Discount on A/V',                             'calc_type': 'note_only',  'notes': 'TBD'},
    {'item_name': 'Group may bring in own laptops and projectors',   'calc_type': 'note_only',  'notes': 'TBD'},
    {'item_name': 'Reduced valet parking fee',                       'calc_type': 'note_only',  'notes': 'TBD'},
    {'item_name': 'Complimentary easels and podiums',                'calc_type': 'note_only',  'notes': 'TBD'},
    {'item_name': 'Complimentary meeting planner room',              'calc_type': 'simple',     'standard_price': 0, 'negotiated_price': 0},
    {'item_name': 'No Resort Fees or Destination Fees',              'calc_type': 'simple',     'standard_price': 0, 'negotiated_price': 0},
    {'item_name': 'Waived storage and handling fees',                'calc_type': 'note_only',  'notes': 'TBD'},
    {'item_name': 'No deposit required',                             'calc_type': 'note_only',  'notes': 'TBD'},
    {'item_name': 'Travel Expenses – Site Visits',                   'calc_type': 'simple',     'standard_price': 0, 'negotiated_price': 0, 'quantity': 1, 'notes': 'Enter total travel cost as Standard, 0 in Negotiated if hotel covered'},
    {'item_name': 'Travel Expenses – Attending Meeting Dates',       'calc_type': 'simple',     'standard_price': 0, 'negotiated_price': 0, 'quantity': 1, 'notes': 'Enter total travel cost as Standard, 0 in Negotiated if hotel covered'},
    {'item_name': 'Meeting Planner Points',                          'calc_type': 'points'},
]

PLANNER_POINTS_BY_BRAND = {
    'Hyatt':     0.018,
    'Hilton':    0.005,
    'Marriott':  0.008,
    'IHG':       0.007,
    'Preferred': 0.005,
}


def _compute_report_totals(report, items):
    rack = (report['gr_rack_rate'] or 0)
    gr_contr = (report['gr_contracted_rate'] or 0)
    gr_nights = (report['gr_total_nights'] or 0)
    gr_savings_per = rack - gr_contr
    gr_total = gr_savings_per * gr_nights

    sr_savings_per = (report['sr_group_rate'] or 0) - (report['sr_contracted_rate'] or 0)
    sr_total = sr_savings_per * (report['sr_total_nights'] or 0)

    comp_negotiated = (report['comp_negotiated_policy'] or 0)
    comp_total = (gr_nights / comp_negotiated) * gr_contr if comp_negotiated else 0

    mr_savings = (report['mr_initial'] or 0) - (report['mr_negotiated'] or 0)
    fb_savings = (report['fb_initial'] or 0) - (report['fb_negotiated'] or 0)

    points_dollar = PLANNER_POINTS_BY_BRAND.get(report['hotel_brand'] or 'Preferred', 0.005)
    points_base = (gr_nights * gr_contr) + (report['mr_negotiated'] or 0)

    item_total = 0
    item_calcs = []
    for it in items:
        ct = it['calc_type']
        b = (it['standard_price'] or 0)
        c = (it['negotiated_price'] or 0)
        d = (it['quantity'] or 0)
        if ct == 'attrition':
            saving = gr_contr * gr_nights * max(0.90 - d, 0)
        elif ct == 'points':
            saving = (points_dollar - c) * points_base
        elif ct == 'note_only':
            saving = 0
        else:
            saving = (b - c) * d
        item_calcs.append({'id': it['id'], 'savings': saving})
        item_total += saving

    grand = gr_total + sr_total + comp_total + mr_savings + fb_savings + item_total
    return {
        'gr_savings_per': gr_savings_per, 'gr_total': gr_total,
        'sr_savings_per': sr_savings_per, 'sr_total': sr_total,
        'comp_total': comp_total,
        'mr_savings': mr_savings, 'fb_savings': fb_savings,
        'item_total': item_total, 'item_calcs': item_calcs,
        'grand_total': grand,
        'points_dollar': points_dollar, 'points_base': points_base,
    }


def _detect_brand_from_hotel(hotel_name):
    h = (hotel_name or '').lower()
    if 'hyatt' in h: return 'Hyatt'
    if 'hilton' in h or 'embassy suites' in h or 'doubletree' in h or 'hampton' in h or 'waldorf' in h or 'conrad' in h or 'curio' in h:
        return 'Hilton'
    if 'marriott' in h or 'ritz' in h or 'sheraton' in h or 'westin' in h or 'renaissance' in h or 'jw ' in h or 'courtyard' in h or 'gaylord' in h:
        return 'Marriott'
    if 'intercontinental' in h or 'crowne plaza' in h or 'holiday inn' in h or 'kimpton' in h or 'staybridge' in h:
        return 'IHG'
    return 'Preferred'


def _ensure_hotel_points_tables(db):
    """Create hotel_points_program + hotel_points_request tables and seed chains."""
    db.executescript('''
        CREATE TABLE IF NOT EXISTS hotel_points_program (
            id                      INTEGER PRIMARY KEY AUTOINCREMENT,
            chain_name              TEXT UNIQUE NOT NULL,
            member_number           TEXT,
            submission_type         TEXT NOT NULL DEFAULT 'manual',
            form_template_data      BLOB,
            form_template_filename  TEXT,
            form_url                TEXT,
            field_mapping_json      TEXT DEFAULT '{}',
            submission_window_days  INTEGER DEFAULT 90,
            receipt_window_days     INTEGER DEFAULT 60,
            notes                   TEXT,
            active                  INTEGER DEFAULT 1,
            created_at              TEXT DEFAULT (datetime('now')),
            updated_at              TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS hotel_points_request (
            id                          INTEGER PRIMARY KEY AUTOINCREMENT,
            pickup_config_id            INTEGER,
            program_id                  INTEGER NOT NULL REFERENCES hotel_points_program(id),
            booking_id                  TEXT,
            form_generated_at           TEXT,
            form_sent_date              TEXT,
            sent_to_name                TEXT,
            sent_to_email               TEXT,
            generated_doc_data          BLOB,
            generated_doc_filename      TEXT,
            points_received_date        TEXT,
            points_awarded              INTEGER,
            status                      TEXT DEFAULT 'pending',
            cvent_rfp_code              TEXT,
            contract_signature_date     TEXT,
            incentive_type              TEXT,
            award_timing                TEXT,
            second_recipient_name       TEXT,
            second_recipient_email      TEXT,
            second_recipient_phone      TEXT,
            second_recipient_number     TEXT,
            resend_count                INTEGER DEFAULT 0,
            last_resend_date            TEXT,
            rewards_form_link           TEXT,
            notes                       TEXT,
            created_at                  TEXT DEFAULT (datetime('now')),
            updated_at                  TEXT DEFAULT (datetime('now'))
        );

        CREATE INDEX IF NOT EXISTS idx_hpr_status     ON hotel_points_request(status);
        CREATE INDEX IF NOT EXISTS idx_hpr_booking    ON hotel_points_request(booking_id);
        CREATE INDEX IF NOT EXISTS idx_hpr_pickup_cfg ON hotel_points_request(pickup_config_id);
        CREATE INDEX IF NOT EXISTS idx_hpr_program    ON hotel_points_request(program_id);
    ''')
    db.commit()

    # Seed seven default chain rows on first run
    existing = db.execute("SELECT COUNT(*) FROM hotel_points_program").fetchone()[0]
    if existing == 0:
        from points_utils import default_chain_seeds
        for seed in default_chain_seeds():
            db.execute('''
                INSERT INTO hotel_points_program
                    (chain_name, submission_type, form_url, field_mapping_json,
                     submission_window_days, receipt_window_days, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (seed['chain_name'], seed['submission_type'], seed['form_url'],
                  seed['field_mapping_json'], seed['submission_window_days'],
                  seed['receipt_window_days'], seed['notes']))
        db.commit()

    # Seed user profile settings if absent (defaults can be edited via /settings)
    for key, default in [('user_full_name', ''),
                         ('user_email',     ''),
                         ('user_phone',     '')]:
        row = db.execute('SELECT 1 FROM Settings WHERE key=?', (key,)).fetchone()
        if not row:
            db.execute('INSERT INTO Settings (key, value) VALUES (?, ?)',
                       (key, default))
    db.commit()

    # ── Per-user loyalty profile (for point-splitting and default-recipient logic) ──
    # All hotel points in CPAinc default to Kristin House. Each user can still
    # enter their own loyalty info to support future point splitting.
    db.executescript('''
        CREATE TABLE IF NOT EXISTS user_loyalty_profile (
            user_id          INTEGER PRIMARY KEY REFERENCES Users(id) ON DELETE CASCADE,
            full_name        TEXT,
            email            TEXT,
            phone            TEXT,
            marriott_number  TEXT,
            hyatt_number     TEXT,
            hilton_number    TEXT,
            ihg_number       TEXT,
            omni_number      TEXT,
            choice_number    TEXT,
            sonesta_number   TEXT,
            created_at       TEXT DEFAULT (datetime('now')),
            updated_at       TEXT DEFAULT (datetime('now'))
        );
    ''')
    db.commit()

    # Seed one row per existing active user, defaulting name/email from Users.
    for u in db.execute("SELECT id, name, email FROM Users WHERE active=1").fetchall():
        existing = db.execute("SELECT 1 FROM user_loyalty_profile WHERE user_id=?",
                              (u['id'],)).fetchone()
        if not existing:
            db.execute('''INSERT INTO user_loyalty_profile
                (user_id, full_name, email)
                VALUES (?, ?, ?)''',
                (u['id'], u['name'] or '', u['email'] or ''))
    db.commit()

    # Default-recipient setting — defaults to Kristin House (id=2). All forms
    # use this user's loyalty profile as the primary recipient until a future
    # point-split feature allows per-request overrides.
    row = db.execute("SELECT 1 FROM Settings WHERE key='hotel_points_default_recipient_user_id'").fetchone()
    if not row:
        kristin = db.execute(
            "SELECT id FROM Users WHERE LOWER(name) LIKE '%kristin%house%' "
            "OR LOWER(username) = 'kristin' LIMIT 1"
        ).fetchone()
        default_uid = str(kristin['id']) if kristin else '2'
        db.execute("INSERT INTO Settings (key, value) VALUES ('hotel_points_default_recipient_user_id', ?)",
                   (default_uid,))
    db.commit()


try:
    with app.app_context():
        ensure_pickup_tables()
except Exception:
    pass


# ── Auth tables & seeding ─────────────────────────────────────────────────────

ALL_PERMISSIONS = [
    ('dashboard',                  'Dashboard'),
    ('admin_panel',                'Admin Panel'),
    ('import_bookings',            'Import → Bookings'),
    ('import_payments',            'Import → Payments'),
    ('import_hhr',                 'Import → HHR'),
    ('import_cancelled',           'Import → Cancelled Meetings'),
    ('reports_commission_kristin', 'Missing Pickup/Comm Report (Kristin)'),
    ('reports_commission_team',    'Missing Pickup/Comm Report (Team)'),
    ('reports_payments',           'Reports → Payment Report'),
    ('reports_customer_summary',   'Reports → Customer Summary'),
    ('bookings_view',              'View Bookings'),
    ('bookings_edit',              'Add / Edit Bookings'),
    ('pickups_payments',           'Add Pickups / Payments'),
    ('contracts',                  'Contracts'),
]

_SEED_USERS = [
    ('Peter Wann',    'peter.wann@conferencedirect.com',     'peter',    'CPAinc2026!', 'admin'),
    ('Kristin House', 'kristin.house@conferencedirect.com',  'kristin',  'CPAinc2026!', 'admin'),
    ('Geralyn Krist', 'geralyn.krist@conferencedirect.com',  'geralyn',  'CPAinc2026!', 'user'),
    ('Morgan Basham', 'morgan.basham@conferencedirect.com',  'morgan',   'CPAinc2026!', 'user'),
    ('Ashleigh Buhr', 'ashleigh.buhr@conferencedirect.com', 'ashleigh', 'CPAinc2026!', 'user'),
]


def ensure_auth_tables():
    db = get_db()
    db.executescript('''
        CREATE TABLE IF NOT EXISTS Users (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            name          TEXT    NOT NULL,
            email         TEXT    NOT NULL UNIQUE,
            username      TEXT    NOT NULL UNIQUE,
            password_hash TEXT    NOT NULL,
            role          TEXT    NOT NULL DEFAULT 'user',
            active        INTEGER NOT NULL DEFAULT 1
        );
        CREATE TABLE IF NOT EXISTS UserPermissions (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id    INTEGER NOT NULL REFERENCES Users(id) ON DELETE CASCADE,
            permission TEXT    NOT NULL,
            enabled    INTEGER NOT NULL DEFAULT 0,
            UNIQUE(user_id, permission)
        );
        CREATE TABLE IF NOT EXISTS UserAccountAccess (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id      INTEGER NOT NULL REFERENCES Users(id) ON DELETE CASCADE,
            account_name TEXT    NOT NULL,
            UNIQUE(user_id, account_name)
        );
        CREATE TABLE IF NOT EXISTS UserMicrosoftTokens (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id       INTEGER NOT NULL UNIQUE REFERENCES Users(id) ON DELETE CASCADE,
            access_token  TEXT    NOT NULL,
            refresh_token TEXT    NOT NULL,
            expires_at    REAL    NOT NULL,
            ms_user_email TEXT,
            connected_at  TEXT    DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS UserPipelineAssociates (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id        INTEGER NOT NULL REFERENCES Users(id) ON DELETE CASCADE,
            associate_name TEXT    NOT NULL,
            UNIQUE(user_id, associate_name)
        );
        CREATE TABLE IF NOT EXISTS status_board_ignore (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id      INTEGER NOT NULL REFERENCES Users(id) ON DELETE CASCADE,
            config_id    INTEGER NOT NULL REFERENCES pickup_config(id) ON DELETE CASCADE,
            issue_type   TEXT NOT NULL,
            ignore_until TEXT,
            created_at   TEXT DEFAULT (datetime('now')),
            UNIQUE(user_id, config_id, issue_type)
        );
        CREATE TABLE IF NOT EXISTS activity_log (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            username    TEXT,
            user_id     INTEGER,
            timestamp   TEXT DEFAULT (datetime('now','localtime')),
            method      TEXT,
            endpoint    TEXT,
            path        TEXT,
            description TEXT,
            detail      TEXT,
            ip_address  TEXT
        );
        CREATE TABLE IF NOT EXISTS pickup_tasks (
            id                  INTEGER PRIMARY KEY AUTOINCREMENT,
            config_id           INTEGER REFERENCES pickup_config(id) ON DELETE CASCADE,
            title               TEXT NOT NULL,
            due_date            TEXT,
            assigned_to_user_id INTEGER REFERENCES Users(id),
            assigned_to_name    TEXT,
            created_by_user_id  INTEGER,
            created_by_name     TEXT,
            created_at          TEXT DEFAULT (datetime('now','localtime')),
            completed_at        TEXT,
            completed_by        TEXT,
            task_notes          TEXT
        );
        CREATE TABLE IF NOT EXISTS pickup_event_notes (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            config_id   INTEGER REFERENCES pickup_config(id) ON DELETE CASCADE,
            user_id     INTEGER,
            username    TEXT,
            note_text   TEXT NOT NULL,
            created_at  TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS pickup_weekly_deleted (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            original_id     INTEGER,
            config_id       INTEGER,
            report_date     TEXT,
            pickup_by_night TEXT,
            total_rooms     INTEGER,
            change_from_last INTEGER,
            pct_of_block    REAL,
            pct_of_attrition REAL,
            ota_rate        REAL,
            label           TEXT,
            notes           TEXT,
            deleted_at      TEXT DEFAULT (datetime('now','localtime')),
            deleted_by      TEXT
        );
        CREATE TABLE IF NOT EXISTS pickup_change_log (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            config_id   INTEGER REFERENCES pickup_config(id) ON DELETE CASCADE,
            user_id     INTEGER,
            username    TEXT,
            timestamp   TEXT DEFAULT (datetime('now','localtime')),
            action      TEXT,
            field       TEXT,
            old_value   TEXT,
            new_value   TEXT
        );
        CREATE TABLE IF NOT EXISTS pickup_amendments (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            config_id     INTEGER NOT NULL REFERENCES pickup_config(id) ON DELETE CASCADE,
            uploaded_at   TEXT DEFAULT (datetime('now')),
            uploaded_by   TEXT,
            filename      TEXT,
            file_data     BLOB,
            description   TEXT,
            changes_json  TEXT DEFAULT '{}',
            applied       INTEGER DEFAULT 1
        );
        CREATE TABLE IF NOT EXISTS pickup_bulk_pending (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            batch_id      TEXT NOT NULL,
            config_id     INTEGER NOT NULL,
            hotel         TEXT,
            filename      TEXT,
            extracted_json TEXT,
            contract_data  BLOB,
            status        TEXT DEFAULT 'ok',
            error_msg     TEXT,
            created_at    TEXT DEFAULT (datetime('now'))
        );
    ''')
    db.commit()
    _seed_users(db)
    _seed_pipeline_associates(db)
    _seed_pickup_placeholders(db)


def _seed_pipeline_associates(db):
    """Ensure Peter and Kristin are mapped to Kristin House + Morgan Klinkradt associates."""
    for username in ('peter', 'kristin'):
        row = db.execute('SELECT id FROM Users WHERE username=?', (username,)).fetchone()
        if not row:
            continue
        uid = row['id']
        for assoc in ('Kristin House', 'Morgan Klinkradt'):
            db.execute(
                'INSERT OR IGNORE INTO UserPipelineAssociates (user_id, associate_name) VALUES (?,?)',
                (uid, assoc)
            )

    # Remove pickup_config records that don't belong in this system
    # Safe delete only — skips records that have pickup history or contact logs
    _safe_delete_where = (
        "AND id NOT IN (SELECT config_id FROM pickup_weekly WHERE config_id IS NOT NULL) "
        "AND id NOT IN (SELECT config_id FROM pickup_contact_log WHERE config_id IS NOT NULL)"
    )
    # Non-system orgs (Peter Wann's FCCS accounts)
    for org in ('FCCS',):
        db.execute(f"DELETE FROM pickup_config WHERE organization=? {_safe_delete_where}", (org,))
    # Specific booking IDs entered in error (2029 ASAS — not in pipeline, no history)
    for bid in ('31226',):
        db.execute(f"DELETE FROM pickup_config WHERE booking_id=? {_safe_delete_where}", (bid,))

    db.commit()


def _seed_pickup_placeholders(db):
    """Create placeholder pickup_config records for known bookings.
    Uses INSERT OR IGNORE so existing records (with contracts/data) are never overwritten."""
    placeholders = [
        # (booking_id, organization, event_name, hotel)
        # ── ASAS 2029 Montreal ───────────────────────────────────────────────
        ('231279', 'American Dairy Science Association and the American Society of Animal Science (ADSA/ASAS)',
         'ASAS 2029 Annual Meeting', 'InterContinental Montreal'),
        ('231280', 'American Dairy Science Association and the American Society of Animal Science (ADSA/ASAS)',
         'ASAS 2029 Annual Meeting', 'Hampton Inn by Hilton Montreal Downtown'),
        # ── National Honey Board 2026 ────────────────────────────────────────
        ('229707', 'National Honey Board',
         'National Honey Board Annual Meeting 2026', 'Hotel Champlain Burlington, Curio Collection by Hilton'),
        # ── IECA 2028 Louisville ─────────────────────────────────────────────
        ('229065', 'International Erosion Control Association',
         '2028 IECA Annual Conference', 'Omni Louisville Hotel'),
        ('229396', 'International Erosion Control Association',
         '2028 IECA Annual Conference', 'The Seelbach Hilton Louisville'),
        # ── SRM 2028 Corpus Christi ──────────────────────────────────────────
        ('230237', 'Society for Range Management',
         '2028 SRM Annual Meeting', 'Holiday Inn Corpus Christi Downtown Marina'),
        # ── IECA 2029 Cleveland ──────────────────────────────────────────────
        ('230882', 'International Erosion Control Association',
         '2029 IECA Annual Conference', 'Cleveland Marriott Downtown at Key Tower'),
        ('230883', 'International Erosion Control Association',
         '2029 IECA Annual Conference', 'Hilton Cleveland Downtown'),
        ('230908', 'International Erosion Control Association',
         '2029 IECA Annual Conference', 'Drury Plaza Hotel Cleveland'),
        # ── ESA 2030 Salt Lake City ──────────────────────────────────────────
        ('231211', 'Entomological Society of America (ESA)',
         '2030 ESA Annual Meeting', 'Salt Lake Marriott Downtown at City Creek'),
        ('231213', 'Entomological Society of America (ESA)',
         '2030 ESA Annual Meeting', 'AC Hotel Salt Lake City Downtown'),
        ('231267', 'Entomological Society of America (ESA)',
         '2030 ESA Annual Meeting', 'Hyatt Regency Salt Lake City'),
        # ── Geosynthetics 2027 Reno ──────────────────────────────────────────
        ('211781', 'Advanced Textiles Association (Former IFAI)',
         'Geosynthetics Conference 2027', 'Atlantis Casino Resort Spa - Reno'),
        ('211782', 'Advanced Textiles Association (Former IFAI)',
         'Geosynthetics Conference 2027', 'Peppermill Resort Spa Casino'),
        # ── 2027 SRM Wichita ─────────────────────────────────────────────────
        ('224460', 'Society for Range Management',
         '2027 SRM Annual Meeting', 'Hyatt Regency Wichita'),
        ('226253', 'Society for Range Management',
         '2027 SRM Annual Meeting (Overflow)', 'Fairfield Inn & Suites Wichita Downtown'),
        ('226256', 'Society for Range Management',
         '2027 SRM Annual Meeting (Overflow)', 'Drury Plaza Hotel Broadview Wichita'),
        ('226252', 'Society for Range Management',
         '2027 SRM Annual Meeting (overflow)', 'Courtyard Wichita at Old Town'),
        ('226255', 'Society for Range Management',
         '2027 SRM Annual Meeting (Overflow)', 'Hilton Garden Inn Wichita Downtown'),
        # ── 2027 NCSL Legislative Summit Atlanta ─────────────────────────────
        ('203371', 'National Conference of State Legislatures (NCSL)',
         '2027 NCSL Legislative Summit', 'Omni Atlanta Hotel at CNN Center'),
        ('203977', 'National Conference of State Legislatures (NCSL)',
         '2027 NCSL Legislative Summit', 'The Westin Peachtree Plaza, Atlanta'),
        ('203979', 'National Conference of State Legislatures (NCSL)',
         '2027 NCSL Legislative Summit', 'The Ritz-Carlton, Atlanta'),
        ('203982', 'National Conference of State Legislatures (NCSL)',
         '2027 NCSL Legislative Summit', 'Courtyard Atlanta Downtown'),
        ('203987', 'National Conference of State Legislatures (NCSL)',
         '2027 NCSL Legislative Summit', 'Atlanta Marriott Marquis'),
        ('204185', 'National Conference of State Legislatures (NCSL)',
         '2027 NCSL Legislative Summit', 'Hyatt Regency Atlanta'),
        ('204830', 'National Conference of State Legislatures (NCSL)',
         '2027 NCSL Legislative Summit', 'Reverb Atlanta by Hard Rock'),
        ('209197', 'National Conference of State Legislatures (NCSL)',
         '2027 NCSL Legislative Summit', 'Hyatt Place Atlanta/Centennial Park'),
        ('209199', 'National Conference of State Legislatures (NCSL)',
         '2027 NCSL Legislative Summit', 'Embassy Suites by Hilton Atlanta at Centennial Olympic Park'),
        ('209201', 'National Conference of State Legislatures (NCSL)',
         '2027 NCSL Legislative Summit', 'Hilton Garden Inn Atlanta Downtown'),
        ('224182', 'National Conference of State Legislatures (NCSL)',
         '2027 NCSL Legislative Summit', 'Signia by Hilton, Atlanta'),
        # ── ESA 2027 Annual Meeting Orlando ──────────────────────────────────
        ('164586', 'Entomological Society of America (ESA)',
         'ESA 2027 Annual Meeting', 'Rosen Shingle Creek'),
        ('164588', 'Entomological Society of America (ESA)',
         'ESA 2027 Annual Meeting - Overflow Rooms', 'Rosen Inn at Pointe Orlando'),
        # ── 2028 SRM Annual Meeting Corpus Christi ───────────────────────────
        ('227474', 'Society for Range Management',
         '2028 SRM Annual Meeting', 'Omni Corpus Christi Hotel'),
        # ── ASAS 2028 Annual Meeting ─────────────────────────────────────────
        ('197900', 'American Dairy Science Association and the American Society of Animal Science (ADSA/ASAS)',
         'ASAS 2028 Annual Meeting', 'Peppermill Resort Spa Casino'),
        # ── 2028 NCSL Legislative Summit Fort Lauderdale ─────────────────────
        ('218484', 'National Conference of State Legislatures (NCSL)',
         '2028 NCSL Legislative Summit', 'Omni Fort Lauderdale'),
        ('219831', 'National Conference of State Legislatures (NCSL)',
         '2028 NCSL Legislative Summit', 'Fort Lauderdale Marriott Harbor Beach Resort & Spa'),
        ('219834', 'National Conference of State Legislatures (NCSL)',
         '2028 NCSL Legislative Summit', 'Sonesta Fort Lauderdale Beach'),
        ('219836', 'National Conference of State Legislatures (NCSL)',
         '2028 NCSL Legislative Summit', 'AC Hotel Fort Lauderdale Beach'),
        ('219859', 'National Conference of State Legislatures (NCSL)',
         '2028 NCSL Legislative Summit', 'The Westin Fort Lauderdale Beach Resort'),
        ('219886', 'National Conference of State Legislatures (NCSL)',
         '2028 NCSL Legislative Summit', 'Renaissance Fort Lauderdale Cruise Port Hotel'),
        ('221734', 'National Conference of State Legislatures (NCSL)',
         '2028 NCSL Legislative Summit', 'B Ocean Resort Fort Lauderdale'),
        ('221735', 'National Conference of State Legislatures (NCSL)',
         '2028 NCSL Legislative Summit', 'Kimpton Goodland Fort Lauderdale Beach'),
        ('221736', 'National Conference of State Legislatures (NCSL)',
         '2028 NCSL Legislative Summit', 'Hyatt Place Fort Lauderdale Airport & Cruise Port'),
        ('221737', 'National Conference of State Legislatures (NCSL)',
         '2028 NCSL Legislative Summit', 'W Fort Lauderdale'),
        ('223806', 'National Conference of State Legislatures (NCSL)',
         '2028 NCSL Legislative Summit', 'Hilton Fort Lauderdale Beach Resort'),
        ('223807', 'National Conference of State Legislatures (NCSL)',
         '2028 NCSL Legislative Summit', 'Hilton Fort Lauderdale Marina'),
        ('223809', 'National Conference of State Legislatures (NCSL)',
         '2028 NCSL Legislative Summit', 'Embassy Suites by Hilton Fort Lauderdale 17th Street'),
        ('224694', 'National Conference of State Legislatures (NCSL)',
         '2028 NCSL Legislative Summit', 'Holiday Inn Express Hotel Fort Lauderdale Convention Center-Cruise'),
        # ── ESA 2028 Joint Annual Meeting Montreal ───────────────────────────
        ('204079', 'Entomological Society of America (ESA)',
         'ESA 2028 Joint Annual Meeting', 'Humaniti Montreal Downtown, Autograph Collection'),
        ('204084', 'Entomological Society of America (ESA)',
         'ESA 2028 Joint Annual Meeting', 'Le Westin Montreal'),
        ('205410', 'Entomological Society of America (ESA)',
         'ESA - 2028 Joint Annual Meeting', 'InterContinental Montreal'),
        ('209446', 'Entomological Society of America (ESA)',
         'ESA 2028 Joint Annual Meeting', 'Doubletree by Hilton Montreal'),
        ('209452', 'Entomological Society of America (ESA)',
         'ESA 2028 Joint Annual Meeting', 'Hotel Monville'),
        ('209453', 'Entomological Society of America (ESA)',
         'ESA 2028 Joint Annual Meeting', 'Hotel Le Dauphin Montreal Centre-Ville'),
        ('209455', 'Entomological Society of America (ESA)',
         'ESA 2028 Joint Annual Meeting', 'Hampton Inn by Hilton Montreal Downtown'),
        ('209457', 'Entomological Society of America (ESA)',
         'ESA 2028 Joint Annual Meeting', 'Homewood Suites by Hilton Montreal Downtown'),
        ('209458', 'Entomological Society of America (ESA)',
         'ESA 2028 Joint Annual Meeting', 'Embassy Suites by Hilton Montreal'),
        ('215959', 'Entomological Society of America (ESA)',
         'ESA - 2028 Joint Annual Meeting', 'Hotel Travelodge Montreal Centre'),
        # ── ASA-CSSA-SSSA 2028 International Annual Meeting Fort Lauderdale ──
        ('192898', 'Alliance of Crop, Soil and Environmental Science Societies (ACSESS)',
         'ASA-CSSA-SSSA 2028 International Annual Meeting', 'Renaissance Fort Lauderdale Cruise Port Hotel'),
        # ── 2029 NCSL Legislative Summit Seattle ─────────────────────────────
        ('224166', 'National Conference of State Legislatures (NCSL)',
         '2029 NCSL Legislative Summit', 'Sheraton Grand Seattle Hotel'),
        ('224169', 'National Conference of State Legislatures (NCSL)',
         '2029 NCSL Legislative Summit', 'Hyatt Regency Seattle'),
        ('224172', 'National Conference of State Legislatures (NCSL)',
         '2029 NCSL Legislative Summit', 'Grand Hyatt Seattle'),
        ('224178', 'National Conference of State Legislatures (NCSL)',
         '2029 NCSL Legislative Summit', 'Hyatt At Olive 8'),
        ('227513', 'National Conference of State Legislatures (NCSL)',
         '2029 NCSL Legislative Summit', 'Coast Seattle Downtown Hotel by APA'),
        ('227514', 'National Conference of State Legislatures (NCSL)',
         '2029 NCSL Legislative Summit', 'Hilton Garden Inn Seattle Downtown'),
        ('227518', 'National Conference of State Legislatures (NCSL)',
         '2029 NCSL Legislative Summit', 'The Westin Seattle'),
        ('227520', 'National Conference of State Legislatures (NCSL)',
         '2029 NCSL Legislative Summit', 'Hilton Motif Seattle'),
        ('227523', 'National Conference of State Legislatures (NCSL)',
         '2029 NCSL Legislative Summit', 'The Paramount Hotel Seattle'),
        ('227526', 'National Conference of State Legislatures (NCSL)',
         '2029 NCSL Legislative Summit', 'Hotel Theodore Seattle'),
        ('227700', 'National Conference of State Legislatures (NCSL)',
         '2029 NCSL Legislative Summit', 'Residence Inn Seattle Downtown/Convention Center'),
        # ── 2029 ASA-CSSA-SSSA Annual Meeting Phoenix ────────────────────────
        ('203547', 'Alliance of Crop, Soil and Environmental Science Societies (ACSESS)',
         '2029 ASA-CSSA-SSSA Annual Meeting', 'Sheraton Phoenix Downtown'),
        ('203549', 'Alliance of Crop, Soil and Environmental Science Societies (ACSESS)',
         '2029 ASA-CSSA-SSSA Annual Meeting', 'Renaissance Phoenix Downtown Hotel'),
        ('209961', 'Alliance of Crop, Soil and Environmental Science Societies (ACSESS)',
         '2029 ASA-CSSA-SSA Annual Meeting', 'Hyatt Regency Phoenix'),
        ('209637', 'Alliance of Crop, Soil and Environmental Science Societies (ACSESS)',
         '2029 ASA-CSSA-SSSA Annual Meeting', 'Hilton Garden Inn Phoenix Downtown'),
        ('209641', 'Alliance of Crop, Soil and Environmental Science Societies (ACSESS)',
         '2029 ASA-CSSA-SSSA Annual Meeting', 'SpringHill Suites Phoenix Downtown'),
        ('209644', 'Alliance of Crop, Soil and Environmental Science Societies (ACSESS)',
         '2029 ASA-CSSA-SSSA Annual Meeting', 'Residence Inn Phoenix Downtown'),
        ('209646', 'Alliance of Crop, Soil and Environmental Science Societies (ACSESS)',
         '2029 ASA-CSSA-SSSA Annual Meeting', 'Courtyard by Marriott Downtown Phoenix'),
        # ── ESA 2029 Annual Meeting Phoenix ──────────────────────────────────
        ('218308', 'Entomological Society of America (ESA)',
         'ESA 2029 Annual Meeting', 'Hyatt Regency Phoenix'),
        ('218500', 'Entomological Society of America (ESA)',
         'ESA 2029 Annual Meeting', 'Sheraton Phoenix Downtown'),
        ('218538', 'Entomological Society of America (ESA)',
         'ESA - 2029 Annual Meeting', 'The Westin Phoenix Downtown'),
        ('218540', 'Entomological Society of America (ESA)',
         'ESA 2029 Annual Meeting', 'Renaissance Phoenix Downtown Hotel'),
        ('223849', 'Entomological Society of America (ESA)',
         'ESA 2029 Annual Meeting', 'Hyatt Place Phoenix/Downtown'),
        ('219443', 'Entomological Society of America (ESA)',
         'ESA - 2029 Annual Meeting', 'Residence Inn Phoenix Downtown'),
        ('219446', 'Entomological Society of America (ESA)',
         'ESA 2029 Annual Meeting', 'Courtyard by Marriott Downtown Phoenix'),
        ('219448', 'Entomological Society of America (ESA)',
         'ESA 2029 Annual Meeting', 'Hilton Garden Inn Phoenix Downtown'),
        ('219483', 'Entomological Society of America (ESA)',
         'ESA 2029 Annual Meeting', 'SpringHill Suites Phoenix Downtown'),
        # ── 2030 NCSL Legislative Summit Salt Lake City ──────────────────────
        ('216325', 'National Conference of State Legislatures (NCSL)',
         '2030 NCSL Legislative Summit', 'Hyatt Regency Salt Lake City'),
        ('216326', 'National Conference of State Legislatures (NCSL)',
         '2030 NCSL Legislative Summit', 'Salt Lake Marriott Downtown at City Creek'),
        ('216327', 'National Conference of State Legislatures (NCSL)',
         '2030 NCSL Legislative Summit', 'Kimpton Hotel Monaco Salt Lake City'),
        ('216328', 'National Conference of State Legislatures (NCSL)',
         '2030 NCSL Legislative Summit', 'Hilton Salt Lake City Center'),
        ('216329', 'National Conference of State Legislatures (NCSL)',
         '2030 NCSL Legislative Summit', 'Salt Lake City Marriott City Center'),
        ('216331', 'National Conference of State Legislatures (NCSL)',
         '2030 NCSL Legislative Summit', 'Radisson Hotel Salt Lake City Downtown'),
        ('216333', 'National Conference of State Legislatures (NCSL)',
         '2030 NCSL Legislative Summit', 'Holiday Inn Express Salt Lake City Downtown'),
        # ── 2030 ASA-CSSA-SSSA Annual Meeting Minneapolis ────────────────────
        ('205462', 'Alliance of Crop, Soil and Environmental Science Societies (ACSESS)',
         '2030 ASA-CSSA-SSSA International Annual Meeting', 'Hilton Minneapolis'),
        ('208433', 'Alliance of Crop, Soil and Environmental Science Societies (ACSESS)',
         '2030 ASA-CSSA-SSSA Annual Meeting', 'Hyatt Regency Minneapolis'),
        ('213917', 'Alliance of Crop, Soil and Environmental Science Societies (ACSESS)',
         '2030 ASA-CSSA-SSSA International Annual Meeting', 'Millennium Hotel Minneapolis'),
        ('211117', 'Alliance of Crop, Soil and Environmental Science Societies (ACSESS)',
         '2030 ASA-CSSA-SSSA Annual Meeting (Overflow)', 'Hilton Garden Inn Minneapolis Downtown'),
        ('211121', 'Alliance of Crop, Soil and Environmental Science Societies (ACSESS)',
         '2030 ASA-CSSA-SSSA Annual Meeting (Overflow)', 'The Marquette Hotel, Curio Collection by Hilton'),
        ('211122', 'Alliance of Crop, Soil and Environmental Science Societies (ACSESS)',
         '2030 ASA-CSSA-SSSA Annual Meeting (Overflow)', 'Holiday Inn Express & Suites Minneapolis-Dwtn (Conv Ctr)'),
        ('211123', 'Alliance of Crop, Soil and Environmental Science Societies (ACSESS)',
         '2030 ASA-CSSA-SSSA Annual Meeting (Overflow)', 'Royal Sonesta Minneapolis Downtown'),
    ]
    for booking_id, org, event_name, hotel in placeholders:
        exists = db.execute(
            'SELECT 1 FROM pickup_config WHERE booking_id=? AND hotel=?',
            (booking_id, hotel)
        ).fetchone()
        if not exists:
            db.execute('''
                INSERT INTO pickup_config
                    (booking_id, organization, event_name, hotel,
                     contracted_block, status,
                     hotel_contacts, cc_emails, force_current, force_past,
                     shoulder_pre, shoulder_post)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
            ''', (booking_id, org, event_name, hotel,
                  '{}', 'active', '[]', '[]', 0, 0, 3, 3))
    db.commit()
    # ── Remove duplicate (booking_id, hotel) rows created by prior bug ──
    # Keep the row with the most data (non-empty block, or rate, or lowest id).
    # Safe: never deletes rows that have pickup history or contact logs.
    dupes = db.execute('''
        SELECT booking_id, hotel, COUNT(*) AS cnt
        FROM pickup_config
        WHERE booking_id IS NOT NULL
        GROUP BY booking_id, hotel
        HAVING COUNT(*) > 1
    ''').fetchall()
    deleted = 0
    for d in dupes:
        rows = db.execute('''
            SELECT id,
                   CASE WHEN contracted_block IS NOT NULL AND contracted_block NOT IN ('{}','')
                        THEN 1 ELSE 0 END AS has_block,
                   CASE WHEN contracted_rate IS NOT NULL THEN 1 ELSE 0 END AS has_rate
            FROM pickup_config
            WHERE booking_id=? AND hotel=?
            ORDER BY has_block DESC, has_rate DESC, id ASC
        ''', (d['booking_id'], d['hotel'])).fetchall()
        # Keep the first (best) row, delete the rest if they have no history
        keep_id = rows[0]['id']
        for r in rows[1:]:
            has_history = db.execute(
                'SELECT 1 FROM pickup_weekly WHERE config_id=? UNION '
                'SELECT 1 FROM pickup_contact_log WHERE config_id=?',
                (r['id'], r['id'])
            ).fetchone()
            if not has_history:
                db.execute('DELETE FROM pickup_config WHERE id=?', (r['id'],))
                deleted += 1
    if deleted:
        db.commit()


def _hash_password(password):
    return generate_password_hash(password, method='pbkdf2:sha256')


def _seed_users(db):
    for name, email, username, password, role in _SEED_USERS:
        existing = db.execute('SELECT id FROM Users WHERE email = ? OR username = ?', (email, username)).fetchone()
        if not existing:
            ph = _hash_password(password)
            try:
                db.execute(
                    'INSERT INTO Users (name, email, username, password_hash, role) VALUES (?,?,?,?,?)',
                    (name, email, username, ph, role)
                )
            except Exception:
                continue
            db.commit()
            user = db.execute('SELECT id FROM Users WHERE email = ?', (email,)).fetchone()
            uid = user['id']
            is_admin = (role == 'admin')
            for perm, _ in ALL_PERMISSIONS:
                enabled = 1 if is_admin or perm in ('bookings_view', 'pickups_payments') else 0
                db.execute(
                    'INSERT OR IGNORE INTO UserPermissions (user_id, permission, enabled) VALUES (?,?,?)',
                    (uid, perm, enabled)
                )
            db.commit()


try:
    with app.app_context():
        ensure_auth_tables()
except Exception:
    pass


# ── Auth helpers ──────────────────────────────────────────────────────────────

def get_current_user():
    uid = session.get('user_id')
    if not uid:
        return None
    return get_db().execute('SELECT * FROM Users WHERE id = ? AND active = 1', (uid,)).fetchone()


def has_permission(user, perm):
    if user is None:
        return False
    if user['role'] == 'admin':
        return True
    row = get_db().execute(
        'SELECT enabled FROM UserPermissions WHERE user_id = ? AND permission = ?',
        (user['id'], perm)
    ).fetchone()
    return bool(row and row['enabled'])


def get_user_account_filter(user=None):
    """Returns a list of account names the user may see, or None for admins (no filter).
    Used for financial reporting — admins always see all."""
    if user is None:
        user = get_current_user()
    if user is None:
        return []
    if user['role'] == 'admin':
        return None
    rows = get_db().execute(
        'SELECT account_name FROM UserAccountAccess WHERE user_id = ?', (user['id'],)
    ).fetchall()
    return [r['account_name'] for r in rows]


def get_pickup_account_filter(user=None, db=None):
    """Returns an account-name filter list for Pickup, RFP, and Status Board views.
    Unlike get_user_account_filter, this applies to ALL users including admins.

    Priority:
      1. UserPipelineAssociates — dynamically queries ReportPipeline for all
         AccountNames belonging to those associates. Used for Peter & Kristin.
      2. UserAccountAccess — explicit list (Morgan, Ashleigh, Geralyn).
      3. None — no filter (see everything). Only if user has neither table populated.
    """
    if user is None:
        user = get_current_user()
    if user is None:
        return []
    if db is None:
        db = get_db()

    # 1. Associate-driven filter (pipeline lookup)
    assoc_rows = db.execute(
        'SELECT associate_name FROM UserPipelineAssociates WHERE user_id=?', (user['id'],)
    ).fetchall()
    if assoc_rows:
        associates = [r['associate_name'] for r in assoc_rows]
        ph = ','.join('?' * len(associates))
        acct_rows = db.execute(
            f"SELECT DISTINCT AccountName FROM ReportPipeline "
            f"WHERE BookingAssociate IN ({ph}) AND AccountName IS NOT NULL",
            associates
        ).fetchall()
        return [r['AccountName'] for r in acct_rows]

    # 2. Explicit UserAccountAccess list
    acct_rows = db.execute(
        'SELECT account_name FROM UserAccountAccess WHERE user_id=?', (user['id'],)
    ).fetchall()
    if acct_rows:
        return [r['account_name'] for r in acct_rows]

    # 3. No restriction (admin with no filter configured)
    if user['role'] == 'admin':
        return None

    return []  # non-admin with no access configured


@app.before_request
def require_login():
    exempt = {'login', 'logout', 'static'}
    if request.endpoint in exempt or request.endpoint is None:
        return
    uid = session.get('user_id')
    if not uid:
        if request.headers.get('X-Requested-With') == 'fetch' or request.is_json:
            return jsonify({'ok': False, 'error': 'Not authenticated'}), 401
        return redirect(url_for('login', next=request.path))
    # ── Session timeout (1 hour of inactivity) ──
    last_active = session.get('last_activity')
    now_ts = datetime.utcnow()
    if last_active:
        try:
            elapsed = (now_ts - datetime.fromisoformat(last_active)).total_seconds()
            if elapsed > SESSION_TIMEOUT_SECONDS:
                session.clear()
                if request.headers.get('X-Requested-With') == 'fetch' or request.is_json:
                    return jsonify({'ok': False, 'error': 'Session expired'}), 401
                flash('Your session expired after 1 hour of inactivity. Please log in again.', 'warning')
                return redirect(url_for('login', next=request.path))
        except Exception:
            pass
    session['last_activity'] = now_ts.isoformat()
    user = get_db().execute('SELECT id, active FROM Users WHERE id = ?', (uid,)).fetchone()
    if not user or not user['active']:
        session.clear()
        return redirect(url_for('login'))


# ── Change log helper ────────────────────────────────────────────────────────

def _log_change(db, config_id, action, changes):
    """Record field-level changes to a pickup_config record.
    changes = list of (field, old_value, new_value) tuples.
    Only logs entries where old != new."""
    uid      = session.get('user_id')
    username = session.get('username') or 'unknown'
    for field, old_val, new_val in changes:
        old_s = str(old_val) if old_val is not None else ''
        new_s = str(new_val) if new_val is not None else ''
        if old_s != new_s:
            db.execute(
                '''INSERT INTO pickup_change_log (config_id, user_id, username, action, field, old_value, new_value)
                   VALUES (?,?,?,?,?,?,?)''',
                (config_id, uid, username, action, field, old_s or None, new_s or None)
            )


# ── Activity logging ──────────────────────────────────────────────────────────

# Human-readable labels for each endpoint (keyed by actual Flask function name)
_ENDPOINT_LABELS = {
    # ── Auth ──
    'login':                            'Logged in',
    'logout':                           'Logged out',
    # ── Home / Bookings ──
    'index':                            'Viewed Home / Dashboard',
    'bookings':                         'Viewed Bookings list',
    'booking_detail':                   'Viewed Booking #{booking_id}',
    'booking_new':                      'Created new Booking',
    'booking_edit':                     'Edited Booking #{booking_id}',
    'booking_cancel':                   'Cancelled Booking #{booking_id}',
    # ── Contracts ──
    'booking_contract_upload':          'Uploaded contract — Booking #{booking_id}',
    'booking_contract_download':        'Downloaded contract — Booking #{booking_id}',
    'booking_contract_delete':          'Deleted contract — Booking #{booking_id}',
    'booking_contract_extract':         'Extracted contract to Pickup — Booking #{booking_id}',
    'booking_contract_select_pickup':   'Selected Pickup for contract — Booking #{booking_id}',
    'booking_contract_parse_for_pickup':'Parsed contract into Pickup (config #{cid})',
    # ── Imports ──
    'import_bookings':                  'Imported bookings from Excel',
    'import_cancelled':                 'Imported cancelled bookings',
    'import_payments':                  'Imported payments',
    'import_voucher':                   'Imported voucher data',
    'import_hhr':                       'Imported Housing History Report (HHR)',
    # ── Pickup Dashboard / Events ──
    'pickup_dashboard':                 'Viewed Pickup Tracking dashboard',
    'pickup_new_event':                 'Created new Pickup event',
    'pickup_event':                     'Viewed Pickup event (config #{cid})',
    'pickup_edit_event':                'Edited Pickup event (config #{cid})',
    'pickup_sync_pipeline':             'Synced Pickup from Pipeline',
    'pickup_fill_missing':              'Filled missing Pickup data',
    'pickup_auto_match':                'Auto-matched Pickup records',
    'pickup_import_xlsx':               'Imported Pickup data from Excel',
    'pickup_import_xlsx_confirm':       'Confirmed Pickup Excel import',
    # ── Pickup Reports ──
    'pickup_weekly_new':                'Added pickup report (config #{cid})',
    'pickup_weekly_edit':               'Edited pickup report (config #{cid})',
    'pickup_weekly_delete':             'Deleted pickup report (config #{cid})',
    'pickup_event_report':              'Viewed Pickup report (config #{primary_cid})',
    'pickup_event_report_xlsx':         'Downloaded Pickup report Excel (config #{primary_cid})',
    # ── Pickup Contracts / HHR ──
    'pickup_upload_contract':           'Uploaded contract — Pickup config #{cid}',
    'pickup_contract_download':         'Downloaded contract — Pickup config #{cid}',
    'pickup_hhr_download':              'Downloaded HHR — Pickup config #{cid}',
    'pickup_housing_form':              'Viewed Housing Form (config #{cid})',
    'pickup_final_history':             'Viewed Final History (config #{cid})',
    # ── Pickup Contacts / Emails ──
    'pickup_contact_log':               'Logged hotel contact (config #{cid})',
    'pickup_email_housing':             'Sent housing email (config #{cid})',
    'pickup_email_hotel':               'Sent hotel email (config #{cid})',
    'pickup_email_client':              'Sent client email (config #{cid})',
    'pickup_email_client_launch_outlook':'Launched Outlook for client email (config #{cid})',
    'pickup_email_client_paste':        'Copied client email to clipboard (config #{cid})',
    'pickup_email_post_report':         'Sent post-event report (config #{cid})',
    # ── Pickup Rooming List ──
    'pickup_rooming_upload':            'Uploaded Rooming List (config #{cid})',
    'pickup_rooming_manual':            'Added manual Rooming List entry (config #{cid})',
    'pickup_rooming_review':            'Reviewed Rooming List (config #{cid})',
    'pickup_rooming_confirm':           'Confirmed Rooming List (config #{cid})',
    'pickup_rooming_download_csv':      'Downloaded Rooming List CSV (config #{cid})',
    # ── Customer Reports ──
    'pickup_customer_report_select':    'Selected accounts for Customer Report',
    'pickup_customer_report_xlsx':      'Downloaded Customer Report Excel',
    'pickup_export_row':                'Exported pickup row (config #{cid})',
    # ── Status Board ──
    'status_board':                     'Viewed Status Board',
    'status_board_ignore':              'Ignored Status Board item',
    # ── Admin ──
    'admin_users':                      'Viewed Admin — User Management',
    'admin_toggle_permission':          'Changed permission for user #{uid}',
    'admin_update_accounts':            'Updated account access for user #{uid}',
    'admin_toggle_active':              'Toggled active status for user #{uid}',
    'admin_reset_password':             'Reset password for user #{uid}',
    'admin_download_db':                'Downloaded full database backup',
    'admin_export_pickup_data':         'Exported pickup data (JSON)',
    'admin_import_pickup_data':         'Imported pickup data (JSON)',
    'admin_activity_log':               'Viewed Activity Log',
    # ── RFP ──
    'rfp_list':                         'Viewed RFP list',
    'rfp_detail':                       'Viewed RFP #{rfp_id}',
    'rfp_create':                       'Created new RFP',
    'rfp_edit':                         'Edited RFP #{rfp_id}',
    'rfp_delete':                       'Deleted RFP #{rfp_id}',
}

# Endpoints to skip logging (too noisy or not meaningful)
_SKIP_LOG_ENDPOINTS = {
    'static', None,
    'admin_activity_log',   # avoid self-referential noise
    'status_board_ignore',  # logged via POST action description instead
}

def _log_activity(response):
    """Record the current request to activity_log. Called from after_request."""
    try:
        endpoint = request.endpoint
        if endpoint in _SKIP_LOG_ENDPOINTS:
            return
        # Only log successful page loads and meaningful POSTs (skip redirects for GET)
        if request.method == 'GET' and response.status_code in (301, 302):
            return
        # Skip non-HTML/non-download responses that are just JSON fragments
        ct = response.content_type or ''
        if 'application/json' in ct and request.method == 'GET':
            return

        uid      = session.get('user_id')
        username = session.get('username')
        if not username and uid:
            # Fall back to DB lookup for sessions that predate the username cache
            try:
                row = get_db().execute('SELECT name FROM Users WHERE id=?', (uid,)).fetchone()
                if row:
                    username = row['name']
                    session['username'] = username  # cache it going forward
            except Exception:
                pass
        username = username or 'unknown'
        path     = request.path

        # Build human-readable description
        label = _ENDPOINT_LABELS.get(endpoint)
        if label:
            # Fill in URL params
            kw = dict(request.view_args or {})
            try:
                description = label.format(**kw)
            except KeyError:
                description = label
        else:
            # Fallback: turn endpoint name into readable text
            description = (endpoint or path).replace('_', ' ').title()
            if request.method == 'POST':
                description = 'Submitted: ' + description

        # Extra detail: form keys (not values) for POSTs, or query string for GETs
        detail = None
        if request.method == 'POST':
            keys = [k for k in (request.form.keys() if request.form else [])
                    if k not in ('password', 'password_hash', 'csrf_token')]
            if keys:
                detail = 'Fields: ' + ', '.join(keys[:8])
        elif request.args:
            detail = 'Query: ' + '&'.join(f'{k}={v}' for k, v in list(request.args.items())[:5])

        ip = request.headers.get('X-Forwarded-For', request.remote_addr or '')
        if ip and ',' in ip:
            ip = ip.split(',')[0].strip()

        db = get_db()
        db.execute(
            '''INSERT INTO activity_log (username, user_id, method, endpoint, path, description, detail, ip_address)
               VALUES (?,?,?,?,?,?,?,?)''',
            (username, uid, request.method, endpoint, path, description, detail, ip)
        )
        db.commit()
    except Exception:
        pass  # Never let logging break the actual response


@app.after_request
def after_request_log(response):
    _log_activity(response)
    return response


@app.context_processor
def inject_user():
    user = get_current_user()
    def _has_perm(perm):
        return has_permission(user, perm)
    return dict(current_user=user, has_perm=_has_perm, all_permissions=ALL_PERMISSIONS)


# ── Login / Logout ────────────────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    if session.get('user_id'):
        return redirect(url_for('status_board'))
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        db = get_db()
        user = db.execute(
            'SELECT * FROM Users WHERE (username = ? OR email = ?) AND active = 1',
            (username, username)
        ).fetchone()
        if user and check_password_hash(user['password_hash'], password):
            session.clear()
            session['user_id']   = user['id']
            session['username']  = user['name']
            session['user_role'] = user['role']
            next_page = request.form.get('next') or request.args.get('next') or url_for('status_board')
            return redirect(next_page)
        error = 'Invalid username or password.'
    return render_template('login.html', error=error, next=request.args.get('next', ''))


@app.route('/help')
def help_page():
    return render_template('help.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# ── Admin User Management ─────────────────────────────────────────────────────

@app.route('/admin/users')
def admin_users():
    user = get_current_user()
    if not has_permission(user, 'admin_panel'):
        flash('Access denied.', 'error')
        return redirect(url_for('pipeline'))
    db = get_db()
    users = db.execute('SELECT * FROM Users ORDER BY role DESC, name').fetchall()
    all_perms_rows = db.execute('SELECT user_id, permission, enabled FROM UserPermissions').fetchall()
    perm_map = {}
    for r in all_perms_rows:
        perm_map.setdefault(r['user_id'], {})[r['permission']] = bool(r['enabled'])
    available_accounts = [r[0] for r in db.execute(
        "SELECT DISTINCT AccountName FROM ReportPipeline "
        "WHERE AccountName IS NOT NULL "
        "AND LOWER(BookingAssociate) = 'kristin house' "
        "ORDER BY AccountName"
    ).fetchall()]
    acc_rows = db.execute('SELECT user_id, account_name FROM UserAccountAccess').fetchall()
    acc_map = {}
    for r in acc_rows:
        acc_map.setdefault(r['user_id'], [])
        acc_map[r['user_id']].append(r['account_name'])
    return render_template('admin_users.html',
        users=users, perm_map=perm_map,
        all_permissions=ALL_PERMISSIONS,
        available_accounts=available_accounts,
        acc_map=acc_map)


@app.route('/admin/users/<int:uid>/permission/<perm>/toggle', methods=['POST'])
def admin_toggle_permission(uid, perm):
    user = get_current_user()
    if not has_permission(user, 'admin_panel'):
        return jsonify({'ok': False, 'error': 'Access denied'}), 403
    db = get_db()
    existing = db.execute(
        'SELECT enabled FROM UserPermissions WHERE user_id = ? AND permission = ?', (uid, perm)
    ).fetchone()
    if existing:
        new_val = 0 if existing['enabled'] else 1
        db.execute('UPDATE UserPermissions SET enabled = ? WHERE user_id = ? AND permission = ?',
                   (new_val, uid, perm))
    else:
        new_val = 1
        db.execute('INSERT INTO UserPermissions (user_id, permission, enabled) VALUES (?,?,?)',
                   (uid, perm, new_val))
    db.commit()
    return jsonify({'ok': True, 'enabled': bool(new_val)})


@app.route('/admin/users/<int:uid>/accounts', methods=['POST'])
def admin_update_accounts(uid):
    user = get_current_user()
    if not has_permission(user, 'admin_panel'):
        return jsonify({'ok': False, 'error': 'Access denied'}), 403
    db = get_db()
    data = request.get_json(silent=True) or {}
    accounts = data.get('accounts', [])
    db.execute('DELETE FROM UserAccountAccess WHERE user_id = ?', (uid,))
    for acct in accounts:
        if acct.strip():
            db.execute('INSERT OR IGNORE INTO UserAccountAccess (user_id, account_name) VALUES (?,?)',
                       (uid, acct.strip()))
    db.commit()
    return jsonify({'ok': True, 'count': len(accounts)})


@app.route('/admin/users/<int:uid>/active/toggle', methods=['POST'])
def admin_toggle_active(uid):
    user = get_current_user()
    if not has_permission(user, 'admin_panel'):
        return jsonify({'ok': False}), 403
    db = get_db()
    existing = db.execute('SELECT active FROM Users WHERE id = ?', (uid,)).fetchone()
    if not existing:
        return jsonify({'ok': False}), 404
    new_val = 0 if existing['active'] else 1
    db.execute('UPDATE Users SET active = ? WHERE id = ?', (new_val, uid))
    db.commit()
    return jsonify({'ok': True, 'active': bool(new_val)})


@app.route('/admin/users/<int:uid>/password', methods=['POST'])
def admin_reset_password(uid):
    user = get_current_user()
    if not has_permission(user, 'admin_panel'):
        return jsonify({'ok': False}), 403
    data = request.get_json(silent=True) or {}
    new_password = data.get('password', '')
    if len(new_password) < 6:
        return jsonify({'ok': False, 'error': 'Password must be at least 6 characters.'}), 400
    ph = _hash_password(new_password)
    db = get_db()
    db.execute('UPDATE Users SET password_hash = ? WHERE id = ?', (ph, uid))
    db.commit()
    return jsonify({'ok': True})


@app.route('/admin/scan-rebates', methods=['GET', 'POST'])
def admin_scan_rebates():
    """Scan all stored contracts for rebate clauses and update pickup_config.rebate_per_room."""
    user = get_current_user()
    if not user or not has_permission(user, 'admin_panel'):
        flash('Admin access required.', 'error')
        return redirect(url_for('pickup_dashboard'))

    from pickup_utils import _extract_text_from_contract, _pdf_pages_to_images, _ai_parse_contract_vision
    import io, re, time as _time

    db = get_db()
    rows = db.execute(
        "SELECT id, hotel, event_name, contract_filename, contract_data, contracted_rate, rebate_per_room "
        "FROM pickup_config WHERE contract_data IS NOT NULL AND status='active'"
    ).fetchall()

    CONTRACT_KEYWORDS = {'rate','room','block','arrival','departure','cutoff','night','hotel',
                         'group','suite','meeting','attrition','reservation','check','guest',
                         'contract','agreement'}
    REBATE_KEYWORDS   = ['rebate','commission rebate','net rate','rebate per room',
                         'per room rebate','allowance','credit per room']

    # Get API key
    api_key = ''
    try:
        import importlib, sys as _sys
        if 'config' in _sys.modules:
            importlib.reload(_sys.modules['config'])
        else:
            import config as _cfg; _sys.modules['config'] = _cfg
        api_key = _sys.modules['config'].ANTHROPIC_API_KEY.strip()
    except Exception:
        pass
    if not api_key:
        import os as _os
        api_key = _os.environ.get('ANTHROPIC_API_KEY','').strip()

    results = []

    for r in rows:
        hotel   = r['hotel'] or r['event_name'] or f"id={r['id']}"
        cid     = r['id']
        already = r['rebate_per_room']
        blob    = bytes(r['contract_data'])

        # Extract text
        text   = _extract_text_from_contract(blob, r['contract_filename'] or 'contract.pdf')
        tl     = (text or '').lower()
        words  = tl.split()
        meaningful = len(words) >= 50 and any(k in tl for k in CONTRACT_KEYWORDS)

        rebate_amount = None
        method = ''

        if meaningful:
            if not any(kw in tl for kw in REBATE_KEYWORDS):
                results.append({'id': cid, 'hotel': hotel, 'status': 'no_rebate',
                                'rebate': None, 'already': already, 'method': 'text'})
                continue
            # Text has rebate keywords — parse with Claude text
            try:
                import anthropic
                client = anthropic.Anthropic(api_key=api_key)
                msg = client.messages.create(
                    model='claude-haiku-4-5', max_tokens=200,
                    messages=[{'role':'user','content':
                        f"Hotel contract for {hotel} (rate: ${r['contracted_rate']}). "
                        f"Find any per-room-per-night rebate, credit, or allowance paid back to the group. "
                        f"Return ONLY JSON: {{\"rebate_per_room\": 10.00}} or {{\"rebate_per_room\": null}}\n\n{text[:3000]}"}])
                raw = msg.content[0].text.strip()
                m2 = re.search(r'\{.*?\}', raw, re.DOTALL)
                if m2:
                    rebate_amount = json.loads(m2.group()).get('rebate_per_room')
                method = 'text+claude'
            except Exception as e:
                results.append({'id': cid, 'hotel': hotel, 'status': 'error',
                                'rebate': None, 'already': already, 'method': 'text', 'error': str(e)})
                continue
        else:
            # Image-only — use vision
            fname = (r['contract_filename'] or '').lower()
            if not fname.endswith('.pdf'):
                results.append({'id': cid, 'hotel': hotel, 'status': 'no_rebate',
                                'rebate': None, 'already': already, 'method': 'non-pdf'})
                continue
            try:
                images = _pdf_pages_to_images(blob, max_pages=10, dpi=100)
                if not images:
                    results.append({'id': cid, 'hotel': hotel, 'status': 'error',
                                    'rebate': None, 'already': already, 'method': 'vision', 'error': 'no images'})
                    continue
                import anthropic, base64
                client = anthropic.Anthropic(api_key=api_key)
                content = []
                for img_b64 in images:
                    content.append({'type':'image','source':{'type':'base64','media_type':'image/jpeg','data':img_b64}})
                content.append({'type':'text','text':
                    f"This is a hotel contract for {hotel} (contracted rate: ${r['contracted_rate']}). "
                    f"Look for any rebate, credit, or allowance per room per night paid back to the group/client. "
                    f"Return ONLY JSON: {{\"rebate_per_room\": 10.00}} or {{\"rebate_per_room\": null}}"})
                msg = client.messages.create(model='claude-opus-4-5', max_tokens=200, messages=[{'role':'user','content':content}])
                raw = msg.content[0].text.strip()
                m2 = re.search(r'\{.*?\}', raw, re.DOTALL)
                if m2:
                    rebate_amount = json.loads(m2.group()).get('rebate_per_room')
                method = 'vision'
                _time.sleep(1)
            except Exception as e:
                results.append({'id': cid, 'hotel': hotel, 'status': 'error',
                                'rebate': None, 'already': already, 'method': 'vision', 'error': str(e)})
                continue

        if rebate_amount:
            db.execute("UPDATE pickup_config SET rebate_per_room=? WHERE id=?", (rebate_amount, cid))
            db.commit()
            results.append({'id': cid, 'hotel': hotel, 'status': 'updated',
                            'rebate': rebate_amount, 'already': already, 'method': method})
        else:
            results.append({'id': cid, 'hotel': hotel, 'status': 'no_rebate',
                            'rebate': None, 'already': already, 'method': method})

    updated  = [r for r in results if r['status'] == 'updated']
    no_rebate= [r for r in results if r['status'] == 'no_rebate']
    errors   = [r for r in results if r['status'] == 'error']
    return render_template('admin_scan_rebates.html',
                           results=results, updated=updated,
                           no_rebate=no_rebate, errors=errors,
                           total=len(rows))


@app.route('/admin/scan-block-review-dates', methods=['POST'])
def admin_scan_block_review_dates():
    """Scan all uploaded contracts and populate block_review_date where found."""
    user = get_current_user()
    if not has_permission(user, 'admin'):
        return 'Forbidden', 403

    import anthropic as _ant
    from pickup_utils import _extract_text_from_contract, _pdf_pages_to_images

    api_key = ''
    try:
        api_key = app.config.get('ANTHROPIC_API_KEY', '').strip()
    except Exception:
        pass
    if not api_key:
        api_key = os.environ.get('ANTHROPIC_API_KEY', '').strip()
    if not api_key:
        flash('Anthropic API key not configured.', 'error')
        return redirect(url_for('pickup_dashboard'))

    db = get_db()
    rows = db.execute(
        "SELECT id, contract_filename, contract_data FROM pickup_config "
        "WHERE contract_data IS NOT NULL AND (block_review_date IS NULL OR block_review_date = '')"
    ).fetchall()

    updated = skipped = errors = 0

    for row in rows:
        try:
            file_bytes = bytes(row['contract_data'])
            filename   = row['contract_filename'] or 'contract.pdf'

            # Try text extraction first; fall back to vision for image-only PDFs
            text = _extract_text_from_contract(file_bytes, filename)
            review_date = None

            if text and len(text.split()) >= 30:
                prompt = (
                    "You are reading a hotel group sales contract. "
                    "Find the room block review clause — a clause that specifies a date when the hotel "
                    "will review group pickup and may reduce the contracted block. "
                    "It may be labeled 'block review date', 'pickup review date', 'review date', "
                    "'room block review', or similar.\n\n"
                    "Return ONLY a valid JSON object with one key:\n"
                    "  block_review_date — the review date in YYYY-MM-DD format, or null if not found\n\n"
                    "No markdown fences, no explanation — just the JSON.\n\n"
                    "Contract text:\n" + text[:12000]
                )
                client  = _ant.Anthropic(api_key=api_key)
                message = client.messages.create(
                    model='claude-haiku-4-5',
                    max_tokens=64,
                    messages=[{'role': 'user', 'content': prompt}]
                )
                raw = message.content[0].text.strip()
                raw = re.sub(r'^```[a-z]*\s*', '', raw)
                raw = re.sub(r'\s*```$', '', raw)
                data = json.loads(raw)
                review_date = data.get('block_review_date') or None
            else:
                # Vision fallback for scanned PDFs
                images = _pdf_pages_to_images(file_bytes, max_pages=8, dpi=80)
                if images:
                    content = [{
                        'type': 'text',
                        'text': (
                            "Find the room block review clause in this hotel contract — a date when "
                            "the hotel will review group pickup and may reduce the contracted block. "
                            "Return ONLY JSON: {\"block_review_date\": \"YYYY-MM-DD\"} or "
                            "{\"block_review_date\": null} if not found."
                        )
                    }]
                    for b64 in images:
                        content.append({
                            'type': 'image',
                            'source': {'type': 'base64', 'media_type': 'image/jpeg', 'data': b64}
                        })
                    client  = _ant.Anthropic(api_key=api_key)
                    message = client.messages.create(
                        model='claude-opus-4-5',
                        max_tokens=64,
                        messages=[{'role': 'user', 'content': content}]
                    )
                    raw = message.content[0].text.strip()
                    raw = re.sub(r'^```[a-z]*\s*', '', raw)
                    raw = re.sub(r'\s*```$', '', raw)
                    data = json.loads(raw)
                    review_date = data.get('block_review_date') or None

            if review_date:
                # Validate YYYY-MM-DD format
                from datetime import datetime as _dtt
                _dtt.strptime(review_date, '%Y-%m-%d')
                db.execute(
                    "UPDATE pickup_config SET block_review_date=? WHERE id=?",
                    (review_date, row['id'])
                )
                db.commit()
                updated += 1
            else:
                skipped += 1

        except Exception:
            errors += 1

    parts = [f'{updated} contract{"s" if updated != 1 else ""} updated with block review date']
    if skipped:
        parts.append(f'{skipped} had no review clause')
    if errors:
        parts.append(f'{errors} error{"s" if errors != 1 else ""}')
    flash(' — '.join(parts) + '.', 'success' if not errors else 'warning')
    return redirect(url_for('pickup_dashboard'))


@app.route('/admin/archive-cancelled-pickups', methods=['POST'])
def admin_archive_cancelled_pickups():
    user = get_current_user()
    if not has_permission(user, 'admin'):
        return 'Forbidden', 403
    import datetime as _dt
    db = get_db()
    cancel_note = f"CANCELLED — archived {_dt.date.today().strftime('%Y-%m-%d')}"
    rows = db.execute("""
        SELECT pc.id, pc.notes
        FROM pickup_config pc
        JOIN ReportPipeline rp
          ON CAST(pc.booking_id AS TEXT) = CAST(rp.BookingId AS TEXT)
        WHERE pc.status != 'archived'
          AND LOWER(rp.BookingStatus) = 'cancelled'
    """).fetchall()
    count = 0
    for r in rows:
        existing = (r['notes'] or '').strip()
        new_notes = cancel_note + ('\n' + existing if existing else '')
        db.execute(
            "UPDATE pickup_config SET status='archived', notes=? WHERE id=?",
            (new_notes, r['id'])
        )
        count += 1
    db.commit()
    flash(f'{count} pickup card{"s" if count != 1 else ""} archived for cancelled bookings.', 'success')
    return redirect(url_for('pickup_dashboard'))


@app.route('/admin/download-db')
def admin_download_db():
    """Stream a consistent snapshot of the live SQLite database (admin only)."""
    user = get_current_user()
    if not user or not has_permission(user, 'admin_panel'):
        return 'Forbidden — admin login required.', 403
    import sqlite3 as _sq3, tempfile, os as _os
    from flask import Response
    with tempfile.NamedTemporaryFile(suffix='.sqlite', delete=False) as tf:
        tmp_path = tf.name
    try:
        src = _sq3.connect(DATABASE)
        dst = _sq3.connect(tmp_path)
        src.backup(dst)          # atomic consistent copy
        src.close()
        dst.close()
        with open(tmp_path, 'rb') as f:
            db_bytes = f.read()
    finally:
        try:
            _os.unlink(tmp_path)
        except Exception:
            pass
    return Response(
        db_bytes,
        mimetype='application/octet-stream',
        headers={'Content-Disposition': 'attachment; filename="CPAinc.sqlite"'}
    )


@app.route('/admin/activity-log')
def admin_activity_log():
    user = get_current_user()
    if not user or user['role'] != 'admin':
        flash('Admin access required.', 'error')
        return redirect(url_for('pipeline'))

    db = get_db()

    # Filters
    filter_user = request.args.get('user', '').strip()
    filter_date = request.args.get('date', '').strip()
    filter_endpoint = request.args.get('endpoint', '').strip()
    page = max(1, int(request.args.get('page', 1)))
    per_page = 100

    where  = []
    params = []
    if filter_user:
        where.append('username = ?')
        params.append(filter_user)
    if filter_date:
        where.append("DATE(timestamp) = ?")
        params.append(filter_date)
    if filter_endpoint:
        where.append('endpoint = ?')
        params.append(filter_endpoint)

    where_sql = ('WHERE ' + ' AND '.join(where)) if where else ''

    total = db.execute(
        f'SELECT COUNT(*) FROM activity_log {where_sql}', params
    ).fetchone()[0]

    logs = db.execute(
        f'''SELECT id, username, timestamp, method, endpoint, path, description, detail, ip_address
            FROM activity_log {where_sql}
            ORDER BY id DESC
            LIMIT ? OFFSET ?''',
        params + [per_page, (page - 1) * per_page]
    ).fetchall()

    # Distinct users and endpoints for filter dropdowns
    all_users     = [r[0] for r in db.execute(
        "SELECT DISTINCT username FROM activity_log WHERE username IS NOT NULL ORDER BY username"
    ).fetchall()]
    all_endpoints = [r[0] for r in db.execute(
        "SELECT DISTINCT endpoint FROM activity_log WHERE endpoint IS NOT NULL ORDER BY endpoint"
    ).fetchall()]

    total_pages = max(1, (total + per_page - 1) // per_page)

    return render_template('activity_log.html',
        logs=logs,
        total=total,
        page=page,
        total_pages=total_pages,
        per_page=per_page,
        filter_user=filter_user,
        filter_date=filter_date,
        filter_endpoint=filter_endpoint,
        all_users=all_users,
        all_endpoints=all_endpoints,
    )


@app.route('/admin/export-pickup-data')
def admin_export_pickup_data():
    """Export all pickup_weekly rows as JSON keyed by booking_id (safe to share)."""
    user = get_current_user()
    if not user or not has_permission(user, 'admin_panel'):
        return 'Forbidden — admin login required.', 403
    db = get_db()
    rows = db.execute('''
        SELECT pc.booking_id, pw.report_date, pw.pickup_by_night, pw.total_rooms,
               pw.pct_of_block, pw.pct_of_attrition, pw.ota_rate, pw.label, pw.notes
        FROM pickup_weekly pw
        JOIN pickup_config pc ON pc.id = pw.config_id
        ORDER BY pc.booking_id, pw.report_date
    ''').fetchall()
    out = [dict(r) for r in rows]
    resp = app.response_class(
        response=json.dumps(out, indent=2),
        mimetype='application/json',
        headers={'Content-Disposition': 'attachment; filename="pickup_weekly_export.json"'}
    )
    return resp


@app.route('/admin/import-pickup-data', methods=['GET', 'POST'])
def admin_import_pickup_data():
    """Import pickup_weekly rows from JSON — matches by booking_id, never overwrites existing rows."""
    user = get_current_user()
    if not user or not has_permission(user, 'admin_panel'):
        return 'Forbidden — admin login required.', 403
    if request.method == 'GET':
        return '''<!doctype html><html><body style="font-family:sans-serif;padding:2rem">
        <h3>Import Pickup Weekly Data</h3>
        <p>Uploads pickup_weekly rows matched by booking_id. Skips duplicates (same booking_id + report_date).</p>
        <form method="post" enctype="multipart/form-data">
          <input type="file" name="json_file" accept=".json"><br><br>
          <button type="submit">Import</button>
        </form></body></html>'''
    # POST — process uploaded JSON file
    f = request.files.get('json_file')
    if not f:
        return 'No file provided.', 400
    try:
        data = json.loads(f.read().decode('utf-8'))
    except Exception as e:
        return f'Invalid JSON: {e}', 400
    db = get_db()
    # Build booking_id → config_id map from THIS database
    cfg_map = {}
    for row in db.execute('SELECT id, booking_id FROM pickup_config'):
        if row['booking_id']:
            cfg_map[str(row['booking_id'])] = row['id']
    # Get existing (config_id, report_date) pairs to skip duplicates
    existing = set()
    for row in db.execute('SELECT config_id, report_date FROM pickup_weekly'):
        existing.add((row['config_id'], row['report_date']))
    inserted = skipped_no_config = skipped_duplicate = 0
    for row in data:
        bid = str(row.get('booking_id', ''))
        cid = cfg_map.get(bid)
        if not cid:
            skipped_no_config += 1
            continue
        rd = row.get('report_date', '')
        if (cid, rd) in existing:
            skipped_duplicate += 1
            continue
        db.execute('''INSERT INTO pickup_weekly
            (config_id, report_date, pickup_by_night, total_rooms,
             pct_of_block, pct_of_attrition, ota_rate, label, notes)
            VALUES (?,?,?,?,?,?,?,?,?)''',
            (cid, rd, row.get('pickup_by_night'), row.get('total_rooms'),
             row.get('pct_of_block'), row.get('pct_of_attrition'),
             row.get('ota_rate'), row.get('label'), row.get('notes')))
        existing.add((cid, rd))
        inserted += 1
    db.commit()
    return (f'<p>Done. Inserted: <strong>{inserted}</strong> rows. '
            f'Skipped duplicates: {skipped_duplicate}. '
            f'Skipped (event not found here): {skipped_no_config}.</p>'
            f'<a href="/pickup">Go to Pickup Tracking</a>')


@app.route('/admin/fix-asa-cssa-sssa-dates')
def admin_fix_asa_dates():
    """One-time fix: shift contracted_block dates and correct booking_ids for ASA-CSSA-SSSA 2027/2028 records."""
    user = get_current_user()
    if not user or not has_permission(user, 'admin_panel'):
        return 'Forbidden — admin login required.', 403

    db = get_db()
    results = []

    # Correct booking_id mapping by hotel name substring (case-insensitive)
    HOTEL_FIXES_2027 = [
        ('hilton columbus',           '185599', 'Hilton Columbus Downtown'),
        ('hyatt regency columbus',    '185603', 'Hyatt Regency Columbus'),
        ('drury',                     '185601', 'Drury Inn & Suites Convention Centre'),
        ('hampton inn',               '185524', 'Hampton Inn & Suites Columbus-Downtown'),
        ('ac hotel columbus',         '185521', 'AC Hotel Columbus Downtown'),
        ('ac columbus',               '185521', 'AC Hotel Columbus Downtown'),
        ('red roof',                  '185879', 'Red Roof PLUS+ Columbus Downtown'),
    ]
    HOTEL_FIXES_2028 = [
        ('omni',                      '193163', 'Omni Fort Lauderdale'),
        ('hilton marina',             '193346', 'Hilton Fort Lauderdale Marina'),
        ('hilton fort lauderdale',    '193346', 'Hilton Fort Lauderdale Marina'),
        ('renaissance',               '75067',  'Renaissance Fort Lauderdale'),
        ('embassy suites',            '193345', 'Embassy Suites by Hilton Fort Lauderdale'),
        ('marriott harbor',           '192896', 'Fort Lauderdale Marriott Harbor Beach'),
        ('hyatt place',               '193347', 'Hyatt Place Fort Lauderdale Airport'),
    ]

    def shift_block_year(block_json, old_year, new_year):
        """Replace old_year with new_year in all date keys of a contracted_block JSON string."""
        try:
            block = json.loads(block_json or '{}')
        except Exception:
            block = {}
        new_block = {}
        for k, v in block.items():
            new_k = k.replace(str(old_year) + '-', str(new_year) + '-', 1)
            new_block[new_k] = v
        return json.dumps(new_block)

    def fix_group(year_label, old_year, new_year, hotel_fixes, event_new_name, org_name):
        rows = db.execute(
            "SELECT id, hotel, contracted_block, booking_id FROM pickup_config "
            "WHERE LOWER(event_name) LIKE ? AND status != 'archived'",
            (f'%asa%cssa%sssa%{year_label}%',)
        ).fetchall()
        for row in rows:
            hotel_lower = (row['hotel'] or '').lower()
            new_bid = None
            new_hotel_name = row['hotel']
            for key, bid, clean_name in hotel_fixes:
                if key in hotel_lower:
                    new_bid = bid
                    new_hotel_name = clean_name
                    break
            new_block = shift_block_year(row['contracted_block'], old_year, new_year)
            db.execute(
                '''UPDATE pickup_config
                   SET booking_id=COALESCE(?,booking_id),
                       hotel=?,
                       event_name=?,
                       organization=?,
                       contracted_block=?,
                       status='active'
                   WHERE id=?''',
                (new_bid, new_hotel_name, event_new_name, org_name, new_block, row['id'])
            )
            results.append(f"ID {row['id']}: {row['hotel']} → {new_hotel_name} | "
                           f"booking_id={new_bid or row['booking_id']} | "
                           f"block shifted {old_year}→{new_year}")

    fix_group('2027', 2022, 2027, HOTEL_FIXES_2027,
              '2027 ASA-CSSA-SSSA Annual Meeting',
              'Alliance of Crop, Soil and Environmental Science Societies')
    fix_group('2028', 2023, 2028, HOTEL_FIXES_2028,
              '2028 ASA-CSSA-SSSA Annual Meeting',
              'Alliance of Crop, Soil and Environmental Science Societies')

    # Also fix any remaining record with 2022 block dates under any ASA CSSA SSSA event
    leftover = db.execute(
        "SELECT id, hotel, contracted_block, event_name FROM pickup_config "
        "WHERE LOWER(event_name) LIKE '%asa%cssa%sssa%2022%'"
    ).fetchall()
    for row in leftover:
        new_block = shift_block_year(row['contracted_block'], 2022, 2027)
        db.execute(
            "UPDATE pickup_config SET contracted_block=?, event_name=?, status='active' WHERE id=?",
            (new_block, '2027 ASA-CSSA-SSSA Annual Meeting', row['id'])
        )
        results.append(f"ID {row['id']}: leftover 2022 event fixed → 2027 block | hotel={row['hotel']}")

    db.commit()

    html = '<h3>ASA-CSSA-SSSA Date Fix — Results</h3>'
    if results:
        html += '<ul>' + ''.join(f'<li>{r}</li>' for r in results) + '</ul>'
        html += f'<p><strong>{len(results)} record(s) updated.</strong></p>'
    else:
        html += '<p>No matching records found — nothing to fix (or already fixed).</p>'
    html += '<p><a href="/pickup">Go to Pickup Tracking</a></p>'
    return html


@app.route('/admin/fix-wrong-year-blocks')
def admin_fix_wrong_year_blocks():
    """One-time fix: clear contracted_block dates that don't match the booking's ReportPipeline year,
    and unarchive the record so it reappears in the status board with correct pipeline-date fallback."""
    user = get_current_user()
    if not user or not has_permission(user, 'admin_panel'):
        return 'Forbidden — admin login required.', 403

    db  = get_db()
    results = []

    # These booking IDs were reported as having date inconsistencies.
    # For each: look up the correct year from ReportPipeline, then clear contracted_block
    # if its date keys belong to a different year.
    TARGET_BIDS = ['170939', '170973', '212752', '231248', '226442', '231993']

    pipeline = {}
    for row in db.execute(
        "SELECT BookingId, StartDate FROM ReportPipeline WHERE BookingId IN (?,?,?,?,?,?)",
        TARGET_BIDS
    ).fetchall():
        s = _to_iso(row['StartDate']) or ''
        if len(s) >= 4:
            pipeline[str(row['BookingId'])] = int(s[:4])  # correct year

    for bid in TARGET_BIDS:
        correct_year = pipeline.get(bid)
        rows = db.execute(
            "SELECT id, event_name, hotel, contracted_block, status FROM pickup_config "
            "WHERE booking_id = ?", (bid,)
        ).fetchall()
        for row in rows:
            cid   = row['id']
            block = {}
            try:
                block = json.loads(row['contracted_block'] or '{}')
            except Exception:
                pass

            stray_keys  = [k for k in block if len(k) >= 4 and correct_year and int(k[:4]) != correct_year]
            good_keys   = [k for k in block if k not in stray_keys]
            need_unarch = row['status'] == 'archived'

            if stray_keys or need_unarch:
                if stray_keys and not good_keys:
                    # All keys are wrong-year → wipe entire block
                    new_block = '{}'
                    block_note = f"cleared entire block (stray dates: {stray_keys})"
                elif stray_keys:
                    # Mixed: remove only the stray keys, keep the correct ones
                    clean = {k: v for k, v in block.items() if k not in stray_keys}
                    new_block = json.dumps(clean)
                    block_note = f"removed stray date(s): {stray_keys}"
                else:
                    new_block = row['contracted_block']
                    block_note = None

                db.execute(
                    "UPDATE pickup_config SET contracted_block=?, status='active' WHERE id=?",
                    (new_block, cid)
                )
                msg = f"BID {bid} CID {cid} ({row['hotel']}): "
                parts = []
                if block_note:
                    parts.append(block_note)
                if need_unarch:
                    parts.append("unarchived")
                results.append(msg + '; '.join(parts))
            else:
                results.append(f"BID {bid} CID {cid} ({row['hotel']}): no change needed")

    db.commit()

    html = '<h3>Wrong-Year Block Fix — Results</h3>'
    if results:
        html += '<ul>' + ''.join(f'<li>{r}</li>' for r in results) + '</ul>'
    else:
        html += '<p>No matching records found.</p>'
    html += '<p><a href="/status-board">Go to Status Board</a></p>'
    return html


@app.route('/admin/fix-stray-block-dates')
def admin_fix_stray_block_dates():
    """Remove individual contracted_block date keys that fall outside a booking's
    ReportPipeline start/end window (catches same-year stray dates missed by fix-wrong-year-blocks)."""
    user = get_current_user()
    if not user or not has_permission(user, 'admin_panel'):
        return 'Forbidden — admin login required.', 403

    db      = get_db()
    results = []

    # Load all active pickup_configs that have a booking_id and a non-empty block
    rows = db.execute(
        "SELECT id, booking_id, hotel, event_name, contracted_block "
        "FROM pickup_config WHERE status != 'archived' AND booking_id IS NOT NULL "
        "AND contracted_block IS NOT NULL AND contracted_block != '{}'"
    ).fetchall()

    # Fetch ReportPipeline start/end for all relevant booking_ids
    bid_list = list({str(r['booking_id']) for r in rows})
    pipeline = {}
    if bid_list:
        ph = ','.join('?' * len(bid_list))
        for pr in db.execute(
            f"SELECT BookingId, StartDate, EndDate FROM ReportPipeline WHERE BookingId IN ({ph})",
            bid_list
        ).fetchall():
            s = _to_iso(pr['StartDate']) or ''
            e = _to_iso(pr['EndDate'])   or ''
            if len(s) == 10:
                pipeline[str(pr['BookingId'])] = {'start': s, 'end': e}

    for row in rows:
        bid   = str(row['booking_id'])
        pdates = pipeline.get(bid)
        if not pdates:
            continue

        try:
            block = json.loads(row['contracted_block'])
        except Exception:
            continue

        p_start = pdates['start']
        p_end   = pdates['end']

        # Keep only dates that fall within [start - 7 days, end + 7 days] of the pipeline window
        # (small buffer for pre/post dates that are legitimately just outside the window)
        from datetime import datetime as _dt, timedelta as _td
        try:
            window_start = (_dt.strptime(p_start, '%Y-%m-%d') - _td(days=7)).strftime('%Y-%m-%d')
            window_end   = (_dt.strptime(p_end,   '%Y-%m-%d') + _td(days=7)).strftime('%Y-%m-%d')
        except Exception:
            continue

        stray  = [k for k in block if k < window_start or k > window_end]
        if not stray:
            continue

        clean     = {k: v for k, v in block.items() if k not in stray}
        new_block = json.dumps(clean)
        db.execute(
            "UPDATE pickup_config SET contracted_block=? WHERE id=?",
            (new_block, row['id'])
        )
        results.append(
            f"BID {bid} CID {row['id']} ({row['hotel']}): "
            f"removed stray date(s) {stray} — kept {sorted(clean.keys())}"
        )

    db.commit()

    html = '<h3>Stray Block Date Fix — Results</h3>'
    if results:
        html += '<ul>' + ''.join(f'<li>{r}</li>' for r in results) + '</ul>'
        html += f'<p><strong>{len(results)} record(s) updated.</strong></p>'
    else:
        html += '<p>No stray dates found — all blocks look clean.</p>'
    html += '<p><a href="/status-board">Go to Status Board</a></p>'
    return html


# ── Status Board ──────────────────────────────────────────────────────────────

@app.route('/status-board')
def status_board():
    """Per-user action-item dashboard for pickup tracking data quality."""
    user = get_current_user()
    if not has_permission(user, 'pickups_payments'):
        flash('Access denied.', 'error')
        return redirect(url_for('pipeline'))

    db          = get_db()
    today       = datetime.today().strftime('%Y-%m-%d')

    acct_filter = get_pickup_account_filter(user, db)

    # --- Load active ignore records for this user ---
    ignored = set()
    for row in db.execute(
        "SELECT config_id, issue_type FROM status_board_ignore "
        "WHERE user_id=? AND (ignore_until IS NULL OR ignore_until > ?)",
        (user['id'], today)
    ).fetchall():
        ignored.add((row['config_id'], row['issue_type']))

    # --- Load all non-archived pickup_config with account filter ---
    # Also exclude records whose booking is Cancelled in ReportPipeline
    base_sql = ("SELECT * FROM pickup_config WHERE status != 'archived' "
                "AND (booking_id IS NULL OR booking_id NOT IN "
                "  (SELECT BookingId FROM ReportPipeline WHERE BookingStatus='Cancelled'))")
    params   = []
    if acct_filter is None:
        pass
    elif acct_filter:
        ph = ','.join('?' * len(acct_filter))
        base_sql += f' AND organization IN ({ph})'
        params.extend(acct_filter)
    else:
        base_sql += ' AND 1=0'
    configs = db.execute(base_sql, params).fetchall()

    # --- ReportPipeline dates fallback (used when contracted_block is empty) ---
    pipeline_dates = {}  # booking_id (str) → {'start': 'YYYY-MM-DD', 'end': 'YYYY-MM-DD'}
    bid_list = [str(cfg['booking_id']) for cfg in configs if cfg['booking_id']]
    if bid_list:
        ph = ','.join('?' * len(bid_list))
        for row in db.execute(
            f"SELECT BookingId, StartDate, EndDate FROM ReportPipeline WHERE BookingId IN ({ph})",
            bid_list
        ).fetchall():
            pipeline_dates[str(row['BookingId'])] = {
                'start': _to_iso(row['StartDate']),
                'end':   _to_iso(row['EndDate']),
            }

    # --- Latest pickup_weekly per config ---
    latest_pickup = {}
    for row in db.execute(
        "SELECT config_id, MAX(report_date) as last_date FROM pickup_weekly GROUP BY config_id"
    ).fetchall():
        latest_pickup[row['config_id']] = row['last_date']

    # --- Latest contact log per config ---
    latest_contact = {}
    for row in db.execute(
        "SELECT config_id, MAX(contact_date) as last_contact FROM pickup_contact_log GROUP BY config_id"
    ).fetchall():
        latest_contact[row['config_id']] = row['last_contact']

    # --- Configs that have ANY pickup_weekly rows ---
    has_history = set()
    for row in db.execute("SELECT DISTINCT config_id FROM pickup_weekly").fetchall():
        has_history.add(row['config_id'])

    # --- Configs that have an HHR uploaded (keyed by booking_id) ---
    has_hhr = set()
    for row in db.execute(
        "SELECT DISTINCT booking_id FROM housing_history_files WHERE booking_id IS NOT NULL"
    ).fetchall():
        has_hhr.add(str(row['booking_id']))

    # --- Issue definitions: (severity, title, icon) ---
    ISSUE_META = {
        'overdue_pickup':        ('danger',  'Overdue Pickup Report',                          'bi-exclamation-circle-fill'),
        'past_cutoff_no_history':('danger',  'Past Cutoff',                                   'bi-exclamation-triangle-fill'),
        'event_ended_no_hhr':    ('danger',  'Event Ended — No HHR Uploaded',                 'bi-file-earmark-x-fill'),
        'rooming_list_due':      ('danger',  'Rooming List Due Within 15 Days',               'bi-list-check'),
        'amenity_order':         ('warning', 'Amenity Order — Meeting in 14 Days',            'bi-gift-fill'),
        'block_review_due':      ('warning', 'Room Block Review Due',                          'bi-calendar2-check-fill'),
        'no_recent_contact':     ('warning', 'No Hotel Contact in 21+ Days',                  'bi-telephone-x-fill'),
        'cutoff_approaching':    ('warning', 'Cutoff Approaching — Block Not Verified',        'bi-alarm-fill'),
        'uniform_block':         ('warning', 'Block Needs Verification (All Nights Identical)','bi-grid-fill'),
        'empty_block':           ('warning', 'No Contracted Block Entered',                    'bi-calendar-x-fill'),
        'missing_hotel_email':   ('info',    'Missing Hotel Contact Email',                    'bi-envelope-x-fill'),
        'missing_client_contact':('info',    'Missing Client Name or Email',                   'bi-person-x-fill'),
        'missing_cutoff':        ('info',    'No Cutoff Date Set',                             'bi-calendar-minus-fill'),
        'missing_rate':          ('info',    'No Contracted Room Rate',                        'bi-currency-dollar'),
        'points_overdue':        ('warning', 'Hotel Points Not Received',                      'bi-award-fill'),
    }

    issues_by_type = {k: [] for k in ISSUE_META}

    # Preload hotel-points programs and existing requests for fast lookup
    from points_utils import detect_chain as _detect_chain_sb
    hp_programs = {row['chain_name']: row for row in db.execute(
        'SELECT id, chain_name, active, receipt_window_days FROM hotel_points_program WHERE active=1'
    ).fetchall()}
    hp_requests = {}
    for r in db.execute(
        'SELECT pickup_config_id, program_id, status, points_received_date '
        'FROM hotel_points_request WHERE pickup_config_id IS NOT NULL'
    ).fetchall():
        hp_requests[(r['pickup_config_id'], r['program_id'])] = r

    for cfg in configs:
        cid   = cfg['id']
        block = {}
        try:
            block = json.loads(cfg['contracted_block'] or '{}')
        except Exception:
            pass

        dates       = sorted(block.keys())
        # Priority: 1) ReportPipeline, 2) manually-set pickup_config.event_start/end, 3) contracted_block keys
        _pdates     = pipeline_dates.get(str(cfg['booking_id'] or ''), {})
        event_start = _pdates.get('start') or cfg['event_start'] or (dates[0]  if dates else None)
        event_end   = _pdates.get('end')   or cfg['event_end']   or (dates[-1] if dates else None)
        total_block = sum(block.values()) if block else 0
        block_vals  = list(block.values())

        def _issue(itype, detail, fix_url=None):
            if (cid, itype) in ignored:
                return
            issues_by_type[itype].append({
                'config_id':   cid,
                'issue_type':  itype,
                'event_name':  cfg['event_name'] or cfg['organization'],
                'organization':cfg['organization'],
                'hotel':       cfg['hotel'] or '—',
                'booking_id':  cfg['booking_id'],
                'event_start': event_start,
                'event_end':   event_end,
                'detail':      detail,
                'fix_url':     fix_url or url_for('pickup_event', cid=cid),
            })

        is_started = event_start and event_start <= today
        is_ended   = event_end   and event_end   <  today
        is_current = is_started and not is_ended

        last_rpt  = latest_pickup.get(cid)
        last_ctct = latest_contact.get(cid)
        cutoff    = (cfg['cutoff_date'] or '').strip()

        # 0a. Room block review due — within 15 days of block_review_date
        block_review_date = (cfg['block_review_date'] or '').strip()
        if block_review_date and not is_ended:
            try:
                days_to_review = (datetime.strptime(block_review_date, '%Y-%m-%d') -
                                  datetime.strptime(today, '%Y-%m-%d')).days
                if days_to_review <= 15:
                    if days_to_review >= 0:
                        detail = f'Block review date is {block_review_date} — {days_to_review} day{"s" if days_to_review != 1 else ""} away.'
                    else:
                        detail = f'Block review date was {block_review_date} — {abs(days_to_review)} day{"s" if abs(days_to_review) != 1 else ""} ago.'
                    _issue('block_review_due', detail, url_for('pickup_event', cid=cid))
            except Exception:
                pass

        # 0. Amenity order — meeting starts within 14 days
        if event_start and not is_started:
            try:
                days_to_start = (datetime.strptime(event_start, '%Y-%m-%d') -
                                 datetime.strptime(today, '%Y-%m-%d')).days
                if days_to_start <= 14:
                    _issue('amenity_order',
                           f'Meeting starts {event_start} — {days_to_start} day{"s" if days_to_start != 1 else ""} away. Order client amenity.',
                           url_for('pickup_event', cid=cid))
            except Exception:
                pass

        # 1. Overdue pickup report (current events only)
        if is_current:
            if not last_rpt:
                _issue('overdue_pickup',
                       'No pickup reports filed yet — event already started.',
                       url_for('pickup_event', cid=cid))
            else:
                days_since = (datetime.strptime(today, '%Y-%m-%d') -
                              datetime.strptime(last_rpt,  '%Y-%m-%d')).days
                if days_since >= 7:
                    _issue('overdue_pickup',
                           f'Last report was {days_since} days ago ({last_rpt}).',
                           url_for('pickup_event', cid=cid))

        # 2. Past cutoff with no pickup history (only while event is still ongoing)
        if cutoff and cutoff < today and cid not in has_history and not is_ended:
            _issue('past_cutoff_no_history',
                   f'Cutoff was {cutoff} — no pickup history entered yet.',
                   url_for('pickup_event', cid=cid))

        # 2b. Event ended with no HHR uploaded
        if is_ended and str(cfg['booking_id'] or '') not in has_hhr:
            _issue('event_ended_no_hhr',
                   f'Event ended {event_end} and no Housing History Report has been uploaded.',
                   url_for('import_hhr'))

        # 2c. Rooming list required and due within 15 days of cutoff
        if cfg['rooming_list_required'] and cutoff and not is_ended:
            try:
                days_to_cutoff = (datetime.strptime(cutoff, '%Y-%m-%d') -
                                  datetime.strptime(today, '%Y-%m-%d')).days
                has_rooming = db.execute(
                    'SELECT id FROM pickup_rooming_list WHERE config_id=? LIMIT 1', (cid,)
                ).fetchone()
                if days_to_cutoff <= 15 and not has_rooming:
                    _issue('rooming_list_due',
                           f'Rooming list required — cutoff is {cutoff} ({days_to_cutoff} days away). '
                           f'Client must provide names before cut-off.',
                           url_for('pickup_event', cid=cid))
            except Exception:
                pass

        # 2d. Hotel points overdue — event ended 60+ days ago, points not received
        if is_ended and event_end:
            chain = _detect_chain_sb(cfg['hotel'])
            if chain and chain in hp_programs:
                prog = hp_programs[chain]
                try:
                    days_since_end = (datetime.strptime(today, '%Y-%m-%d') -
                                      datetime.strptime(event_end, '%Y-%m-%d')).days
                except Exception:
                    days_since_end = None
                if days_since_end is not None and days_since_end >= 60:
                    req = hp_requests.get((cid, prog['id']))
                    if not req:
                        _issue('points_overdue',
                               f'Event ended {days_since_end} days ago — no {chain} points request created.',
                               url_for('points_generate', cid=cid))
                    elif req['status'] not in ('received', 'cancelled', 'disallowed'):
                        _issue('points_overdue',
                               f'Event ended {days_since_end} days ago — {chain} points still not received '
                               f'(status: {req["status"] or "pending"}).')

        # 3. No hotel contact in 21+ days (current events only)
        if is_current:
            if not last_ctct:
                _issue('no_recent_contact',
                       'No contact with hotel has ever been logged.',
                       url_for('pickup_event', cid=cid))
            else:
                days_since = (datetime.strptime(today, '%Y-%m-%d') -
                              datetime.strptime(last_ctct, '%Y-%m-%d')).days
                if days_since >= 21:
                    _issue('no_recent_contact',
                           f'Last hotel contact logged {days_since} days ago ({last_ctct}).',
                           url_for('pickup_event', cid=cid))

        # 4. Uniform block (all nights identical — needs real nightly breakdown)
        if len(block) > 1 and len(set(block_vals)) == 1 and block_vals[0] > 0:
            _issue('uniform_block',
                   f'Every night shows exactly {block_vals[0]} rooms — likely auto-filled.')

        # 5. Empty block
        if not block or total_block == 0:
            _issue('empty_block', 'contracted_block is empty or all zeros.')

        # 6. Missing hotel email
        if not (cfg['hotel_contact_email'] or '').strip():
            _issue('missing_hotel_email', 'No hotel contact email on file.')

        # 7. Missing client contact
        no_name  = not (cfg['group_contact'] or '').strip()
        no_email = not (cfg['group_contact_email'] or '').strip()
        if no_name or no_email:
            parts = []
            if no_name:  parts.append('name')
            if no_email: parts.append('email')
            _issue('missing_client_contact',
                   'Missing client ' + ' and '.join(parts) + '.')

        # 8. Missing cutoff
        if not cutoff:
            _issue('missing_cutoff', 'No cutoff date has been set.')

        # 8b. Cutoff approaching (≤30 days) but block not yet verified
        if cutoff and cutoff >= today and not is_ended:
            try:
                days_to_cutoff = (datetime.strptime(cutoff, '%Y-%m-%d') -
                                  datetime.strptime(today, '%Y-%m-%d')).days
                block_unverified = (not block or total_block == 0 or
                                    (len(block) > 1 and len(set(block_vals)) == 1))
                if days_to_cutoff <= 30 and block_unverified:
                    _issue('cutoff_approaching',
                           f'Cutoff is in {days_to_cutoff} day(s) ({cutoff}) — '
                           f'contracted block has not been verified.',
                           url_for('pickup_edit_event', cid=cid))
            except Exception:
                pass

        # 9. Missing room rate
        if not cfg['contracted_rate'] or cfg['contracted_rate'] == 0:
            _issue('missing_rate', 'No contracted room rate entered.')

    total_issues = sum(len(v) for v in issues_by_type.values())

    return render_template('status_board.html',
                           issues_by_type=issues_by_type,
                           issue_meta=ISSUE_META,
                           total_issues=total_issues,
                           today=today)


@app.route('/status-board/ignore', methods=['POST'])
def status_board_ignore():
    """Set or update an ignore record for a status board issue."""
    user = get_current_user()
    if not has_permission(user, 'pickups_payments'):
        return jsonify({'ok': False, 'error': 'Access denied'}), 403

    data       = request.get_json(silent=True) or {}
    config_id  = int(data.get('config_id', 0))
    issue_type = data.get('issue_type', '')
    mode       = data.get('mode', '1month')   # '1month' | '2months' | 'custom' | 'permanent'
    custom_dt  = data.get('custom_date', '')

    if not config_id or not issue_type:
        return jsonify({'ok': False, 'error': 'Missing params'}), 400

    today = datetime.today()
    if mode == 'permanent':
        ignore_until = None
    elif mode == '2months':
        ignore_until = (today + timedelta(days=60)).strftime('%Y-%m-%d')
    elif mode == 'custom' and custom_dt:
        ignore_until = custom_dt
    else:  # default: 1 month
        ignore_until = (today + timedelta(days=30)).strftime('%Y-%m-%d')

    db = get_db()
    db.execute(
        '''INSERT INTO status_board_ignore (user_id, config_id, issue_type, ignore_until)
           VALUES (?,?,?,?)
           ON CONFLICT(user_id, config_id, issue_type)
           DO UPDATE SET ignore_until=excluded.ignore_until,
                         created_at=datetime('now')''',
        (user['id'], config_id, issue_type, ignore_until)
    )
    db.commit()
    label = 'permanently' if ignore_until is None else f'until {ignore_until}'
    return jsonify({'ok': True, 'label': label})


@app.route('/status-board/ignore-bulk', methods=['POST'])
def status_board_ignore_bulk():
    """Bulk-ignore multiple status board items in one call."""
    user = get_current_user()
    if not has_permission(user, 'pickups_payments'):
        return jsonify({'ok': False, 'error': 'Access denied'}), 403
    data    = request.get_json(silent=True) or {}
    items   = data.get('items', [])   # list of {config_id, issue_type}
    mode    = data.get('mode', '1month')
    custom_dt = data.get('custom_date', '')
    if not items:
        return jsonify({'ok': False, 'error': 'No items provided'}), 400
    today = datetime.today()
    if mode == 'permanent':
        ignore_until = None
    elif mode == '2months':
        ignore_until = (today + timedelta(days=60)).strftime('%Y-%m-%d')
    elif mode == 'custom' and custom_dt:
        ignore_until = custom_dt
    else:
        ignore_until = (today + timedelta(days=30)).strftime('%Y-%m-%d')
    db = get_db()
    count = 0
    for item in items:
        cid = int(item.get('config_id', 0))
        itype = item.get('issue_type', '')
        if not cid or not itype:
            continue
        db.execute(
            '''INSERT INTO status_board_ignore (user_id, config_id, issue_type, ignore_until)
               VALUES (?,?,?,?)
               ON CONFLICT(user_id, config_id, issue_type)
               DO UPDATE SET ignore_until=excluded.ignore_until, created_at=datetime('now')''',
            (user['id'], cid, itype, ignore_until)
        )
        count += 1
    db.commit()
    return jsonify({'ok': True, 'ignored': count})


# ── Tasks ────────────────────────────────────────────────────────────────────

@app.route('/pickup/<int:cid>/tasks/new', methods=['POST'])
def pickup_task_new(cid):
    db   = get_db()
    user = get_current_user()
    f    = request.form
    title    = f.get('title', '').strip()
    due_date = f.get('due_date', '').strip() or None
    assigned_uid   = f.get('assigned_to_user_id', '').strip() or None
    assigned_name  = f.get('assigned_to_name', '').strip() or None
    task_notes     = f.get('task_notes', '').strip() or None
    if not title:
        flash('Task title is required.', 'error')
        return redirect(url_for('pickup_event', cid=cid))
    if assigned_uid:
        row = db.execute("SELECT name FROM Users WHERE id=?", (assigned_uid,)).fetchone()
        if row:
            assigned_name = row['name']
    db.execute(
        '''INSERT INTO pickup_tasks
           (config_id, title, due_date, assigned_to_user_id, assigned_to_name,
            created_by_user_id, created_by_name, task_notes)
           VALUES (?,?,?,?,?,?,?,?)''',
        (cid, title, due_date, assigned_uid, assigned_name,
         user['id'] if user else None, user['name'] if user else None, task_notes)
    )
    db.commit()
    flash('Task added.', 'success')
    return redirect(url_for('pickup_event', cid=cid))


@app.route('/pickup/<int:cid>/tasks/<int:tid>/complete', methods=['POST'])
def pickup_task_complete(cid, tid):
    db   = get_db()
    user = get_current_user()
    db.execute(
        "UPDATE pickup_tasks SET completed_at=datetime('now','localtime'), completed_by=? "
        "WHERE id=? AND config_id=?",
        (user['name'] if user else 'unknown', tid, cid)
    )
    db.commit()
    flash('Task marked complete.', 'success')
    return redirect(url_for('pickup_event', cid=cid))


@app.route('/pickup/<int:cid>/tasks/<int:tid>/reopen', methods=['POST'])
def pickup_task_reopen(cid, tid):
    db = get_db()
    db.execute("UPDATE pickup_tasks SET completed_at=NULL, completed_by=NULL WHERE id=? AND config_id=?", (tid, cid))
    db.commit()
    flash('Task reopened.', 'success')
    return redirect(url_for('pickup_event', cid=cid))


@app.route('/pickup/<int:cid>/tasks/<int:tid>/delete', methods=['POST'])
def pickup_task_delete(cid, tid):
    db = get_db()
    db.execute("DELETE FROM pickup_tasks WHERE id=? AND config_id=?", (tid, cid))
    db.commit()
    flash('Task deleted.', 'success')
    return redirect(url_for('pickup_event', cid=cid))


# ── Event Notes ───────────────────────────────────────────────────────────────

@app.route('/pickup/<int:cid>/notes/add', methods=['POST'])
def pickup_note_add(cid):
    db   = get_db()
    user = get_current_user()
    text = request.form.get('note_text', '').strip()
    if not text:
        flash('Note cannot be empty.', 'error')
        return redirect(url_for('pickup_event', cid=cid))
    db.execute(
        "INSERT INTO pickup_event_notes (config_id, user_id, username, note_text) VALUES (?,?,?,?)",
        (cid, user['id'] if user else None, user['name'] if user else 'unknown', text)
    )
    db.commit()
    flash('Note added.', 'success')
    return redirect(url_for('pickup_event', cid=cid))


@app.route('/pickup/<int:cid>/notes/<int:nid>/delete', methods=['POST'])
def pickup_note_delete(cid, nid):
    user = get_current_user()
    if not user or user['role'] != 'admin':
        flash('Admin access required to delete notes.', 'error')
        return redirect(url_for('pickup_event', cid=cid))
    db = get_db()
    db.execute("DELETE FROM pickup_event_notes WHERE id=? AND config_id=?", (nid, cid))
    db.commit()
    flash('Note deleted.', 'success')
    return redirect(url_for('pickup_event', cid=cid))


def _create_pickup_config_from_booking(db, bid, account, event, hotel,
                                       start_str, end_str, peak_rooms, room_rate):
    """Create a pickup_config entry for a booking if one doesn't already exist.
    Builds the contracted_block day-by-day from StartDate→EndDate at PeakRooms per night.
    Returns True if created, False if skipped (already exists or missing data).
    """
    if not bid:
        return False
    existing = db.execute(
        "SELECT id FROM pickup_config WHERE booking_id = ?", (str(bid),)
    ).fetchone()
    if existing:
        return False

    contracted_block = {}
    if start_str and end_str:
        try:
            from datetime import date as _d, timedelta as _td
            start = _d.fromisoformat(str(start_str)[:10])
            end   = _d.fromisoformat(str(end_str)[:10])
            peak  = max(1, int(float(peak_rooms))) if peak_rooms else 1
            day   = start
            while day < end:
                contracted_block[day.isoformat()] = peak
                day += _td(days=1)
        except Exception:
            pass

    try:
        rate = float(room_rate) if room_rate else None
    except Exception:
        rate = None

    db.execute('''
        INSERT INTO pickup_config
        (booking_id, organization, event_name, hotel, contracted_block, contracted_rate, status)
        VALUES (?,?,?,?,?,?,'active')
    ''', (
        str(bid),
        account or 'Unknown',
        event or None,
        hotel or None,
        json.dumps(contracted_block),
        rate,
    ))
    return True


def sync_pickup_from_pipeline(db=None):
    """Update pickup_config organization/event_name from ReportPipeline for matched booking IDs."""
    close_after = db is None
    if db is None:
        db = get_db()
    configs = db.execute(
        "SELECT id, booking_id FROM pickup_config WHERE booking_id IS NOT NULL AND booking_id != ''"
    ).fetchall()
    updated = 0
    for c in configs:
        row = db.execute(
            'SELECT AccountName, EventName FROM ReportPipeline WHERE CAST(BookingId AS INTEGER) = CAST(? AS INTEGER) LIMIT 1',
            (c['booking_id'],)
        ).fetchone()
        if row:
            db.execute(
                'UPDATE pickup_config SET organization=COALESCE(?,organization), event_name=COALESCE(?,event_name) WHERE id=?',
                (row['AccountName'] or None, row['EventName'] or None, c['id'])
            )
            if db.execute('SELECT changes()').fetchone()[0]:
                updated += 1
    db.commit()
    if close_after:
        db.close()
    return updated


def create_pickup_configs_from_pipeline(conn=None):
    """
    Create pickup_config records for future ReportPipeline bookings that don't
    have one yet. Builds a day-by-day contracted_block from StartDate → EndDate
    using PeakRooms each night. Returns the number of new records created.
    """
    from datetime import datetime as _dt, timedelta as _td
    close_after = conn is None
    if conn is None:
        conn = get_db()
    today = _dt.today().strftime('%Y-%m-%d')
    rows = conn.execute('''
        SELECT BookingId AS booking_id, BookingName AS booking_name,
               EventName AS event_name, AccountName AS org,
               StartDate AS start_date, EndDate AS end_date,
               PeakRooms AS peak_rooms, RoomRate AS room_rate, Customer AS hotel
        FROM ReportPipeline
        WHERE EndDate > ?
          AND COALESCE(BookingStatus,'') NOT IN ('Lost','Cancelled','Turned Down')
          AND PeakRooms IS NOT NULL AND PeakRooms != ''
          AND CAST(PeakRooms AS REAL) > 0
          AND NOT EXISTS (
              SELECT 1 FROM pickup_config pc
              WHERE CAST(pc.booking_id AS INTEGER) = CAST(BookingId AS INTEGER)
          )
    ''', (today,)).fetchall()
    created = 0
    seen_bids = set()  # guard against duplicate booking_ids in ReportPipeline
    for r in rows:
        try:
            start_str = (r['start_date'] or '')[:10]
            end_str   = (r['end_date']   or '')[:10]
            if not start_str or not end_str:
                continue
            start_d = _dt.strptime(start_str, '%Y-%m-%d')
            end_d   = _dt.strptime(end_str,   '%Y-%m-%d')
            peak    = int(float(r['peak_rooms']))
            rate    = float(r['room_rate']) if r['room_rate'] else None
            block   = {}
            cur = start_d
            while cur < end_d:
                block[cur.strftime('%Y-%m-%d')] = peak
                cur += _td(days=1)
            if not block:
                continue
            cutoff     = (start_d - _td(days=30)).strftime('%Y-%m-%d')
            org        = r['org'] or ''
            event_name = r['event_name'] or r['booking_name'] or org
            hotel      = r['hotel'] or ''
            booking_id = str(r['booking_id']).split('.')[0]
            if booking_id in seen_bids:
                continue
            seen_bids.add(booking_id)
            conn.execute('''
                INSERT INTO pickup_config
                    (booking_id, organization, event_name, hotel, contracted_block,
                     contracted_rate, cutoff_date, attrition_pct, status,
                     hotel_contacts, cc_emails, force_current, force_past)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'active', '[]', '[]', 0, 0)
            ''', (booking_id, org, event_name, hotel,
                  json.dumps(block), rate, cutoff, 0.80))
            created += 1
        except Exception:
            continue
    if created:
        conn.commit()
    if close_after:
        conn.close()
    return created


# ── API: booking rate lookup (for pickup config form) ─────────────────────────

@app.route('/api/booking-rate')
def api_booking_rate():
    booking_id = request.args.get('booking_id', '')
    if not booking_id:
        return jsonify({'room_rate': None})
    db = get_db()
    row = db.execute(
        'SELECT RoomRate, AccountName, EventName FROM ReportPipeline WHERE CAST(BookingId AS INTEGER) = CAST(? AS INTEGER) LIMIT 1',
        (booking_id,)
    ).fetchone()
    if row:
        return jsonify({
            'room_rate':  round(float(row['RoomRate']), 2) if row['RoomRate'] else None,
            'account':    row['AccountName'] or '',
            'event_name': row['EventName'] or '',
        })
    return jsonify({'room_rate': None, 'account': '', 'event_name': ''})


# ── Pickup Tracking Routes ────────────────────────────────────────────────────

@app.route('/pickup/sync-pipeline', methods=['POST'])
def pickup_sync_pipeline():
    synced  = sync_pickup_from_pipeline()
    created = create_pickup_configs_from_pipeline()
    parts = []
    if created:
        parts.append(f'{created} new event{"s" if created != 1 else ""} created')
    if synced:
        parts.append(f'{synced} event{"s" if synced != 1 else ""} updated')
    if parts:
        flash('Sync complete: ' + ', '.join(parts) + '.', 'success')
    else:
        flash('Already up to date — nothing to sync.', 'info')
    return redirect(url_for('pickup_dashboard'))


@app.route('/pickup/fill-missing', methods=['GET', 'POST'])
def pickup_fill_missing():
    """Find future bookings that have no pickup_config entry and create them."""
    user = get_current_user()
    if not has_permission(user, 'pickups_payments'):
        flash('Access denied.', 'error')
        return redirect(url_for('pipeline'))

    db           = get_db()
    today        = datetime.today().strftime('%Y-%m-%d')
    acct_filter  = get_pickup_account_filter(user)   # None = no filter, [] = none, [...] = allowed list

    # Base query: future non-cancelled Kristin House bookings without a pickup_config entry
    base_sql = '''
        SELECT r.BookingId, r.BookingName, r.EventName, r.AccountName,
               r.Customer, r.StartDate, r.EndDate, r.PeakRooms, r.RoomRate,
               r.BookingAssociate, r.BookingStatus
        FROM ReportPipeline r
        WHERE (r.BookingStatus IS NULL OR r.BookingStatus NOT LIKE '%Cancel%')
          AND r.EndDate >= ?
          AND r.BookingAssociate = 'Kristin House'
          AND NOT EXISTS (
              SELECT 1 FROM pickup_config p WHERE p.booking_id = CAST(r.BookingId AS TEXT)
          )
    '''
    params = [today]

    if acct_filter is None:
        # Admin — no account restriction
        pass
    elif acct_filter:
        ph = ','.join('?' * len(acct_filter))
        base_sql += f' AND r.AccountName IN ({ph})'
        params.extend(acct_filter)
    else:
        # User has no accounts assigned — show nothing
        base_sql += ' AND 1=0'

    base_sql += ' ORDER BY r.StartDate'
    missing = db.execute(base_sql, params).fetchall()

    if request.method == 'POST':
        selected_ids = request.form.getlist('booking_ids')
        created = 0
        for row in missing:
            bid = str(row['BookingId'])
            if bid not in selected_ids:
                continue
            ok = _create_pickup_config_from_booking(
                db, bid,
                account    = row['AccountName'],
                event      = row['BookingName'] or row['EventName'],
                hotel      = row['Customer'],
                start_str  = row['StartDate'],
                end_str    = row['EndDate'],
                peak_rooms = row['PeakRooms'],
                room_rate  = row['RoomRate'],
            )
            if ok:
                created += 1
        db.commit()
        flash(f'{created} pickup tracking event{"s" if created != 1 else ""} created.', 'success')
        return redirect(url_for('pickup_dashboard'))

    return render_template('pickup_fill_missing.html', missing=missing, today=today)


# ── Contact-book helpers ──────────────────────────────────────────────────────

def _upsert_contacts(db, hotel_contacts, hotel_name, client_contacts, organization):
    """Save hotel and client contacts to their respective lookup tables."""
    for c in hotel_contacts:
        name  = (c.get('name')  or '').strip()
        email = (c.get('email') or '').strip().lower()
        if not email:
            continue
        db.execute('''
            INSERT INTO hotel_contact_list (name, email, hotel_name, created_at, updated_at)
            VALUES (?, ?, ?, datetime('now'), datetime('now'))
            ON CONFLICT(email) DO UPDATE SET
                name       = excluded.name,
                hotel_name = COALESCE(excluded.hotel_name, hotel_name),
                updated_at = datetime('now')
        ''', (name, email, hotel_name or None))

    for c in client_contacts:
        name  = (c.get('name')  or '').strip()
        email = (c.get('email') or '').strip().lower()
        if not email:
            continue
        db.execute('''
            INSERT INTO client_contacts (name, email, organization, created_at, updated_at)
            VALUES (?, ?, ?, datetime('now'), datetime('now'))
            ON CONFLICT(email) DO UPDATE SET
                name         = excluded.name,
                organization = COALESCE(excluded.organization, organization),
                updated_at   = datetime('now')
        ''', (name, email, organization or None))


@app.route('/api/contacts/clients')
def api_client_contacts():
    q = request.args.get('q', '').strip()
    if len(q) < 1:
        return jsonify([])
    db = get_db()
    rows = db.execute(
        '''SELECT name, email, organization FROM client_contacts
           WHERE name LIKE ? OR email LIKE ? OR organization LIKE ?
           ORDER BY name LIMIT 20''',
        (f'%{q}%', f'%{q}%', f'%{q}%')
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/contacts/hotels')
def api_hotel_contacts():
    q = request.args.get('q', '').strip()
    if len(q) < 1:
        return jsonify([])
    db = get_db()
    rows = db.execute(
        '''SELECT name, email, hotel_name FROM hotel_contact_list
           WHERE name LIKE ? OR email LIKE ? OR hotel_name LIKE ?
           ORDER BY name LIMIT 20''',
        (f'%{q}%', f'%{q}%', f'%{q}%')
    ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/contacts')
def contacts_page():
    db = get_db()
    clients = [dict(r) for r in db.execute('SELECT * FROM client_contacts ORDER BY name').fetchall()]
    hotels  = [dict(r) for r in db.execute('SELECT * FROM hotel_contact_list ORDER BY name').fetchall()]
    return render_template('contacts.html', clients=clients, hotels=hotels)


@app.route('/contacts/delete/<string:tbl>/<int:cid>', methods=['POST'])
def contact_delete(tbl, cid):
    if tbl not in ('client', 'hotel'):
        abort(400)
    table = 'client_contacts' if tbl == 'client' else 'hotel_contact_list'
    db = get_db()
    db.execute(f'DELETE FROM {table} WHERE id=?', (cid,))
    db.commit()
    flash('Contact deleted.', 'success')
    return redirect(url_for('contacts_page') + ('?tab=hotels' if tbl == 'hotel' else ''))


_CONTACT_FIELDS = ['name', 'email', 'title', 'phone_office', 'phone_cell',
                   'fax', 'address', 'city', 'state', 'zip', 'website', 'linkedin', 'notes']


@app.route('/contacts/add/<string:tbl>', methods=['POST'])
def contact_add(tbl):
    if tbl not in ('client', 'hotel'):
        abort(400)
    db = get_db()
    f = request.form
    name  = (f.get('name')  or '').strip()
    email = (f.get('email') or '').strip().lower()
    if not name or not email:
        flash('Name and email are required.', 'danger')
        return redirect(url_for('contacts_page') + ('?tab=hotels' if tbl == 'hotel' else ''))
    if tbl == 'client':
        db.execute('''
            INSERT INTO client_contacts
              (name, email, organization, title, phone_office, phone_cell, fax,
               address, city, state, zip, website, linkedin, notes, created_at, updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,datetime('now'),datetime('now'))
            ON CONFLICT(email) DO UPDATE SET
              name=excluded.name, organization=excluded.organization,
              title=excluded.title, phone_office=excluded.phone_office,
              phone_cell=excluded.phone_cell, fax=excluded.fax,
              address=excluded.address, city=excluded.city, state=excluded.state,
              zip=excluded.zip, website=excluded.website, linkedin=excluded.linkedin,
              notes=excluded.notes, updated_at=datetime('now')
        ''', (name, email, f.get('organization','').strip() or None,
              f.get('title','').strip() or None, f.get('phone_office','').strip() or None,
              f.get('phone_cell','').strip() or None, f.get('fax','').strip() or None,
              f.get('address','').strip() or None, f.get('city','').strip() or None,
              f.get('state','').strip() or None, f.get('zip','').strip() or None,
              f.get('website','').strip() or None, f.get('linkedin','').strip() or None,
              f.get('notes','').strip() or None))
    else:
        db.execute('''
            INSERT INTO hotel_contact_list
              (name, email, hotel_name, title, phone_office, phone_cell, fax,
               address, city, state, zip, website, linkedin, notes, created_at, updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,datetime('now'),datetime('now'))
            ON CONFLICT(email) DO UPDATE SET
              name=excluded.name, hotel_name=excluded.hotel_name,
              title=excluded.title, phone_office=excluded.phone_office,
              phone_cell=excluded.phone_cell, fax=excluded.fax,
              address=excluded.address, city=excluded.city, state=excluded.state,
              zip=excluded.zip, website=excluded.website, linkedin=excluded.linkedin,
              notes=excluded.notes, updated_at=datetime('now')
        ''', (name, email, f.get('hotel_name','').strip() or None,
              f.get('title','').strip() or None, f.get('phone_office','').strip() or None,
              f.get('phone_cell','').strip() or None, f.get('fax','').strip() or None,
              f.get('address','').strip() or None, f.get('city','').strip() or None,
              f.get('state','').strip() or None, f.get('zip','').strip() or None,
              f.get('website','').strip() or None, f.get('linkedin','').strip() or None,
              f.get('notes','').strip() or None))
    db.commit()
    flash('Contact saved.', 'success')
    return redirect(url_for('contacts_page') + ('?tab=hotels' if tbl == 'hotel' else ''))


@app.route('/contacts/edit/<string:tbl>/<int:cid>', methods=['POST'])
def contact_edit(tbl, cid):
    if tbl not in ('client', 'hotel'):
        abort(400)
    db = get_db()
    f = request.form
    name  = (f.get('name')  or '').strip()
    email = (f.get('email') or '').strip().lower()
    if not name or not email:
        flash('Name and email are required.', 'danger')
        return redirect(url_for('contacts_page') + ('?tab=hotels' if tbl == 'hotel' else ''))
    extra_col = 'organization' if tbl == 'client' else 'hotel_name'
    extra_val = f.get(extra_col, '').strip() or None
    db.execute(f'''
        UPDATE {"client_contacts" if tbl=="client" else "hotel_contact_list"}
        SET name=?, email=?, {extra_col}=?, title=?, phone_office=?, phone_cell=?,
            fax=?, address=?, city=?, state=?, zip=?, website=?, linkedin=?, notes=?,
            updated_at=datetime('now')
        WHERE id=?
    ''', (name, email, extra_val,
          f.get('title','').strip() or None, f.get('phone_office','').strip() or None,
          f.get('phone_cell','').strip() or None, f.get('fax','').strip() or None,
          f.get('address','').strip() or None, f.get('city','').strip() or None,
          f.get('state','').strip() or None, f.get('zip','').strip() or None,
          f.get('website','').strip() or None, f.get('linkedin','').strip() or None,
          f.get('notes','').strip() or None, cid))
    db.commit()
    flash('Contact updated.', 'success')
    return redirect(url_for('contacts_page') + ('?tab=hotels' if tbl == 'hotel' else ''))


@app.route('/pickup/auto-match', methods=['GET', 'POST'])
def pickup_auto_match():
    """Find pickup configs missing booking_id and suggest matches from ReportPipeline."""
    user = get_current_user()
    if not has_permission(user, 'pickups_payments'):
        flash('Access denied.', 'error')
        return redirect(url_for('pickup_dashboard'))
    db = get_db()

    if request.method == 'POST':
        # Apply confirmed matches
        applied = 0
        for key, val in request.form.items():
            if key.startswith('match_') and val:
                cid = int(key.replace('match_', ''))
                db.execute('UPDATE pickup_config SET booking_id=? WHERE id=? AND (booking_id IS NULL OR booking_id="")',
                           (val, cid))
                applied += 1
        db.commit()
        flash(f'Applied {applied} booking ID match(es).', 'success')
        return redirect(url_for('pickup_dashboard'))

    # Find unmatched configs
    unmatched = db.execute(
        "SELECT * FROM pickup_config WHERE (booking_id IS NULL OR booking_id='') AND status='active'"
    ).fetchall()

    results = []
    for c in unmatched:
        org   = (c['organization'] or '').strip().lower()
        event = (c['event_name'] or '').strip().lower()
        hotel = (c['hotel'] or '').strip().lower()

        # Score each booking: +3 org match, +3 event match, +2 hotel match
        bookings = db.execute(
            "SELECT BookingId, AccountName, EventName, Customer, StartDate, EndDate "
            "FROM ReportPipeline WHERE BookingStatus NOT LIKE '%Cancel%' OR BookingStatus IS NULL"
        ).fetchall()

        scored = []
        for b in bookings:
            score = 0
            b_org   = (b['AccountName'] or '').strip().lower()
            b_event = (b['EventName'] or '').strip().lower()
            b_hotel = (b['Customer'] or '').strip().lower()
            if org   and b_org   and (org in b_org   or b_org in org):   score += 3
            if event and b_event and (event in b_event or b_event in event): score += 3
            if hotel and b_hotel and (hotel in b_hotel or b_hotel in hotel): score += 2
            if score >= 3:
                scored.append((score, dict(b)))

        scored.sort(key=lambda x: -x[0])
        results.append({'config': c, 'suggestions': scored[:5]})

    return render_template('pickup_auto_match.html', results=results)


@app.route('/pickup')
def pickup_dashboard():
    from datetime import date, timedelta
    user = get_current_user()
    db = get_db()
    today = date.today()
    today_str = today.isoformat()
    future_cutoff = today + timedelta(days=120)

    # Account-level filter (applies to all users including admins for Pickup)
    acct_filter = get_pickup_account_filter(user)
    if acct_filter is None:
        # no restriction
        acct_where = ''
        acct_params = []
    elif acct_filter:
        ph = ','.join('?' * len(acct_filter))
        acct_where = f' AND (p.AccountName IN ({ph}) OR c.organization IN ({ph}))'
        acct_params = acct_filter + acct_filter
    else:
        acct_where = ' AND 1=0'
        acct_params = []

    configs = db.execute(f"""
        SELECT c.*,
               p.StartDate  AS bk_start,
               p.EndDate    AS bk_end,
               p.EventName  AS bk_event,
               p.AccountName AS bk_org,
               p.Customer   AS bk_hotel,
               p.PeakRooms  AS bk_peak_rooms,
               p.TotalRoomNights AS bk_total_rn
        FROM pickup_config c
        LEFT JOIN ReportPipeline p ON CAST(c.booking_id AS TEXT)=CAST(p.BookingId AS TEXT)
        WHERE c.status='active'{acct_where} ORDER BY c.cutoff_date
    """, acct_params).fetchall()
    archived_configs = db.execute(f"""
        SELECT c.*,
               p.StartDate AS bk_start, p.EndDate AS bk_end
        FROM pickup_config c
        LEFT JOIN ReportPipeline p ON CAST(c.booking_id AS TEXT)=CAST(p.BookingId AS TEXT)
        WHERE c.status='archived'{acct_where} ORDER BY c.cutoff_date
    """, acct_params).fetchall()
    past_rows, current_rows, future_rows, archived_rows, missing_hhr_rows = [], [], [], [], []

    # Pre-compute which booking_ids have HHR uploaded (needed for categorization)
    has_hhr_bids = {str(r[0]) for r in db.execute(
        "SELECT DISTINCT booking_id FROM housing_history_files WHERE booking_id IS NOT NULL"
    ).fetchall()}

    for c in configs:
        last = db.execute(
            """SELECT * FROM pickup_weekly WHERE config_id=?
               AND (total_rooms IS NOT NULL AND total_rooms > 0)
               AND (label IS NULL OR (label NOT LIKE 'ASAS %' AND label NOT LIKE 'Historical%'
                                      AND label NOT LIKE '%Final%'))
               ORDER BY report_date DESC LIMIT 1""",
            (c['id'],)
        ).fetchone()
        last_any = db.execute(
            "SELECT report_date FROM pickup_weekly WHERE config_id=? ORDER BY report_date DESC LIMIT 1",
            (c['id'],)
        ).fetchone()
        all_weekly = db.execute("SELECT label FROM pickup_weekly WHERE config_id=?", (c['id'],)).fetchall()
        last_contact = db.execute(
            "SELECT contact_date, contact_type FROM pickup_contact_log WHERE config_id=? AND contact_type='email_sent' ORDER BY contact_date DESC, id DESC LIMIT 1",
            (c['id'],)
        ).fetchone()
        responded_row = db.execute(
            "SELECT contact_date FROM pickup_contact_log WHERE config_id=? AND contact_type='responded' ORDER BY contact_date DESC LIMIT 1",
            (c['id'],)
        ).fetchone()
        has_response  = bool(responded_row)
        responded_date = responded_row['contact_date'] if responded_row else None
        block = json.loads(c['contracted_block'] or '{}')
        contracted_total = sum(block.values())
        last_total = last['total_rooms'] if last else 0
        last_pct_block = last['pct_of_block'] if last else None
        last_pct_attr  = last['pct_of_attrition'] if last else None
        last_date      = last_any['report_date'] if last_any else None
        last_ota       = last['ota_rate'] if last else None
        if last_pct_block is None:
            badge, badge_label = 'secondary', 'No data'
        elif last_pct_block >= 80:
            badge, badge_label = 'success', 'On Pace'
        elif last_pct_block >= 60:
            badge, badge_label = 'warning', 'Watch'
        else:
            badge, badge_label = 'danger', 'At Risk'
        days_to_cutoff = None
        if c['cutoff_date']:
            try:
                days_to_cutoff = (date.fromisoformat(c['cutoff_date']) - today).days
            except Exception:
                pass
        has_final_history = any(w['label'] and 'final' in w['label'].lower() for w in all_weekly)
        all_dates = sorted(block.keys())
        # Priority: 1) ReportPipeline (bk_start/bk_end), 2) manually-set pickup_config.event_start/end, 3) contracted_block keys
        # bk_start may come as ISO datetime "2026-07-24T00:00:00.000Z" — strip time part
        event_start = _to_iso(c['bk_start']) or c['event_start'] or (all_dates[0]  if all_dates else None)
        event_end   = _to_iso(c['bk_end'])   or c['event_end']   or (all_dates[-1] if all_dates else None)
        force_past = bool(c['force_past']) if c['force_past'] else False
        row = {
            'config': c, 'contracted_total': contracted_total,
            'last_total': last_total, 'last_pct_block': last_pct_block,
            'last_pct_attr': last_pct_attr, 'last_date': last_date,
            'last_ota': last_ota, 'days_to_cutoff': days_to_cutoff,
            'badge': badge, 'badge_label': badge_label,
            'has_final_history': has_final_history,
            'event_start': event_start, 'event_end': event_end,
            'force_current': bool(c['force_current']),
            'force_past': force_past,
            'last_contact': last_contact,
            'has_response': has_response, 'responded_date': responded_date,
        }
        start_date = None
        for _sd in [_to_iso(c['bk_start']),
                    c['event_start'],
                    all_dates[0] if all_dates else None]:
            if _sd:
                try:
                    start_date = date.fromisoformat(_sd)
                    break
                except Exception:
                    pass
        end_date = None
        for _ed in [_to_iso(c['bk_end']),
                    c['event_end'],
                    all_dates[-1] if all_dates else None]:
            if _ed:
                try:
                    end_date = date.fromisoformat(_ed)
                    break
                except Exception:
                    pass
        meeting_ended = end_date is not None and end_date < today
        has_hhr = str(c['booking_id'] or '') in has_hhr_bids
        if force_past or (has_final_history and meeting_ended):
            past_rows.append(row)
        elif meeting_ended and has_hhr:
            past_rows.append(row)
        elif meeting_ended and not has_hhr:
            missing_hhr_rows.append(row)
        elif c['force_current']:
            current_rows.append(row)
        elif start_date is not None and start_date > future_cutoff:
            future_rows.append(row)
        else:
            current_rows.append(row)

    for c in archived_configs:
        block = json.loads(c['contracted_block'] or '{}')
        all_dates = sorted(block.keys())
        archived_rows.append({
            'config': c,
            'event_start': _to_iso(c['bk_start']) or c['event_start'] or (all_dates[0]  if all_dates else None),
            'event_end':   _to_iso(c['bk_end'])   or c['event_end']   or (all_dates[-1] if all_dates else None),
        })

    sort_mode = request.args.get('sort', 'date')
    if sort_mode == 'customer':
        _cur_sort  = lambda r: ((r['config']['organization'] or r['config']['bk_org'] or '').lower(),
                                r.get('event_start') or '')
        current_rows.sort(key=_cur_sort)
    else:
        _sort_key = lambda r: r.get('event_start') or ''
        current_rows.sort(key=_sort_key)
    _sort_key = lambda r: r.get('event_start') or ''
    past_rows.sort(key=_sort_key)
    missing_hhr_rows.sort(key=_sort_key)
    future_rows.sort(key=_sort_key)
    archived_rows.sort(key=_sort_key)

    # ── Build event groups (events with >1 hotel in the same section) ──────────
    import re as _re_group
    def _slugify(s):
        return _re_group.sub(r'[^a-z0-9]+', '-', (s or '').lower()).strip('-')

    def _build_groups(rows):
        """Return (groups_dict, group_order) for events with >1 hotel in rows."""
        from collections import defaultdict
        bucket = defaultdict(list)
        for r in rows:
            key = (r['config']['event_name'] or '').strip().lower()
            bucket[key].append(r)
        groups = {}
        group_order = []
        for key, grp_rows in bucket.items():
            if len(grp_rows) > 1:
                combined_block = sum(r['contracted_total'] for r in grp_rows)
                combined_total = sum(r['last_total'] for r in grp_rows)
                if combined_block:
                    combined_pct = round(combined_total / combined_block * 100, 1)
                else:
                    combined_pct = None
                if combined_pct is None:
                    grp_badge, grp_badge_label = 'secondary', 'No data'
                elif combined_pct >= 80:
                    grp_badge, grp_badge_label = 'success', 'On Pace'
                elif combined_pct >= 60:
                    grp_badge, grp_badge_label = 'warning', 'Watch'
                else:
                    grp_badge, grp_badge_label = 'danger', 'At Risk'
                starts = [r['event_start'] for r in grp_rows if r['event_start']]
                ends   = [r['event_end']   for r in grp_rows if r['event_end']]
                dates  = [r['last_date']   for r in grp_rows if r['last_date']]
                rep_config = grp_rows[0]['config']
                display_name = (rep_config['event_name'] or '').strip() or key
                # earliest cutoff across hotels
                cutoffs = [r['config']['cutoff_date'] for r in grp_rows if r['config']['cutoff_date']]
                groups[key] = {
                    'event_name':     display_name,
                    'organization':   rep_config['organization'] or rep_config['bk_org'] or '',
                    'hotels':         grp_rows,
                    'combined_block': combined_block,
                    'combined_total': combined_total,
                    'combined_pct':   combined_pct,
                    'event_start':    min(starts) if starts else None,
                    'event_end':      max(ends)   if ends   else None,
                    'last_date':      max(dates)  if dates  else None,
                    'cutoff_date':    min(cutoffs) if cutoffs else None,
                    'badge':          grp_badge,
                    'badge_label':    grp_badge_label,
                    'safe_id':        _slugify(display_name),
                    'primary_cid':    min(r['config']['id'] for r in grp_rows),
                }
                group_order.append(key)
        return groups, group_order

    current_groups,     current_group_order     = _build_groups(current_rows)
    future_groups,      future_group_order      = _build_groups(future_rows)
    past_groups,        past_group_order        = _build_groups(past_rows)
    missing_hhr_groups, missing_hhr_group_order = _build_groups(missing_hhr_rows)

    # ── KPI metrics ──────────────────────────────────────────────────────────
    def _sum_block(rows):
        t = 0
        for r in rows:
            try:
                t += sum(json.loads(r['config']['contracted_block'] or '{}').values())
            except Exception:
                pass
        return t

    kpi_active_events      = len(current_rows)
    kpi_future_events      = len(future_rows)
    kpi_contracted_rn      = _sum_block(current_rows) + _sum_block(future_rows)
    kpi_current_pickup     = sum((r['last_total'] or 0) for r in current_rows)
    kpi_contracted_rn_cur  = _sum_block(current_rows)
    kpi_pickup_pct         = round(kpi_current_pickup / kpi_contracted_rn_cur * 100, 1) if kpi_contracted_rn_cur else None

    # Status board open items — count from issues_by_type logic (quick query)
    kpi_status_items = db.execute(
        "SELECT COUNT(*) FROM pickup_config WHERE status='active'"
    ).fetchone()[0]  # placeholder — actual count needs status board logic; show total active

    # Events missing HHR (ended with no HHR uploaded) — now its own section
    kpi_missing_hhr = len(missing_hhr_rows)

    kpis = {
        'active_events':   kpi_active_events,
        'future_events':   kpi_future_events,
        'contracted_rn':   kpi_contracted_rn,
        'current_pickup':  kpi_current_pickup,
        'pickup_pct':      kpi_pickup_pct,
        'missing_hhr':     kpi_missing_hhr,
    }

    return render_template('pickup_dashboard.html',
                           past_rows=past_rows, current_rows=current_rows,
                           future_rows=future_rows, archived_rows=archived_rows,
                           missing_hhr_rows=missing_hhr_rows,
                           current_groups=current_groups, current_group_order=current_group_order,
                           future_groups=future_groups,   future_group_order=future_group_order,
                           past_groups=past_groups,       past_group_order=past_group_order,
                           missing_hhr_groups=missing_hhr_groups,
                           missing_hhr_group_order=missing_hhr_group_order,
                           today=today_str, sort_mode=sort_mode, kpis=kpis)


# ── Import pickup data from Excel (NCSL-style master spreadsheet) ─────────────

@app.route('/pickup/import-xlsx', methods=['GET', 'POST'])
def pickup_import_xlsx():
    if request.method == 'GET':
        return render_template('pickup_import_xlsx.html')

    f = request.files.get('xlsx_file')
    if not f or not f.filename:
        flash('Please select an Excel file.', 'warning')
        return redirect(url_for('pickup_import_xlsx'))

    file_bytes = f.read()
    try:
        from pickup_utils import parse_pickup_xlsx
        grids = parse_pickup_xlsx(file_bytes)
    except Exception as e:
        flash(f'Error parsing file: {e}', 'danger')
        return redirect(url_for('pickup_import_xlsx'))

    if not grids:
        flash('No valid hotel grids found in the file.', 'warning')
        return redirect(url_for('pickup_import_xlsx'))

    # Check which booking IDs already have pickup_config records
    db = get_db()
    existing = {}
    for g in grids:
        bid = g['booking_id']
        if bid:
            row = db.execute(
                'SELECT id, event_name, hotel FROM pickup_config WHERE booking_id=?', (bid,)
            ).fetchone()
            if row:
                existing[bid] = dict(row)

    import json as _json
    grids_json = _json.dumps(grids)
    return render_template('pickup_import_xlsx_review.html',
                           grids=grids, grids_json=grids_json, existing=existing,
                           filename=f.filename)


@app.route('/pickup/import-xlsx/confirm', methods=['POST'])
def pickup_import_xlsx_confirm():
    import json as _json
    from datetime import date
    grids = _json.loads(request.form.get('grids_json', '[]'))
    selected_indices = set(request.form.getlist('selected'))
    db = get_db()
    today = date.today().isoformat()

    created = updated = skipped = pickups_added = 0
    for i, g in enumerate(grids):
        if str(i) not in selected_indices:
            skipped += 1
            continue

        bid = g.get('booking_id') or None
        contracted_block = g.get('contracted_block') or {}
        contracted_rate  = g.get('contracted_rate')
        attrition_pct    = g.get('attrition_pct')
        pickups          = g.get('pickups', [])

        # Contacts
        hc_name   = g.get('contact_name', '') or ''
        hc_phone  = g.get('contact_phone', '') or ''
        hc_email  = g.get('contact_email', '') or ''
        gc_name   = g.get('gc_name', '') or ''
        gc_email  = g.get('gc_email', '') or ''

        # Derive cutoff_date from last date in contracted_block
        cutoff_date = None
        if contracted_block:
            cutoff_date = max(contracted_block.keys())

        # Build hotel_contacts JSON (includes phone)
        hc = []
        if hc_name or hc_email:
            entry = {'name': hc_name, 'email': hc_email}
            if hc_phone:
                entry['phone'] = hc_phone
            hc = [entry]

        # Check existing
        existing_row = None
        if bid:
            existing_row = db.execute(
                'SELECT id FROM pickup_config WHERE booking_id=?', (bid,)
            ).fetchone()

        if existing_row:
            cid = existing_row['id']
            db.execute('''UPDATE pickup_config SET
                contracted_block=?, contracted_rate=?, attrition_pct=?, cutoff_date=?,
                hotel_contact=?, hotel_contact_email=?, hotel_contacts=?,
                group_contact=?, group_contact_email=?
                WHERE id=?''',
                (_json.dumps(contracted_block), contracted_rate, attrition_pct,
                 cutoff_date,
                 hc_name or None, hc_email or None, _json.dumps(hc),
                 gc_name or None, gc_email or None,
                 cid))
            updated += 1
        else:
            db.execute('''INSERT INTO pickup_config
                (booking_id, organization, event_name, hotel, contracted_block, contracted_rate,
                 attrition_pct, cutoff_date, hotel_contact, hotel_contact_email, hotel_contacts,
                 group_contact, group_contact_email, cc_emails, status, force_current, force_past)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                (bid, g.get('organization', 'NCSL'), g.get('event_name', ''),
                 g.get('hotel', ''),
                 _json.dumps(contracted_block), contracted_rate, attrition_pct,
                 cutoff_date,
                 hc_name or None, hc_email or None, _json.dumps(hc),
                 gc_name or None, gc_email or None,
                 '[]', 'active', 0, 0))
            cid = db.execute('SELECT last_insert_rowid()').fetchone()[0]
            created += 1

        # Insert pickup_weekly records (skip duplicates by report_date)
        existing_dates = {r['report_date'] for r in db.execute(
            'SELECT report_date FROM pickup_weekly WHERE config_id=?', (cid,)).fetchall()}

        for p in pickups:
            rd = p.get('report_date')
            if not rd or rd in existing_dates:
                continue
            db.execute('''INSERT INTO pickup_weekly
                (config_id, report_date, pickup_by_night, total_rooms, change_from_last,
                 pct_of_block, pct_of_attrition, label, notes)
                VALUES (?,?,?,?,?,?,?,?,?)''',
                (cid, rd, _json.dumps(p.get('pickup_by_night', {})),
                 p.get('total_rooms'), p.get('change_from_last'),
                 p.get('pct_of_block'), p.get('pct_of_attrition'), None, None))
            pickups_added += 1
            existing_dates.add(rd)

    db.commit()
    parts = []
    if created:   parts.append(f'{created} event{"s" if created != 1 else ""} created')
    if updated:   parts.append(f'{updated} event{"s" if updated != 1 else ""} updated')
    if skipped:   parts.append(f'{skipped} skipped')
    if pickups_added: parts.append(f'{pickups_added} pickup record{"s" if pickups_added != 1 else ""} added')
    flash('Import complete: ' + ', '.join(parts) + '.', 'success')
    return redirect(url_for('pickup_dashboard'))


@app.route('/pickup/new', methods=['GET', 'POST'])
def pickup_new_event():
    if request.method == 'POST':
        f = request.form
        contracted_block = {}
        for d, r in zip(f.getlist('block_date'), f.getlist('block_rooms')):
            if d and r:
                contracted_block[d] = int(r)
        attrition_raw = f.get('attrition_pct', '')
        attrition = float(attrition_raw) / 100 if attrition_raw else None
        ota_url = f.get('ota_url', '').strip() or None
        hc_name   = f.get('hotel_contact', '').strip() or None
        hc_email  = f.get('hotel_contact_email', '').strip() or None
        hc2_name  = f.get('hotel_contact2', '').strip() or hc_name   # default = primary
        hc2_email = f.get('hotel_contact2_email', '').strip() or hc_email
        gc_name  = f.get('group_contact', '').strip() or None
        gc_email = f.get('group_contact_email', '').strip() or None
        hotel_contacts = [{'name': hc_name or '', 'email': hc_email or ''}] if (hc_name or hc_email) else []
        # Additional CC recipients
        cc_emails = []
        for n, e in zip(f.getlist('cc_name[]'), f.getlist('cc_email[]')):
            if n.strip() or e.strip():
                cc_emails.append({'name': n.strip(), 'email': e.strip()})
        db = get_db()

        # ── Duplicate detection ──────────────────────────────────────────────
        new_bid   = (f.get('booking_id') or '').strip()
        new_event = (f.get('event_name') or '').strip().lower()
        new_hotel = (f.get('hotel') or '').strip().lower()
        duplicates = []
        if new_bid:
            bid_dups = db.execute(
                "SELECT id, event_name, hotel, status FROM pickup_config "
                "WHERE CAST(booking_id AS TEXT)=CAST(? AS TEXT)",
                (new_bid,)
            ).fetchall()
            for d in bid_dups:
                duplicates.append(
                    f"Booking ID {new_bid} already exists: "
                    f"<strong>{d['event_name']}</strong> @ {d['hotel']} "
                    f"({'archived' if d['status']=='archived' else 'active'}) — "
                    f"<a href=\"{url_for('pickup_event', cid=d['id'])}\">View</a>"
                )
        if new_event and new_hotel:
            name_dups = db.execute(
                "SELECT id, event_name, hotel, status FROM pickup_config "
                "WHERE LOWER(TRIM(event_name))=? AND LOWER(TRIM(hotel))=? "
                "AND (? = '' OR CAST(booking_id AS TEXT) != ?)",
                (new_event, new_hotel, new_bid, new_bid)
            ).fetchall()
            for d in name_dups:
                duplicates.append(
                    f"Same event name + hotel already exists: "
                    f"<strong>{d['event_name']}</strong> @ {d['hotel']} "
                    f"({'archived' if d['status']=='archived' else 'active'}) — "
                    f"<a href=\"{url_for('pickup_event', cid=d['id'])}\">View</a>"
                )
        if duplicates and not f.get('confirm_duplicate'):
            # Re-render the form with warnings — user must tick confirm to proceed
            pipeline_rate = pipeline_org = pipeline_event = None
            pipeline_start = pipeline_end = None
            if new_bid:
                row = db.execute(
                    'SELECT RoomRate, AccountName, EventName, StartDate, EndDate FROM ReportPipeline WHERE CAST(BookingId AS INTEGER)=CAST(? AS INTEGER) LIMIT 1',
                    (new_bid,)
                ).fetchone()
                if row:
                    pipeline_rate  = round(float(row['RoomRate']), 2) if row['RoomRate'] else None
                    pipeline_org   = row['AccountName'] or None
                    pipeline_event = row['EventName'] or None
                    if row['StartDate']:
                        pipeline_start = str(row['StartDate'])[:10]
                    if row['EndDate']:
                        pipeline_end = str(row['EndDate'])[:10]
            return render_template('pickup_config_form.html', config=None,
                                   action=url_for('pickup_new_event'),
                                   cancel_url=url_for('pickup_dashboard'),
                                   pipeline_rate=pipeline_rate,
                                   pipeline_org=pipeline_org,
                                   pipeline_event=pipeline_event,
                                   pipeline_booking_id=new_bid,
                                   pipeline_start=pipeline_start,
                                   pipeline_end=pipeline_end,
                                   duplicate_warnings=duplicates,
                                   form_data=f)

        db.execute('''
            INSERT INTO pickup_config
            (booking_id, tab_name, organization, event_name, hotel, hotel_contact,
             hotel_contact_email, hotel_contact2, hotel_contact2_email, hotel_contacts,
             group_contact, group_contact_email,
             cutoff_date, attrition_pct, contracted_block, contracted_rate, shoulder_pre,
             shoulder_post, hotel_booking_link, notes, ota_url, cc_emails)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ''', (
            f.get('booking_id'), f.get('tab_name'), f['organization'],
            f.get('event_name'), f.get('hotel'), hc_name,
            hc_email, hc2_name, hc2_email, json.dumps(hotel_contacts), gc_name,
            gc_email, f.get('cutoff_date'),
            attrition, json.dumps(contracted_block),
            float(f['contracted_rate']) if f.get('contracted_rate') else None,
            int(f.get('shoulder_pre', 3)), int(f.get('shoulder_post', 3)),
            f.get('hotel_booking_link'), f.get('notes'), ota_url, json.dumps(cc_emails)
        ))
        db.commit()
        _upsert_contacts(db, hotel_contacts, f.get('hotel', ''), cc_emails, f['organization'])
        db.commit()
        flash('Event created.', 'success')
        return redirect(url_for('pickup_dashboard'))
    pipeline_rate = pipeline_org = pipeline_event = None
    pipeline_start = pipeline_end = None
    booking_id_qs = request.args.get('booking_id')
    if booking_id_qs:
        db = get_db()
        row = db.execute(
            'SELECT RoomRate, AccountName, EventName, StartDate, EndDate FROM ReportPipeline WHERE CAST(BookingId AS INTEGER) = CAST(? AS INTEGER) LIMIT 1',
            (booking_id_qs,)
        ).fetchone()
        if row:
            if row['RoomRate']:
                pipeline_rate = round(float(row['RoomRate']), 2)
            pipeline_org   = row['AccountName'] or None
            pipeline_event = row['EventName'] or None
            if row['StartDate']:
                pipeline_start = str(row['StartDate'])[:10]
            if row['EndDate']:
                pipeline_end = str(row['EndDate'])[:10]
    return render_template('pickup_config_form.html', config=None,
                           action=url_for('pickup_new_event'),
                           cancel_url=url_for('pickup_dashboard'),
                           pipeline_rate=pipeline_rate,
                           pipeline_org=pipeline_org,
                           pipeline_event=pipeline_event,
                           pipeline_booking_id=booking_id_qs,
                           pipeline_start=pipeline_start,
                           pipeline_end=pipeline_end)


@app.route('/pickup/<int:cid>')
def pickup_event(cid):
    from datetime import datetime as _dt, date as _date
    db = get_db()
    config = db.execute("SELECT * FROM pickup_config WHERE id=?", (cid,)).fetchone()
    if not config:
        flash('Event not found.', 'error')
        return redirect(url_for('pickup_dashboard'))
    all_weekly = db.execute("SELECT * FROM pickup_weekly WHERE config_id=? ORDER BY report_date DESC", (cid,)).fetchall()
    # Most recent real entry for OTA rate comparison (excluding historical/placeholder rows)
    last_real = db.execute(
        """SELECT * FROM pickup_weekly WHERE config_id=?
           AND (total_rooms IS NOT NULL AND total_rooms > 0)
           AND (label IS NULL OR (label NOT LIKE 'ASAS %' AND label NOT LIKE 'Historical%'
                                  AND label NOT LIKE '%Final%'))
           ORDER BY report_date DESC LIMIT 1""",
        (cid,)
    ).fetchone()
    rooming = db.execute(
        "SELECT id, upload_date, filename, total_guests, reconciliation_status, discrepancy_notes FROM pickup_rooming_list WHERE config_id=? ORDER BY upload_date DESC", (cid,)
    ).fetchall()
    contact_log = db.execute(
        "SELECT * FROM pickup_contact_log WHERE config_id=? ORDER BY contact_date DESC, id DESC", (cid,)
    ).fetchall()
    block = json.loads(config['contracted_block'] or '{}')
    # Include shoulder nights from any weekly entry that fall outside the contracted block
    all_date_set = set(block.keys())
    for _w in all_weekly:
        try:
            _pbn = json.loads(_w['pickup_by_night'] or '{}')
            all_date_set.update(d for d in _pbn if d != 'historical_total')
        except Exception:
            pass
    all_dates = sorted(all_date_set)
    contracted_total = sum(v for v in block.values() if v)
    day_map = {}
    for d in all_dates:
        try:
            day_map[d] = _dt.strptime(d, '%Y-%m-%d').strftime('%a')
        except Exception:
            day_map[d] = ''
    attrition_pct   = config['attrition_pct'] or 0
    attrition_rooms = round(contracted_total * attrition_pct, 1)

    # Split rows into historical pace rows (label = 'ASAS YYYY') vs current reporting rows
    raw_historical = [w for w in all_weekly if w['label'] and (
        w['label'].startswith('Historical') or w['label'].startswith('ASAS ')
    )]
    raw_current    = [w for w in all_weekly if not (w['label'] and (
        w['label'].startswith('Historical') or w['label'].startswith('ASAS ')
    ))]

    # Build historical display list (for backward-compat; sorted most recent year first)
    historical_years = []
    for w in sorted(raw_historical, key=lambda x: x['report_date'], reverse=True):
        pbn = json.loads(w['pickup_by_night'] or '{}')
        total = pbn.get('historical_total', w['total_rooms'] or 0)
        row = dict(w)
        row['total_rooms'] = total
        historical_years.append(row)

    # Build pace comparison structure for multi-year chart
    pace_comparison = None
    if raw_historical:
        # Event start dates for ASAS historical years (confirmed from Excel data)
        ASAS_EVENT_STARTS = {
            2021: '2021-07-11',  # Louisville, KY
            2022: '2022-06-23',  # OKC
            2023: '2023-07-15',  # Albuquerque
            2024: '2024-07-17',  # Calgary
            2025: '2025-06-30',  # The Diplomat
            2026: '2026-07-17',  # Madison
        }
        from datetime import date as _ddate

        def _weeks_out(event_start_str, report_date_str):
            try:
                es = _ddate.fromisoformat(event_start_str)
                rd = _ddate.fromisoformat(report_date_str)
                return (es - rd).days // 7
            except Exception:
                return None

        # Build per-year dict: year -> list of {weeks_out, total_rooms}
        hist_by_year = {}
        for w in raw_historical:
            lbl = w['label'] or ''
            # Support both 'ASAS YYYY' and legacy 'Historical YYYY...' labels
            import re as _re
            m = _re.search(r'(\d{4})', lbl)
            if not m:
                continue
            yr = int(m.group(1))
            es = ASAS_EVENT_STARTS.get(yr)
            if not es:
                continue
            wo = _weeks_out(es, w['report_date'])
            if wo is None or wo < 0:
                continue
            pbn = json.loads(w['pickup_by_night'] or '{}')
            total = pbn.get('historical_total', w['total_rooms'] or 0)
            if yr not in hist_by_year:
                hist_by_year[yr] = {}
            # Keep the entry with highest weeks_out match; for duplicate weeks, keep most rooms
            if wo not in hist_by_year[yr] or total > hist_by_year[yr][wo]:
                hist_by_year[yr][wo] = total

        # Add 2026 current data
        current_event_start = ASAS_EVENT_STARTS.get(2026)
        if current_event_start and raw_current:
            hist_by_year[2026] = {}
            for w in raw_current:
                pbn = json.loads(w['pickup_by_night'] or '{}')
                total = sum(v for v in pbn.values() if isinstance(v, (int, float)))
                wo = _weeks_out(current_event_start, w['report_date'])
                if wo is None or wo < 0:
                    continue
                if wo not in hist_by_year[2026] or total > hist_by_year[2026][wo]:
                    hist_by_year[2026][wo] = total

        # Collect all unique weeks_out values and sort descending (most weeks out first)
        all_weeks = sorted(
            set(wo for yr_data in hist_by_year.values() for wo in yr_data.keys()),
            reverse=True
        )

        # Build by_weeks_out lookup
        by_weeks_out = {}
        for wo in all_weeks:
            by_weeks_out[wo] = {}
            for yr, yr_data in hist_by_year.items():
                if wo in yr_data:
                    by_weeks_out[wo][str(yr)] = yr_data[wo]

        years_available = sorted(hist_by_year.keys(), reverse=True)

        pace_comparison = {
            'years': years_available,
            'event_starts': ASAS_EVENT_STARTS,
            'by_weeks_out': by_weeks_out,
            'all_weeks': all_weeks,
        }

    # Build current weekly display with pct_of_block / WoW calculations
    weekly_display = []
    prev_total = None
    for w in reversed(raw_current):
        pbn = json.loads(w['pickup_by_night'] or '{}')
        computed_total = sum(v for v in pbn.values() if v is not None and v != '')
        pct_blk  = round(computed_total / contracted_total * 100, 1) if contracted_total else None
        pct_attr = round(computed_total / attrition_rooms * 100, 1) if attrition_rooms else None
        wow      = (computed_total - prev_total) if prev_total is not None else None
        row = dict(w)
        row['total_rooms']      = computed_total
        row['pct_of_block']     = pct_blk
        row['pct_of_attrition'] = pct_attr
        row['change_from_last'] = wow
        weekly_display.append(row)
        prev_total = computed_total
    weekly_display.reverse()

    past_cutoff = False
    meeting_ended = False
    has_final_history = any(w['label'] and 'final' in w['label'].lower() for w in all_weekly)
    if all_dates:
        try:
            if _dt.strptime(all_dates[0], '%Y-%m-%d') < _dt.today():
                past_cutoff = True
        except Exception:
            pass
        try:
            if _dt.strptime(all_dates[-1], '%Y-%m-%d') < _dt.today():
                meeting_ended = True
        except Exception:
            pass
    hhr_row = db.execute(
        'SELECT id FROM housing_history_files WHERE booking_id=? ORDER BY id DESC LIMIT 1',
        (config['booking_id'],)
    ).fetchone() if config['booking_id'] else None
    has_hhr = bool(hhr_row)

    # Check if this is part of a multi-hotel event
    sibling_count = db.execute(
        "SELECT COUNT(*) FROM pickup_config WHERE LOWER(TRIM(event_name))=LOWER(TRIM(?)) AND status='active'",
        ((config['event_name'] or ''),)
    ).fetchone()[0]
    is_multi_hotel = sibling_count > 1
    # Find primary_cid for the event report link
    if is_multi_hotel:
        primary_cid_for_report = db.execute(
            "SELECT MIN(id) FROM pickup_config WHERE LOWER(TRIM(event_name))=LOWER(TRIM(?)) AND status='active'",
            ((config['event_name'] or ''),)
        ).fetchone()[0]
    else:
        primary_cid_for_report = None

    # Tasks for this event
    tasks = db.execute(
        "SELECT * FROM pickup_tasks WHERE config_id=? ORDER BY completed_at IS NOT NULL, due_date ASC NULLS LAST, id DESC",
        (cid,)
    ).fetchall()

    # Notes for this event
    event_notes = db.execute(
        "SELECT * FROM pickup_event_notes WHERE config_id=? ORDER BY id DESC",
        (cid,)
    ).fetchall()

    # All active users (for task assignment)
    all_users = db.execute(
        "SELECT id, name FROM Users WHERE active=1 ORDER BY name"
    ).fetchall()

    # Recently deleted weekly reports (kept 90 days, restorable)
    deleted_weekly = db.execute(
        """SELECT * FROM pickup_weekly_deleted
           WHERE config_id=?
             AND deleted_at >= datetime('now','-90 days','localtime')
           ORDER BY deleted_at DESC""",
        (cid,)
    ).fetchall()

    last_ota_rate = last_real['ota_rate'] if last_real and 'ota_rate' in last_real.keys() else None
    last_ota_url  = last_real['ota_url']  if last_real and 'ota_url'  in last_real.keys() else None
    show_rate_issue = bool(
        last_ota_rate and config['contracted_rate'] and
        float(last_ota_rate) < float(config['contracted_rate'])
    )

    # Amendments / addendum history
    amendments = db.execute(
        "SELECT * FROM pickup_amendments WHERE config_id=? ORDER BY uploaded_at DESC", (cid,)
    ).fetchall()

    # Hotel Points integration
    from points_utils import detect_chain
    points_program = None
    points_request = None
    if config['contract_data']:
        chain = detect_chain(config['hotel'])
        if chain:
            points_program = db.execute(
                'SELECT * FROM hotel_points_program WHERE chain_name=? AND active=1',
                (chain,)).fetchone()
            if points_program:
                points_request = db.execute(
                    'SELECT * FROM hotel_points_request WHERE pickup_config_id=? AND program_id=? '
                    'ORDER BY id DESC LIMIT 1',
                    (cid, points_program['id'])).fetchone()

    # RFP critical dates lookup (for Email Dates button)
    rfp_for_dates = None
    if config['booking_id']:
        rfp_for_dates = db.execute(
            "SELECT id, critical_dates_json, critical_dates_sent_at "
            "FROM rfp WHERE booking_id=? LIMIT 1",
            (config['booking_id'],)
        ).fetchone()

    return render_template('pickup_event.html',
                           config=config, weekly=weekly_display, historical_years=historical_years,
                           pace_comparison=pace_comparison,
                           rooming=rooming,
                           contact_log=contact_log, block=block, all_dates=all_dates,
                           day_map=day_map, contracted_total=contracted_total,
                           attrition_pct=attrition_pct, attrition_rooms=attrition_rooms,
                           past_cutoff=past_cutoff, has_final_history=has_final_history,
                           meeting_ended=meeting_ended,
                           has_hhr=has_hhr, today=_date.today().isoformat(),
                           is_multi_hotel=is_multi_hotel,
                           primary_cid_for_report=primary_cid_for_report,
                           deleted_weekly=deleted_weekly,
                           tasks=tasks, event_notes=event_notes, all_users=all_users,
                           last_ota_rate=last_ota_rate, last_ota_url=last_ota_url,
                           show_rate_issue=show_rate_issue,
                           amendments=amendments,
                           points_program=points_program,
                           points_request=points_request,
                           rfp_for_dates=rfp_for_dates)


# ── Event Report: combined view across all hotels for an event ────────────────

@app.route('/pickup/event-report/<int:primary_cid>')
def pickup_event_report(primary_cid):
    """Show combined pickup data for all hotels sharing the same event_name."""
    from datetime import date as _date2, datetime as _dt2
    db = get_db()

    # Get the primary config to find the event_name
    primary = db.execute("SELECT * FROM pickup_config WHERE id=?", (primary_cid,)).fetchone()
    if not primary:
        flash('Event not found.', 'error')
        return redirect(url_for('pickup_dashboard'))

    event_name = (primary['event_name'] or '').strip()
    if not event_name:
        return redirect(url_for('pickup_event', cid=primary_cid))

    # Find all configs with the same event_name (case-insensitive, active only)
    all_configs = db.execute(
        "SELECT * FROM pickup_config WHERE LOWER(TRIM(event_name))=LOWER(TRIM(?)) AND status='active' ORDER BY id",
        (event_name,)
    ).fetchall()

    if len(all_configs) <= 1:
        # Only one hotel — redirect to normal event page
        return redirect(url_for('pickup_event', cid=primary_cid))

    organization = primary['organization'] or primary['event_name'] or ''

    # ── Per-hotel summaries ────────────────────────────────────────────────────
    hotel_summaries = []
    combined_block = 0
    all_date_union = set()

    for c in all_configs:
        block = json.loads(c['contracted_block'] or '{}')
        contracted_total = sum(block.values())
        combined_block += contracted_total
        all_date_union.update(block.keys())
        # Also include any shoulder nights from weekly entries
        for _w in db.execute("SELECT pickup_by_night FROM pickup_weekly WHERE config_id=?", (c['id'],)).fetchall():
            try:
                _pbn = json.loads(_w['pickup_by_night'] or '{}')
                all_date_union.update(d for d in _pbn if d != 'historical_total')
            except Exception:
                pass

        # Latest non-historical weekly entry (label IS NULL counts as current)
        last = db.execute(
            """SELECT * FROM pickup_weekly WHERE config_id=?
               AND (label IS NULL OR (label NOT LIKE 'ASAS %' AND label NOT LIKE 'Historical%'))
               ORDER BY report_date DESC LIMIT 1""",
            (c['id'],)
        ).fetchone()
        last_total = last['total_rooms'] if last else 0
        last_pct = last['pct_of_block'] if last else None
        last_date = last['report_date'] if last else None

        if last_pct is None:
            h_badge, h_badge_label = 'secondary', 'No data'
        elif last_pct >= 80:
            h_badge, h_badge_label = 'success', 'On Pace'
        elif last_pct >= 60:
            h_badge, h_badge_label = 'warning', 'Watch'
        else:
            h_badge, h_badge_label = 'danger', 'At Risk'

        hotel_summaries.append({
            'cid':              c['id'],
            'hotel_name':       c['hotel'] or c['event_name'] or '—',
            'contracted_total': contracted_total,
            'last_total':       last_total,
            'last_pct':         last_pct,
            'last_date':        last_date,
            'badge':            h_badge,
            'badge_label':      h_badge_label,
            'cutoff_date':      c['cutoff_date'],
        })

    all_dates = sorted(all_date_union)

    # ── Combined weekly timeline ───────────────────────────────────────────────
    # Gather all non-historical weekly rows across all hotels
    from collections import defaultdict
    date_totals = defaultdict(int)   # report_date -> sum of total_rooms
    date_hotels = defaultdict(set)   # report_date -> set of cids that reported

    for c in all_configs:
        rows_w = db.execute(
            """SELECT * FROM pickup_weekly WHERE config_id=?
               AND (label IS NULL OR (label NOT LIKE 'ASAS %' AND label NOT LIKE 'Historical%'))
               ORDER BY report_date""",
            (c['id'],)
        ).fetchall()
        for w in rows_w:
            total = w['total_rooms'] or 0
            date_totals[w['report_date']] += total
            date_hotels[w['report_date']].add(c['id'])

    # Build sorted weekly display with WoW change
    sorted_dates = sorted(date_totals.keys())
    combined_weekly = []
    prev_total = None
    for rd in sorted_dates:
        total = date_totals[rd]
        pct = round(total / combined_block * 100, 1) if combined_block else None
        wow = (total - prev_total) if prev_total is not None else None
        combined_weekly.append({
            'report_date':   rd,
            'total_rooms':   total,
            'pct_of_block':  pct,
            'wow_change':    wow,
            'hotels_count':  len(date_hotels[rd]),
        })
        prev_total = total
    combined_weekly.reverse()  # newest first

    # ── Historical pace comparison (if any hotel has ASAS historical rows) ────
    pace_comparison = None
    ASAS_EVENT_STARTS = {
        2021: '2021-07-11',
        2022: '2022-06-23',
        2023: '2023-07-15',
        2024: '2024-07-17',
        2025: '2025-06-30',
        2026: '2026-07-17',
    }

    hist_rows_all = []
    for c in all_configs:
        hist_rows = db.execute(
            """SELECT * FROM pickup_weekly WHERE config_id=?
               AND (label LIKE 'ASAS %' OR label LIKE 'Historical%')
               ORDER BY report_date""",
            (c['id'],)
        ).fetchall()
        hist_rows_all.extend(hist_rows)

    if hist_rows_all:
        import re as _re2
        def _weeks_out2(event_start_str, report_date_str):
            try:
                es = _date2.fromisoformat(event_start_str)
                rd = _date2.fromisoformat(report_date_str)
                return (es - rd).days // 7
            except Exception:
                return None

        hist_by_year = {}
        for w in hist_rows_all:
            lbl = w['label'] or ''
            m = _re2.search(r'(\d{4})', lbl)
            if not m:
                continue
            yr = int(m.group(1))
            es = ASAS_EVENT_STARTS.get(yr)
            if not es:
                continue
            wo = _weeks_out2(es, w['report_date'])
            if wo is None or wo < 0:
                continue
            pbn = json.loads(w['pickup_by_night'] or '{}')
            total = pbn.get('historical_total', w['total_rooms'] or 0)
            if yr not in hist_by_year:
                hist_by_year[yr] = {}
            if wo not in hist_by_year[yr] or total > hist_by_year[yr][wo]:
                hist_by_year[yr][wo] = total

        # Add current year from combined_weekly (reversed to get chronological)
        current_event_start = ASAS_EVENT_STARTS.get(2026)
        if current_event_start and date_totals:
            hist_by_year[2026] = {}
            for rd, total in date_totals.items():
                wo = _weeks_out2(current_event_start, rd)
                if wo is None or wo < 0:
                    continue
                if wo not in hist_by_year[2026] or total > hist_by_year[2026][wo]:
                    hist_by_year[2026][wo] = total

        if hist_by_year:
            all_weeks_hist = sorted(
                set(wo for yr_data in hist_by_year.values() for wo in yr_data.keys()),
                reverse=True
            )
            by_weeks_out = {}
            for wo in all_weeks_hist:
                by_weeks_out[wo] = {}
                for yr, yr_data in hist_by_year.items():
                    if wo in yr_data:
                        by_weeks_out[wo][str(yr)] = yr_data[wo]
            years_available = sorted(hist_by_year.keys(), reverse=True)
            pace_comparison = {
                'years':        years_available,
                'event_starts': ASAS_EVENT_STARTS,
                'by_weeks_out': by_weeks_out,
                'all_weeks':    all_weeks_hist,
            }

    # ── Cutoff range ──────────────────────────────────────────────────────────
    cutoffs = [c['cutoff_date'] for c in all_configs if c['cutoff_date']]
    cutoff_earliest = min(cutoffs) if cutoffs else None
    cutoff_latest   = max(cutoffs) if cutoffs else None

    return render_template('pickup_event_report.html',
                           event_name=event_name,
                           organization=organization,
                           configs=all_configs,
                           combined_block=combined_block,
                           combined_weekly=combined_weekly,
                           hotel_summaries=hotel_summaries,
                           pace_comparison=pace_comparison,
                           all_dates=all_dates,
                           cutoff_earliest=cutoff_earliest,
                           cutoff_latest=cutoff_latest,
                           primary_cid=primary_cid)


@app.route('/pickup/event-report/<int:primary_cid>/download-xlsx')
def pickup_event_report_xlsx(primary_cid):
    """Download a combined XLSX pace report for all hotels sharing the same event_name."""
    import io
    import re as _re_fn
    from datetime import date as _date_xlsx
    from collections import defaultdict
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    db = get_db()
    primary = db.execute("SELECT * FROM pickup_config WHERE id=?", (primary_cid,)).fetchone()
    if not primary:
        flash('Event not found.', 'error')
        return redirect(url_for('pickup_dashboard'))

    event_name = (primary['event_name'] or '').strip()
    if not event_name:
        return redirect(url_for('pickup_event', cid=primary_cid))

    organization = primary['organization'] or primary['event_name'] or ''

    # Find all active configs with the same event_name
    all_configs = db.execute(
        "SELECT * FROM pickup_config WHERE LOWER(TRIM(event_name))=LOWER(TRIM(?)) AND status='active' ORDER BY id",
        (event_name,)
    ).fetchall()

    ASAS_EVENT_STARTS = {
        2021: '2021-07-11',
        2022: '2022-06-23',
        2023: '2023-07-15',
        2024: '2024-07-17',
        2025: '2025-06-30',
        2026: '2026-07-17',
    }

    def weeks_out(event_start_str, report_date_str):
        try:
            es = _date_xlsx.fromisoformat(event_start_str)
            rd = _date_xlsx.fromisoformat(report_date_str)
            return (es - rd).days // 7
        except Exception:
            return None

    # ── Per-hotel data ─────────────────────────────────────────────────────────
    hotel_data = []
    combined_block = 0
    all_event_date_set = set()

    for c in all_configs:
        block = json.loads(c['contracted_block'] or '{}')
        contracted_total = sum(block.values())
        combined_block += contracted_total
        all_event_date_set.update(block.keys())

        weekly_rows = db.execute(
            """SELECT * FROM pickup_weekly WHERE config_id=?
               AND (label IS NULL OR (label NOT LIKE 'ASAS %' AND label NOT LIKE 'Historical%'))
               ORDER BY report_date""",
            (c['id'],)
        ).fetchall()

        # Also include any shoulder nights recorded in pickup_by_night
        for w in weekly_rows:
            try:
                pbn = json.loads(w['pickup_by_night'] or '{}')
                for d in pbn.keys():
                    if d:
                        all_event_date_set.add(d)
            except Exception:
                pass

        hotel_data.append({
            'config': c,
            'hotel_name': c['hotel'] or c['event_name'] or '—',
            'contracted_block': block,
            'contracted_total': contracted_total,
            'contracted_rate': c['contracted_rate'] or 0,
            'attrition_pct': c['attrition_pct'],
            'weekly': [dict(w) for w in weekly_rows],
        })

    event_dates = sorted(all_event_date_set)
    n_nights = len(event_dates)
    n_hotels = len(all_configs)

    # ── Historical rows across all configs ────────────────────────────────────
    import re as _re2
    hist_rows_all = []
    for c in all_configs:
        rows = db.execute(
            """SELECT * FROM pickup_weekly WHERE config_id=?
               AND (label LIKE 'ASAS %' OR label LIKE 'Historical%')
               ORDER BY report_date""",
            (c['id'],)
        ).fetchall()
        hist_rows_all.extend(rows)

    hist_by_year = {}
    for w in hist_rows_all:
        lbl = w['label'] or ''
        m = _re2.search(r'(\d{4})', lbl)
        if not m:
            continue
        yr = int(m.group(1))
        es = ASAS_EVENT_STARTS.get(yr)
        if not es:
            continue
        wo = weeks_out(es, w['report_date'])
        if wo is None or wo < 0:
            continue
        pbn = json.loads(w['pickup_by_night'] or '{}')
        total = pbn.get('historical_total', w['total_rooms'] or 0)
        if yr not in hist_by_year:
            hist_by_year[yr] = {}
        if wo not in hist_by_year[yr] or total > hist_by_year[yr][wo]:
            hist_by_year[yr][wo] = total

    # ── Build per-date aggregate (Tab 2 data) ─────────────────────────────────
    all_report_date_set = set()
    for hd in hotel_data:
        for w in hd['weekly']:
            all_report_date_set.add(w['report_date'])
    all_report_dates = sorted(all_report_date_set)

    # aggregate_by_date: report_date -> sum of total_rooms (for 2026 hist)
    aggregate_by_date = defaultdict(int)
    total_by_night = {}       # report_date -> {event_date: int}
    total_rooms_agg = {}      # report_date -> int
    hotel_pbn = {}            # cid -> report_date -> {event_date: int}

    for hd in hotel_data:
        cid = hd['config']['id']
        hotel_pbn[cid] = {}
        for w in hd['weekly']:
            rd = w['report_date']
            pbn = json.loads(w['pickup_by_night'] or '{}')
            hotel_pbn[cid][rd] = pbn
            aggregate_by_date[rd] += (w['total_rooms'] or 0)
            if rd not in total_by_night:
                total_by_night[rd] = defaultdict(int)
            for ed, rooms in pbn.items():
                if isinstance(rooms, (int, float)):
                    total_by_night[rd][ed] += int(rooms)

    for rd in all_report_dates:
        total_rooms_agg[rd] = aggregate_by_date[rd]

    # Add 2026 current to hist_by_year
    current_event_start_2026 = ASAS_EVENT_STARTS.get(2026)
    if current_event_start_2026 and aggregate_by_date:
        hist_by_year[2026] = {}
        for rd, total in aggregate_by_date.items():
            wo = weeks_out(current_event_start_2026, rd)
            if wo is None or wo < 0:
                continue
            if wo not in hist_by_year[2026] or total > hist_by_year[2026][wo]:
                hist_by_year[2026][wo] = total

    all_weeks_hist = sorted(
        set(wo for yr_data in hist_by_year.values() for wo in yr_data.keys()),
        reverse=True
    )
    years_available = sorted(hist_by_year.keys(), reverse=True)

    # ── Cutoff info ───────────────────────────────────────────────────────────
    cutoffs = [c['cutoff_date'] for c in all_configs if c['cutoff_date']]
    cutoff_earliest = min(cutoffs) if cutoffs else None

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = Workbook()

    # ════════════════════════════════════════════════════════════════════════
    # TAB 1 — Historical Comparison
    # ════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Historical Comparison"

    header_fill = PatternFill("solid", fgColor="1A3A5C")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center")
    alt_fill = PatternFill("solid", fgColor="EAF0FB")

    # A1: event name
    ws1['A1'] = event_name
    ws1['A1'].font = Font(bold=True, size=14)
    ws1.merge_cells('A1:G1')

    # A2: organization
    ws1['A2'] = organization
    ws1['A2'].font = Font(italic=True, size=10)

    # A3: subtitle
    ws1['A3'] = "All Hotels Combined — Historical Pickup by Weeks Out"
    ws1['A3'].font = Font(size=10)

    # A4: blank
    ws1['A4'] = ""

    # Row 5: headers
    ws1.cell(5, 1).value = "Weeks Out"
    ws1.cell(5, 1).font = header_font
    ws1.cell(5, 1).fill = header_fill
    ws1.cell(5, 1).alignment = center_align

    for ci, yr in enumerate(years_available, start=2):
        lbl = f"{yr} (All Hotels)" if yr == 2026 else str(yr)
        cell = ws1.cell(5, ci)
        cell.value = lbl
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # Data rows
    for ri, wo in enumerate(all_weeks_hist):
        row_num = 6 + ri
        is_even = (ri % 2 == 0)
        row_fill = alt_fill if is_even else None

        wk_cell = ws1.cell(row_num, 1)
        wk_cell.value = f"{wo} wks"
        wk_cell.alignment = center_align
        if row_fill:
            wk_cell.fill = row_fill

        for ci, yr in enumerate(years_available, start=2):
            val = hist_by_year.get(yr, {}).get(wo)
            cell = ws1.cell(row_num, ci)
            cell.value = val if val is not None else "—"
            if val is not None:
                cell.number_format = "#,##0"
            cell.alignment = center_align
            if yr == 2026:
                cell.font = Font(bold=True)
            if row_fill:
                cell.fill = row_fill

    # Column widths Tab 1
    ws1.column_dimensions['A'].width = 12
    for ci in range(2, 2 + len(years_available)):
        ws1.column_dimensions[get_column_letter(ci)].width = 16

    # ════════════════════════════════════════════════════════════════════════
    # TAB 2 — Pace 2026
    # ════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Pace 2026")

    start_col = 5   # col E
    end_col = start_col + n_nights - 1
    total_col = end_col + 1
    pct_col = total_col + 1
    last_col_idx = pct_col + 1

    end_col_letter = get_column_letter(end_col)
    total_col_letter = get_column_letter(total_col)
    pct_col_letter = get_column_letter(pct_col)

    # ── Header block (rows 2-5) ───────────────────────────────────────────
    # Row 2: event name
    ws2.merge_cells(start_row=2, start_column=4, end_row=2, end_column=total_col)
    ws2.cell(2, 4).value = event_name
    ws2.cell(2, 4).font = Font(bold=True, size=18)
    ws2.cell(2, 4).alignment = Alignment(horizontal="center")

    # Row 3: date range
    if event_dates:
        try:
            d0 = _date_xlsx.fromisoformat(event_dates[0])
            d1 = _date_xlsx.fromisoformat(event_dates[-1])
            date_range_str = f"{d0.strftime('%B')} {d0.day}–{d1.day}, {d1.year}"
        except Exception:
            date_range_str = f"{event_dates[0]} to {event_dates[-1]}"
    else:
        date_range_str = ""
    ws2.merge_cells(start_row=3, start_column=4, end_row=3, end_column=total_col)
    ws2.cell(3, 4).value = date_range_str
    ws2.cell(3, 4).font = Font(bold=True, size=18)
    ws2.cell(3, 4).alignment = Alignment(horizontal="center")

    # Row 4: "Pace Report"
    ws2.merge_cells(start_row=4, start_column=4, end_row=4, end_column=total_col)
    ws2.cell(4, 4).value = "Pace Report"
    ws2.cell(4, 4).font = Font(bold=True, size=18)
    ws2.cell(4, 4).alignment = Alignment(horizontal="center")

    # Row 5: labels
    ws2.cell(5, 4).value = "Total Contracted Block"
    ws2.cell(5, 4).font = Font(italic=True, size=11)
    ws2.cell(5, 4).alignment = Alignment(horizontal="center")

    if cutoff_earliest:
        cutoff_cell = ws2.cell(5, last_col_idx)
        cutoff_cell.value = f"Cut off date: {cutoff_earliest}"
        cutoff_cell.font = Font(bold=True, size=10, color="FFFF0000")

    # ── Date header row (row 6) ───────────────────────────────────────────
    date_hdr_fill = PatternFill("solid", fgColor="1A3A5C")
    date_hdr_font = Font(bold=True, size=11, color="FFFFFF")
    day_hdr_fill  = PatternFill("solid", fgColor="2D5986")

    ws2.cell(6, 4).value = "Date:"
    ws2.cell(6, 4).font = Font(bold=True, size=11, color="FFFFFF")
    ws2.cell(6, 4).fill = date_hdr_fill

    for ni, ed in enumerate(event_dates):
        col_idx = start_col + ni
        try:
            d = _date_xlsx.fromisoformat(ed)
            ws2.cell(6, col_idx).value = d
            ws2.cell(6, col_idx).number_format = "d-mmm"
        except Exception:
            ws2.cell(6, col_idx).value = ed
        ws2.cell(6, col_idx).alignment = Alignment(horizontal="center")
        ws2.cell(6, col_idx).font = date_hdr_font
        ws2.cell(6, col_idx).fill = date_hdr_fill
        ws2.cell(6, col_idx).border = Border(left=Side(style='thin', color='FFFFFF'),
                                              right=Side(style='thin', color='FFFFFF'))

    ws2.cell(6, total_col).value = "Total"
    ws2.cell(6, total_col).font = date_hdr_font
    ws2.cell(6, total_col).alignment = Alignment(horizontal="center")
    ws2.cell(6, total_col).fill = date_hdr_fill

    # ── Day-of-week row (row 7) ───────────────────────────────────────────
    ws2.cell(7, 4).value = "Day:"
    ws2.cell(7, 4).font = Font(bold=True, size=11, color="FFFFFF")
    ws2.cell(7, 4).fill = day_hdr_fill

    day_abbrevs = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
    for ni, ed in enumerate(event_dates):
        col_idx = start_col + ni
        try:
            d = _date_xlsx.fromisoformat(ed)
            ws2.cell(7, col_idx).value = day_abbrevs[d.weekday()]
        except Exception:
            ws2.cell(7, col_idx).value = ""
        ws2.cell(7, col_idx).alignment = Alignment(horizontal="center")
        ws2.cell(7, col_idx).font = Font(bold=True, size=11, color="FFFFFF")
        ws2.cell(7, col_idx).fill = day_hdr_fill
        ws2.cell(7, col_idx).border = Border(left=Side(style='thin', color='FFFFFF'),
                                              right=Side(style='thin', color='FFFFFF'))
    ws2.cell(7, total_col).fill = day_hdr_fill

    # ── Contracted block rows (rows 8 to 7+n_hotels) ─────────────────────
    hotel_row_fill  = PatternFill("solid", fgColor="F8F9FA")
    hotel_row_fill2 = PatternFill("solid", fgColor="FFFFFF")
    thin = Side(style='thin')
    med  = Side(style='medium')
    def cell_border(top=None, bottom=None, left=None, right=None):
        return Border(top=top, bottom=bottom, left=left, right=right)

    for hi, hd in enumerate(hotel_data):
        r = 8 + hi
        c_cfg = hd['config']
        # cols 1 (A) and 2 (B) intentionally left empty; org (col 3/C) also left empty
        # hotel name goes in col 4 (D) — cols A-C deleted later
        ws2.cell(r, 4).value = f"{hd['hotel_name']} (${hd['contracted_rate']:.0f})"
        ws2.cell(r, 4).font = Font(size=11)
        row_fill = hotel_row_fill if hi % 2 == 0 else hotel_row_fill2

        for ni, ed in enumerate(event_dates):
            col_idx = start_col + ni
            rooms = hd['contracted_block'].get(ed, 0)
            ws2.cell(r, col_idx).value = rooms if rooms else None
            ws2.cell(r, col_idx).alignment = Alignment(horizontal="center")
            ws2.cell(r, col_idx).font = Font(size=11)
            ws2.cell(r, col_idx).fill = row_fill
            ws2.cell(r, col_idx).border = cell_border(top=Side(style='hair'), bottom=Side(style='hair'),
                                                       left=Side(style='hair'), right=Side(style='hair'))

        ws2.cell(r, total_col).value = hd['contracted_total']
        ws2.cell(r, total_col).font = Font(bold=True, size=11)
        ws2.cell(r, total_col).alignment = Alignment(horizontal="center")
        ws2.cell(r, total_col).fill = row_fill
        ws2.cell(r, total_col).border = cell_border(left=Side(style='thin'), right=Side(style='thin'),
                                                      top=Side(style='hair'), bottom=Side(style='hair'))

        # Attrition column
        atr = c_cfg['attrition_pct']
        if atr:
            atr_rooms = round(hd['contracted_total'] * atr)
            ws2.cell(r, pct_col).value = f"Attrition {round(atr*100):.0f}% = {atr_rooms}"
        else:
            ws2.cell(r, pct_col).value = "Attrition Waived"
        ws2.cell(r, pct_col).font = Font(size=10)
        ws2.cell(r, pct_col).fill = row_fill

    # ── Total block row (row 8+n_hotels) ─────────────────────────────────
    total_block_row = 8 + n_hotels
    red_font_bold = Font(bold=True, size=11, color="FFFF0000")
    total_blk_fill = PatternFill("solid", fgColor="FCE4EC")  # pale red
    total_blk_border_mid = Border(top=Side(style='medium'), bottom=Side(style='medium'))
    ws2.cell(total_block_row, 4).value = "Total Block"
    ws2.cell(total_block_row, 4).font = red_font_bold
    ws2.cell(total_block_row, 4).alignment = Alignment(horizontal="center")
    ws2.cell(total_block_row, 4).fill = total_blk_fill
    ws2.cell(total_block_row, 4).border = Border(top=Side(style='medium'), bottom=Side(style='medium'),
                                                   left=Side(style='medium'))

    for ni, ed in enumerate(event_dates):
        col_idx = start_col + ni
        night_total = sum(hd['contracted_block'].get(ed, 0) for hd in hotel_data)
        ws2.cell(total_block_row, col_idx).value = night_total if night_total else None
        ws2.cell(total_block_row, col_idx).font = red_font_bold
        ws2.cell(total_block_row, col_idx).alignment = Alignment(horizontal="center")
        ws2.cell(total_block_row, col_idx).fill = total_blk_fill
        ws2.cell(total_block_row, col_idx).border = total_blk_border_mid

    ws2.cell(total_block_row, total_col).value = combined_block
    ws2.cell(total_block_row, total_col).font = red_font_bold
    ws2.cell(total_block_row, total_col).alignment = Alignment(horizontal="center")
    ws2.cell(total_block_row, total_col).fill = total_blk_fill
    ws2.cell(total_block_row, total_col).border = Border(top=Side(style='medium'), bottom=Side(style='medium'),
                                                          right=Side(style='medium'))

    # ── Spacer row ────────────────────────────────────────────────────────
    spacer_row = total_block_row + 1  # = 9 + n_hotels

    # ── Weekly snapshot blocks ────────────────────────────────────────────
    body_start = spacer_row + 1  # = 10 + n_hotels
    yellow_fill  = PatternFill("solid", fgColor="FFFF00")
    blk_hdr_fill = PatternFill("solid", fgColor="D9E1F2")   # slate-blue header
    blk_tot_fill = PatternFill("solid", fgColor="E2EFDA")   # pale green total row
    blk_h1_fill  = PatternFill("solid", fgColor="F2F7FF")   # hotel alt 1
    blk_h2_fill  = PatternFill("solid", fgColor="FFFFFF")   # hotel alt 2
    hair_b = Border(top=Side(style='hair'), bottom=Side(style='hair'),
                    left=Side(style='hair'), right=Side(style='hair'))
    edge_l = Border(top=Side(style='hair'), bottom=Side(style='hair'),
                    left=Side(style='thin'), right=Side(style='hair'))
    edge_r = Border(top=Side(style='hair'), bottom=Side(style='hair'),
                    left=Side(style='hair'), right=Side(style='thin'))

    for i, report_date in enumerate(all_report_dates):
        block_row = body_start + (i * 8)
        is_latest = (report_date == all_report_dates[-1])
        hdr_fill = yellow_fill if is_latest else blk_hdr_fill
        hdr_color = "000000" if is_latest else "FF0000"

        # Row 0 of block: header ─────────────────────────────────────────
        ws2.cell(block_row, 3).value = "Actual Pick Up"
        ws2.cell(block_row, 3).font = Font(bold=True, size=11, color=hdr_color)
        ws2.cell(block_row, 3).fill = hdr_fill

        try:
            ws2.cell(block_row, 4).value = _date_xlsx.fromisoformat(report_date)
            ws2.cell(block_row, 4).number_format = "d-mmm-yy"
        except Exception:
            ws2.cell(block_row, 4).value = report_date
        ws2.cell(block_row, 4).font = Font(size=10, color=hdr_color)
        ws2.cell(block_row, 4).fill = hdr_fill

        ws2.merge_cells(start_row=block_row, start_column=4, end_row=block_row, end_column=total_col)

        wo = weeks_out('2026-07-17', report_date)
        wo_label = f"{wo} weeks" if wo is not None else ""
        ws2.cell(block_row, last_col_idx).value = wo_label
        ws2.cell(block_row, last_col_idx).font = Font(italic=True, size=11, color=hdr_color)
        ws2.cell(block_row, last_col_idx).fill = hdr_fill

        # Row 1 of block: Total row ──────────────────────────────────────
        tr = block_row + 1
        ws2.cell(tr, 4).value = "Total"
        ws2.cell(tr, 4).font = Font(bold=True, italic=True, size=10)
        ws2.cell(tr, 4).alignment = Alignment(horizontal="center")
        ws2.cell(tr, 4).fill = blk_tot_fill
        ws2.cell(tr, 4).border = Border(top=Side(style='thin'), bottom=Side(style='thin'),
                                         left=Side(style='thin'), right=Side(style='hair'))

        for ni, ed in enumerate(event_dates):
            night_total = total_by_night.get(report_date, {}).get(ed, 0)
            col_idx = start_col + ni
            ws2.cell(tr, col_idx).value = night_total if night_total else None
            ws2.cell(tr, col_idx).number_format = "0"
            ws2.cell(tr, col_idx).font = Font(bold=True, size=11)
            ws2.cell(tr, col_idx).alignment = Alignment(horizontal="center")
            ws2.cell(tr, col_idx).fill = blk_tot_fill
            ws2.cell(tr, col_idx).border = Border(top=Side(style='thin'), bottom=Side(style='thin'),
                                                   left=Side(style='hair'), right=Side(style='hair'))

        ws2.cell(tr, total_col).value = total_rooms_agg.get(report_date, 0) or None
        ws2.cell(tr, total_col).number_format = "0"
        ws2.cell(tr, total_col).font = Font(bold=True, size=11)
        ws2.cell(tr, total_col).fill = blk_tot_fill
        ws2.cell(tr, total_col).border = Border(top=Side(style='thin'), bottom=Side(style='thin'),
                                                  left=Side(style='hair'), right=Side(style='thin'))

        pct_val = total_rooms_agg.get(report_date, 0) / combined_block if combined_block else None
        ws2.cell(tr, pct_col).value = pct_val
        ws2.cell(tr, pct_col).number_format = "0%"
        ws2.cell(tr, pct_col).font = Font(bold=True, size=11)
        ws2.cell(tr, pct_col).fill = blk_tot_fill

        # Rows 2-N of block: per-hotel rows ─────────────────────────────
        for hi, hd in enumerate(hotel_data):
            hr = block_row + 2 + hi
            cid = hd['config']['id']
            h_row_fill = blk_h1_fill if hi % 2 == 0 else blk_h2_fill
            ws2.cell(hr, 4).value = hd['hotel_name']
            ws2.cell(hr, 4).font = Font(bold=True, italic=True, size=10)
            ws2.cell(hr, 4).alignment = Alignment(horizontal="center")
            ws2.cell(hr, 4).fill = h_row_fill
            ws2.cell(hr, 4).border = edge_l

            hotel_pbn_rd = hotel_pbn.get(cid, {}).get(report_date, {})
            for ni, ed in enumerate(event_dates):
                rooms = hotel_pbn_rd.get(ed, None)
                col_idx = start_col + ni
                ws2.cell(hr, col_idx).value = rooms if rooms else None
                ws2.cell(hr, col_idx).number_format = "0"
                ws2.cell(hr, col_idx).alignment = Alignment(horizontal="center")
                ws2.cell(hr, col_idx).font = Font(size=11)
                ws2.cell(hr, col_idx).fill = h_row_fill
                ws2.cell(hr, col_idx).border = hair_b

            # Use stored total_rooms for the hotel
            w_entry = next((w for w in hd['weekly'] if w['report_date'] == report_date), None)
            hotel_total = (w_entry['total_rooms'] or 0) if w_entry else 0

            ws2.cell(hr, total_col).value = hotel_total if hotel_total else None
            ws2.cell(hr, total_col).number_format = "0"
            ws2.cell(hr, total_col).font = Font(bold=True, size=11)
            ws2.cell(hr, total_col).fill = h_row_fill
            ws2.cell(hr, total_col).border = edge_r

            h_pct = hotel_total / hd['contracted_total'] if hd['contracted_total'] else None
            ws2.cell(hr, pct_col).value = h_pct
            ws2.cell(hr, pct_col).number_format = "0%"
            ws2.cell(hr, pct_col).font = Font(bold=True, size=11)
            ws2.cell(hr, pct_col).fill = h_row_fill

    # ── Remove empty cols A–B, then set column widths ────────────────────
    ws2.delete_cols(1, 2)  # cols A (booking_id) + B (event_name) → now C→A, D→B, E→C …
    # After deletion: A=org(empty), B=hotel name, C=first night, … total/pct/weeks-label shift -2
    ws2.column_dimensions['A'].width = 4   # empty spacer
    ws2.column_dimensions['B'].width = 26  # hotel name
    for ni in range(n_nights):
        ws2.column_dimensions[get_column_letter(start_col - 2 + ni)].width = 9
    ws2.column_dimensions[get_column_letter(total_col - 2)].width = 8
    ws2.column_dimensions[get_column_letter(pct_col - 2)].width = 14
    ws2.column_dimensions[get_column_letter(last_col_idx - 2)].width = 10

    # ── Row heights Tab 2 ─────────────────────────────────────────────────
    for rn in range(2, 5):
        ws2.row_dimensions[rn].height = 23.75
    for rn in range(5, ws2.max_row + 1):
        ws2.row_dimensions[rn].height = 14.25

    # ── Return file ───────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = _re_fn.sub(r'[^a-zA-Z0-9_]+', '_', event_name)
    return send_file(buf, as_attachment=True,
                     download_name=f"{safe_name}_Pace_Report.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def _get_current_pickup_events(db, account_filter=None):
    """Shared helper: return current-section events sorted by org → event_start.
    account_filter: None = admin (no filter), [] = no access, [list] = allowed accounts."""
    from datetime import date as _d, timedelta
    today = _d.today()
    future_cutoff = today + timedelta(days=120)
    if account_filter is None:
        af_where, af_params = '', []
    elif account_filter:
        ph = ','.join('?' * len(account_filter))
        af_where = f' AND (p.AccountName IN ({ph}) OR c.organization IN ({ph}))'
        af_params = account_filter + account_filter
    else:
        af_where, af_params = ' AND 1=0', []
    configs = db.execute(f"""
        SELECT c.*,
               p.StartDate AS bk_start, p.EndDate AS bk_end,
               p.EventName AS bk_event, p.AccountName AS bk_org,
               p.Customer  AS bk_hotel
        FROM pickup_config c
        LEFT JOIN ReportPipeline p ON CAST(c.booking_id AS TEXT)=CAST(p.BookingId AS TEXT)
        WHERE c.status='active'{af_where}
    """, af_params).fetchall()
    current_events = []
    for c in configs:
        block            = json.loads(c['contracted_block'] or '{}')
        all_dates        = sorted(block.keys())
        contracted_total = sum(block.values())
        start_date = None
        if all_dates:
            try:
                start_date = _d.fromisoformat(all_dates[0])
            except Exception:
                pass
        if start_date is None and c['bk_start']:
            try:
                _bs = _to_iso(c['bk_start'])
                if _bs:
                    start_date = _d.fromisoformat(_bs)
            except Exception:
                pass
        has_final = db.execute(
            "SELECT COUNT(*) FROM pickup_weekly WHERE config_id=? AND LOWER(label) LIKE '%final%'",
            (c['id'],)
        ).fetchone()[0] > 0
        if has_final or bool(c['force_past']):
            continue
        if not c['force_current']:
            if start_date is not None and start_date < today:
                continue
            if start_date is not None and start_date > future_cutoff:
                continue
        last_w = db.execute(
            """SELECT total_rooms, pct_of_block FROM pickup_weekly
               WHERE config_id=? AND (label IS NULL OR label NOT LIKE '%final%')
               ORDER BY report_date DESC LIMIT 1""",
            (c['id'],)
        ).fetchone()
        weekly_rows = db.execute(
            """SELECT report_date, total_rooms, pct_of_block, pct_of_attrition, pickup_by_night
               FROM pickup_weekly
               WHERE config_id=?
                 AND (label IS NULL
                      OR (label NOT LIKE 'ASAS %'
                          AND label NOT LIKE 'Historical%'
                          AND LOWER(label) NOT LIKE '%final%'))
               ORDER BY report_date DESC""",
            (c['id'],)
        ).fetchall()
        org        = c['organization'] or c['bk_org'] or ''
        event_name = c['event_name']   or c['bk_event'] or org
        end_date   = None
        if all_dates:
            try:
                end_date = _d.fromisoformat(all_dates[-1])
            except Exception:
                pass
        # Extend event_dates to include any shoulder nights recorded in pickup_by_night
        all_event_date_set = set(all_dates)
        for w in weekly_rows:
            try:
                pbn = json.loads(w['pickup_by_night'] or '{}')
                for d in pbn.keys():
                    if d:
                        all_event_date_set.add(d)
            except Exception:
                pass
        all_event_dates = sorted(all_event_date_set)

        current_events.append({
            'config':           c,
            'block':            block,
            'event_dates':      all_event_dates,
            'contracted_total': contracted_total,
            'org':              org,
            'event_name':       event_name,
            'hotel':            c['hotel'] or c['bk_hotel'] or '',
            'start_date':       start_date,
            'end_date':         end_date,
            'weekly':           [dict(w) for w in weekly_rows],
            'last_total':       last_w['total_rooms']  if last_w else 0,
            'last_pct':         last_w['pct_of_block'] if last_w else None,
        })
    current_events.sort(key=lambda r: (r['org'].lower(), str(r['start_date'] or '')))
    return current_events


@app.route('/pickup/customer-report', methods=['GET'])
def pickup_customer_report_select():
    """Selection page — choose which customers/events to include in the report."""
    user = get_current_user()
    db = get_db()
    current_events = _get_current_pickup_events(db, account_filter=get_pickup_account_filter(user))

    # Group by organisation for the UI
    from collections import OrderedDict
    orgs = OrderedDict()
    for ev in current_events:
        key = ev['org'] or '(No Organisation)'
        orgs.setdefault(key, []).append(ev)

    return render_template('pickup_customer_report_select.html',
                           orgs=orgs, total=len(current_events))


@app.route('/pickup/customer-report-xlsx', methods=['GET', 'POST'])
def pickup_customer_report_xlsx():
    """Generate multi-tab XLSX: Tab 1 = summary sorted by customer, Tab N = per-event pace report."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import date as _d
    import io, re as _re

    user  = get_current_user()
    db    = get_db()
    today = _d.today()

    # Load all current events via shared helper (filtered by account access)
    all_events = _get_current_pickup_events(db, account_filter=get_pickup_account_filter(user))

    # If POSTed from the selection page, filter to chosen config IDs only
    if request.method == 'POST':
        selected_ids = set(request.form.getlist('cids'))
        if selected_ids:
            current_events = [ev for ev in all_events
                              if str(ev['config']['id']) in selected_ids]
        else:
            current_events = all_events   # nothing checked → include all
    else:
        current_events = all_events       # direct GET → include all

    if not current_events:
        flash('No events selected — nothing to download.', 'warning')
        return redirect(url_for('pickup_customer_report_select'))

    # ── Colour / style constants ─────────────────────────────────────────────
    navy       = "1A3A5C"
    mid_navy   = "2D5986"
    setup_fg   = "BFBFBF"   # rows 9–12 gray
    green_fg   = "92D050"   # final / remaining rows
    org_hdr_fg = "D9E1F2"   # org separator in summary

    def _fill(hex6):      return PatternFill("solid", fgColor=hex6)
    def _font(bold=False, size=11, color="000000", italic=False):
        return Font(bold=bold, size=size, color=color, italic=italic)

    hair  = Side(style='hair')
    thin  = Side(style='thin')
    def _border(**kw): return Border(**kw)

    wb = Workbook()

    # ════════════════════════════════════════════════════════════════════════
    # TAB 1 — SUMMARY
    # ════════════════════════════════════════════════════════════════════════
    ws1       = wb.active
    ws1.title = "Summary"

    SUMM_COLS = 13
    last_ltr  = get_column_letter(SUMM_COLS)

    # Row 1 – title banner
    ws1.merge_cells(f"A1:{last_ltr}1")
    ws1["A1"].value     = f"PICK-UP SUMMARY REPORT — {today.year}"
    ws1["A1"].font      = Font(bold=True, size=18, color="FFFFFF")
    ws1["A1"].fill      = _fill(navy)
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 30

    # Row 2 – sub-header
    ws1.merge_cells(f"A2:{last_ltr}2")
    ws1["A2"].value     = (f"Generated: {today.strftime('%B %d, %Y')}   |   "
                           f"Current events sorted by customer")
    ws1["A2"].font      = Font(italic=True, size=10, color="555555")
    ws1["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws1.row_dimensions[2].height = 16

    # Row 3 – column headers
    hdrs = ["Organization", "Event Name", "Hotel", "Start", "End",
            "Block", "Pickup", "WoW Chg", "% Block", "% Attrition",
            "Cutoff", "Days Left", "Status"]
    for ci, h in enumerate(hdrs, 1):
        cell            = ws1.cell(3, ci, h)
        cell.font       = Font(bold=True, color="FFFFFF", size=11)
        cell.fill       = _fill(mid_navy)
        cell.alignment  = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[3].height = 18

    # Column widths for summary
    for ci, w in enumerate([30,30,22,11,11,10,10,10,9,12,12,9,10], 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    # Data rows
    row_idx  = 4
    prev_org = None
    alt_fills = [_fill("F2F7FF"), _fill("FFFFFF")]
    green_txt  = Font(bold=True, size=11, color="006400")
    yellow_txt = Font(bold=True, size=11, color="7D6608")
    red_txt    = Font(bold=True, size=11, color="CC0000")

    for ev in current_events:
        # Org separator row when organisation changes
        if ev['org'] != prev_org:
            ws1.merge_cells(f"A{row_idx}:{last_ltr}{row_idx}")
            cell            = ws1.cell(row_idx, 1, ev['org'] or '—')
            cell.font       = Font(bold=True, size=11, color=navy)
            cell.fill       = _fill(org_hdr_fg)
            cell.alignment  = Alignment(horizontal="left", vertical="center")
            ws1.row_dimensions[row_idx].height = 16
            row_idx += 1
            prev_org = ev['org']

        weekly = ev['weekly']
        last_w = weekly[0] if weekly else None
        last_total     = last_w['total_rooms']     if last_w else 0
        last_pct_block = last_w['pct_of_block']    if last_w else None
        last_pct_attr  = last_w['pct_of_attrition'] if last_w else None
        wow = None
        if len(weekly) >= 2 and last_w:
            wow = (last_w['total_rooms'] or 0) - (weekly[1]['total_rooms'] or 0)
        days_left = None
        if ev['config']['cutoff_date']:
            try:
                days_left = (_d.fromisoformat(ev['config']['cutoff_date']) - today).days
            except Exception:
                pass

        status = ('No data' if last_pct_block is None
                  else 'On Pace' if last_pct_block >= 80
                  else 'Watch'   if last_pct_block >= 60
                  else 'At Risk')

        rf = alt_fills[row_idx % 2]
        base_font = _font(size=11)

        def _s(col, val):
            c = ws1.cell(row_idx, col, val)
            c.font = base_font; c.fill = rf
            return c

        _s(1, ev['org'])
        _s(2, ev['event_name'])
        _s(3, ev['hotel'])
        d4 = _s(4, ev['start_date'])
        d4.number_format = "mm-dd-yy"
        d5 = _s(5, ev['end_date'])
        d5.number_format = "mm-dd-yy"
        _s(6, ev['contracted_total']).number_format = "#,##0"
        _s(7, last_total).number_format = "#,##0"
        if wow is not None:
            _s(8, wow).number_format = "+#,##0;-#,##0;0"
        else:
            _s(8, None)

        # % Block — colour-coded
        pct_cell = ws1.cell(row_idx, 9,
                             (last_pct_block / 100) if last_pct_block is not None else None)
        pct_cell.fill          = rf
        pct_cell.number_format = "0%"
        if last_pct_block is not None:
            pct_cell.font = (green_txt  if last_pct_block >= 80
                             else yellow_txt if last_pct_block >= 60
                             else red_txt)

        # % Attrition
        a_cell = ws1.cell(row_idx, 10,
                          (last_pct_attr / 100) if last_pct_attr is not None else None)
        a_cell.fill = rf
        a_cell.number_format = "0%"
        a_cell.font = base_font

        # Cutoff date
        cutoff_cell = ws1.cell(row_idx, 11)
        cutoff_cell.fill = rf
        if ev['config']['cutoff_date']:
            try:
                cutoff_cell.value = _d.fromisoformat(ev['config']['cutoff_date'])
                cutoff_cell.number_format = "mm-dd-yy"
            except Exception:
                cutoff_cell.value = ev['config']['cutoff_date']
        cutoff_cell.font = base_font

        dl_cell = ws1.cell(row_idx, 12, days_left)
        dl_cell.fill = rf
        dl_cell.font = (Font(bold=True, size=11, color="CC0000")
                        if days_left is not None and days_left < 0
                        else Font(bold=True, size=11, color="B8860B")
                        if days_left is not None and days_left <= 14
                        else base_font)

        st_cell = ws1.cell(row_idx, 13, status)
        st_cell.fill = rf
        st_cell.font = (green_txt  if status == 'On Pace'
                        else yellow_txt if status == 'Watch'
                        else red_txt    if status == 'At Risk'
                        else base_font)

        ws1.row_dimensions[row_idx].height = 15
        row_idx += 1

    # ════════════════════════════════════════════════════════════════════════
    # TABS 2+ — PER-EVENT PICK-UP UPDATE (matches example format)
    # ════════════════════════════════════════════════════════════════════════
    setup_fill = _fill(setup_fg)
    green_fill = _fill(green_fg)
    used_names = {"Summary"}
    day_abbrevs = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']

    for ev in current_events:
        # ── Tab name ────────────────────────────────────────────────────────
        raw = (ev['event_name'] or ev['org'] or 'Event')
        tab = _re.sub(r'[:\\/?*\[\]]', '', raw).strip()[:31]
        orig_tab = tab
        i = 2
        while tab in used_names:
            tab = orig_tab[:28] + f' {i}'; i += 1
        used_names.add(tab)

        ws = wb.create_sheet(title=tab)
        c  = ev['config']
        event_dates      = ev['event_dates']
        block            = ev['block']
        contracted_total = ev['contracted_total']
        n_nights         = len(event_dates)

        # ── No dates: stub ───────────────────────────────────────────────────
        if n_nights == 0:
            ws["A1"].value = "PICK-UP UPDATE"
            ws["A1"].font  = Font(bold=True, size=18)
            ws["A2"].value = "ORGANIZATION:"
            ws["A2"].font  = Font(bold=True, size=11)
            ws["B2"].value = ev['org']
            ws["A3"].value = "(Contracted block dates not yet defined)"
            ws["A3"].font  = Font(italic=True, size=10, color="888888")
            continue

        # ── Column index helpers ─────────────────────────────────────────────
        S_COL          = 2                            # first night col (B)
        E_COL          = S_COL + n_nights - 1         # last night col
        TOT_COL        = E_COL + 1                    # Total
        CHG_COL        = TOT_COL + 1                  # Change
        PCT_BLK_COL    = CHG_COL + 1                  # % Block
        PCT_ATR_COL    = PCT_BLK_COL + 1              # % Attrition
        LAST_COL       = PCT_ATR_COL
        s_ltr          = get_column_letter(S_COL)
        e_ltr          = get_column_letter(E_COL)
        tot_ltr        = get_column_letter(TOT_COL)
        chg_ltr        = get_column_letter(CHG_COL)
        pct_blk_ltr    = get_column_letter(PCT_BLK_COL)
        pct_atr_ltr    = get_column_letter(PCT_ATR_COL)
        last_ltr_ev    = get_column_letter(LAST_COL)

        # ── Column widths ────────────────────────────────────────────────────
        ws.column_dimensions['A'].width = 24.5
        for ci in range(S_COL, LAST_COL + 1):
            ws.column_dimensions[get_column_letter(ci)].width = (
                10 if ci == PCT_ATR_COL else 9)

        # ── Row 1: Title ─────────────────────────────────────────────────────
        ws.merge_cells(f"A1:{last_ltr_ev}1")
        ws["A1"].value     = "PICK-UP UPDATE"
        ws["A1"].font      = Font(bold=True, size=18)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # ── Helper: label in A, merged value in B–last ──────────────────────
        def _hdr(row, label, val):
            ws.cell(row, 1, label).font      = Font(bold=True, size=11)
            ws.cell(row, 1).alignment        = Alignment(horizontal="right", vertical="center")
            ws.merge_cells(f"B{row}:{last_ltr_ev}{row}")
            ws.cell(row, 2, val).font        = Font(size=11)
            ws.cell(row, 2).alignment        = Alignment(horizontal="left", wrap_text=True)

        _hdr(2, "ORGANIZATION:", ev['org'])
        _hdr(3, "HOTEL & LOCATION:", ev['hotel'])

        # Event name + date range
        if event_dates:
            try:
                ds = _d.fromisoformat(event_dates[0]).strftime('%B %d')
                de = _d.fromisoformat(event_dates[-1]).strftime('%B %d, %Y')
                ev_date_str = f"{ds} – {de}"
            except Exception:
                ev_date_str = f"{event_dates[0]} – {event_dates[-1]}"
        else:
            ev_date_str = ""
        _hdr(4, "NAME & DATE OF EVENT:", f"{ev['event_name']}   |   {ev_date_str}")

        # Row 5: Group contact
        ws.cell(5, 1, "GROUP CONTACT").font      = Font(bold=True, size=11)
        ws.cell(5, 1).alignment                  = Alignment(horizontal="right")
        ws.cell(5, 2, c['group_contact'] or '').font = Font(size=11)
        if c['group_contact_email']:
            ec = ws.cell(5, min(4, LAST_COL), c['group_contact_email'])
            ec.font = Font(size=10, color="0070C0")

        _hdr(6, "HOTEL CONTACT:", c['hotel_contact'] or '')
        _hdr(7, "CONTACT'S NUMBER:", c['hotel_contact_email'] or '')
        ws.cell(8, 1, "Booking").font      = Font(bold=True, size=11)
        ws.cell(8, 1).alignment            = Alignment(horizontal="right")
        ws.cell(8, 2, c['booking_id']).font = Font(size=11)

        # ── Row 9: Cut-off date + attrition (gray) ───────────────────────────
        ws.row_dimensions[9].height = 45

        def _setup(row):
            for ci in range(1, LAST_COL + 1):
                ws.cell(row, ci).fill = setup_fill

        _setup(9)
        ws.cell(9, 1, "Cut-off Date:").font      = Font(bold=True, size=11)
        ws.cell(9, 1).alignment                  = Alignment(horizontal="right", vertical="center")
        merge_end = min(S_COL + 2, E_COL)
        ws.merge_cells(start_row=9, start_column=S_COL, end_row=9, end_column=merge_end)
        if c['cutoff_date']:
            try:
                ws.cell(9, S_COL).value = _d.fromisoformat(c['cutoff_date'])
                ws.cell(9, S_COL).number_format = "[$-409]mmmm d, yyyy"
            except Exception:
                ws.cell(9, S_COL).value = c['cutoff_date']
        ws.cell(9, S_COL).font      = Font(size=11)
        ws.cell(9, S_COL).alignment = Alignment(horizontal="left", vertical="center")

        attrition_pct   = c['attrition_pct'] or 0
        attrition_rooms = round(contracted_total * attrition_pct) if contracted_total and attrition_pct else 0

        ws.cell(9, TOT_COL, "Attrition:").font      = Font(bold=True, size=11)
        ws.cell(9, CHG_COL).value                   = attrition_pct or None
        ws.cell(9, CHG_COL).number_format            = "0%"
        ws.cell(9, CHG_COL).font                     = Font(size=11)
        ws.cell(9, PCT_BLK_COL).value               = attrition_rooms or None
        ws.cell(9, PCT_BLK_COL).number_format        = "#,##0"
        ws.cell(9, PCT_BLK_COL).font                 = Font(size=11)
        if c['contracted_rate']:
            ws.cell(9, PCT_ATR_COL).value            = c['contracted_rate']
            ws.cell(9, PCT_ATR_COL).number_format    = '"$"#,##0.00'
            ws.cell(9, PCT_ATR_COL).font             = Font(size=11)

        # ── Row 10: Dates (gray) ─────────────────────────────────────────────
        _setup(10)
        ws.cell(10, 1, "Dates:").font      = Font(bold=True, size=11)
        ws.cell(10, 1).alignment           = Alignment(horizontal="right")
        for ni, ed in enumerate(event_dates):
            ci = S_COL + ni
            try:
                ws.cell(10, ci).value = _d.fromisoformat(ed)
                ws.cell(10, ci).number_format = "mm-dd-yy"
            except Exception:
                ws.cell(10, ci).value = ed
            ws.cell(10, ci).alignment = Alignment(horizontal="center")
            ws.cell(10, ci).font      = Font(size=11)
        ws.cell(10, TOT_COL, "Total").font      = Font(bold=True, size=11)
        ws.cell(10, TOT_COL).alignment          = Alignment(horizontal="center")

        # ── Row 11: Day of week (gray) ───────────────────────────────────────
        _setup(11)
        ws.cell(11, 1, "Day:").font    = Font(bold=True, size=11)
        ws.cell(11, 1).alignment       = Alignment(horizontal="right")
        for ni, ed in enumerate(event_dates):
            ci = S_COL + ni
            try:
                ws.cell(11, ci).value = day_abbrevs[_d.fromisoformat(ed).weekday()]
            except Exception:
                pass
            ws.cell(11, ci).alignment = Alignment(horizontal="center")
            ws.cell(11, ci).font      = Font(size=11)

        # ── Row 12: Block (gray) ─────────────────────────────────────────────
        _setup(12)
        ws.row_dimensions[12].height = 18
        ws.cell(12, 1, "Block:").font      = Font(bold=True, size=11)
        ws.cell(12, 1).alignment           = Alignment(horizontal="right")
        for ni, ed in enumerate(event_dates):
            ci = S_COL + ni
            ws.cell(12, ci).value         = block.get(ed, 0) or None
            ws.cell(12, ci).number_format = "#,##0_);[Red](#,##0)"
            ws.cell(12, ci).alignment     = Alignment(horizontal="center")
            ws.cell(12, ci).font          = Font(size=11)
        ws.cell(12, TOT_COL).value         = f"=SUM({s_ltr}12:{e_ltr}12)"
        ws.cell(12, TOT_COL).number_format = "#,##0_);[Red](#,##0)"
        ws.cell(12, TOT_COL).font          = Font(bold=True, size=11)
        ws.cell(12, TOT_COL).alignment     = Alignment(horizontal="center")
        for lbl, ci in [("Change", CHG_COL), ("% Block", PCT_BLK_COL), ("% Attrition", PCT_ATR_COL)]:
            ws.cell(12, ci, lbl).font      = Font(bold=True, size=11)
            ws.cell(12, ci).alignment      = Alignment(horizontal="center")

        # ── Row 13: Pending History (green) ──────────────────────────────────
        ws.row_dimensions[13].height = 18
        ws.cell(13, 1, "Pending History").font      = Font(bold=True, size=11)
        ws.cell(13, 1).fill                         = green_fill
        ws.cell(13, 1).alignment                    = Alignment(horizontal="left")
        for ci in range(S_COL, LAST_COL + 1):
            ws.cell(13, ci).fill = green_fill

        # ── Rows 14+: Weekly pickup entries (most recent first) ───────────────
        weekly     = ev['weekly']
        n_data     = len(weekly)

        for wi, w in enumerate(weekly):
            rn = 14 + wi
            ws.row_dimensions[rn].height = 15

            # Date label
            try:
                ws.cell(rn, 1).value = _d.fromisoformat(w['report_date'])
                ws.cell(rn, 1).number_format = "mm-dd-yy"
            except Exception:
                ws.cell(rn, 1).value = w['report_date']
            ws.cell(rn, 1).font      = Font(size=11)
            ws.cell(rn, 1).alignment = Alignment(horizontal="left")

            # Per-night pickup
            pbn = {}
            try:
                pbn = json.loads(w['pickup_by_night'] or '{}')
            except Exception:
                pass
            for ni, ed in enumerate(event_dates):
                ci    = S_COL + ni
                rooms = pbn.get(ed)
                ws.cell(rn, ci).value         = rooms if rooms else None
                ws.cell(rn, ci).number_format = "#,##0_);[Red](#,##0)"
                ws.cell(rn, ci).alignment     = Alignment(horizontal="center")
                ws.cell(rn, ci).font          = Font(size=11)

            # Total (SUM formula)
            ws.cell(rn, TOT_COL).value         = f"=SUM({s_ltr}{rn}:{e_ltr}{rn})"
            ws.cell(rn, TOT_COL).number_format = "#,##0_);[Red](#,##0)"
            ws.cell(rn, TOT_COL).font          = Font(bold=True, size=11)
            ws.cell(rn, TOT_COL).alignment     = Alignment(horizontal="center")

            # Change vs next (older) row
            if wi < n_data - 1:
                ws.cell(rn, CHG_COL).value = f"={tot_ltr}{rn}-{tot_ltr}{rn+1}"
            else:
                ws.cell(rn, CHG_COL).value = f"={tot_ltr}{rn}"
            ws.cell(rn, CHG_COL).number_format = "#,##0_);[Red](#,##0)"
            ws.cell(rn, CHG_COL).font          = Font(size=11)
            ws.cell(rn, CHG_COL).alignment     = Alignment(horizontal="center")

            # % of contracted block
            ws.cell(rn, PCT_BLK_COL).value         = f"={tot_ltr}{rn}/{tot_ltr}$12"
            ws.cell(rn, PCT_BLK_COL).number_format = "0%"
            ws.cell(rn, PCT_BLK_COL).font          = Font(size=11)
            ws.cell(rn, PCT_BLK_COL).alignment     = Alignment(horizontal="center")

            # % of attrition requirement
            if attrition_rooms:
                ws.cell(rn, PCT_ATR_COL).value         = f"={tot_ltr}{rn}/{pct_blk_ltr}$9"
                ws.cell(rn, PCT_ATR_COL).number_format = "0%"
                ws.cell(rn, PCT_ATR_COL).font          = Font(size=11)
                ws.cell(rn, PCT_ATR_COL).alignment     = Alignment(horizontal="center")

        # ── Last row: Remaining (green) ───────────────────────────────────────
        rem_row = 14 + n_data
        ws.row_dimensions[rem_row].height = 18
        ws.cell(rem_row, 1, "Remaining").font      = Font(bold=True, size=11)
        ws.cell(rem_row, 1).fill                   = green_fill
        ws.cell(rem_row, 1).alignment              = Alignment(horizontal="left")

        latest_data_row = 14  # most recent weekly entry
        for ni, ed in enumerate(event_dates):
            ci      = S_COL + ni
            col_ltr = get_column_letter(ci)
            if n_data > 0:
                ws.cell(rem_row, ci).value = f"={col_ltr}12-{col_ltr}{latest_data_row}"
            else:
                ws.cell(rem_row, ci).value = f"={col_ltr}12"
            ws.cell(rem_row, ci).number_format = "#,##0_);[Red](#,##0)"
            ws.cell(rem_row, ci).alignment     = Alignment(horizontal="center")
            ws.cell(rem_row, ci).font          = Font(size=11)
            ws.cell(rem_row, ci).fill          = green_fill

        # Remaining total
        ws.cell(rem_row, TOT_COL).value = (
            f"={tot_ltr}12-{tot_ltr}{latest_data_row}" if n_data > 0
            else f"={tot_ltr}12")
        ws.cell(rem_row, TOT_COL).number_format = "#,##0_);[Red](#,##0)"
        ws.cell(rem_row, TOT_COL).font          = Font(bold=True, size=11)
        ws.cell(rem_row, TOT_COL).alignment     = Alignment(horizontal="center")
        ws.cell(rem_row, TOT_COL).fill          = green_fill

    # ── Return file ───────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f"Pickup_Customer_Report_{today.strftime('%Y%m%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fn,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/pickup/<int:cid>/upload-contract', methods=['GET', 'POST'])
def pickup_upload_contract(cid):
    """Upload a PDF/Word contract, extract room block with AI, review then save."""
    from pickup_utils import parse_contract_document
    db = get_db()
    config = db.execute("SELECT * FROM pickup_config WHERE id=?", (cid,)).fetchone()
    if not config:
        flash('Event not found.', 'error')
        return redirect(url_for('pickup_dashboard'))

    if request.method == 'POST':
        action = request.form.get('action', 'upload')

        # ── Step 2: user confirmed the extracted data ──────────────────────
        if action == 'confirm':
            try:
                raw_block = request.form.get('contracted_block', '{}')
                block     = json.loads(raw_block)
                rate_str   = request.form.get('contracted_rate', '').strip()
                cutoff     = request.form.get('cutoff_date', '').strip() or None
                atr_str    = request.form.get('attrition_pct', '').strip()
                rebate_str = request.form.get('rebate_per_room', '').strip()
                rate       = float(rate_str)   if rate_str   else config['contracted_rate']
                atr        = float(atr_str)    if atr_str    else config['attrition_pct']
                rebate     = float(rebate_str) if rebate_str else None
                # Contact fields are never updated on contract import —
                # block/rate/cutoff/attrition only.

                contract_data     = request.form.get('_contract_data_b64', '')
                contract_filename = request.form.get('_contract_filename', '')
                block_review_date = request.form.get('block_review_date', '').strip() or None
                import base64
                file_blob = base64.b64decode(contract_data) if contract_data else None

                db.execute('''
                    UPDATE pickup_config
                    SET contracted_block     = ?,
                        contracted_rate      = ?,
                        rebate_per_room      = COALESCE(?, rebate_per_room),
                        cutoff_date          = ?,
                        attrition_pct        = ?,
                        block_is_estimated   = 0,
                        contract_filename    = ?,
                        contract_data        = ?,
                        block_review_date    = COALESCE(?, block_review_date)
                    WHERE id = ?
                ''', (json.dumps(block), rate, rebate, cutoff, atr,
                      contract_filename or None, file_blob,
                      block_review_date, cid))

                # Cascade rich extraction → cost_savings_report rows
                cs_flash = None
                if file_blob:
                    cs_res = cascade_contract_cost_savings(
                        db, file_blob, contract_filename,
                        booking_id=config['booking_id'],
                        pickup_config_id=cid,
                    )
                    if cs_res and not cs_res.get('error'):
                        cs_flash = 'Cost Savings reports updated with contract values.'

                db.commit()
                flash('Contract data saved — room block updated.', 'success')
                if cs_flash:
                    flash(cs_flash, 'success')
                return redirect(url_for('pickup_event', cid=cid))
            except Exception as e:
                flash(f'Error saving contract data: {e}', 'error')
                return redirect(url_for('pickup_event', cid=cid))

        # ── Step 1: parse the uploaded file ───────────────────────────────
        f = request.files.get('contract_file')
        if not f or not f.filename:
            flash('No file selected.', 'error')
            return redirect(url_for('pickup_upload_contract', cid=cid))

        file_bytes = f.read()
        extracted  = parse_contract_document(file_bytes, filename=f.filename)

        if extracted.get('error'):
            flash(f'Could not parse contract: {extracted["error"]}', 'error')
            return redirect(url_for('pickup_event', cid=cid))

        # Don't round-trip the PDF through the form — file_b64 stays empty
        return render_template('pickup_contract_review.html',
                               config=config, extracted=extracted,
                               filename=f.filename, file_b64='')

    return render_template('pickup_contract_upload.html', config=config)


@app.route('/pickup/<int:cid>/contract/download')
def pickup_contract_download(cid):
    """Download the stored contract file."""
    db = get_db()
    row = db.execute(
        "SELECT contract_filename, contract_data FROM pickup_config WHERE id=?", (cid,)
    ).fetchone()
    if not row or not row['contract_data']:
        flash('No contract on file.', 'error')
        return redirect(url_for('pickup_event', cid=cid))
    buf = io.BytesIO(row['contract_data'])
    return send_file(buf, as_attachment=True,
                     download_name=row['contract_filename'] or f'contract_{cid}.pdf')


@app.route('/pickup/bulk-upload/<int:primary_cid>', methods=['GET', 'POST'])
def pickup_bulk_upload(primary_cid):
    """Bulk contract upload for multi-hotel events."""
    import uuid
    from pickup_utils import parse_contract_document
    db = get_db()
    primary = db.execute("SELECT * FROM pickup_config WHERE id=?", (primary_cid,)).fetchone()
    if not primary:
        flash('Event not found.', 'error')
        return redirect(url_for('pickup_dashboard'))

    event_name = primary['event_name'] or primary['organization']

    # All siblings (same event_name, active)
    siblings = db.execute(
        """SELECT id, hotel, contract_filename FROM pickup_config
           WHERE LOWER(TRIM(event_name))=LOWER(TRIM(?)) AND status='active'
           ORDER BY hotel""",
        (event_name,)
    ).fetchall()

    if request.method == 'POST':
        batch_id = uuid.uuid4().hex
        for sib in siblings:
            cid = sib['id']
            f = request.files.get(f'file_{cid}')
            if not f or not f.filename:
                continue
            file_bytes = f.read()
            extracted = parse_contract_document(file_bytes, filename=f.filename)
            if extracted.get('error'):
                db.execute(
                    """INSERT INTO pickup_bulk_pending
                       (batch_id, config_id, hotel, filename, status, error_msg)
                       VALUES (?,?,?,?,?,?)""",
                    (batch_id, cid, sib['hotel'], f.filename, 'failed', extracted['error'])
                )
            else:
                db.execute(
                    """INSERT INTO pickup_bulk_pending
                       (batch_id, config_id, hotel, filename, extracted_json, contract_data, status)
                       VALUES (?,?,?,?,?,?,?)""",
                    (batch_id, cid, sib['hotel'], f.filename,
                     json.dumps(extracted), file_bytes, 'ok')
                )
        db.commit()
        return redirect(url_for('pickup_bulk_review', batch_id=batch_id))

    return render_template('pickup_bulk_upload.html',
                           primary=primary,
                           event_name=event_name,
                           siblings=siblings,
                           primary_cid=primary_cid)


@app.route('/pickup/bulk-review')
def pickup_bulk_review():
    """Review extracted contracts from a bulk upload batch."""
    batch_id = request.args.get('batch_id', '')
    if not batch_id:
        flash('No batch ID provided.', 'error')
        return redirect(url_for('pickup_dashboard'))

    db = get_db()
    rows = db.execute(
        "SELECT * FROM pickup_bulk_pending WHERE batch_id=? ORDER BY hotel",
        (batch_id,)
    ).fetchall()

    if not rows:
        flash('Batch not found or already processed.', 'error')
        return redirect(url_for('pickup_dashboard'))

    ok_items = []
    failed_items = []
    primary_cid = None

    for row in rows:
        config = db.execute("SELECT * FROM pickup_config WHERE id=?", (row['config_id'],)).fetchone()
        item = dict(row)
        item['config'] = config
        if row['status'] == 'ok':
            item['extracted'] = json.loads(row['extracted_json'] or '{}')
            ok_items.append(item)
        else:
            failed_items.append(item)
        if primary_cid is None and config:
            # Find the primary cid for the event
            event_name = config['event_name'] or config['organization']
            pc = db.execute(
                """SELECT MIN(id) as pid FROM pickup_config
                   WHERE LOWER(TRIM(event_name))=LOWER(TRIM(?)) AND status='active'""",
                (event_name,)
            ).fetchone()
            primary_cid = pc['pid'] if pc else config['id']

    # Determine event_name from first row
    event_name = ''
    if rows:
        first_config = db.execute("SELECT * FROM pickup_config WHERE id=?", (rows[0]['config_id'],)).fetchone()
        if first_config:
            event_name = first_config['event_name'] or first_config['organization']

    return render_template('pickup_bulk_review.html',
                           batch_id=batch_id,
                           event_name=event_name,
                           ok_items=ok_items,
                           failed_items=failed_items,
                           primary_cid=primary_cid)


@app.route('/pickup/bulk-confirm', methods=['POST'])
def pickup_bulk_confirm():
    """Confirm and save all bulk-uploaded contracts."""
    db = get_db()
    batch_id = request.form.get('batch_id', '')
    primary_cid = request.form.get('primary_cid', '')

    if not batch_id:
        flash('Missing batch ID.', 'error')
        return redirect(url_for('pickup_dashboard'))

    rows = db.execute(
        "SELECT * FROM pickup_bulk_pending WHERE batch_id=? AND status='ok'",
        (batch_id,)
    ).fetchall()

    saved_count = 0
    for row in rows:
        pending_id = row['id']
        # Skip if checkbox checked
        if request.form.get(f'skip_{pending_id}'):
            continue

        try:
            extracted = json.loads(row['extracted_json'] or '{}')

            rate_str = request.form.get(f'rate_{pending_id}', '').strip()
            cutoff   = request.form.get(f'cutoff_{pending_id}', '').strip() or None
            block_raw = request.form.get(f'block_json_{pending_id}', '{}')
            block    = json.loads(block_raw)

            rate = float(rate_str) if rate_str else extracted.get('contracted_rate') or None

            hotel_contact       = request.form.get(f'hotel_contact_{pending_id}', '').strip() or None
            hotel_contact_email = request.form.get(f'hotel_contact_email_{pending_id}', '').strip() or None
            group_contact       = request.form.get(f'group_contact_{pending_id}', '').strip() or None
            group_contact_email = request.form.get(f'group_contact_email_{pending_id}', '').strip() or None

            db.execute('''
                UPDATE pickup_config SET
                    contracted_block      = ?,
                    contracted_rate       = ?,
                    cutoff_date           = ?,
                    attrition_pct         = 0.8,
                    block_is_estimated    = 0,
                    contract_filename     = ?,
                    contract_data         = ?,
                    hotel_contact         = COALESCE(?, hotel_contact),
                    hotel_contact_email   = COALESCE(?, hotel_contact_email),
                    group_contact         = COALESCE(?, group_contact),
                    group_contact_email   = COALESCE(?, group_contact_email)
                WHERE id = ?
            ''', (json.dumps(block), rate, cutoff,
                  row['filename'], row['contract_data'],
                  hotel_contact, hotel_contact_email,
                  group_contact, group_contact_email,
                  row['config_id']))
            saved_count += 1
        except Exception as e:
            flash(f'Error saving contract for {row["hotel"]}: {e}', 'error')

    # Clean up pending rows
    db.execute("DELETE FROM pickup_bulk_pending WHERE batch_id=?", (batch_id,))
    db.commit()

    flash(f'{saved_count} contract(s) saved successfully.', 'success')
    if primary_cid:
        return redirect(url_for('pickup_event', cid=int(primary_cid)))
    return redirect(url_for('pickup_dashboard'))


@app.route('/pickup/<int:cid>/import-amendment', methods=['GET', 'POST'])
def pickup_import_amendment(cid):
    """Upload a contract amendment/addendum PDF, AI-extract changes, review then apply."""
    from pickup_utils import parse_amendment_document
    db = get_db()
    config = db.execute("SELECT * FROM pickup_config WHERE id=?", (cid,)).fetchone()
    if not config:
        flash('Event not found.', 'error')
        return redirect(url_for('pickup_dashboard'))

    if request.method == 'POST':
        action = request.form.get('action', 'upload')

        # ── Step 2: user confirmed the diff ───────────────────────────────────
        if action == 'confirm':
            try:
                import base64
                current_block = json.loads(config['contracted_block'] or '{}')

                # Collect changes from form
                new_rate_str    = request.form.get('new_rate', '').strip()
                new_hotel       = request.form.get('new_hotel', '').strip() or None
                new_cutoff      = request.form.get('new_cutoff', '').strip() or None
                description     = request.form.get('description', '').strip()
                block_changes   = json.loads(request.form.get('block_changes_json', '{}'))

                filename_stored = request.form.get('_amend_filename', '')
                tmp_id   = request.form.get('_tmp_id', '')
                tmp_path = os.path.join(tempfile.gettempdir(), f'amend_{tmp_id}') if tmp_id else ''
                file_blob = None
                if tmp_path and os.path.exists(tmp_path):
                    with open(tmp_path, 'rb') as _fh:
                        file_blob = _fh.read()
                    try:
                        os.remove(tmp_path)
                    except Exception:
                        pass

                # Build audit diff
                changes = {}
                if new_rate_str:
                    new_rate = round(float(new_rate_str), 2)
                    changes['contracted_rate'] = {
                        'old': config['contracted_rate'],
                        'new': new_rate
                    }
                else:
                    new_rate = None

                if new_hotel:
                    changes['hotel'] = {'old': config['hotel'], 'new': new_hotel}

                if new_cutoff:
                    changes['cutoff_date'] = {'old': config['cutoff_date'], 'new': new_cutoff}

                # Apply block changes
                replace_block  = request.form.get('replace_block') == '1'
                merged_block   = dict(current_block)
                block_diff     = {}

                # If replace mode: nights in current block NOT in amendment → drop them
                if replace_block:
                    for d in list(current_block.keys()):
                        if d not in block_changes:
                            block_diff[d] = {'old': current_block[d], 'new': None}
                            merged_block.pop(d, None)

                for d, rooms in block_changes.items():
                    rooms = int(rooms)
                    old_rooms = current_block.get(d)
                    if rooms == 0:
                        # Remove this night
                        block_diff[d] = {'old': old_rooms, 'new': None}
                        merged_block.pop(d, None)
                    else:
                        block_diff[d] = {'old': old_rooms, 'new': rooms}
                        merged_block[d] = rooms

                if block_diff:
                    changes['contracted_block'] = block_diff

                # Apply to pickup_config
                update_parts = []
                update_vals  = []
                if new_rate is not None:
                    update_parts.append('contracted_rate = ?')
                    update_vals.append(new_rate)
                if new_hotel:
                    update_parts.append('hotel = ?')
                    update_vals.append(new_hotel)
                if new_cutoff:
                    update_parts.append('cutoff_date = ?')
                    update_vals.append(new_cutoff)
                if merged_block != current_block:
                    update_parts.append('contracted_block = ?')
                    update_vals.append(json.dumps(merged_block))

                if update_parts:
                    update_vals.append(cid)
                    db.execute(
                        f"UPDATE pickup_config SET {', '.join(update_parts)} WHERE id=?",
                        update_vals
                    )

                # Store amendment record
                user = session.get('user', {})
                db.execute('''
                    INSERT INTO pickup_amendments
                        (config_id, uploaded_by, filename, file_data, description, changes_json)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (cid,
                      user.get('name') or user.get('username') or 'Unknown',
                      filename_stored or None,
                      file_blob,
                      description,
                      json.dumps(changes)))

                db.commit()
                flash('Amendment applied and recorded.', 'success')
                return redirect(url_for('pickup_event', cid=cid))

            except Exception as e:
                flash(f'Error applying amendment: {e}', 'error')
                return redirect(url_for('pickup_event', cid=cid))

        # ── Step 1: parse the uploaded file ───────────────────────────────────
        f = request.files.get('amendment_file')
        if not f or not f.filename:
            flash('No file selected.', 'error')
            return redirect(url_for('pickup_event', cid=cid))

        file_bytes = f.read()
        extracted  = parse_amendment_document(file_bytes, filename=f.filename)

        if extracted.get('error'):
            flash(f'Could not parse amendment: {extracted["error"]}', 'error')
            return redirect(url_for('pickup_event', cid=cid))

        # Write file to temp — avoids round-tripping large PDFs through the form
        import uuid as _uuid
        tmp_id   = _uuid.uuid4().hex
        tmp_path = os.path.join(tempfile.gettempdir(), f'amend_{tmp_id}')
        with open(tmp_path, 'wb') as _fh:
            _fh.write(file_bytes)

        current_block = json.loads(config['contracted_block'] or '{}')

        # ── Date-shift: re-key the existing block to the new dates ───────────
        date_shift_applied = False
        date_shift = extracted.get('date_shift') or {}
        if date_shift and not extracted.get('contracted_block') and current_block:
            try:
                from datetime import date as _ddate, timedelta as _td
                old_start = _ddate.fromisoformat(date_shift['old_start'])
                new_start = _ddate.fromisoformat(date_shift['new_start'])
                delta = new_start - old_start
                shifted = {}
                for d, rooms in sorted(current_block.items()):
                    orig = _ddate.fromisoformat(d)
                    shifted[(orig + delta).isoformat()] = rooms
                extracted['contracted_block'] = shifted
                # Also shift the cutoff date by the same offset
                if config['cutoff_date'] and not extracted.get('cutoff_date'):
                    try:
                        shifted_cutoff = (_ddate.fromisoformat(config['cutoff_date']) + delta).isoformat()
                        extracted['cutoff_date'] = shifted_cutoff
                    except Exception:
                        pass
                date_shift_applied = True
            except Exception:
                pass  # leave contracted_block as None; user sees empty table

        return render_template('pickup_amendment_review.html',
                               config=config,
                               extracted=extracted,
                               current_block=current_block,
                               filename=f.filename,
                               tmp_id=tmp_id,
                               date_shift=date_shift,
                               date_shift_applied=date_shift_applied)

    # GET — redirect back (upload happens via modal POST)
    return redirect(url_for('pickup_event', cid=cid))


@app.route('/pickup/<int:cid>/amendment/<int:amid>/download')
def pickup_amendment_download(cid, amid):
    """Download a stored amendment file."""
    db = get_db()
    row = db.execute(
        "SELECT filename, file_data FROM pickup_amendments WHERE id=? AND config_id=?",
        (amid, cid)
    ).fetchone()
    if not row or not row['file_data']:
        flash('Amendment file not found.', 'error')
        return redirect(url_for('pickup_event', cid=cid))
    buf = io.BytesIO(row['file_data'])
    return send_file(buf, as_attachment=True,
                     download_name=row['filename'] or f'amendment_{amid}.pdf')


def _build_housing_form_wb(config, pipeline, pickup_dates=None):
    """Return a filled openpyxl Workbook for the Housing History Form.

    pickup_dates: optional list/set of ISO date strings from the latest pickup
    report. Shoulder nights present in the pickup but not in the contracted block
    are added as padding columns (0 contracted block) so the hotel can fill them in.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    from datetime import datetime as _dt

    block        = json.loads(config['contracted_block'] or '{}')
    sorted_dates = sorted(block.keys())

    # Build the full date range: contracted dates + any shoulder nights from pickup
    if pickup_dates:
        all_dates = sorted(set(sorted_dates) | set(pickup_dates))
    elif sorted_dates:
        all_dates = sorted_dates
    else:
        all_dates = sorted_dates
    n_dates = len(all_dates)

    comm_pct   = float(pipeline['CommissionPercent']) if pipeline and pipeline['CommissionPercent'] else 0.0
    currency   = (pipeline.get('Currency') or 'USD') if pipeline else 'USD'
    booking_id = config['booking_id'] or ''
    org_name   = (pipeline['AccountName'] or '') if pipeline else (config['organization'] or '')
    hotel_name = config['hotel'] or ''
    event_name = config['event_name'] or config['organization'] or ''

    if sorted_dates:
        start_str  = _dt.strptime(sorted_dates[0],  '%Y-%m-%d').strftime('%m/%d/%Y')
        end_str    = _dt.strptime(sorted_dates[-1], '%Y-%m-%d').strftime('%m/%d/%Y')
        event_cell = f"{event_name}  {start_str} – {end_str}"
    else:
        event_cell = event_name

    tmpl_path = os.path.join(os.path.dirname(__file__), 'static', 'housing_history_template.xlsx')
    wb = load_workbook(tmpl_path)
    ws = wb.active

    n_extra   = max(0, n_dates - 10)
    if n_extra > 0:
        ws.insert_cols(13, n_extra)

    total_col = 13 + n_extra
    rate_col  = 14 + n_extra

    ws['J1'] = comm_pct
    ws.cell(row=1, column=13 + n_extra).value = booking_id
    ws.cell(row=1, column=16 + n_extra).value = currency

    ws['C2'] = org_name
    ws['C3'] = hotel_name
    ws['C4'] = event_cell

    days_abbr = ['Mon','Tue','Wed','Thu','Fri','Sat','Sun']
    center    = Alignment(horizontal='center', vertical='center')
    col_start = 3

    for col_offset in range(n_dates):
        col = col_start + col_offset
        for row in (5, 6, 7):
            ws.cell(row=row, column=col).value = None

    for i, d in enumerate(all_dates):
        col      = col_start + i
        date_obj = _dt.strptime(d, '%Y-%m-%d')
        c5 = ws.cell(row=5, column=col)
        c6 = ws.cell(row=6, column=col)
        c7 = ws.cell(row=7, column=col)
        c5.value = date_obj;  c5.alignment = center
        c6.value = days_abbr[date_obj.weekday()]; c6.alignment = center
        c7.value = block.get(d, 0); c7.alignment = center   # 0 for shoulder nights

    ws.cell(row=7, column=total_col).value = sum(block.get(d, 0) for d in sorted_dates)
    ws.cell(row=7, column=rate_col).value  = config['contracted_rate'] or 0

    return wb, event_name


@app.route('/pickup/<int:cid>/housing-form')
def pickup_housing_form(cid):
    """Generate a pre-filled Housing History Form Excel and return as download."""
    db = get_db()
    config = db.execute("SELECT * FROM pickup_config WHERE id=?", (cid,)).fetchone()
    if not config:
        flash('Event not found.', 'error')
        return redirect(url_for('pickup_dashboard'))

    pipeline = None
    if config['booking_id']:
        pipeline = db.execute(
            'SELECT * FROM ReportPipeline WHERE BookingId = ?', (config['booking_id'],)
        ).fetchone()
    latest_weekly = db.execute(
        "SELECT pickup_by_night FROM pickup_weekly WHERE config_id=? ORDER BY report_date DESC LIMIT 1",
        (cid,)
    ).fetchone()
    pickup_dates = list(json.loads(latest_weekly['pickup_by_night'] or '{}').keys()) if latest_weekly else None
    wb, event_name = _build_housing_form_wb(config, pipeline, pickup_dates=pickup_dates)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = event_name.replace('/', '-').replace(' ', '_')[:50]
    return send_file(buf, as_attachment=True,
                     download_name=f"Housing_History_{safe_name}.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/pickup/<int:cid>/final-history', methods=['GET', 'POST'])
def pickup_final_history(cid):
    """Enter or update the Final History (actual final pickup numbers) for a past event."""
    db = get_db()
    config = db.execute("SELECT * FROM pickup_config WHERE id=?", (cid,)).fetchone()
    if not config:
        flash('Event not found.', 'error')
        return redirect(url_for('pickup_dashboard'))

    last = db.execute(
        """SELECT * FROM pickup_weekly WHERE config_id=?
           AND (total_rooms IS NOT NULL AND total_rooms > 0)
           AND (label IS NULL OR (label NOT LIKE '%Pending%' AND label NOT LIKE '%placeholder%'))
           ORDER BY report_date DESC LIMIT 1""",
        (cid,)
    ).fetchone()

    existing_fh = db.execute(
        "SELECT * FROM pickup_weekly WHERE config_id=? AND label='Final History' ORDER BY id DESC LIMIT 1",
        (cid,)
    ).fetchone()

    # Build full date set: contracted block + shoulder nights from any weekly entry
    block = json.loads(config['contracted_block'] or '{}')
    all_weekly = db.execute(
        "SELECT pickup_by_night FROM pickup_weekly WHERE config_id=? AND label != 'Final History'",
        (cid,)
    ).fetchall()
    all_date_set = set(block.keys())
    for _w in all_weekly:
        try:
            _pbn = json.loads(_w['pickup_by_night'] or '{}')
            all_date_set.update(d for d in _pbn if d != 'historical_total')
        except Exception:
            pass
    all_dates = sorted(all_date_set)

    if request.method == 'POST':
        f = request.form
        pickup_by_night = {}
        # Collect known dates (contracted block + prior weekly shoulder nights)
        for d in all_dates:
            val = f.get(f'night_{d}', '').strip()
            if val != '':
                pickup_by_night[d] = int(val)
        # Also collect any dynamically-added shoulder dates submitted via JS
        import re as _re
        for key in f:
            m = _re.match(r'^night_(\d{4}-\d{2}-\d{2})$', key)
            if m:
                d = m.group(1)
                if d not in pickup_by_night:
                    val = f.get(key, '').strip()
                    if val != '':
                        pickup_by_night[d] = int(val)
        total_rooms      = sum(pickup_by_night.values())
        contracted_total = sum(block.values())
        pct_of_block     = round(total_rooms / contracted_total * 100, 1) if contracted_total else None
        attrition_floor  = contracted_total * (config['attrition_pct'] or 0)
        pct_of_attrition = round(total_rooms / attrition_floor * 100, 1) if attrition_floor else None
        prev = db.execute(
            "SELECT total_rooms FROM pickup_weekly WHERE config_id=? AND label IS NULL "
            "ORDER BY report_date DESC, id DESC LIMIT 1", (cid,)
        ).fetchone()
        change_from_last = (total_rooms - prev['total_rooms']) if prev and prev['total_rooms'] else None
        report_date = f.get('report_date') or datetime.now().strftime('%Y-%m-%d')
        notes = f.get('notes', '').strip() or None

        if existing_fh:
            db.execute('''UPDATE pickup_weekly
                SET report_date=?, pickup_by_night=?, total_rooms=?,
                    change_from_last=?, pct_of_block=?, pct_of_attrition=?, notes=?
                WHERE id=?''',
                (report_date, json.dumps(pickup_by_night), total_rooms,
                 change_from_last, pct_of_block, pct_of_attrition, notes, existing_fh['id']))
        else:
            db.execute('''INSERT INTO pickup_weekly
                (config_id, report_date, pickup_by_night, total_rooms,
                 change_from_last, pct_of_block, pct_of_attrition, label, notes)
                VALUES (?,?,?,?,?,?,?,?,?)''',
                (cid, report_date, json.dumps(pickup_by_night), total_rooms,
                 change_from_last, pct_of_block, pct_of_attrition, 'Final History', notes))
        db.commit()
        flash('Final History saved — event moved to Past Events.', 'success')
        return redirect(url_for('pickup_event', cid=cid))

    prefill_entry = existing_fh or last
    last_pickup = json.loads(prefill_entry['pickup_by_night'] or '{}') if prefill_entry else {}

    # If an HHR import just happened, use its richer pickup data (includes shoulder nights)
    import flask as _flask
    hhr_import_pickup = None
    if _flask.session.get('hhr_import_cid') == cid and _flask.session.get('hhr_import_pickup'):
        try:
            hhr_import_pickup = json.loads(_flask.session.pop('hhr_import_pickup'))
            _flask.session.pop('hhr_import_cid', None)
            last_pickup = hhr_import_pickup
        except Exception:
            pass

    return render_template('pickup_final_history_form.html',
                           config=config, block=block, all_dates=all_dates,
                           last_pickup=last_pickup, existing_fh=existing_fh,
                           today=datetime.now().strftime('%Y-%m-%d'),
                           hhr_imported=(hhr_import_pickup is not None))


@app.route('/pickup/<int:cid>/import-completed-hhr', methods=['POST'])
def pickup_import_completed_hhr(cid):
    """Accept a completed HHR (PDF or Excel) returned by the hotel.
    Parses the FINAL TOTAL PICKUP row and redirects to Final History form pre-filled.
    """
    db = get_db()
    config = db.execute("SELECT * FROM pickup_config WHERE id=?", (cid,)).fetchone()
    if not config:
        flash('Event not found.', 'error')
        return redirect(url_for('pickup_dashboard'))

    f = request.files.get('hhr_file')
    if not f or not f.filename:
        flash('No file uploaded.', 'error')
        return redirect(url_for('pickup_event', cid=cid))

    file_bytes = f.read()
    filename   = f.filename.lower()

    parsed = {}
    if filename.endswith('.pdf'):
        from pickup_utils import parse_hhr_pdf as _parse_pdf
        parsed = _parse_pdf(file_bytes)
    elif filename.endswith(('.xlsx', '.xlsm', '.xls')):
        from pickup_utils import parse_hhr_excel as _parse_xl
        parsed = _parse_xl(file_bytes)
    else:
        flash('Unsupported file type. Upload a PDF or Excel file.', 'error')
        return redirect(url_for('pickup_event', cid=cid))

    if parsed.get('error'):
        flash(f'Could not parse HHR: {parsed["error"]}', 'error')
        return redirect(url_for('pickup_event', cid=cid))

    pickup_by_night = parsed.get('pickup_by_night') or {}
    if not pickup_by_night:
        flash('No per-night pickup data found in the uploaded file.', 'error')
        return redirect(url_for('pickup_event', cid=cid))

    # Keep contracted nights always; keep shoulder nights only if they have actual pickup
    block_keys = set(json.loads(config['contracted_block'] or '{}').keys())
    pickup_by_night = {d: n for d, n in pickup_by_night.items()
                       if n > 0 or d in block_keys}

    import flask
    flask.session['hhr_import_pickup'] = json.dumps(pickup_by_night)
    flask.session['hhr_import_cid']    = cid
    if parsed.get('hotel_approver'):
        flask.session['hhr_import_approver']       = parsed['hotel_approver']
        flask.session['hhr_import_approver_email'] = parsed.get('hotel_approver_email', '')

    total = parsed.get('final_total_pickup') or sum(pickup_by_night.values())
    flash(
        f'HHR parsed — {total} total rooms across {len(pickup_by_night)} nights. '
        f'Review and save below.',
        'info'
    )
    return redirect(url_for('pickup_final_history', cid=cid))


@app.route('/pickup/<int:cid>/config/edit', methods=['GET', 'POST'])
def pickup_edit_event(cid):
    db = get_db()
    config = db.execute("SELECT * FROM pickup_config WHERE id=?", (cid,)).fetchone()
    if not config:
        flash('Event not found.', 'error')
        return redirect(url_for('pickup_dashboard'))
    if request.method == 'POST':
        f = request.form
        contracted_block = {}
        for d, r in zip(f.getlist('block_date'), f.getlist('block_rooms')):
            if d and r:
                contracted_block[d] = int(r)
        attrition_raw = f.get('attrition_pct', '')
        attrition = float(attrition_raw) / 100 if attrition_raw else None
        ota_url = f.get('ota_url', '').strip() or None
        hc_name   = f.get('hotel_contact', '').strip() or None
        hc_email  = f.get('hotel_contact_email', '').strip() or None
        hc2_name  = f.get('hotel_contact2', '').strip() or None
        hc2_email = f.get('hotel_contact2_email', '').strip() or None
        gc_name  = f.get('group_contact', '').strip() or None
        gc_email = f.get('group_contact_email', '').strip() or None
        hotel_contacts = [{'name': hc_name or '', 'email': hc_email or ''}] if (hc_name or hc_email) else []
        # Additional CC recipients
        cc_emails = []
        for n, e in zip(f.getlist('cc_name[]'), f.getlist('cc_email[]')):
            if n.strip() or e.strip():
                cc_emails.append({'name': n.strip(), 'email': e.strip()})
        rooming_list_required = 1 if f.get('rooming_list_required') else 0
        rebate_raw = f.get('rebate_per_room', '').strip()
        rebate = float(rebate_raw) if rebate_raw else None
        _changes = [
            ('organization',          config['organization'],          f['organization']),
            ('event_name',            config['event_name'],            f.get('event_name')),
            ('hotel',                 config['hotel'],                 f.get('hotel')),
            ('cutoff_date',           config['cutoff_date'],           f.get('cutoff_date')),
            ('contracted_rate',       config['contracted_rate'],       f.get('contracted_rate')),
            ('rebate_per_room',       config['rebate_per_room'],       rebate),
            ('attrition_pct',         config['attrition_pct'],        attrition),
            ('contracted_block',      config['contracted_block'],      json.dumps(contracted_block)),
            ('hotel_contact',         config['hotel_contact'],        hc_name),
            ('hotel_contact_email',   config['hotel_contact_email'],  hc_email),
            ('hotel_contact2',        config['hotel_contact2'],       hc2_name),
            ('hotel_contact2_email',  config['hotel_contact2_email'], hc2_email),
            ('group_contact',         config['group_contact'],        gc_name),
            ('group_contact_email',   config['group_contact_email'],  gc_email),
            ('notes',                 config['notes'],                f.get('notes')),
            ('rooming_list_required', config['rooming_list_required'], rooming_list_required),
            ('block_review_date',     config['block_review_date'],    f.get('block_review_date') or None),
        ]
        db.execute('''
            UPDATE pickup_config SET
            booking_id=?, tab_name=?, organization=?, event_name=?, hotel=?,
            hotel_contact=?, hotel_contact_email=?, hotel_contact2=?, hotel_contact2_email=?,
            hotel_contacts=?, group_contact=?, group_contact_email=?, cutoff_date=?, attrition_pct=?,
            contracted_block=?, contracted_rate=?, rebate_per_room=?, shoulder_pre=?,
            shoulder_post=?, hotel_booking_link=?, notes=?, ota_url=?, cc_emails=?,
            event_start=?, event_end=?, rooming_list_required=?, block_review_date=?
            WHERE id=?
        ''', (
            f.get('booking_id'), f.get('tab_name'), f['organization'],
            f.get('event_name'), f.get('hotel'), hc_name,
            hc_email, hc2_name, hc2_email, json.dumps(hotel_contacts), gc_name,
            gc_email, f.get('cutoff_date'),
            attrition, json.dumps(contracted_block),
            float(f['contracted_rate']) if f.get('contracted_rate') else None,
            rebate,
            int(f.get('shoulder_pre', 3)), int(f.get('shoulder_post', 3)),
            f.get('hotel_booking_link'), f.get('notes'), ota_url, json.dumps(cc_emails),
            f.get('event_start') or None, f.get('event_end') or None,
            rooming_list_required, f.get('block_review_date') or None, cid
        ))
        _log_change(db, cid, 'edit_event', _changes)
        db.commit()
        _upsert_contacts(db, hotel_contacts, f.get('hotel', ''), cc_emails, f['organization'])
        db.commit()
        flash('Event updated.', 'success')
        return redirect(url_for('pickup_event', cid=cid))
    pipeline_rate = pipeline_org = pipeline_event = None
    pipeline_start = pipeline_end = None
    if config['booking_id']:
        row = db.execute(
            'SELECT RoomRate, AccountName, EventName, StartDate, EndDate FROM ReportPipeline WHERE CAST(BookingId AS INTEGER) = CAST(? AS INTEGER) LIMIT 1',
            (config['booking_id'],)
        ).fetchone()
        if row:
            if row['RoomRate']:
                pipeline_rate = round(float(row['RoomRate']), 2)
            pipeline_org   = row['AccountName'] or None
            pipeline_event = row['EventName'] or None
            # Convert ISO timestamp to YYYY-MM-DD for the date input
            if row['StartDate']:
                pipeline_start = str(row['StartDate'])[:10]
            if row['EndDate']:
                pipeline_end = str(row['EndDate'])[:10]
    return render_template('pickup_config_form.html', config=config,
                           action=url_for('pickup_edit_event', cid=cid),
                           cancel_url=url_for('pickup_event', cid=cid),
                           pipeline_rate=pipeline_rate,
                           pipeline_org=pipeline_org,
                           pipeline_event=pipeline_event,
                           pipeline_booking_id=config['booking_id'],
                           pipeline_start=pipeline_start,
                           pipeline_end=pipeline_end)


@app.route('/pickup/<int:cid>/extract-pickup-pdf', methods=['POST'])
def pickup_extract_pickup_pdf(cid):
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'No file uploaded', 'pairs': [], 'text': ''})
    filename = (f.filename or '').lower()
    data = f.read()
    if filename.endswith('.xlsx') or filename.endswith('.xls'):
        from pickup_utils import parse_columnar_pickup_xlsx
        result = parse_columnar_pickup_xlsx(data, filename=filename)
    else:
        from pickup_utils import parse_columnar_pickup_pdf
        result = parse_columnar_pickup_pdf(data, filename=filename)
    return jsonify(result)


@app.route('/pickup/<int:cid>/weekly/new', methods=['GET', 'POST'])
def pickup_weekly_new(cid):
    db = get_db()
    config = db.execute("SELECT * FROM pickup_config WHERE id=?", (cid,)).fetchone()
    if not config:
        flash('Event not found.', 'error')
        return redirect(url_for('pickup_dashboard'))
    last = db.execute(
        "SELECT * FROM pickup_weekly WHERE config_id=? ORDER BY report_date DESC LIMIT 1", (cid,)
    ).fetchone()
    if request.method == 'POST':
        f = request.form
        block = json.loads(config['contracted_block'] or '{}')
        pickup_by_night = {}
        # Collect contracted block nights AND any shoulder nights added via the form
        for key in f:
            if key.startswith('night_'):
                val = f.get(key, '').strip()
                if val != '':
                    d = key[6:]  # strip 'night_'
                    try:
                        pickup_by_night[d] = int(val)
                    except ValueError:
                        pass
        total_rooms = sum(pickup_by_night.values())
        contracted_total = sum(block.values())
        change_from_last = (total_rooms - (last['total_rooms'] or 0)) if last else None
        pct_of_block     = round(total_rooms / contracted_total * 100, 1) if contracted_total else None
        attrition_floor  = contracted_total * (config['attrition_pct'] or 0)
        pct_of_attrition = round(total_rooms / attrition_floor * 100, 1) if attrition_floor else None
        ota_rate = float(f['ota_rate']) if f.get('ota_rate') else None
        weekly_ota_url = f.get('weekly_ota_url', '').strip() or config['ota_url'] or None
        db.execute('''
            INSERT INTO pickup_weekly
            (config_id, report_date, pickup_by_night, total_rooms, change_from_last,
             pct_of_block, pct_of_attrition, ota_rate, ota_url, label, notes)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)
        ''', (cid, f['report_date'], json.dumps(pickup_by_night), total_rooms,
              change_from_last, pct_of_block, pct_of_attrition, ota_rate, weekly_ota_url,
              f.get('label'), f.get('notes')))
        _log_change(db, cid, 'add_report', [
            ('report_date', None, f['report_date']),
            ('total_rooms', last['total_rooms'] if last else None, total_rooms),
            ('pct_of_block', last['pct_of_block'] if last else None, pct_of_block),
        ])
        db.commit()
        flash('Weekly entry saved.', 'success')
        return redirect(url_for('pickup_event', cid=cid))
    block = json.loads(config['contracted_block'] or '{}')
    last_pickup = json.loads(last['pickup_by_night']) if last else {}
    today_str = datetime.today().strftime('%Y-%m-%d')
    # Shoulder nights from last entry that aren't in the contracted block
    extra_dates = sorted(set(last_pickup.keys()) - set(block.keys()))
    return render_template('pickup_weekly_form.html', config=config, block=block,
                           last=last, last_pickup=last_pickup, today=today_str,
                           is_edit=False, entry=None, entry_pickup={}, wid=None,
                           extra_dates=extra_dates)


@app.route('/pickup/<int:cid>/weekly/<int:wid>/edit', methods=['GET', 'POST'])
def pickup_weekly_edit(cid, wid):
    db = get_db()
    config = db.execute("SELECT * FROM pickup_config WHERE id=?", (cid,)).fetchone()
    entry  = db.execute("SELECT * FROM pickup_weekly WHERE id=? AND config_id=?", (wid, cid)).fetchone()
    if not config or not entry:
        flash('Entry not found.', 'error')
        return redirect(url_for('pickup_event', cid=cid))
    if request.method == 'POST':
        f = request.form
        block = json.loads(config['contracted_block'] or '{}')
        pickup_by_night = {}
        # Collect contracted block nights AND any shoulder nights added via the form
        for key in f:
            if key.startswith('night_'):
                val = f.get(key, '').strip()
                if val != '':
                    d = key[6:]  # strip 'night_'
                    try:
                        pickup_by_night[d] = int(val)
                    except ValueError:
                        pass
        total_rooms = sum(pickup_by_night.values())
        contracted_total = sum(block.values())
        pct_of_block     = round(total_rooms / contracted_total * 100, 1) if contracted_total else None
        attrition_floor  = contracted_total * (config['attrition_pct'] or 0)
        pct_of_attrition = round(total_rooms / attrition_floor * 100, 1) if attrition_floor else None
        ota_rate = float(f['ota_rate']) if f.get('ota_rate') else None
        weekly_ota_url = f.get('weekly_ota_url', '').strip() or config['ota_url'] or None
        db.execute('''
            UPDATE pickup_weekly
            SET report_date=?, pickup_by_night=?, total_rooms=?,
                pct_of_block=?, pct_of_attrition=?, ota_rate=?, ota_url=?, label=?, notes=?
            WHERE id=? AND config_id=?
        ''', (f['report_date'], json.dumps(pickup_by_night), total_rooms,
              pct_of_block, pct_of_attrition, ota_rate, weekly_ota_url,
              f.get('label') or None, f.get('notes') or None, wid, cid))
        # Recompute change_from_last for all entries
        all_w = db.execute(
            "SELECT id, total_rooms, label FROM pickup_weekly WHERE config_id=? ORDER BY report_date ASC, id ASC", (cid,)
        ).fetchall()
        prev_total = None
        for w_row in all_w:
            is_label = bool(w_row['label'] and w_row['label'].strip())
            if is_label or w_row['total_rooms'] is None:
                db.execute("UPDATE pickup_weekly SET change_from_last=NULL WHERE id=?", (w_row['id'],))
            else:
                change = (w_row['total_rooms'] - prev_total) if prev_total is not None else None
                db.execute("UPDATE pickup_weekly SET change_from_last=? WHERE id=?", (change, w_row['id']))
                prev_total = w_row['total_rooms']
        db.commit()
        flash('Weekly entry updated.', 'success')
        return redirect(url_for('pickup_event', cid=cid))
    block = json.loads(config['contracted_block'] or '{}')
    entry_pickup = json.loads(entry['pickup_by_night']) if entry['pickup_by_night'] else {}
    # Shoulder nights saved on this entry that aren't in the contracted block
    extra_dates = sorted(set(entry_pickup.keys()) - set(block.keys()))
    prev_entry = db.execute(
        "SELECT * FROM pickup_weekly WHERE config_id=? AND report_date < ? ORDER BY report_date DESC LIMIT 1",
        (cid, entry['report_date'])
    ).fetchone()
    last_pickup = json.loads(prev_entry['pickup_by_night']) if prev_entry and prev_entry['pickup_by_night'] else {}
    return render_template('pickup_weekly_form.html',
                           config=config, block=block,
                           last=prev_entry, last_pickup=last_pickup,
                           entry_pickup=entry_pickup, entry=entry,
                           today=entry['report_date'], is_edit=True, wid=wid,
                           extra_dates=extra_dates)


@app.route('/pickup/<int:cid>/weekly/<int:wid>/delete', methods=['POST'])
def pickup_weekly_delete(cid, wid):
    db = get_db()
    row = db.execute("SELECT * FROM pickup_weekly WHERE id=? AND config_id=?", (wid, cid)).fetchone()
    if row:
        username = session.get('username') or 'unknown'
        db.execute('''
            INSERT INTO pickup_weekly_deleted
            (original_id, config_id, report_date, pickup_by_night, total_rooms,
             change_from_last, pct_of_block, pct_of_attrition, ota_rate, label, notes, deleted_by)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
        ''', (row['id'], row['config_id'], row['report_date'], row['pickup_by_night'],
              row['total_rooms'], row['change_from_last'], row['pct_of_block'],
              row['pct_of_attrition'], row['ota_rate'], row['label'], row['notes'], username))
        db.execute("DELETE FROM pickup_weekly WHERE id=? AND config_id=?", (wid, cid))
        _log_change(db, cid, 'delete_report', [
            ('report_date', row['report_date'], None),
            ('total_rooms', row['total_rooms'], None),
        ])
    db.commit()
    flash('Weekly entry deleted. You can restore it within 90 days from the pickup event page.', 'success')
    return redirect(url_for('pickup_event', cid=cid))


@app.route('/pickup/<int:cid>/weekly/deleted/<int:did>/restore', methods=['POST'])
def pickup_weekly_restore(cid, did):
    db = get_db()
    row = db.execute(
        "SELECT * FROM pickup_weekly_deleted WHERE id=? AND config_id=?", (did, cid)
    ).fetchone()
    if row:
        db.execute('''
            INSERT INTO pickup_weekly
            (config_id, report_date, pickup_by_night, total_rooms, change_from_last,
             pct_of_block, pct_of_attrition, ota_rate, label, notes)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        ''', (row['config_id'], row['report_date'], row['pickup_by_night'], row['total_rooms'],
              row['change_from_last'], row['pct_of_block'], row['pct_of_attrition'],
              row['ota_rate'], row['label'], row['notes']))
        db.execute("DELETE FROM pickup_weekly_deleted WHERE id=?", (did,))
        _log_change(db, cid, 'restore_report', [
            ('report_date', None, row['report_date']),
            ('total_rooms', None, row['total_rooms']),
        ])
        db.commit()
        flash(f'Report for {row["report_date"]} restored.', 'success')
    else:
        flash('Deleted entry not found.', 'error')
    return redirect(url_for('pickup_event', cid=cid))


@app.route('/pickup/<int:cid>/rooming-list', methods=['GET', 'POST'])
def pickup_rooming_upload(cid):
    from pickup_utils import parse_rooming_list_pdf, parse_rooming_list_csv, parse_rooming_list_xls, reconcile, _build_guest_result as _bgr
    db = get_db()
    config = db.execute("SELECT * FROM pickup_config WHERE id=?", (cid,)).fetchone()
    if not config:
        flash('Event not found.', 'error')
        return redirect(url_for('pickup_dashboard'))
    if request.method == 'POST':
        uploaded_files = [f for f in request.files.getlist('rooming_pdf') if f and f.filename]
        labels = request.form.getlist('list_label')
        if not uploaded_files:
            flash('No file uploaded.', 'error')
            return redirect(request.url)
        all_guests, combined_nights, filenames, errors = [], {}, [], []
        generated_pdfs = []   # (pdf_filename, pdf_bytes) when parser auto-generates PDF
        any_ai_parsed = False
        primary_file_bytes = None   # bytes of first successfully parsed file, for storage
        for i, f in enumerate(uploaded_files):
            label = labels[i].strip() if i < len(labels) and labels[i].strip() else f.filename
            file_bytes = f.read()
            ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else ''
            if ext == 'pdf':
                parsed = parse_rooming_list_pdf(file_bytes)
            elif ext == 'csv':
                parsed = parse_rooming_list_csv(file_bytes)
            elif ext in ('xls', 'xlsx'):
                parsed = parse_rooming_list_xls(file_bytes, filename=f.filename)
            else:
                errors.append(f"{f.filename}: unsupported file type (.{ext}). Use PDF, CSV, XLS, or XLSX.")
                continue
            if parsed.get('error'):
                errors.append(f"{f.filename}: {parsed['error']}")
                continue
            if parsed.get('ai_parsed'):
                any_ai_parsed = True
            if parsed.get('ai_error'):
                errors.append(f"AI parsing failed for {f.filename}: {parsed['ai_error']}")
            filenames.append(f.filename)
            if primary_file_bytes is None:
                primary_file_bytes = file_bytes
            # If the parser generated a clean PDF (e.g. Omni XLSX → PDF), capture it
            if parsed.get('pdf_bytes') and not generated_pdfs:
                pdf_name = f.filename.rsplit('.', 1)[0] + '_rooming_list.pdf'
                generated_pdfs.append((pdf_name, parsed['pdf_bytes']))
                primary_file_bytes = parsed['pdf_bytes']
            for g in parsed['guests']:
                g['source'] = label
            all_guests.extend(parsed['guests'])
            for date_key, rooms in parsed['nights_by_date'].items():
                combined_nights[date_key] = combined_nights.get(date_key, 0) + rooms
        if errors:
            flash('Parse errors: ' + '; '.join(errors), 'error')
        if not all_guests:
            flash('No guest records could be read. The format may not be recognized.', 'error')
            return redirect(request.url)
        last = db.execute(
            "SELECT * FROM pickup_weekly WHERE config_id=? ORDER BY report_date DESC LIMIT 1", (cid,)
        ).fetchone()
        pickup_by_night = json.loads(last['pickup_by_night']) if last else {}
        recon = reconcile(pickup_by_night, combined_nights)
        combined_result = _bgr(all_guests)
        unique_rooms  = combined_result['unique_rooms']
        combined_nights = combined_result['nights_by_date']
        # Use the generated PDF filename if available
        if generated_pdfs:
            combined_filename = generated_pdfs[0][0]
        else:
            combined_filename = ', '.join(filenames)
        rl_id = db.execute('''
            INSERT INTO pickup_rooming_list
            (config_id, weekly_id, filename, file_data, total_guests,
             nights_by_date, reconciliation_status, discrepancy_notes, guests_json)
            VALUES (?,?,?,?,?,?,?,?,?)
        ''', (cid, last['id'] if last else None, combined_filename,
              primary_file_bytes, unique_rooms, json.dumps(combined_nights),
              recon['status'], recon.get('notes', ''),
              json.dumps(all_guests))).lastrowid
        db.commit()
        return redirect(url_for('pickup_rooming_review', cid=cid, rl_id=rl_id,
                                ai_parsed=1 if any_ai_parsed else 0))
    return render_template('pickup_rooming_upload.html', config=config)


@app.route('/pickup/<int:cid>/rooming-list/manual', methods=['POST'])
def pickup_rooming_manual(cid):
    from pickup_utils import reconcile
    db = get_db()
    config = db.execute("SELECT * FROM pickup_config WHERE id=?", (cid,)).fetchone()
    if not config:
        flash('Event not found.', 'error')
        return redirect(url_for('pickup_dashboard'))
    dates  = request.form.getlist('manual_date')
    rooms  = request.form.getlist('manual_rooms')
    label  = request.form.get('manual_label', '').strip() or 'Manual entry'
    nights_by_date = {}
    for d, r in zip(dates, rooms):
        if d and r:
            try:
                nights_by_date[d] = int(r)
            except ValueError:
                pass
    if not nights_by_date:
        flash('No room counts entered.', 'error')
        return redirect(url_for('pickup_rooming_upload', cid=cid))
    total_rooms = sum(nights_by_date.values())
    last = db.execute(
        "SELECT * FROM pickup_weekly WHERE config_id=? ORDER BY report_date DESC LIMIT 1", (cid,)
    ).fetchone()
    pickup_by_night = json.loads(last['pickup_by_night']) if last else {}
    recon = reconcile(pickup_by_night, nights_by_date)
    rl_id = db.execute('''
        INSERT INTO pickup_rooming_list
        (config_id, weekly_id, filename, total_guests,
         nights_by_date, reconciliation_status, discrepancy_notes, guests_json)
        VALUES (?,?,?,?,?,?,?,?)
    ''', (cid, last['id'] if last else None, label,
          total_rooms, json.dumps(nights_by_date),
          recon['status'], recon.get('notes', ''), json.dumps([]))).lastrowid
    db.commit()
    return redirect(url_for('pickup_rooming_review', cid=cid, rl_id=rl_id))


@app.route('/pickup/<int:cid>/rooming-list/<int:rl_id>/review')
def pickup_rooming_review(cid, rl_id):
    from pickup_utils import _build_guest_result as _bgr
    from datetime import datetime as _dt, timedelta as _td
    db = get_db()
    config = db.execute("SELECT * FROM pickup_config WHERE id=?", (cid,)).fetchone()
    rl     = db.execute("SELECT * FROM pickup_rooming_list WHERE id=? AND config_id=?", (rl_id, cid)).fetchone()
    last   = db.execute("SELECT * FROM pickup_weekly WHERE config_id=? ORDER BY report_date DESC LIMIT 1", (cid,)).fetchone()
    if not rl:
        flash('Rooming list not found.', 'error')
        return redirect(url_for('pickup_event', cid=cid))
    guests = json.loads(rl['guests_json'] or '[]')
    nights_by_date  = json.loads(rl['nights_by_date'] or '{}')
    pickup_by_night = json.loads(last['pickup_by_night']) if last else {}
    block = json.loads(config['contracted_block'] or '{}')
    dedup_result  = _bgr(guests)
    dedup_summary = dedup_result['dedup_summary']
    unique_rooms  = dedup_result['unique_rooms']
    sources_seen, guests_by_source, nights_by_source = [], {}, {}
    for g in guests:
        src = g.get('source', 'Unknown')
        if src not in sources_seen:
            sources_seen.append(src)
            guests_by_source[src] = []
            nights_by_source[src] = {}
        guests_by_source[src].append(g)
        try:
            arr = _dt.strptime(g['arrival'],   '%Y-%m-%d')
            dep = _dt.strptime(g['departure'], '%Y-%m-%d')
            cur = arr
            while cur < dep:
                key = cur.strftime('%Y-%m-%d')
                nights_by_source[src][key] = nights_by_source[src].get(key, 0) + g.get('rooms', 1)
                cur += _td(days=1)
        except Exception:
            pass
    ai_parsed = bool(int(request.args.get('ai_parsed', 0)))
    return render_template('pickup_rooming_review.html', config=config, rl=rl,
                           guests=guests, nights_by_date=nights_by_date,
                           pickup_by_night=pickup_by_night, block=block,
                           sources=sources_seen, guests_by_source=guests_by_source,
                           nights_by_source=nights_by_source,
                           dedup_summary=dedup_summary, unique_rooms=unique_rooms,
                           ai_parsed=ai_parsed)


@app.route('/pickup/<int:cid>/rooming-list/<int:rl_id>/confirm', methods=['POST'])
def pickup_rooming_confirm(cid, rl_id):
    db = get_db()
    db.execute("UPDATE pickup_rooming_list SET reconciliation_status=? WHERE id=?",
               (request.form.get('status', 'match'), rl_id))
    db.commit()
    flash('Rooming list confirmed.', 'success')
    return redirect(url_for('pickup_event', cid=cid))


@app.route('/pickup/<int:cid>/rooming-list/<int:rl_id>/download-csv')
def pickup_rooming_download_csv(cid, rl_id):
    import csv
    from flask import Response
    db = get_db()
    rl = db.execute("SELECT filename, guests_json FROM pickup_rooming_list WHERE id=? AND config_id=?",
                    (rl_id, cid)).fetchone()
    if not rl:
        flash('Rooming list not found.', 'error')
        return redirect(url_for('pickup_event', cid=cid))
    guests = json.loads(rl['guests_json'] or '[]')
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Name', 'Confirmation #', 'Arrival', 'Departure', 'Nights', 'Rooms', 'List'])
    for g in guests:
        writer.writerow([g.get('name',''), g.get('conf_no',''), g.get('arrival',''),
                         g.get('departure',''), g.get('nights',''), g.get('rooms',1), g.get('source','')])
    csv_bytes = output.getvalue().encode('utf-8-sig')
    base = (rl['filename'] or 'rooming_list').rsplit('.', 1)[0]
    return Response(csv_bytes, mimetype='text/csv',
                    headers={'Content-Disposition': f'attachment; filename="{base}_parsed.csv"'})


def _open_in_outlook(cid, subject, to_addr, cc_list, body_html, email_type,
                     attachments=None):
    """Open an Outlook compose window with subject, recipients, HTML body, and optional attachments.
    Returns (True, None) on success or (False, error_str) on failure.
    Only works on macOS (osascript).
    """
    import subprocess, tempfile, os
    from datetime import date as _date

    def esc(s):
        return str(s or '').replace('\\', '\\\\').replace('"', '\\"')

    tmp = tempfile.NamedTemporaryFile(mode='w', suffix='.html', delete=False, encoding='utf-8')
    tmp.write(body_html)
    tmp.close()

    to_line = (
        f'make new to recipient at theMsg with properties '
        f'{{email address:{{address:"{esc(to_addr)}"}}}}'
        if to_addr else ''
    )
    cc_lines = '\n    '.join(
        f'make new cc recipient at theMsg with properties {{email address:{{address:"{esc(a)}"}}}}'
        for a in cc_list if a
    )
    attach_lines = '\n    '.join(
        f'make new attachment at theMsg with properties {{file:(POSIX file "{esc(p)}") as alias}}'
        for p in (attachments or []) if p and os.path.exists(p)
    )

    script = f'''
tell application "Microsoft Outlook"
    set htmlContent to (read (POSIX file "{tmp.name}") as «class utf8»)
    set theMsg to make new outgoing message with properties {{subject:"{esc(subject)}", content:htmlContent}}
    {to_line}
    {cc_lines}
    {attach_lines}
    open theMsg
    activate
end tell
'''
    try:
        subprocess.run(['/usr/bin/osascript', '-e', script], timeout=20)
        attach_note = f' + {len(attachments)} attachment(s)' if attachments else ''
        db = get_db()
        db.execute(
            "INSERT INTO pickup_contact_log (config_id, contact_date, contact_type, notes) VALUES (?,?,?,?)",
            (cid, _date.today().isoformat(), 'email_sent',
             f'Opened in Outlook — {email_type} email{attach_note} — To: {to_addr}')
        )
        db.commit()
        return True, None
    except Exception as e:
        return False, str(e)
    finally:
        try:
            os.unlink(tmp.name)
        except Exception:
            pass


def _create_outlook_draft(user_id, subject, to_addr, cc_list, body_html,
                           attachment_bytes=None, attachment_filename=None):
    """Create a draft in the user's Outlook Drafts folder via Microsoft Graph API.
    Returns (draft_id, None) on success or (None, error_str) on failure.
    """
    import time, base64, requests as _req
    from config import MS_CLIENT_ID, MS_TENANT_ID, MS_CLIENT_SECRET

    db = get_db()
    row = db.execute(
        'SELECT access_token, refresh_token, expires_at FROM UserMicrosoftTokens WHERE user_id = ?',
        (user_id,)
    ).fetchone()
    if not row:
        return None, 'No Microsoft account connected. Visit My Account to connect.'

    access_token  = row['access_token']
    refresh_token = row['refresh_token']
    expires_at    = row['expires_at']

    # Refresh token if expired
    if time.time() >= expires_at:
        r = _req.post(
            f'https://login.microsoftonline.com/{MS_TENANT_ID}/oauth2/v2.0/token',
            data={
                'grant_type':    'refresh_token',
                'client_id':     MS_CLIENT_ID,
                'client_secret': MS_CLIENT_SECRET,
                'refresh_token': refresh_token,
                'scope':         'Mail.ReadWrite User.Read offline_access',
            }
        )
        if r.status_code != 200:
            db.execute('DELETE FROM UserMicrosoftTokens WHERE user_id = ?', (user_id,))
            db.commit()
            return None, 'Microsoft token expired. Please reconnect via My Account.'
        tok           = r.json()
        access_token  = tok['access_token']
        refresh_token = tok.get('refresh_token', refresh_token)
        expires_at    = time.time() + tok.get('expires_in', 3600) - 60
        db.execute(
            'UPDATE UserMicrosoftTokens SET access_token=?, refresh_token=?, expires_at=? WHERE user_id=?',
            (access_token, refresh_token, expires_at, user_id)
        )
        db.commit()

    headers = {'Authorization': f'Bearer {access_token}', 'Content-Type': 'application/json'}

    message = {
        'subject': subject,
        'body': {'contentType': 'HTML', 'content': body_html},
        'toRecipients':  [{'emailAddress': {'address': to_addr}}] if to_addr else [],
        'ccRecipients':  [{'emailAddress': {'address': a}} for a in (cc_list or []) if a],
    }

    resp = _req.post('https://graph.microsoft.com/v1.0/me/messages', headers=headers, json=message)
    if resp.status_code not in (200, 201):
        return None, f'Graph API error {resp.status_code}: {resp.text[:300]}'

    draft_id = resp.json().get('id')

    if attachment_bytes and attachment_filename and draft_id:
        attach_payload = {
            '@odata.type':  '#microsoft.graph.fileAttachment',
            'name':         attachment_filename,
            'contentBytes': base64.b64encode(attachment_bytes).decode('ascii'),
            'contentType':  'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        }
        _req.post(
            f'https://graph.microsoft.com/v1.0/me/messages/{draft_id}/attachments',
            headers=headers, json=attach_payload
        )

    return draft_id, None


@app.route('/pickup/<int:cid>/email/housing')
def pickup_email_housing(cid):
    """Generate the Housing History Form, save to temp file, and open in Outlook with attachment."""
    import platform
    from datetime import datetime as dt

    db = get_db()
    config = db.execute("SELECT * FROM pickup_config WHERE id=?", (cid,)).fetchone()
    if not config:
        flash('Event not found.', 'error')
        return redirect(url_for('pickup_event', cid=cid))

    pipeline = None
    if config['booking_id']:
        pipeline = db.execute(
            'SELECT * FROM ReportPipeline WHERE BookingId = ?', (config['booking_id'],)
        ).fetchone()
    latest_weekly = db.execute(
        "SELECT pickup_by_night FROM pickup_weekly WHERE config_id=? ORDER BY report_date DESC LIMIT 1",
        (cid,)
    ).fetchone()
    pickup_dates  = list(json.loads(latest_weekly['pickup_by_night'] or '{}').keys()) if latest_weekly else None

    sorted_dates  = sorted(json.loads(config['contracted_block'] or '{}').keys())
    org_name      = (pipeline['AccountName'] or '') if pipeline else (config['organization'] or '')
    event_name    = config['event_name'] or config['organization'] or ''
    hotel_email   = config['hotel_contact_email'] or ''
    hotel_contact = config['hotel_contact'] or ''

    if sorted_dates:
        start_str  = dt.strptime(sorted_dates[0],  '%Y-%m-%d').strftime('%m/%d/%Y')
        end_str    = dt.strptime(sorted_dates[-1], '%Y-%m-%d').strftime('%m/%d/%Y')
        date_range = f"{start_str} – {end_str}"
    else:
        date_range = ''

    first_name = hotel_contact.split(',')[-1].strip().split()[0] if hotel_contact else 'Team'
    subject    = f"Final Housing History Form for {event_name}, {org_name}, {date_range}"
    body_text  = (
        f"Dear {first_name},\n\n"
        f"Please see the attached Housing History form for {event_name}. "
        f"Please fill in all the relevant areas highlighted in yellow. "
        f"Please include the relevant pre and post days in the count. "
        f"Attach the final rooming list to your email response to me. "
        f"If you have any questions please reach out."
    )
    body_html  = body_text.replace('\n\n', '<br><br>').replace('\n', '<br>')

    user = get_current_user()

    # ── Local Mac: auto-open Outlook with form attached ────────────────────
    if platform.system() == 'Darwin':
        wb, _ = _build_housing_form_wb(config, pipeline, pickup_dates=pickup_dates)
        safe_name   = event_name.replace('/', '-').replace(' ', '_')[:50]
        attach_path = f'/tmp/Housing_History_{safe_name}.xlsx'
        wb.save(attach_path)
        ok, err = _open_in_outlook(cid, subject, hotel_email, [],
                                   body_html, 'housing', attachments=[attach_path])
        if ok:
            flash('Housing History email opened in Outlook with form attached — review and send.', 'success')
        else:
            flash(f'Could not open Outlook: {err}', 'error')
        return redirect(url_for('pickup_event', cid=cid))

    # ── Web: try Graph API if user has connected Microsoft account ─────────
    ms_row = db.execute(
        'SELECT user_id FROM UserMicrosoftTokens WHERE user_id = ?', (user['id'],)
    ).fetchone() if user else None

    if ms_row:
        from datetime import date as _date
        wb, _ = _build_housing_form_wb(config, pipeline, pickup_dates=pickup_dates)
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        safe_name = event_name.replace('/', '-').replace(' ', '_')[:50]
        draft_id, err = _create_outlook_draft(
            user_id=user['id'], subject=subject, to_addr=hotel_email, cc_list=[],
            body_html=body_html, attachment_bytes=buf.getvalue(),
            attachment_filename=f'Housing_History_{safe_name}.xlsx',
        )
        if draft_id:
            db.execute(
                "INSERT INTO pickup_contact_log (config_id, contact_date, contact_type, notes) VALUES (?,?,?,?)",
                (cid, _date.today().isoformat(), 'email_sent',
                 f'Outlook draft created via Graph API — To: {hotel_email} (with attachment)')
            )
            db.commit()
            flash('Draft created in your Outlook Drafts folder with the form attached. Open Outlook to review and send.', 'success')
            return redirect(url_for('pickup_event', cid=cid))
        else:
            flash(f'Could not create Outlook draft: {err}', 'error')

    # ── Fallback: preview page ─────────────────────────────────────────────
    email = {'to': hotel_email, 'cc': '', 'subject': subject, 'body': body_text}
    return render_template('pickup_email_preview.html', config=config, email=email,
                           email_type='hotel', show_housing_download=True,
                           ms_connected=bool(ms_row))


@app.route('/pickup/<int:cid>/email/hotel')
def pickup_email_hotel(cid):
    """Open a hotel pickup-status email directly in Outlook."""
    import platform
    from pickup_utils import build_hotel_email

    db = get_db()
    config = db.execute("SELECT * FROM pickup_config WHERE id=?", (cid,)).fetchone()
    if not config:
        flash('Event not found.', 'error')
        return redirect(url_for('pickup_dashboard'))

    email       = build_hotel_email(config)
    hotel_email = config['hotel_contact_email'] or ''
    cc_list     = [a.strip() for a in (email.get('cc') or '').replace(';', ',').split(',') if a.strip()]
    body_html   = email.get('body', '').replace('\n', '<br>')
    user        = get_current_user()

    # ── Local Mac: auto-open Outlook ───────────────────────────────────────
    if platform.system() == 'Darwin':
        ok, err = _open_in_outlook(cid, email.get('subject', ''), hotel_email,
                                   cc_list, body_html, 'hotel')
        if ok:
            flash('Hotel email opened in Outlook — review and send when ready.', 'success')
        else:
            flash(f'Could not open Outlook: {err}', 'error')
        return redirect(url_for('pickup_event', cid=cid))

    # ── Web: try Graph API if user has connected Microsoft account ─────────
    ms_row = db.execute(
        'SELECT user_id FROM UserMicrosoftTokens WHERE user_id = ?', (user['id'],)
    ).fetchone() if user else None

    if ms_row:
        from datetime import date as _date
        draft_id, err = _create_outlook_draft(
            user_id=user['id'], subject=email.get('subject', ''),
            to_addr=hotel_email, cc_list=cc_list, body_html=body_html,
        )
        if draft_id:
            db.execute(
                "INSERT INTO pickup_contact_log (config_id, contact_date, contact_type, notes) VALUES (?,?,?,?)",
                (cid, _date.today().isoformat(), 'email_sent',
                 f'Outlook draft created via Graph API — To: {hotel_email}')
            )
            db.commit()
            flash('Draft created in your Outlook Drafts folder. Open Outlook to review and send.', 'success')
            return redirect(url_for('pickup_event', cid=cid))
        else:
            flash(f'Could not create Outlook draft: {err}', 'error')

    # ── Fallback: preview page ─────────────────────────────────────────────
    return render_template('pickup_email_preview.html', config=config, email=email,
                           email_type='hotel', ms_connected=bool(ms_row))


@app.route('/pickup/<int:cid>/email/hotel/rate-issue')
def pickup_email_hotel_rate_issue(cid):
    """Open a hotel rate-parity issue email in Outlook."""
    import platform
    from pickup_utils import build_hotel_rate_issue_email

    db = get_db()
    config = db.execute("SELECT * FROM pickup_config WHERE id=?", (cid,)).fetchone()
    if not config:
        flash('Event not found.', 'error')
        return redirect(url_for('pickup_dashboard'))

    # Get the most recent real weekly entry for OTA data
    last_real = db.execute(
        """SELECT * FROM pickup_weekly WHERE config_id=?
           AND (total_rooms IS NOT NULL AND total_rooms > 0)
           AND (label IS NULL OR (label NOT LIKE 'ASAS %' AND label NOT LIKE 'Historical%'
                                  AND label NOT LIKE '%Final%'))
           ORDER BY report_date DESC LIMIT 1""",
        (cid,)
    ).fetchone()

    if not last_real or not last_real['ota_rate']:
        flash('No OTA rate found in the latest weekly entry.', 'error')
        return redirect(url_for('pickup_event', cid=cid))

    # Use weekly entry's ota_url first, fall back to config-level ota_url
    ota_url = (last_real['ota_url'] or config['ota_url'] or '').strip() or None
    email = build_hotel_rate_issue_email(config, last_real['ota_rate'], ota_url)
    hotel_email = config['hotel_contact_email'] or ''
    cc_list     = [a.strip() for a in (email.get('cc') or '').replace(';', ',').split(',') if a.strip()]
    body_html   = email.get('body_html') or email.get('body', '').replace('\n', '<br>')
    user        = get_current_user()

    if platform.system() == 'Darwin':
        ok, err = _open_in_outlook(cid, email.get('subject', ''), hotel_email,
                                   cc_list, body_html, 'hotel_rate_issue')
        if ok:
            flash('Rate issue email opened in Outlook — review and send when ready.', 'success')
        else:
            flash(f'Could not open Outlook: {err}', 'error')
        return redirect(url_for('pickup_event', cid=cid))

    ms_row = db.execute(
        'SELECT user_id FROM UserMicrosoftTokens WHERE user_id = ?', (user['id'],)
    ).fetchone() if user else None

    if ms_row:
        from datetime import date as _date
        draft_id, err = _create_outlook_draft(
            user_id=user['id'], subject=email.get('subject', ''),
            to_addr=hotel_email, cc_list=cc_list, body_html=body_html,
        )
        if draft_id:
            flash('Rate issue draft created in your Outlook Drafts folder.', 'success')
            return redirect(url_for('pickup_event', cid=cid))
        else:
            flash(f'Could not create Outlook draft: {err}', 'error')

    return render_template('pickup_email_preview.html', config=config, email=email,
                           email_type='hotel_rate_issue', ms_connected=bool(ms_row))


@app.route('/pickup/<int:cid>/email/client')
def pickup_email_client(cid):
    import platform
    from datetime import date as _date
    from pickup_utils import build_client_email, _build_cc_recipients
    db = get_db()
    config = db.execute("SELECT * FROM pickup_config WHERE id=?", (cid,)).fetchone()
    last = db.execute("SELECT * FROM pickup_weekly WHERE config_id=? ORDER BY report_date DESC LIMIT 1", (cid,)).fetchone()
    weekly_list = db.execute("SELECT * FROM pickup_weekly WHERE config_id=? ORDER BY report_date DESC", (cid,)).fetchall()
    rl = db.execute("SELECT id, reconciliation_status, filename, total_guests, upload_date FROM pickup_rooming_list WHERE config_id=? ORDER BY upload_date DESC LIMIT 1", (cid,)).fetchone()
    if not config:
        flash('Event not found.', 'error')
        return redirect(url_for('pickup_dashboard'))
    rl_status    = rl['reconciliation_status'] if rl else None
    weekly_dicts = [dict(w) for w in weekly_list]
    email        = build_client_email(config, last, rl_status, weekly_list=weekly_dicts)

    user   = get_current_user()
    ms_row = db.execute('SELECT user_id FROM UserMicrosoftTokens WHERE user_id=?',
                        (user['id'],)).fetchone() if user else None

    # ── Web: try Graph API if user has connected Microsoft account ─────────
    if platform.system() != 'Darwin' and ms_row:
        html_body     = email.get('html_body') or email.get('body', '').replace('\n', '<br>')
        to_addr       = (email.get('to') or '').strip()
        cc_recipients = _build_cc_recipients(config)
        cc_list       = [r['email'] for r in cc_recipients if r.get('email')]
        draft_id, err = _create_outlook_draft(
            user_id=user['id'], subject=email.get('subject', ''),
            to_addr=to_addr, cc_list=cc_list, body_html=html_body,
        )
        if draft_id:
            db.execute(
                "INSERT INTO pickup_contact_log (config_id, contact_date, contact_type, notes) VALUES (?,?,?,?)",
                (cid, _date.today().isoformat(), 'email_sent',
                 f'Client draft created via Graph API — To: {to_addr}')
            )
            db.commit()
            flash('Draft created in your Outlook Drafts folder. Open Outlook to review and send.', 'success')
            return redirect(url_for('pickup_event', cid=cid))
        else:
            flash(f'Could not create Outlook draft: {err}', 'error')

    return render_template('pickup_email_preview.html', config=config, email=email,
                           email_type='client', rooming_list=rl,
                           ms_connected=bool(ms_row))


@app.route('/pickup/<int:cid>/email/client/launch-outlook')
def pickup_email_client_launch_outlook(cid):
    import subprocess, tempfile, platform
    from datetime import date as _date
    from pickup_utils import build_client_email, _build_cc_recipients
    db = get_db()
    config = db.execute("SELECT * FROM pickup_config WHERE id=?", (cid,)).fetchone()
    last = db.execute("SELECT * FROM pickup_weekly WHERE config_id=? ORDER BY report_date DESC LIMIT 1", (cid,)).fetchone()
    weekly_list = db.execute("SELECT * FROM pickup_weekly WHERE config_id=? ORDER BY report_date DESC", (cid,)).fetchall()
    rl = db.execute("SELECT reconciliation_status FROM pickup_rooming_list WHERE config_id=? ORDER BY upload_date DESC LIMIT 1", (cid,)).fetchone()
    if not config:
        flash('Event not found.', 'error')
        return redirect(url_for('pickup_dashboard'))
    rl_status    = rl['reconciliation_status'] if rl else None
    weekly_dicts = [dict(w) for w in weekly_list]
    email         = build_client_email(config, last, rl_status, weekly_list=weekly_dicts)
    html_body     = email.get('html_body', '')
    subject       = (email.get('subject') or '').strip()
    to_addr       = (email.get('to') or '').strip()
    cc_recipients = _build_cc_recipients(config)
    cc_list       = [r['email'] for r in cc_recipients if r.get('email')]

    # ── Web: Graph API draft ───────────────────────────────────────────────
    if platform.system() != 'Darwin':
        user   = get_current_user()
        ms_row = db.execute('SELECT user_id FROM UserMicrosoftTokens WHERE user_id=?',
                            (user['id'],)).fetchone() if user else None
        if ms_row:
            draft_id, err = _create_outlook_draft(
                user_id=user['id'], subject=subject,
                to_addr=to_addr, cc_list=cc_list, body_html=html_body,
            )
            if draft_id:
                db.execute(
                    "INSERT INTO pickup_contact_log (config_id, contact_date, contact_type, notes) VALUES (?,?,?,?)",
                    (cid, _date.today().isoformat(), 'email_sent',
                     f'Client draft created via Graph API — To: {to_addr}')
                )
                db.commit()
                flash('Draft created in your Outlook Drafts folder. Open Outlook to review and send.', 'success')
                return redirect(url_for('pickup_event', cid=cid))
            else:
                flash(f'Could not create Outlook draft: {err}', 'error')
        return redirect(url_for('pickup_email_client', cid=cid))

    # ── Local Mac: JXA clipboard + AppleScript ────────────────────────────
    def write_tmp(content, suffix):
        t = tempfile.NamedTemporaryFile(mode='w', suffix=suffix, delete=False, encoding='utf-8')
        t.write(content); t.close(); return t.name

    def esc(s):
        return s.replace('\\', '\\\\').replace('"', '\\"')

    tmp = write_tmp(html_body, '.html')
    clip_jxa = f"""ObjC.import('AppKit'); ObjC.import('Foundation');
var nsStr = $.NSString.alloc.initWithContentsOfFileEncodingError('{tmp}', $.NSUTF8StringEncoding, null);
var html = ObjC.unwrap(nsStr);
var pb = $.NSPasteboard.generalPasteboard; pb.clearContents;
pb.setStringForType($.NSString.alloc.initWithUTF8String(html), $.NSPasteboardTypeHTML);"""

    to_line  = (f'make new to recipient at theMsg with properties {{email address:{{name:"", address:"{esc(to_addr)}"}}}}' if to_addr else '')
    cc_lines = '\n    '.join(f'make new cc recipient at theMsg with properties {{email address:{{name:"{esc(r["name"])}", address:"{esc(r["email"])}"}}}}' for r in cc_recipients)
    outlook_script = (
        'tell application "Microsoft Outlook"\n'
        f'    set theMsg to make new outgoing message with properties {{subject:"{esc(subject)}"}}\n'
        + (f'    {to_line}\n' if to_line else '')
        + (f'    {cc_lines}\n' if cc_lines else '')
        + '    open theMsg\n    activate\nend tell\n'
    )
    try:
        clip_path    = write_tmp(clip_jxa, '.js')
        outlook_path = write_tmp(outlook_script, '.applescript')
        subprocess.run(['/usr/bin/osascript', '-l', 'JavaScript', clip_path], check=True)
        subprocess.Popen(['/usr/bin/osascript', outlook_path])
    except Exception as exc:
        flash(f'Could not launch Outlook: {exc}', 'error')
        return redirect(url_for('pickup_email_client', cid=cid))
    return redirect(url_for('pickup_email_client_paste', cid=cid))


@app.route('/pickup/<int:cid>/email/client/paste-instructions')
def pickup_email_client_paste(cid):
    db = get_db()
    config = db.execute("SELECT * FROM pickup_config WHERE id=?", (cid,)).fetchone()
    return render_template('pickup_outlook_paste.html', config=config)


@app.route('/pickup/<int:cid>/export-row')
def pickup_export_row(cid):
    from pickup_utils import format_export_row
    db = get_db()
    config = db.execute("SELECT * FROM pickup_config WHERE id=?", (cid,)).fetchone()
    last   = db.execute("SELECT * FROM pickup_weekly WHERE config_id=? ORDER BY report_date DESC LIMIT 1", (cid,)).fetchone()
    if not config or not last:
        flash('No data to export.', 'error')
        return redirect(url_for('pickup_event', cid=cid))
    row_text = format_export_row(config, last)
    return render_template('pickup_export_row.html', config=config, row_text=row_text)


@app.route('/pickup/<int:cid>/hhr/download')
def pickup_hhr_download(cid):
    import io as _io
    db = get_db()
    config = db.execute('SELECT * FROM pickup_config WHERE id=?', (cid,)).fetchone()
    if not config:
        flash('Event not found.', 'error')
        return redirect(url_for('pickup_event', cid=cid))
    hhr = db.execute(
        'SELECT filename, file_data FROM housing_history_files WHERE booking_id=? ORDER BY id DESC LIMIT 1',
        (config['booking_id'],)
    ).fetchone()
    if not hhr or not hhr['file_data']:
        flash('No Housing History Report on file.', 'error')
        return redirect(url_for('pickup_event', cid=cid))
    file_bytes = bytes(hhr['file_data'])
    is_pdf = file_bytes[:4] == b'%PDF'
    if not is_pdf:
        try:
            from pickup_utils import populate_hhr_template
            file_bytes = populate_hhr_template(file_bytes)
        except Exception:
            try:
                from pickup_utils import strip_hhr_commission_rows, clean_hhr_for_client
                file_bytes = clean_hhr_for_client(strip_hhr_commission_rows(file_bytes))
            except Exception:
                pass
    stored_name = hhr['filename'] or ('housing_history.pdf' if is_pdf else 'housing_history.xlsx')
    return send_file(_io.BytesIO(file_bytes),
                     download_name=stored_name,
                     as_attachment=True)


def _get_post_report_data(cid):
    """Assemble stats and config_dict for the Post Report from correct DB sources."""
    db = get_db()
    config = db.execute('SELECT * FROM pickup_config WHERE id=?', (cid,)).fetchone()
    if not config:
        return None, None, None

    fh = db.execute(
        "SELECT * FROM pickup_weekly WHERE config_id=? AND label='Final History' ORDER BY id DESC LIMIT 1",
        (cid,)
    ).fetchone()
    if not fh:
        return config, None, None

    hhr_row = db.execute(
        'SELECT filename, file_data FROM housing_history_files WHERE booking_id=? ORDER BY id DESC LIMIT 1',
        (config['booking_id'],)
    ).fetchone()

    pickup_row = db.execute(
        'SELECT ActualPickup, TotalRevenue FROM Pickup WHERE BookingID=?',
        (config['booking_id'],)
    ).fetchone()

    hhr_stats = {}
    if hhr_row and hhr_row['file_data']:
        try:
            from pickup_utils import parse_hhr_excel
            hhr_stats = parse_hhr_excel(bytes(hhr_row['file_data']))
        except Exception:
            pass

    block = json.loads(config['contracted_block'] or '{}')
    contracted_total = sum(block.values()) or hhr_stats.get('contracted_total') or 0
    pbn = json.loads(fh['pickup_by_night'] or '{}') if fh['pickup_by_night'] else {}
    final_total_pickup = (fh['total_rooms'] or hhr_stats.get('final_total_pickup')
                          or (float(pickup_row['ActualPickup']) if pickup_row and pickup_row['ActualPickup'] else 0))

    stats = {
        'organization':         config['organization'] or hhr_stats.get('organization', ''),
        'hotel':                config['hotel']        or hhr_stats.get('hotel', ''),
        'event_name':           config['event_name']   or hhr_stats.get('event_name', ''),
        'contracted_total':     contracted_total,
        'contracted_block':     block,
        'contracted_rate':      hhr_stats.get('contracted_rate') or config['contracted_rate'],
        'final_total_pickup':   final_total_pickup,
        'final_pickup_by_night': pbn,
        'pct_of_block':         fh['pct_of_block'] or hhr_stats.get('pct_of_block'),
        'pct_of_attrition':     fh['pct_of_attrition'],
        'room_revenue':         hhr_stats.get('room_revenue') or (float(pickup_row['TotalRevenue']) if pickup_row and pickup_row['TotalRevenue'] else None),
        'fb_revenue':           hhr_stats.get('fb_revenue'),
        'audit_pickup':         hhr_stats.get('audit_pickup'),
        'no_shows':             hhr_stats.get('no_shows'),
        'cancellations':        hhr_stats.get('cancellations'),
        'hotel_approver':       hhr_stats.get('hotel_approver'),
        'hotel_approver_email': hhr_stats.get('hotel_approver_email'),
    }

    config_dict = dict(config)
    import re as _re
    _event = (config['event_name'] or config['organization'] or 'Housing History Report').strip()
    _safe  = _re.sub(r'[\\/*?:"<>|]', '', _event)

    raw_bytes = bytes(hhr_row['file_data']) if (hhr_row and hhr_row['file_data']) else None
    _is_pdf = raw_bytes and raw_bytes[:4] == b'%PDF'

    if _is_pdf:
        # Hotel returned a PDF — generate a proper client-facing Excel using the
        # same template. Write Final History pickup into Rate 1 date columns and
        # rate into N8. All SUM/formula cells (M8, O8, P8, Q8, rows 15-19, etc.)
        # are left intact — they recalculate from the values written.
        config_dict['hhr_filename'] = f"{_safe} — Housing History Report.xlsx"
        pipeline = None
        if config['booking_id']:
            try:
                pipeline = db.execute(
                    'SELECT * FROM ReportPipeline WHERE BookingId = ?', (config['booking_id'],)
                ).fetchone()
            except Exception:
                pass
        try:
            import io as _io
            from openpyxl.styles import Alignment as _Align
            _center = _Align(horizontal='center', vertical='center')

            _pickup_dates = list(pbn.keys()) if pbn else []
            _wb, _ = _build_housing_form_wb(config, pipeline, pickup_dates=_pickup_dates)
            _ws = _wb.active

            _blk    = json.loads(config['contracted_block'] or '{}')
            _sdates = sorted(_blk.keys())
            _all_d  = sorted(set(_sdates) | set(_pickup_dates)) if _pickup_dates else _sdates
            _n_xtra = max(0, len(_all_d) - 10)
            _rcol   = 14 + _n_xtra   # N column: rate for Rate 1

            # Write per-night pickup into Rate 1 date columns only.
            # M8 has =SUM(C8:L8), O8 has =N8*M8 — do NOT overwrite those formulas.
            for _i, _d in enumerate(_all_d):
                _v = pbn.get(_d, 0)
                try:
                    _c = _ws.cell(row=8, column=3 + _i)
                    _c.value     = _v if _v else None
                    _c.alignment = _center
                except Exception:
                    pass

            # Write contracted rate into N8 so O8 (=N8*M8) calculates revenue
            try:
                _ws.cell(row=8, column=_rcol).value = (
                    float(config['contracted_rate']) if config['contracted_rate'] else None
                )
            except Exception:
                pass

            _buf = _io.BytesIO()
            _wb.save(_buf)
            raw_bytes = _buf.getvalue()
        except Exception:
            config_dict['hhr_filename'] = f"{_safe} — Housing History Report.pdf"
    else:
        config_dict['hhr_filename'] = f"{_safe} — Housing History Report.xlsx"
        if raw_bytes:
            try:
                from pickup_utils import strip_hhr_commission_rows, clean_hhr_for_client
                raw_bytes = clean_hhr_for_client(strip_hhr_commission_rows(raw_bytes))
            except Exception:
                pass
    config_dict['_hhr_file_data'] = raw_bytes

    # Most recent rooming list — inclusion decided by checkbox on summary screen
    config_dict['_rl_concurrent'] = None
    rl_latest = db.execute(
        """SELECT id, filename, file_data, upload_date
           FROM pickup_rooming_list
           WHERE config_id=?
           ORDER BY upload_date DESC LIMIT 1""",
        (cid,)
    ).fetchone()
    if rl_latest and rl_latest['file_data']:
        config_dict['_rl_concurrent'] = dict(rl_latest)

    return config_dict, stats, fh


@app.route('/pickup/<int:cid>/email/post-report')
def pickup_email_post_report(cid):
    config_dict, stats, fh = _get_post_report_data(cid)
    if config_dict is None:
        flash('Event not found.', 'error')
        return redirect(url_for('pickup_dashboard'))
    if stats is None:
        flash('No Final History on file — import the HHR first via Import → HHR.', 'warning')
        return redirect(url_for('pickup_event', cid=cid))
    email = _build_post_report_email(config_dict, stats)
    return render_template('pickup_post_report_email.html', config=config_dict, stats=stats, email=email)


@app.route('/pickup/<int:cid>/email/post-report/launch-outlook')
def pickup_email_post_report_outlook(cid):
    import subprocess, tempfile, platform
    from datetime import date as _date

    config_dict, stats, fh = _get_post_report_data(cid)
    if config_dict is None:
        flash('Event not found.', 'error')
        return redirect(url_for('pickup_dashboard'))
    if stats is None:
        flash('No Final History on file — import the HHR first via Import → HHR.', 'warning')
        return redirect(url_for('pickup_event', cid=cid))

    # ── Web: Graph API draft with HHR attachment ───────────────────────────
    if platform.system() not in ('Darwin', 'Windows'):
        user   = get_current_user()
        db     = get_db()
        ms_row = db.execute('SELECT user_id FROM UserMicrosoftTokens WHERE user_id=?',
                            (user['id'],)).fetchone() if user else None
        if ms_row:
            email        = _build_post_report_email(config_dict, stats)
            html_body    = email.get('html_body', '')
            subject      = email.get('subject', '')
            to_addr      = email.get('to', '')
            file_data    = config_dict.get('_hhr_file_data')
            hhr_filename = config_dict.get('hhr_filename') or 'Housing History Report.xlsx'
            include_rl   = request.args.get('include_rl') == '1'
            rl_row       = config_dict.get('_rl_concurrent') if include_rl else None
            from pickup_utils import _build_cc_recipients
            cc_list = [r['email'] for r in _build_cc_recipients(config_dict) if r.get('email')]
            draft_id, err = _create_outlook_draft(
                user_id=user['id'], subject=subject,
                to_addr=to_addr, cc_list=cc_list, body_html=html_body,
                attachment_bytes=file_data, attachment_filename=hhr_filename,
            )
            if draft_id and rl_row and rl_row['file_data']:
                # Attach rooming list as a second file
                import base64, requests as _req
                token_row = db.execute('SELECT access_token FROM UserMicrosoftTokens WHERE user_id=?', (user['id'],)).fetchone()
                if token_row:
                    rl_filename = rl_row['filename'] or 'rooming_list.pdf'
                    rl_ct = 'application/pdf' if rl_filename.lower().endswith('.pdf') else 'application/octet-stream'
                    _req.post(
                        f'https://graph.microsoft.com/v1.0/me/messages/{draft_id}/attachments',
                        headers={'Authorization': f'Bearer {token_row["access_token"]}', 'Content-Type': 'application/json'},
                        json={'@odata.type': '#microsoft.graph.fileAttachment', 'name': rl_filename,
                              'contentBytes': base64.b64encode(bytes(rl_row['file_data'])).decode('ascii'),
                              'contentType': rl_ct}
                    )
            if draft_id:
                rl_note = ' with rooming list attached' if (rl_row and rl_row['file_data']) else ''
                flash(f'Post Report draft created in your Outlook Drafts folder with HHR{rl_note}. Open Outlook to review and send.', 'success')
                return redirect(url_for('pickup_event', cid=config_dict['id']))
            else:
                flash(f'Could not create Outlook draft: {err}', 'error')
        else:
            flash('Connect your Microsoft account (My Account) to create Outlook drafts from the website.', 'warning')
        return redirect(url_for('pickup_email_post_report', cid=config_dict['id']))

    email     = _build_post_report_email(config_dict, stats)
    html_body = email.get('html_body', '')
    subject   = email.get('subject', '')
    to_addr   = email.get('to', '')

    from pickup_utils import _build_cc_recipients
    cc_recipients = _build_cc_recipients(config_dict)

    file_data    = config_dict.get('_hhr_file_data')
    hhr_filename = config_dict.get('hhr_filename') or 'Housing History Report.xlsx'
    include_rl   = request.args.get('include_rl') == '1'
    rl_row       = config_dict.get('_rl_concurrent') if include_rl else None

    def write_tmp(content, suffix, mode='w', encoding='utf-8'):
        t = tempfile.NamedTemporaryFile(mode=mode, suffix=suffix, delete=False,
                                        encoding=encoding if mode == 'w' else None)
        t.write(content); t.close(); return t.name

    def write_named_att(data, filename):
        """Write attachment bytes to a temp dir using the proper display filename."""
        import os as _os
        d = tempfile.mkdtemp()
        p = _os.path.join(d, filename)
        with open(p, 'wb') as f:
            f.write(data)
        return p

    try:
        if platform.system() == 'Windows':
            # ── Windows: PowerShell + Outlook COM ─────────────────────────────
            tmp_html = write_tmp(html_body, '.html')

            att_path = ''
            if file_data:
                att_path = write_named_att(file_data, hhr_filename).replace('\\', '\\\\')
            rl_att_path = ''
            if rl_row and rl_row['file_data']:
                rl_filename = rl_row['filename'] or 'rooming_list.pdf'
                rl_att_path = write_named_att(bytes(rl_row['file_data']), rl_filename).replace('\\', '\\\\')

            cc_str = '; '.join(r['email'] for r in cc_recipients if r.get('email'))

            ps_lines = [
                '$html = Get-Content -Path \'' + tmp_html.replace("'", "''") + '\' -Raw -Encoding UTF8',
                '$ol = New-Object -ComObject Outlook.Application',
                '$mail = $ol.CreateItem(0)',
                f'$mail.Subject = \'{subject.replace(chr(39), chr(39)*2)}\'',
                '$mail.HTMLBody = $html',
            ]
            if to_addr:
                ps_lines.append(f'$mail.To = \'{to_addr.replace(chr(39), chr(39)*2)}\'')
            if cc_str:
                ps_lines.append(f'$mail.CC = \'{cc_str.replace(chr(39), chr(39)*2)}\'')
            if att_path:
                ps_lines.append(f'$mail.Attachments.Add(\'{att_path}\') | Out-Null')
            if rl_att_path:
                ps_lines.append(f'$mail.Attachments.Add(\'{rl_att_path}\') | Out-Null')
            ps_lines.append('$mail.Display()')

            ps_script = '\r\n'.join(ps_lines)
            ps_path   = write_tmp(ps_script, '.ps1')
            subprocess.Popen(['powershell.exe', '-ExecutionPolicy', 'Bypass',
                               '-WindowStyle', 'Hidden', '-File', ps_path])

        else:
            # ── macOS: AppleScript + System Events auto-paste ─────────────────
            def esc(s):
                return s.replace('\\', '\\\\').replace('"', '\\"')

            tmp_html = write_tmp(html_body, '.html')
            clip_jxa = (
                "ObjC.import('AppKit'); ObjC.import('Foundation');\n"
                f"var nsStr = $.NSString.alloc.initWithContentsOfFileEncodingError('{tmp_html}', $.NSUTF8StringEncoding, null);\n"
                "var html = ObjC.unwrap(nsStr);\n"
                "var pb = $.NSPasteboard.generalPasteboard; pb.clearContents;\n"
                "pb.setStringForType($.NSString.alloc.initWithUTF8String(html), $.NSPasteboardTypeHTML);"
            )

            attach_line = ''
            if file_data:
                att_path = write_named_att(file_data, hhr_filename)
                attach_line = f'make new attachment at theMsg with properties {{file:POSIX file "{att_path}"}}'
            rl_attach_line = ''
            if rl_row and rl_row['file_data']:
                rl_filename = rl_row['filename'] or 'rooming_list.pdf'
                rl_att_path = write_named_att(bytes(rl_row['file_data']), rl_filename)
                rl_attach_line = f'make new attachment at theMsg with properties {{file:POSIX file "{rl_att_path}"}}'

            to_line = (
                f'make new to recipient at theMsg with properties {{email address:{{name:"", address:"{esc(to_addr)}"}}}}'
                if to_addr else ''
            )
            cc_lines = '\n    '.join(
                f'make new cc recipient at theMsg with properties {{email address:{{name:"{esc(r["name"])}", address:"{esc(r["email"])}"}}}}'
                for r in cc_recipients
            )

            outlook_script = (
                'tell application "Microsoft Outlook"\n'
                f'    set theMsg to make new outgoing message with properties {{subject:"{esc(subject)}"}}\n'
                + (f'    {to_line}\n' if to_line else '')
                + (f'    {cc_lines}\n' if cc_lines else '')
                + (f'    {attach_line}\n' if attach_line else '')
                + (f'    {rl_attach_line}\n' if rl_attach_line else '')
                + '    open theMsg\n'
                + '    activate\n'
                + 'end tell\n'
                + 'delay 4\n'
                + 'tell application "Microsoft Outlook"\n'
                + '    activate\n'
                + 'end tell\n'
                + 'delay 1\n'
                + 'tell application "System Events"\n'
                + '    keystroke "v" using {command down}\n'
                + 'end tell\n'
            )

            clip_path    = write_tmp(clip_jxa, '.js')
            outlook_path = write_tmp(outlook_script, '.applescript')
            subprocess.run(['osascript', '-l', 'JavaScript', clip_path], check=True)
            subprocess.Popen(['osascript', outlook_path])

    except Exception as exc:
        flash(f'Could not launch Outlook: {exc}', 'error')
        return redirect(url_for('pickup_event', cid=cid))

    rl_note = ' + rooming list' if (include_rl and rl_row) else ''
    flash(f'Post Report email opened in Outlook with HHR{rl_note} attached.', 'success')
    return redirect(url_for('pickup_event', cid=cid))


def _build_post_report_email(config, stats):
    """Build the post-event housing history email dict."""
    from datetime import date as _date
    org        = config['organization'] or stats.get('organization', '')
    event_name = config['event_name']   or stats.get('event_name', '')
    hotel      = config['hotel']        or stats.get('hotel', '')
    to_addr    = config['group_contact_email'] or ''

    ct       = stats.get('contracted_total') or 0
    fp       = stats.get('final_total_pickup') or 0
    pct      = stats.get('pct_of_block')
    rate     = stats.get('contracted_rate')
    rev      = stats.get('room_revenue')
    fb_rev   = stats.get('fb_revenue')
    audit    = stats.get('audit_pickup') or 0
    no_shows = stats.get('no_shows') or 0
    cancels  = stats.get('cancellations') or 0

    attrition_rooms = round(ct * float(config['attrition_pct'] or 0)) if ct and config.get('attrition_pct') else None
    pct_attrition   = round(fp / attrition_rooms * 100, 1) if attrition_rooms and fp else None
    if not pct_attrition:
        pct_attrition_val = stats.get('pct_of_attrition')
        if pct_attrition_val:
            pct_attrition = pct_attrition_val

    # ── Attrition performance statement (Style B — warm, client-friendly) ─────
    # Inline text appended to the intro paragraph. Only built when we know the
    # attrition floor (contracted block × commitment) and the final pickup;
    # otherwise omitted, same as the table's attrition row.
    attrition_statement = ''
    if attrition_rooms and fp:
        _commit_pct = round(float(config['attrition_pct'] or 0) * 100)
        _floor      = int(round(attrition_rooms))
        _pickup     = int(round(fp))
        if _pickup >= _floor:
            _verb  = 'exceeded' if _pickup > _floor else 'met'
            _clear = 'comfortably clearing' if _pickup >= _floor * 1.03 else 'clearing'
            attrition_statement = (
                f'Great news — your group <strong>{_verb} its attrition commitment</strong>. '
                f'You needed {_floor:,} room nights ({_commit_pct}% of the contracted block) and '
                f'picked up {_pickup:,}, {_clear} the threshold with no attrition charges.'
            )
        else:
            _short = _floor - _pickup
            attrition_statement = (
                f'Your group came in <strong>just below its attrition commitment</strong>. '
                f'The contract required {_floor:,} room nights ({_commit_pct}% of the block) and '
                f'final pickup was {_pickup:,} — a shortfall of {_short:,} room night'
                f'{"s" if _short != 1 else ""} we can review with the hotel on your behalf.'
            )

    subject = f'{org} — {event_name} | Final Housing History Report'

    _dow = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']

    block_nights  = stats.get('contracted_block', {})
    pickup_nights = stats.get('final_pickup_by_night', {})
    all_dates = sorted(set(list(block_nights.keys()) + list(pickup_nights.keys())))
    night_rows = ''
    for d in all_dates:
        b = block_nights.get(d, 0)
        p = pickup_nights.get(d, 0)
        if b == 0 and p == 0:
            continue
        diff = p - b
        diff_str   = (f'+{diff}' if diff > 0 else str(diff)) if diff != 0 else '—'
        diff_color = '#16a34a' if diff > 0 else ('#dc2626' if diff < 0 else '#6b7280')
        try:
            dow = _dow[_date.fromisoformat(d).weekday()]
        except Exception:
            dow = ''
        try:
            _d_obj = _date.fromisoformat(d)
            _d_fmt = f'{_d_obj.month}/{_d_obj.day}/{_d_obj.year}'
        except Exception:
            _d_fmt = d[5:].replace('-', '/')
        date_label = f'{_d_fmt} ({dow})' if dow else _d_fmt
        night_rows += (
            f'<tr>'
            f'<td style="padding:4px 10px;border-bottom:1px solid #e5e7eb">{date_label}</td>'
            f'<td style="padding:4px 10px;text-align:center;border-bottom:1px solid #e5e7eb">{b or "—"}</td>'
            f'<td style="padding:4px 10px;text-align:center;border-bottom:1px solid #e5e7eb">{p or "—"}</td>'
            f'<td style="padding:4px 10px;text-align:center;border-bottom:1px solid #e5e7eb;color:{diff_color}">{diff_str}</td>'
            f'</tr>'
        )

    pct_str     = f'{pct:.1f}%' if pct is not None else 'N/A'
    pct_att_str = f'{pct_attrition:.1f}%' if pct_attrition is not None else 'N/A'
    rate_str    = f'${rate:,.2f}' if rate else 'N/A'
    rev_str     = f'${rev:,.2f}' if rev else 'N/A'
    fb_str      = f'${fb_rev:,.2f}' if fb_rev else 'N/A'
    pct_color   = '#16a34a' if (pct or 0) >= 100 else '#d97706'
    att_color   = '#16a34a' if (pct_attrition or 0) >= 100 else '#d97706'

    _contact    = (config.get('group_contact') or '').strip()
    _first_name = _contact.split()[0] if _contact else ''

    html_body = f'''<div style="font-family:Arial,sans-serif;max-width:640px;color:#1f2937">
<p>Hi {_first_name},</p>
<p>Please find attached the final housing history report for <strong>{event_name}</strong> at <strong>{hotel}</strong>. Below is a summary of the pickup performance.{('&nbsp;&nbsp; ' + attrition_statement) if attrition_statement else ''}</p>

<table style="width:100%;border-collapse:collapse;margin:16px 0;background:#f9fafb;border-radius:8px;overflow:hidden">
  <tr style="background:#1a3a5c;color:#fff">
    <th style="padding:8px 10px;text-align:left" colspan="2">Pickup Performance Summary</th>
  </tr>
  <tr><td style="padding:6px 10px;border-bottom:1px solid #e5e7eb">Contracted Block</td>
      <td style="padding:6px 10px;border-bottom:1px solid #e5e7eb;font-weight:600">{int(ct):,} room nights</td></tr>
  <tr><td style="padding:6px 10px;border-bottom:1px solid #e5e7eb">Final Total Pickup</td>
      <td style="padding:6px 10px;border-bottom:1px solid #e5e7eb;font-weight:600">{int(fp):,} room nights</td></tr>
  <tr><td style="padding:6px 10px;border-bottom:1px solid #e5e7eb">% of Contracted Block</td>
      <td style="padding:6px 10px;border-bottom:1px solid #e5e7eb;font-weight:600;color:{pct_color}">{pct_str}</td></tr>
  {'<tr><td style="padding:6px 10px;border-bottom:1px solid #e5e7eb">% of Attrition Commitment</td><td style="padding:6px 10px;border-bottom:1px solid #e5e7eb;font-weight:600;color:' + att_color + '">' + pct_att_str + '</td></tr>' if pct_attrition else ''}
  <tr><td style="padding:6px 10px;border-bottom:1px solid #e5e7eb">Contracted Rate</td>
      <td style="padding:6px 10px;border-bottom:1px solid #e5e7eb">{rate_str}</td></tr>
  <tr><td style="padding:6px 10px;border-bottom:1px solid #e5e7eb">Total Room Revenue</td>
      <td style="padding:6px 10px;border-bottom:1px solid #e5e7eb">{rev_str}</td></tr>
  {'<tr><td style="padding:6px 10px;border-bottom:1px solid #e5e7eb">Food &amp; Beverage Revenue</td><td style="padding:6px 10px;border-bottom:1px solid #e5e7eb">' + fb_str + '</td></tr>' if fb_rev else ''}
  {'<tr><td style="padding:6px 10px;border-bottom:1px solid #e5e7eb">Audit Pickup (Outside Block)</td><td style="padding:6px 10px;border-bottom:1px solid #e5e7eb">' + f'{int(audit):,} rooms' + '</td></tr>' if audit else ''}
  {'<tr><td style="padding:6px 10px">No Shows / Cancellations</td><td style="padding:6px 10px">' + f'{int(no_shows):,} / {int(cancels):,}' + '</td></tr>' if (no_shows or cancels) else ''}
</table>
{'<table style="width:100%;border-collapse:collapse;margin:16px 0"><thead><tr style="background:#e5e7eb"><th style="padding:6px 10px;text-align:left">Night</th><th style="padding:6px 10px;text-align:center">Block</th><th style="padding:6px 10px;text-align:center">Pickup</th><th style="padding:6px 10px;text-align:center">+/−</th></tr></thead><tbody>' + night_rows + '</tbody></table>' if night_rows else ''}
<p>The full Housing History Report is attached. Please don't hesitate to reach out with any questions.</p>
<p>Best regards,</p>
</div>'''

    return {'to': to_addr, 'subject': subject, 'html_body': html_body}


@app.route('/pickup/<int:cid>/archive', methods=['POST'])
def pickup_archive(cid):
    db = get_db()
    db.execute("UPDATE pickup_config SET status='archived' WHERE id=?", (cid,))
    db.commit()
    flash('Event archived.', 'success')
    return redirect(url_for('pickup_dashboard'))


@app.route('/pickup/<int:cid>/cancel-meeting', methods=['POST'])
def pickup_cancel_meeting(cid):
    import datetime as _dt
    db = get_db()
    config = db.execute("SELECT * FROM pickup_config WHERE id=?", (cid,)).fetchone()
    if not config:
        flash('Event not found.', 'error')
        return redirect(url_for('pickup_dashboard'))
    # Cancel the linked pipeline booking
    if config['booking_id']:
        db.execute(
            "UPDATE ReportPipeline SET BookingStatus='Cancelled' WHERE CAST(BookingId AS TEXT)=CAST(? AS TEXT)",
            (config['booking_id'],)
        )
        db.execute(
            "UPDATE ChkRegNote SET Cancelled=1 WHERE BookingID=? AND (Cancelled IS NULL OR Cancelled=0)",
            (config['booking_id'],)
        )
    # Archive the pickup card with a cancellation note
    cancel_note = f"CANCELLED — {_dt.date.today().strftime('%Y-%m-%d')}"
    existing = (config['notes'] or '').strip()
    new_notes = cancel_note + ('\n' + existing if existing else '')
    db.execute(
        "UPDATE pickup_config SET status='archived', notes=? WHERE id=?",
        (new_notes, cid)
    )
    db.commit()
    flash('Meeting cancelled and pickup card archived.', 'success')
    return redirect(url_for('pickup_dashboard'))


@app.route('/pickup/<int:cid>/unarchive', methods=['POST'])
def pickup_unarchive(cid):
    db = get_db()
    db.execute("UPDATE pickup_config SET status='active' WHERE id=?", (cid,))
    db.commit()
    flash('Event restored to active.', 'success')
    return redirect(url_for('pickup_dashboard'))


@app.route('/pickup/<int:cid>/toggle-current', methods=['POST'])
def pickup_toggle_current(cid):
    db = get_db()
    row = db.execute("SELECT force_current FROM pickup_config WHERE id=?", (cid,)).fetchone()
    new_val = 0 if (row and row[0]) else 1
    db.execute("UPDATE pickup_config SET force_current=? WHERE id=?", (new_val, cid))
    db.commit()
    return redirect(url_for('pickup_dashboard'))


@app.route('/pickup/<int:cid>/promote-past', methods=['POST'])
def pickup_promote_past(cid):
    db = get_db()
    row = db.execute("SELECT force_past FROM pickup_config WHERE id=?", (cid,)).fetchone()
    new_val = 0 if (row and row['force_past']) else 1
    db.execute("UPDATE pickup_config SET force_past=? WHERE id=?", (new_val, cid))
    db.commit()
    return redirect(url_for('pickup_dashboard'))


@app.route('/pickup/<int:cid>/contact-log/add', methods=['POST'])
def pickup_contact_log_add(cid):
    contact_date = request.form.get('contact_date', '').strip()
    contact_type = request.form.get('contact_type', 'email_sent')
    notes        = request.form.get('notes', '').strip()
    if contact_date:
        db = get_db()
        db.execute("INSERT INTO pickup_contact_log (config_id, contact_date, contact_type, notes) VALUES (?,?,?,?)",
                   (cid, contact_date, contact_type, notes or None))
        db.commit()
    return redirect(url_for('pickup_event', cid=cid))


@app.route('/pickup/<int:cid>/contact-log/<int:log_id>/delete', methods=['POST'])
def pickup_contact_log_delete(cid, log_id):
    db = get_db()
    db.execute("DELETE FROM pickup_contact_log WHERE id=? AND config_id=?", (log_id, cid))
    db.commit()
    return redirect(url_for('pickup_event', cid=cid))


@app.route('/pickup/<int:cid>/log-email-sent', methods=['POST'])
def pickup_log_email_sent(cid):
    from datetime import date as _date
    db = get_db()
    db.execute("INSERT INTO pickup_contact_log (config_id, contact_date, contact_type, notes) VALUES (?,?,?,?)",
               (cid, _date.today().isoformat(), 'email_sent', 'Logged automatically on Open in Mail App'))
    db.commit()
    return jsonify({'ok': True}), 200, {'Content-Type': 'application/json'}


@app.route('/pickup/<int:cid>/quick-respond', methods=['POST'])
def pickup_quick_respond(cid):
    from datetime import date as _date
    db = get_db()
    existing = db.execute("SELECT id FROM pickup_contact_log WHERE config_id=? AND contact_type='responded' LIMIT 1", (cid,)).fetchone()
    today = _date.today().isoformat()
    if not existing:
        db.execute("INSERT INTO pickup_contact_log (config_id, contact_date, contact_type, notes) VALUES (?,?,?,?)",
                   (cid, today, 'responded', 'Marked responded from dashboard'))
        db.commit()
    if request.headers.get('X-Requested-With') == 'fetch':
        return jsonify({'ok': True, 'date': today})
    return redirect(url_for('pickup_dashboard'))


@app.route('/pickup/<int:cid>/undo-respond', methods=['POST'])
def pickup_undo_respond(cid):
    db = get_db()
    db.execute("DELETE FROM pickup_contact_log WHERE config_id=? AND contact_type='responded'", (cid,))
    db.commit()
    if request.headers.get('X-Requested-With') == 'fetch':
        return jsonify({'ok': True})
    return redirect(url_for('pickup_dashboard'))


@app.route('/pickup/<int:cid>/rooming-list/debug-file', methods=['POST'])
def pickup_debug_file(cid):
    f = request.files.get('rooming_pdf')
    if not f:
        return 'No file uploaded', 400
    ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else ''
    raw_bytes = f.read()
    try:
        if ext == 'pdf':
            import pdfplumber, io as _io
            pages = []
            with pdfplumber.open(_io.BytesIO(raw_bytes)) as pdf:
                for i, page in enumerate(pdf.pages):
                    pages.append(f'=== PAGE {i+1} ===\n' + (page.extract_text() or '(no text)'))
            raw = '\n'.join(pages)
        elif ext == 'csv':
            raw = raw_bytes.decode('utf-8-sig', errors='replace')
        else:
            raw = f'Binary file ({len(raw_bytes)} bytes) — cannot display'
        return f'<pre style="font-size:.8rem;white-space:pre-wrap;word-break:break-all">{raw[:10000]}</pre>'
    except Exception as e:
        return f'Error: {e}', 500


# ── Contract Templates ────────────────────────────────────────────────────────

@app.route('/contracts/templates')
def contract_templates():
    user = get_current_user()
    if not has_permission(user, 'contracts'):
        flash('Access denied.', 'error')
        return redirect(url_for('pipeline'))
    db = get_db()
    templates = db.execute(
        "SELECT * FROM contract_template WHERE is_active=1 ORDER BY chain, template_name"
    ).fetchall()
    from collections import defaultdict
    by_chain = defaultdict(list)
    for t in templates:
        by_chain[t['chain']].append(t)
    return render_template('contract_templates.html', by_chain=dict(by_chain))


@app.route('/contracts/templates/upload', methods=['GET', 'POST'])
def contract_template_upload():
    user = get_current_user()
    if not has_permission(user, 'contracts'):
        flash('Access denied.', 'error')
        return redirect(url_for('pipeline'))
    from pickup_utils import extract_template_metadata
    if request.method == 'GET':
        return render_template('contract_template_upload.html')
    f = request.files.get('template_file')
    if not f or not f.filename:
        flash('No file selected.', 'error')
        return redirect(url_for('contract_templates'))
    file_bytes = f.read()
    chain = request.form.get('chain', '').strip()
    template_name = request.form.get('template_name', '').strip() or f.filename
    description = request.form.get('description', '').strip() or None
    try:
        meta = extract_template_metadata(file_bytes)
    except Exception as e:
        meta = {'merge_fields': [], 'sections': []}
        flash(f'Warning: could not parse template metadata ({e}).', 'warning')
    db = get_db()
    db.execute('''
        INSERT INTO contract_template
            (chain, template_name, description, filename, file_data, sections, merge_fields)
        VALUES (?,?,?,?,?,?,?)
    ''', (chain, template_name, description, f.filename, file_bytes,
          json.dumps(meta['sections']), json.dumps(meta['merge_fields'])))
    db.commit()
    flash(f'Template "{template_name}" uploaded ({len(meta["sections"])} sections, '
          f'{len(meta["merge_fields"])} merge fields).', 'success')
    return redirect(url_for('contract_templates'))


@app.route('/contracts/templates/<int:tid>/download')
def contract_template_download(tid):
    user = get_current_user()
    if not has_permission(user, 'contracts'):
        flash('Access denied.', 'error')
        return redirect(url_for('pipeline'))
    import io as _io
    db = get_db()
    tmpl = db.execute("SELECT * FROM contract_template WHERE id=?", (tid,)).fetchone()
    if not tmpl or not tmpl['file_data']:
        flash('Template not found.', 'error')
        return redirect(url_for('contract_templates'))
    return send_file(_io.BytesIO(tmpl['file_data']),
                     download_name=tmpl['filename'] or 'template.docx',
                     as_attachment=True)


@app.route('/contracts/templates/<int:tid>/delete', methods=['POST'])
def contract_template_delete(tid):
    user = get_current_user()
    if not has_permission(user, 'contracts'):
        flash('Access denied.', 'error')
        return redirect(url_for('pipeline'))
    db = get_db()
    db.execute("UPDATE contract_template SET is_active=0 WHERE id=?", (tid,))
    db.commit()
    flash('Template removed.', 'success')
    return redirect(url_for('contract_templates'))


# ── RFP Tracking ──────────────────────────────────────────────────────────────

RFP_STATUSES = [
    ('sourcing',           'secondary', 'Sourcing'),
    ('proposals_received', 'info',      'Proposals Received'),
    ('negotiating',        'warning',   'Negotiating'),
    ('hotel_selected',     'primary',   'Hotel Selected'),
    ('contracting',        'primary',   'Contracting'),
    ('contracted',         'success',   'Contracted'),
    ('dead',               'danger',    'Dead'),
]
RFP_STATUS_MAP = {s[0]: (s[1], s[2]) for s in RFP_STATUSES}


@app.route('/rfp/<int:rid>/checklist/cell', methods=['POST'])
def rfp_checklist_cell(rid):
    """AJAX — update a single checklist key (Master View inline edit)."""
    data = request.get_json(silent=True) or {}
    key  = data.get('key', '').strip()
    val  = (data.get('value') or '').strip() or None
    if not key:
        return jsonify({'ok': False, 'error': 'missing key'}), 400
    db = get_db()
    row = db.execute("SELECT checklist FROM rfp WHERE id=?", (rid,)).fetchone()
    if not row:
        return jsonify({'ok': False, 'error': 'not found'}), 404
    try:
        cl = json.loads(row['checklist'] or '{}')
    except Exception:
        cl = {}
    cl[key] = val
    db.execute(
        "UPDATE rfp SET checklist=?, updated_at=datetime('now') WHERE id=?",
        (json.dumps(cl), rid)
    )
    db.commit()
    return jsonify({'ok': True, 'key': key, 'value': val})


@app.route('/rfp')
def rfp_dashboard():
    user = get_current_user()
    db = get_db()
    filter_tab = request.args.get('tab', 'master')
    acct_filter = get_pickup_account_filter(user)

    # Fetch all non-archived RFPs in scope
    if acct_filter is None:
        all_rfps = db.execute(
            'SELECT r.*, (SELECT COUNT(*) FROM rfp_hotel h WHERE h.rfp_id=r.id) AS hotel_count '
            'FROM rfp r WHERE r.archived=0 ORDER BY r.start_date, r.created_at DESC'
        ).fetchall()
    elif acct_filter:
        ph = ','.join('?' * len(acct_filter))
        all_rfps = db.execute(
            f'SELECT r.*, (SELECT COUNT(*) FROM rfp_hotel h WHERE h.rfp_id=r.id) AS hotel_count '
            f'FROM rfp r WHERE r.archived=0 AND r.client_org IN ({ph}) ORDER BY r.start_date, r.created_at DESC',
            tuple(acct_filter)
        ).fetchall()
    else:
        all_rfps = []

    active_statuses = {'sourcing','proposals_received','negotiating','hotel_selected','contracting'}

    # Filter rfps for standard tabs
    if filter_tab == 'active':
        rfps = [r for r in all_rfps if r['status'] in active_statuses]
    elif filter_tab == 'contracted':
        # Contracted tab shows archived contracted RFPs (completed deals)
        if acct_filter is None:
            contracted_archived = db.execute(
                "SELECT r.*, (SELECT COUNT(*) FROM rfp_hotel h WHERE h.rfp_id=r.id) AS hotel_count "
                "FROM rfp r WHERE r.archived=1 AND r.status='contracted' ORDER BY r.start_date DESC"
            ).fetchall()
        elif acct_filter:
            ph = ','.join('?' * len(acct_filter))
            contracted_archived = db.execute(
                f"SELECT r.*, (SELECT COUNT(*) FROM rfp_hotel h WHERE h.rfp_id=r.id) AS hotel_count "
                f"FROM rfp r WHERE r.archived=1 AND r.status='contracted' AND r.client_org IN ({ph}) ORDER BY r.start_date DESC",
                tuple(acct_filter)
            ).fetchall()
        else:
            contracted_archived = []
        rfps = contracted_archived
    elif filter_tab == 'dead':
        rfps = [r for r in all_rfps if r['status'] == 'dead']
    else:
        rfps = list(all_rfps)  # 'master' uses full list

    # Build master view data
    master_by_year = {}
    if filter_tab == 'master':
        for r in all_rfps:
            sel_hotel = db.execute(
                "SELECT * FROM rfp_hotel WHERE rfp_id=? AND status='selected' LIMIT 1", (r['id'],)
            ).fetchone()
            if not sel_hotel:
                sel_hotel = db.execute(
                    "SELECT * FROM rfp_hotel WHERE rfp_id=? LIMIT 1", (r['id'],)
                ).fetchone()
            booking_num = r['booking_id'] or ''
            try:
                cl = json.loads(r['checklist'] or '{}')
            except Exception:
                cl = {}
            contracted_total = None
            if sel_hotel and sel_hotel['proposed_rate'] and r['total_room_nights']:
                comm = sel_hotel['commission_pct'] or 0.10
                contracted_total = sel_hotel['proposed_rate'] * r['total_room_nights'] * comm
            year = 'No Date'
            if r['start_date']:
                try:
                    year = str(r['start_date'])[:4]
                except Exception:
                    year = 'No Date'
            # Notes (most recent 10)
            notes = db.execute(
                "SELECT id, note_date, note_type, note_text FROM rfp_note "
                "WHERE rfp_id=? ORDER BY note_date DESC, id DESC LIMIT 10",
                (r['id'],)
            ).fetchall()

            if year not in master_by_year:
                master_by_year[year] = []
            st = RFP_STATUS_MAP.get(r['status'], ('secondary', r['status']))
            master_by_year[year].append({
                'rfp': r,
                'hotel': sel_hotel,
                'booking_num': booking_num,
                'checklist': cl,
                'contracted_total': contracted_total,
                'badge': st[0],
                'badge_label': st[1],
                'notes': [dict(n) for n in notes],
            })
        master_by_year = dict(sorted(master_by_year.items(), reverse=True))

    return render_template('rfp_dashboard.html', rfps=rfps, statuses=RFP_STATUS_MAP,
                           filter_tab=filter_tab, all_statuses=RFP_STATUSES,
                           master_by_year=master_by_year)


@app.route('/rfp/<int:rid>/note/quick', methods=['POST'])
def rfp_note_quick(rid):
    """AJAX — add a quick note from the Master View dialog."""
    data = request.get_json(silent=True) or {}
    text = (data.get('note_text') or '').strip()
    if not text:
        return jsonify({'ok': False, 'error': 'empty note'}), 400
    from datetime import date as _date
    today = _date.today().isoformat()
    db = get_db()
    cur = db.execute(
        "INSERT INTO rfp_note (rfp_id, note_date, note_type, note_text) VALUES (?,?,?,?)",
        (rid, today, 'internal', text)
    )
    db.commit()
    nid = cur.lastrowid
    return jsonify({'ok': True, 'note': {'id': nid, 'note_date': today, 'note_type': 'internal', 'note_text': text}})


@app.route('/rfp/new', methods=['GET', 'POST'])
def rfp_new():
    if request.method == 'POST':
        db = get_db()
        f = request.form
        db.execute('''
            INSERT INTO rfp (rfp_code, client_org, event_name, rfp_name, booking_id,
                start_date, end_date, alt_start_date, alt_end_date,
                alt_start_date_2, alt_end_date_2, alt_start_date_3, alt_end_date_3,
                peak_rooms, total_room_nights, total_attendees, f_and_b_budget,
                response_due_date, decision_due_date, status, notes)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ''', (
            f.get('rfp_code', '').strip() or None,
            f.get('client_org', '').strip(),
            f.get('event_name', '').strip() or None,
            f.get('rfp_name', '').strip() or None,
            f.get('booking_id', '').strip() or None,
            f.get('start_date') or None,
            f.get('end_date') or None,
            f.get('alt_start_date') or None,
            f.get('alt_end_date') or None,
            f.get('alt_start_date_2') or None,
            f.get('alt_end_date_2') or None,
            f.get('alt_start_date_3') or None,
            f.get('alt_end_date_3') or None,
            f.get('peak_rooms') or None,
            f.get('total_room_nights') or None,
            f.get('total_attendees') or None,
            f.get('f_and_b_budget') or None,
            f.get('response_due_date') or None,
            f.get('decision_due_date') or None,
            f.get('status', 'sourcing'),
            f.get('notes', '').strip() or None,
        ))
        db.commit()
        new_id = db.execute('SELECT last_insert_rowid()').fetchone()[0]
        # Save uploaded RFP document if provided
        rfp_file = request.files.get('rfp_file')
        if rfp_file and rfp_file.filename:
            db.execute("UPDATE rfp SET rfp_filename=?, rfp_data=? WHERE id=?",
                       (rfp_file.filename, rfp_file.read(), new_id))
            db.commit()
        flash('RFP created.', 'success')
        return redirect(url_for('rfp_detail', rid=new_id))
    return render_template('rfp_form.html', rfp=None, all_statuses=RFP_STATUSES)


@app.route('/rfp/<int:rid>')
def rfp_detail(rid):
    user = get_current_user()
    db = get_db()
    rfp = db.execute('SELECT * FROM rfp WHERE id=?', (rid,)).fetchone()
    if not rfp:
        flash('RFP not found.', 'error')
        return redirect(url_for('rfp_dashboard'))
    acct_filter = get_pickup_account_filter(user)
    if acct_filter is not None and rfp['client_org'] not in acct_filter:
        flash('You do not have access to this RFP.', 'error')
        return redirect(url_for('rfp_dashboard'))
    _hotel_order = {'selected':0,'shortlisted':1,'proposal_received':2,'pending':3,'eliminated':4,'declined':5}
    hotels = db.execute('SELECT * FROM rfp_hotel WHERE rfp_id=?', (rid,)).fetchall()
    hotels = sorted(hotels, key=lambda h: _hotel_order.get(h['status'] or 'pending', 3))
    notes = db.execute('SELECT * FROM rfp_note WHERE rfp_id=? ORDER BY note_date DESC', (rid,)).fetchall()
    return render_template('rfp_detail.html', rfp=rfp, hotels=hotels, notes=notes,
                           statuses=RFP_STATUS_MAP, all_statuses=RFP_STATUSES)


@app.route('/rfp/<int:rid>/edit', methods=['GET', 'POST'])
def rfp_edit(rid):
    user = get_current_user()
    db = get_db()
    rfp = db.execute('SELECT * FROM rfp WHERE id=?', (rid,)).fetchone()
    if not rfp:
        flash('RFP not found.', 'error')
        return redirect(url_for('rfp_dashboard'))
    acct_filter = get_pickup_account_filter(user)
    if acct_filter is not None and rfp['client_org'] not in acct_filter:
        flash('You do not have access to this RFP.', 'error')
        return redirect(url_for('rfp_dashboard'))
    if request.method == 'POST':
        f = request.form
        db.execute('''
            UPDATE rfp SET rfp_code=?, client_org=?, event_name=?, rfp_name=?,
                booking_id=?, start_date=?, end_date=?, alt_start_date=?, alt_end_date=?,
                alt_start_date_2=?, alt_end_date_2=?, alt_start_date_3=?, alt_end_date_3=?,
                peak_rooms=?, total_room_nights=?, total_attendees=?, f_and_b_budget=?,
                response_due_date=?, decision_due_date=?, status=?, notes=?,
                updated_at=datetime('now')
            WHERE id=?
        ''', (
            f.get('rfp_code', '').strip() or None,
            f.get('client_org', '').strip(),
            f.get('event_name', '').strip() or None,
            f.get('rfp_name', '').strip() or None,
            f.get('booking_id', '').strip() or None,
            f.get('start_date') or None,
            f.get('end_date') or None,
            f.get('alt_start_date') or None,
            f.get('alt_end_date') or None,
            f.get('alt_start_date_2') or None,
            f.get('alt_end_date_2') or None,
            f.get('alt_start_date_3') or None,
            f.get('alt_end_date_3') or None,
            f.get('peak_rooms') or None,
            f.get('total_room_nights') or None,
            f.get('total_attendees') or None,
            f.get('f_and_b_budget') or None,
            f.get('response_due_date') or None,
            f.get('decision_due_date') or None,
            f.get('status', 'sourcing'),
            f.get('notes', '').strip() or None,
            rid,
        ))
        # Save uploaded RFP document if provided
        rfp_file = request.files.get('rfp_file')
        if rfp_file and rfp_file.filename:
            db.execute("UPDATE rfp SET rfp_filename=?, rfp_data=? WHERE id=?",
                       (rfp_file.filename, rfp_file.read(), rid))
        db.commit()
        flash('RFP updated.', 'success')
        return redirect(url_for('rfp_detail', rid=rid))
    return render_template('rfp_form.html', rfp=rfp, all_statuses=RFP_STATUSES)


@app.route('/rfp/<int:rid>/import-crf', methods=['POST'])
def rfp_import_crf(rid):
    from pickup_utils import parse_crf_excel
    db = get_db()
    rfp = db.execute("SELECT * FROM rfp WHERE id=?", (rid,)).fetchone()
    if not rfp:
        flash('RFP not found.', 'error')
        return redirect(url_for('rfp_dashboard'))
    f = request.files.get('crf_file')
    if not f or not f.filename:
        flash('No file selected.', 'error')
        return redirect(url_for('rfp_detail', rid=rid))
    file_bytes = f.read()
    crf_filename = f.filename
    db.execute("UPDATE rfp SET crf_filename=?, crf_data=?, updated_at=datetime('now') WHERE id=?",
               (crf_filename, file_bytes, rid))
    db.commit()
    try:
        result = parse_crf_excel(file_bytes)
    except Exception as e:
        flash(f'Error parsing CRF: {e}', 'error')
        return redirect(url_for('rfp_detail', rid=rid))
    # Fill blank meta fields from CRF
    meta = result.get('rfp_meta', {})
    updates, vals = [], []
    for field in ('event_name', 'response_due_date', 'decision_due_date'):
        if meta.get(field) and not rfp[field]:
            updates.append(f'{field}=?')
            vals.append(meta[field])
    if updates:
        vals.append(rid)
        db.execute(f"UPDATE rfp SET {', '.join(updates)}, updated_at=datetime('now') WHERE id=?", vals)
        db.commit()
    hotels = result.get('hotels', [])
    n_proposals = sum(1 for h in hotels if h['status'] == 'proposal_received')
    n_declined  = sum(1 for h in hotels if h['status'] == 'declined')
    return render_template('rfp_crf_review.html', rfp=rfp, hotels=hotels,
                           n_proposals=n_proposals, n_declined=n_declined,
                           crf_filename=crf_filename)


@app.route('/rfp/<int:rid>/import-crf/confirm', methods=['POST'])
def rfp_import_crf_confirm(rid):
    from pickup_utils import parse_crf_excel
    db = get_db()
    rfp = db.execute("SELECT * FROM rfp WHERE id=?", (rid,)).fetchone()
    if not rfp:
        flash('RFP not found.', 'error')
        return redirect(url_for('rfp_dashboard'))
    crf_data = rfp['crf_data'] if rfp['crf_data'] else None
    if not crf_data:
        flash('No CRF file stored — please re-upload the CRF.', 'error')
        return redirect(url_for('rfp_detail', rid=rid))
    try:
        hotels = parse_crf_excel(bytes(crf_data)).get('hotels', [])
    except Exception as e:
        flash(f'Error re-parsing CRF: {e}', 'error')
        return redirect(url_for('rfp_detail', rid=rid))

    def _fv(s):
        try: return float(s) if s not in (None,'','None') else None
        except: return None
    def _iv(s):
        try: return int(s) if s not in (None,'','None') else None
        except: return None

    inserted = updated = 0
    any_proposal = False
    _GENERIC = {
        'hotel','hotels','resort','resorts','suites','suite','inn','inns',
        'motel','lodge','lodges','hilton','marriott','hyatt','sheraton',
        'westin','omni','embassy','extended','stay','holiday','best','western',
        'comfort','quality','sleep','springhill','the','and','spa','conference',
        'center','collection','autograph','renaissance','tapestry','curio',
        'home2','homewood','hampton','doubletree','courtyard','fairfield',
        'towneplace','graduate','thompson','kimpton','indigo','vignette','by','at','of',
    }

    for idx, h in enumerate(hotels):
        if '1' not in request.form.getlist(f'include_{idx}'):
            continue
        hotel_name = (h.get('hotel_name') or '').strip()
        if not hotel_name:
            continue
        status = h.get('status', 'pending')
        if status == 'proposal_received':
            any_proposal = True
        proposed_rate   = _fv(request.form.get(f'rate_{idx}'))  or _fv(h.get('proposed_rate'))
        f_and_b_minimum = _fv(request.form.get(f'fab_{idx}'))   or _fv(h.get('f_and_b_minimum'))
        comm_raw        = request.form.get(f'comm_{idx}')
        commission_pct  = (_fv(comm_raw) / 100.0 if _fv(comm_raw) else None) or _fv(h.get('commission_pct'))
        notes_val       = request.form.get(f'notes_{idx}','').strip() or h.get('notes') or None
        city  = h.get('city') or None
        state = h.get('state') or None
        crf_row_data = h.get('crf_row_data') or None

        # Match existing hotel by exact name, then word overlap
        existing = db.execute(
            "SELECT id, crf_version FROM rfp_hotel WHERE rfp_id=? AND LOWER(hotel_name)=LOWER(?)",
            (rid, hotel_name)
        ).fetchone()
        if not existing:
            all_hotels = db.execute("SELECT id, hotel_name, crf_version FROM rfp_hotel WHERE rfp_id=?", (rid,)).fetchall()
            h_words = set(w.lower() for w in hotel_name.split() if len(w) > 3 and w.lower() not in _GENERIC)
            for ah in all_hotels:
                ah_words = set(w.lower() for w in ah['hotel_name'].split() if len(w) > 3 and w.lower() not in _GENERIC)
                if h_words and ah_words and (h_words & ah_words):
                    existing = ah
                    break

        new_ver = (existing['crf_version'] or 0) + 1 if existing else 1
        if existing:
            db.execute('''
                UPDATE rfp_hotel SET
                    city=COALESCE(?,city), state=COALESCE(?,state),
                    contact_name=COALESCE(?,contact_name),
                    contact_email=COALESCE(?,contact_email),
                    contact_phone=COALESCE(?,contact_phone),
                    contact_title=COALESCE(?,contact_title),
                    status=?, proposed_rate=?, commission_pct=?,
                    f_and_b_minimum=?, attrition_pct=?, cutoff_days=?,
                    concessions=?, notes=COALESCE(?,notes),
                    crf_row_data=?, crf_version=?, updated_at=datetime('now')
                WHERE id=?
            ''', (city, state,
                  h.get('contact_name') or None, h.get('contact_email') or None,
                  h.get('contact_phone') or None, h.get('contact_title') or None,
                  status, proposed_rate, commission_pct, f_and_b_minimum,
                  _fv(h.get('attrition_pct')), _iv(h.get('cutoff_days')),
                  h.get('concessions') or None, notes_val,
                  crf_row_data, new_ver, existing['id']))
            updated += 1
        else:
            db.execute('''
                INSERT INTO rfp_hotel (rfp_id, hotel_name, city, state,
                    contact_name, contact_email, contact_phone, contact_title,
                    status, proposed_rate, commission_pct, f_and_b_minimum,
                    attrition_pct, cutoff_days, concessions, notes,
                    crf_row_data, crf_version, updated_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,datetime('now'))
            ''', (rid, hotel_name, city, state,
                  h.get('contact_name') or None, h.get('contact_email') or None,
                  h.get('contact_phone') or None, h.get('contact_title') or None,
                  status, proposed_rate, commission_pct, f_and_b_minimum,
                  _fv(h.get('attrition_pct')), _iv(h.get('cutoff_days')),
                  h.get('concessions') or None, notes_val,
                  crf_row_data, new_ver))
            inserted += 1

    if any_proposal and rfp['status'] == 'sourcing':
        db.execute("UPDATE rfp SET status='proposals_received', updated_at=datetime('now') WHERE id=?", (rid,))
    db.commit()
    parts = []
    if inserted: parts.append(f'{inserted} hotel{"s" if inserted!=1 else ""} added')
    if updated:  parts.append(f'{updated} hotel{"s" if updated!=1 else ""} updated')
    flash((', '.join(parts) + '.') if parts else 'No changes made.', 'success' if parts else 'info')
    return redirect(url_for('rfp_detail', rid=rid))


@app.route('/rfp/parse-document', methods=['POST'])
def rfp_parse_document():
    """AJAX — parse an uploaded Cvent RFP .docx and return field values as JSON."""
    from pickup_utils import parse_rfp_docx
    f = request.files.get('rfp_file')
    if not f or not f.filename.lower().endswith('.docx'):
        return jsonify({})
    try:
        return jsonify(parse_rfp_docx(f.read()))
    except Exception as e:
        return jsonify({'_error': str(e)})


def _build_critical_dates_email(rfp, client_email='', hotel_name=''):
    """Return (subject, body) for the critical-dates client email."""
    event_name = rfp['event_name'] or rfp['client_org'] or 'Your Event'
    start_date = rfp['start_date'] or ''
    if start_date:
        try:
            from datetime import datetime as _dta
            start_date_fmt = _dta.strptime(start_date, '%Y-%m-%d').strftime('%-m/%-d/%Y')
        except Exception:
            start_date_fmt = start_date
    else:
        start_date_fmt = ''
    subject = event_name
    if start_date_fmt:
        subject += f' — {start_date_fmt}'

    raw_cd = rfp['critical_dates_json'] if rfp['critical_dates_json'] else '[]'
    try:
        critical_dates = sorted(json.loads(raw_cd), key=lambda x: x.get('date', ''))
    except Exception:
        critical_dates = []

    hotel_phrase = f'the {hotel_name}' if hotel_name else 'your contracted hotel'
    body_lines = [
        'Hi,',
        '',
        f'I wanted to share some important dates related to your upcoming event at {hotel_phrase}.',
        'Please add these to your calendar and set reminders in Outlook so nothing is missed.',
        '',
        '── CRITICAL CONTRACT DATES ──',
        '',
    ]
    if critical_dates:
        for cd in critical_dates:
            date_str = cd.get('date', '')
            try:
                from datetime import datetime as _dtb
                date_str = _dtb.strptime(date_str, '%Y-%m-%d').strftime('%-m/%-d/%Y')
            except Exception:
                pass
            label  = cd.get('label', '')
            amount = cd.get('amount')
            line   = f'  {date_str}  —  {label}'
            if amount:
                try:
                    line += f'  (${float(amount):,.2f})'
                except Exception:
                    line += f'  ({amount})'
            body_lines.append(line)
    else:
        body_lines.append('  (No critical dates on file — upload the signed contract to extract them)')

    body_lines += [
        '',
        '── OUTLOOK REMINDER TIP ──',
        '',
        'For each date above, we recommend creating a calendar reminder in Outlook at least 2 weeks',
        'in advance so you have time to take action if needed.',
        '',
        'A copy of the signed contract is attached for your reference.',
        '',
        'Please reach out if you have any questions.',
    ]
    body = '\r\n'.join(body_lines)
    return subject, body


def _rfp_email_lookup(db, rfp):
    """Return (client_email, hotel_name) for an rfp row."""
    client_email = ''
    hotel_name   = ''
    if rfp['booking_id']:
        pc = db.execute(
            "SELECT group_contact_email, hotel FROM pickup_config "
            "WHERE booking_id=? LIMIT 1", (rfp['booking_id'],)
        ).fetchone()
        if pc:
            client_email = pc['group_contact_email'] or ''
            hotel_name   = pc['hotel'] or ''
    if not hotel_name:
        sel = db.execute(
            "SELECT hotel_name FROM rfp_hotel WHERE rfp_id=? AND status='selected' "
            "ORDER BY updated_at DESC LIMIT 1", (rfp['id'],)
        ).fetchone()
        if sel:
            hotel_name = sel['hotel_name'] or ''
    return client_email, hotel_name


@app.route('/rfp/<int:rid>/email-dates')
def rfp_client_dates_email(rid):
    """Show email preview page."""
    db  = get_db()
    rfp = db.execute("SELECT * FROM rfp WHERE id=?", (rid,)).fetchone()
    if not rfp:
        db.close()
        flash('RFP not found.', 'error')
        return redirect(url_for('rfp_dashboard'))
    client_email, hotel_name = _rfp_email_lookup(db, rfp)
    db.close()

    subject, body = _build_critical_dates_email(rfp, client_email, hotel_name)
    raw_cd = rfp['critical_dates_json'] if rfp['critical_dates_json'] else '[]'
    try:
        critical_dates = sorted(json.loads(raw_cd), key=lambda x: x.get('date', ''))
    except Exception:
        critical_dates = []

    return render_template('rfp_email_dates.html',
                           rfp=rfp,
                           subject=subject,
                           body=body,
                           critical_dates=critical_dates,
                           client_email=client_email,
                           hotel_name=hotel_name)


@app.route('/rfp/<int:rid>/email-dates/launch', methods=['POST'])
def rfp_dates_email_launch(rid):
    """Open Outlook compose window with body + contract attached via osascript/PowerShell."""
    import subprocess, tempfile, platform, os as _os

    db  = get_db()
    rfp = db.execute("SELECT * FROM rfp WHERE id=?", (rid,)).fetchone()
    if not rfp:
        db.close()
        flash('RFP not found.', 'error')
        return redirect(url_for('rfp_dashboard'))
    client_email, hotel_name = _rfp_email_lookup(db, rfp)
    contract_data     = bytes(rfp['contract_data']) if rfp['contract_data'] else None
    contract_filename = rfp['contract_filename'] or 'contract.pdf'
    # Mark as sent at the same time
    db.execute(
        "UPDATE rfp SET critical_dates_sent_at=datetime('now'), updated_at=datetime('now') WHERE id=?",
        (rid,)
    )
    db.commit()
    db.close()

    subject, body = _build_critical_dates_email(rfp, client_email, hotel_name)

    def write_tmp(content, suffix, mode='w', encoding='utf-8'):
        t = tempfile.NamedTemporaryFile(mode=mode, suffix=suffix, delete=False,
                                        encoding=encoding if mode == 'w' else None)
        t.write(content); t.close(); return t.name

    def write_named_att(data, filename):
        d = tempfile.mkdtemp()
        p = _os.path.join(d, filename)
        with open(p, 'wb') as fh:
            fh.write(data)
        return p

    try:
        if platform.system() == 'Windows':
            ps_lines = [
                f'$body = Get-Content -Path \'{write_tmp(body, ".txt")}\' -Raw -Encoding UTF8',
                '$ol = New-Object -ComObject Outlook.Application',
                '$mail = $ol.CreateItem(0)',
                f'$mail.Subject = \'{subject.replace(chr(39), chr(39)*2)}\'',
                '$mail.Body = $body',
            ]
            if client_email:
                ps_lines.append(f'$mail.To = \'{client_email.replace(chr(39), chr(39)*2)}\'')
            if contract_data:
                att_path = write_named_att(contract_data, contract_filename).replace('\\', '\\\\')
                ps_lines.append(f'$mail.Attachments.Add(\'{att_path}\') | Out-Null')
            ps_lines.append('$mail.Display()')
            ps_path = write_tmp('\r\n'.join(ps_lines), '.ps1')
            subprocess.Popen(['powershell.exe', '-ExecutionPolicy', 'Bypass',
                               '-WindowStyle', 'Hidden', '-File', ps_path])
        else:
            # macOS: write body to temp file, JXA reads it onto clipboard as plain text,
            # AppleScript opens Outlook with To/Subject/attachment then Cmd+V pastes body.
            def esc(s):
                return s.replace('\\', '\\\\').replace('"', '\\"')

            tmp_txt = write_tmp(body, '.txt')

            clip_jxa = (
                "ObjC.import('AppKit'); ObjC.import('Foundation');\n"
                f"var nsStr = $.NSString.alloc.initWithContentsOfFileEncodingError('{tmp_txt}', $.NSUTF8StringEncoding, null);\n"
                "var txt = ObjC.unwrap(nsStr);\n"
                "var pb = $.NSPasteboard.generalPasteboard; pb.clearContents;\n"
                "pb.setStringForType($.NSString.alloc.initWithUTF8String(txt), $.NSPasteboardTypeString);"
            )

            attach_line = ''
            if contract_data:
                att_path = write_named_att(contract_data, contract_filename)
                attach_line = f'make new attachment at theMsg with properties {{file:POSIX file "{esc(att_path)}"}}'

            to_line = (
                f'make new to recipient at theMsg with properties {{email address:{{name:"", address:"{esc(client_email)}"}}}}'
                if client_email else ''
            )

            outlook_script = (
                'tell application "Microsoft Outlook"\n'
                f'    set theMsg to make new outgoing message with properties {{subject:"{esc(subject)}"}}\n'
                + (f'    {to_line}\n' if to_line else '')
                + (f'    {attach_line}\n' if attach_line else '')
                + '    open theMsg\n'
                + '    activate\n'
                + 'end tell\n'
                + 'delay 4\n'
                + 'tell application "Microsoft Outlook"\n'
                + '    activate\n'
                + 'end tell\n'
                + 'delay 1\n'
                + 'tell application "System Events"\n'
                + '    keystroke "v" using {command down}\n'
                + 'end tell\n'
            )

            clip_path    = write_tmp(clip_jxa,      '.js')
            outlook_path = write_tmp(outlook_script, '.applescript')
            subprocess.run(['osascript', '-l', 'JavaScript', clip_path], check=True)
            subprocess.Popen(['osascript', outlook_path])

    except Exception as exc:
        flash(f'Could not launch Outlook: {exc}', 'error')
        return redirect(url_for('rfp_client_dates_email', rid=rid))

    att_note = f' with {contract_filename} attached' if contract_data else ''
    flash(f'Email opened in Outlook{att_note}.', 'success')
    return redirect(url_for('rfp_detail', rid=rid))


@app.route('/rfp/<int:rid>/mark-dates-sent', methods=['POST'])
def rfp_mark_dates_sent(rid):
    """Record that the critical-dates email was sent (manual override)."""
    db = get_db()
    db.execute(
        "UPDATE rfp SET critical_dates_sent_at=datetime('now'), updated_at=datetime('now') WHERE id=?",
        (rid,)
    )
    db.commit()
    db.close()
    flash('Email marked as sent.', 'success')
    return redirect(url_for('rfp_detail', rid=rid))


@app.route('/rfp/<int:rid>/rescan-contract-dates')
def rfp_rescan_contract_dates(rid):
    """Re-parse the stored contract to extract critical dates."""
    from pickup_utils import parse_contract_document
    db  = get_db()
    rfp = db.execute(
        "SELECT contract_data, contract_filename FROM rfp WHERE id=?", (rid,)
    ).fetchone()
    if not rfp or not rfp['contract_data']:
        db.close()
        flash('No contract file stored — upload the signed contract first.', 'error')
        return redirect(url_for('rfp_detail', rid=rid))

    file_bytes = bytes(rfp['contract_data'])
    filename   = rfp['contract_filename'] or 'contract.pdf'
    db.close()

    extracted = parse_contract_document(file_bytes, filename=filename)
    critical_dates = extracted.get('critical_dates', [])

    if not critical_dates:
        flash('No critical dates found in the contract. You can add them manually by re-importing the contract.', 'warning')
        return redirect(url_for('rfp_detail', rid=rid))

    db2 = get_db()
    db2.execute(
        "UPDATE rfp SET critical_dates_json=?, updated_at=datetime('now') WHERE id=?",
        (json.dumps(critical_dates), rid)
    )
    db2.commit()
    db2.close()
    flash(f'Found {len(critical_dates)} critical date(s) from the stored contract.', 'success')
    return redirect(url_for('rfp_detail', rid=rid))


@app.route('/rfp/<int:rid>/archive', methods=['POST'])
def rfp_archive(rid):
    db = get_db()
    rfp = db.execute("SELECT contract_filename FROM rfp WHERE id=?", (rid,)).fetchone()
    if not rfp or not rfp['contract_filename']:
        db.close()
        flash('Cannot archive: no signed contract has been uploaded for this RFP.', 'error')
        return redirect(url_for('rfp_detail', rid=rid))
    db.execute("UPDATE rfp SET archived=1 WHERE id=?", (rid,))
    db.commit()
    flash('RFP archived.', 'success')
    return redirect(url_for('rfp_dashboard'))


@app.route('/rfp/<int:rid>/unarchive', methods=['POST'])
def rfp_unarchive(rid):
    db = get_db()
    db.execute("UPDATE rfp SET archived=0 WHERE id=?", (rid,))
    db.commit()
    flash('RFP restored.', 'success')
    return redirect(url_for('rfp_dashboard', archived=1))


@app.route('/rfp/<int:rid>/hotel/add', methods=['GET', 'POST'])
def rfp_hotel_add(rid):
    db = get_db()
    rfp = db.execute('SELECT * FROM rfp WHERE id=?', (rid,)).fetchone()
    if not rfp:
        flash('RFP not found.', 'error')
        return redirect(url_for('rfp_dashboard'))
    if request.method == 'POST':
        db.execute('''
            INSERT INTO rfp_hotel (rfp_id, hotel_name, brand, city, state,
                contact_name, contact_email, contact_phone, contact_title,
                proposed_rate, commission_pct, f_and_b_minimum, meeting_room_rental,
                attrition_pct, cutoff_days, concessions, notes)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ''', (
            rid,
            request.form.get('hotel_name', '').strip(),
            request.form.get('brand', '').strip() or None,
            request.form.get('city', '').strip() or None,
            request.form.get('state', '').strip() or None,
            request.form.get('contact_name', '').strip() or None,
            request.form.get('contact_email', '').strip() or None,
            request.form.get('contact_phone', '').strip() or None,
            request.form.get('contact_title', '').strip() or None,
            request.form.get('proposed_rate') or None,
            request.form.get('commission_pct') or None,
            request.form.get('f_and_b_minimum') or None,
            request.form.get('meeting_room_rental') or None,
            request.form.get('attrition_pct') or None,
            request.form.get('cutoff_days') or None,
            request.form.get('concessions', '').strip() or None,
            request.form.get('notes', '').strip() or None,
        ))
        db.commit()
        flash('Hotel added.', 'success')
        return redirect(url_for('rfp_detail', rid=rid))
    return render_template('rfp_hotel_form.html', rfp=rfp, hotel=None)


@app.route('/rfp/<int:rid>/hotel/<int:hid>/edit', methods=['GET', 'POST'])
def rfp_hotel_edit(rid, hid):
    db = get_db()
    rfp = db.execute('SELECT * FROM rfp WHERE id=?', (rid,)).fetchone()
    hotel = db.execute('SELECT * FROM rfp_hotel WHERE id=? AND rfp_id=?', (hid, rid)).fetchone()
    if not rfp or not hotel:
        flash('Not found.', 'error')
        return redirect(url_for('rfp_dashboard'))
    if request.method == 'POST':
        db.execute('''
            UPDATE rfp_hotel SET hotel_name=?, brand=?, city=?, state=?,
                contact_name=?, contact_email=?, contact_phone=?, contact_title=?,
                proposed_rate=?, commission_pct=?, f_and_b_minimum=?,
                meeting_room_rental=?, attrition_pct=?, cutoff_days=?,
                concessions=?, notes=?, updated_at=datetime('now')
            WHERE id=? AND rfp_id=?
        ''', (
            request.form.get('hotel_name', '').strip(),
            request.form.get('brand', '').strip() or None,
            request.form.get('city', '').strip() or None,
            request.form.get('state', '').strip() or None,
            request.form.get('contact_name', '').strip() or None,
            request.form.get('contact_email', '').strip() or None,
            request.form.get('contact_phone', '').strip() or None,
            request.form.get('contact_title', '').strip() or None,
            request.form.get('proposed_rate') or None,
            request.form.get('commission_pct') or None,
            request.form.get('f_and_b_minimum') or None,
            request.form.get('meeting_room_rental') or None,
            request.form.get('attrition_pct') or None,
            request.form.get('cutoff_days') or None,
            request.form.get('concessions', '').strip() or None,
            request.form.get('notes', '').strip() or None,
            hid, rid,
        ))
        db.commit()
        flash('Hotel updated.', 'success')
        return redirect(url_for('rfp_detail', rid=rid))
    return render_template('rfp_hotel_form.html', rfp=rfp, hotel=hotel)


@app.route('/rfp/<int:rid>/hotel/<int:hid>/delete', methods=['POST'])
def rfp_hotel_delete(rid, hid):
    db = get_db()
    db.execute('DELETE FROM rfp_hotel WHERE id=? AND rfp_id=?', (hid, rid))
    db.commit()
    flash('Hotel removed.', 'success')
    return redirect(url_for('rfp_detail', rid=rid))


@app.route('/rfp/<int:rid>/hotel/<int:hid>/select', methods=['POST'])
def rfp_hotel_select(rid, hid):
    db = get_db()
    # Select this hotel and update the RFP status
    db.execute("UPDATE rfp_hotel SET status='selected' WHERE id=? AND rfp_id=?", (hid, rid))
    db.execute("UPDATE rfp SET status='hotel_selected', updated_at=datetime('now') WHERE id=?", (rid,))
    # Handle other hotels: remove or mark eliminated
    others_action = request.form.get('others_action', 'keep')
    if others_action == 'remove':
        db.execute("DELETE FROM rfp_hotel WHERE rfp_id=? AND id!=?", (rid, hid))
    else:
        db.execute("UPDATE rfp_hotel SET status='eliminated' WHERE rfp_id=? AND id!=? AND status NOT IN ('selected','declined')", (rid, hid))
    # Optional proposal upload
    pf = request.files.get('proposal_file')
    if pf and pf.filename:
        db.execute(
            "UPDATE rfp_hotel SET proposal_filename=?, proposal_data=?, updated_at=datetime('now') WHERE id=? AND rfp_id=?",
            (pf.filename, pf.read(), hid, rid)
        )
    db.commit()
    # Auto-create / refresh Cost Savings Report for the selected hotel.
    cs_id = _cs_seed_from_selected_hotel(db, rid, hid)
    flash('Hotel selected.', 'success')
    if cs_id:
        flash('Cost Savings Report initialized from CRF — open it under Reports → Cost Savings.', 'success')
    return redirect(url_for('rfp_detail', rid=rid))


@app.route('/rfp/<int:rid>/hotel/<int:hid>/upload-proposal', methods=['POST'])
def rfp_hotel_upload_proposal(rid, hid):
    f = request.files.get('proposal_file')
    if not f or not f.filename:
        flash('No file selected.', 'error')
        return redirect(url_for('rfp_detail', rid=rid))
    db = get_db()
    db.execute(
        "UPDATE rfp_hotel SET proposal_filename=?, proposal_data=?, updated_at=datetime('now') WHERE id=? AND rfp_id=?",
        (f.filename, f.read(), hid, rid)
    )
    db.commit()
    flash(f'Proposal "{f.filename}" uploaded.', 'success')
    return redirect(url_for('rfp_detail', rid=rid))


@app.route('/rfp/<int:rid>/hotel/<int:hid>/proposal/download')
def rfp_hotel_proposal_download(rid, hid):
    import io as _io
    db = get_db()
    h = db.execute('SELECT * FROM rfp_hotel WHERE id=? AND rfp_id=?', (hid, rid)).fetchone()
    if not h or not h['proposal_data']:
        flash('Proposal not found.', 'error')
        return redirect(url_for('rfp_detail', rid=rid))
    return send_file(_io.BytesIO(h['proposal_data']),
                     download_name=h['proposal_filename'] or 'proposal.pdf',
                     as_attachment=True)


@app.route('/rfp/<int:rid>/upload-contract', methods=['POST'])
def rfp_upload_contract(rid):
    """Step 1 — AI-parse an uploaded hotel contract and show a review screen."""
    import uuid, tempfile
    from pickup_utils import parse_contract_document
    db = get_db()
    rfp = db.execute('SELECT * FROM rfp WHERE id=?', (rid,)).fetchone()
    if not rfp:
        flash('RFP not found.', 'error')
        return redirect(url_for('rfp_dashboard'))
    # Save booking_id to rfp record if provided (required by modal)
    booking_id = request.form.get('booking_id', '').strip()
    if not booking_id:
        flash('Booking ID is required before importing a contract.', 'error')
        return redirect(url_for('rfp_detail', rid=rid))
    if booking_id != (rfp['booking_id'] or '').strip():
        db.execute("UPDATE rfp SET booking_id=?, updated_at=datetime('now') WHERE id=?",
                   (booking_id, rid))
        db.commit()
        rfp = db.execute('SELECT * FROM rfp WHERE id=?', (rid,)).fetchone()

    f = request.files.get('contract_file')
    if not f or not f.filename:
        flash('No file selected.', 'error')
        return redirect(url_for('rfp_detail', rid=rid))
    file_bytes = f.read()

    # Find the selected (or best available) hotel to pre-fill contacts
    selected_hotel = db.execute(
        "SELECT * FROM rfp_hotel WHERE rfp_id=? AND status='selected' "
        "ORDER BY updated_at DESC LIMIT 1", (rid,)
    ).fetchone()
    if not selected_hotel:
        selected_hotel = db.execute(
            "SELECT * FROM rfp_hotel WHERE rfp_id=? "
            "AND status NOT IN ('declined','eliminated') "
            "ORDER BY updated_at DESC LIMIT 1", (rid,)
        ).fetchone()

    extracted = parse_contract_document(file_bytes, filename=f.filename)
    err = extracted.get('error') or ''
    # Hard error: no block data at all — redirect back
    if err and not extracted.get('contracted_block') and not extracted.get('years'):
        flash(f'Could not parse contract: {err}', 'error')
        return redirect(url_for('rfp_detail', rid=rid))
    # Soft error: AI failed but direct parse succeeded — show warning on review page
    if err:
        flash(f'Note: {err}', 'warning')

    # Save to temp file so confirm step doesn't need a re-upload
    tmp_id   = str(uuid.uuid4())
    tmp_path = os.path.join(tempfile.gettempdir(), f'rfp_contract_{tmp_id}')
    with open(tmp_path, 'wb') as fh:
        fh.write(file_bytes)

    return render_template('rfp_contract_review.html',
                           rfp=rfp,
                           selected_hotel=selected_hotel,
                           extracted=extracted,
                           filename=f.filename,
                           tmp_id=tmp_id)


@app.route('/rfp/<int:rid>/upload-contract/confirm', methods=['POST'])
def rfp_upload_contract_confirm(rid):
    """Step 2 — save AI-extracted contract data to rfp, hotel, and pickup_config."""
    import tempfile
    from datetime import date as _date
    db = get_db()
    rfp = db.execute('SELECT * FROM rfp WHERE id=?', (rid,)).fetchone()
    if not rfp:
        flash('RFP not found.', 'error')
        return redirect(url_for('rfp_dashboard'))

    # ── Parse form ────────────────────────────────────────────────────────────
    raw_block    = request.form.get('contracted_block', '{}')
    block        = json.loads(raw_block)
    rate_str     = request.form.get('contracted_rate', '').strip()
    cutoff_str   = request.form.get('cutoff_date', '').strip()
    atr_str      = request.form.get('attrition_pct', '').strip()
    hotel_contact        = request.form.get('hotel_contact', '').strip() or None
    hotel_contact_email  = request.form.get('hotel_contact_email', '').strip() or None
    hotel_contact2       = request.form.get('hotel_contact2', '').strip() or hotel_contact
    hotel_contact2_email = request.form.get('hotel_contact2_email', '').strip() or hotel_contact_email
    group_contact        = request.form.get('group_contact', '').strip() or None
    group_contact_email  = request.form.get('group_contact_email', '').strip() or None
    critical_dates_raw   = request.form.get('critical_dates_json', '[]').strip()
    # Validate JSON; default to empty list on failure
    try:
        json.loads(critical_dates_raw)
    except Exception:
        critical_dates_raw = '[]'

    contracted_rate = float(rate_str) if rate_str else None
    attrition_pct   = float(atr_str)  if atr_str  else None

    # ── Derive rfp-level fields from block ────────────────────────────────────
    dates             = sorted(block.keys())
    start_date        = dates[0]  if dates else None
    end_date          = dates[-1] if dates else None
    peak_rooms        = max(block.values()) if block else None
    total_room_nights = sum(block.values()) if block else None

    cutoff_days = None
    if cutoff_str and start_date:
        try:
            cutoff_days = (_date.fromisoformat(start_date) -
                           _date.fromisoformat(cutoff_str)).days
        except Exception:
            pass

    # ── File blob from temp file written in step 1 ────────────────────────────
    contract_filename = request.form.get('_contract_filename', '')
    tmp_id   = request.form.get('_tmp_id', '')
    tmp_path = os.path.join(tempfile.gettempdir(), f'rfp_contract_{tmp_id}') if tmp_id else ''
    file_blob = None
    if tmp_path and os.path.exists(tmp_path):
        with open(tmp_path, 'rb') as fh:
            file_blob = fh.read()
        try:
            os.remove(tmp_path)
        except Exception:
            pass

    # ── Update rfp record ─────────────────────────────────────────────────────
    db.execute('''
        UPDATE rfp SET
            start_date          = COALESCE(?, start_date),
            end_date            = COALESCE(?, end_date),
            peak_rooms          = COALESCE(?, peak_rooms),
            total_room_nights   = COALESCE(?, total_room_nights),
            contract_filename   = ?,
            contract_data       = ?,
            critical_dates_json = ?,
            status              = CASE WHEN status NOT IN ('contracted','dead')
                                       THEN 'contracted' ELSE status END,
            updated_at          = datetime('now')
        WHERE id=?
    ''', (start_date, end_date, peak_rooms, total_room_nights,
          contract_filename or None, file_blob,
          critical_dates_raw if critical_dates_raw != '[]' else None,
          rid))

    # ── Update selected hotel record ──────────────────────────────────────────
    selected_hotel = db.execute(
        "SELECT * FROM rfp_hotel WHERE rfp_id=? AND status='selected' "
        "ORDER BY updated_at DESC LIMIT 1", (rid,)
    ).fetchone()
    if not selected_hotel:
        selected_hotel = db.execute(
            "SELECT * FROM rfp_hotel WHERE rfp_id=? "
            "AND status NOT IN ('declined','eliminated') "
            "ORDER BY updated_at DESC LIMIT 1", (rid,)
        ).fetchone()
    # If still no hotel but the contract named one, create it automatically
    extracted_hotel_name = request.form.get('_extracted_hotel', '').strip()
    if not selected_hotel and extracted_hotel_name:
        db.execute('''
            INSERT INTO rfp_hotel (rfp_id, hotel_name, status, proposed_rate, attrition_pct,
                                   cutoff_days, contact_name, contact_email, updated_at)
            VALUES (?, ?, 'selected', ?, ?, ?, ?, ?, datetime('now'))
        ''', (rid, extracted_hotel_name, contracted_rate, attrition_pct, cutoff_days,
              hotel_contact, hotel_contact_email))
        db.commit()
        selected_hotel = db.execute(
            "SELECT * FROM rfp_hotel WHERE rfp_id=? ORDER BY id DESC LIMIT 1", (rid,)
        ).fetchone()
    if selected_hotel:
        db.execute('''
            UPDATE rfp_hotel SET
                proposed_rate = COALESCE(?, proposed_rate),
                attrition_pct = COALESCE(?, attrition_pct),
                cutoff_days   = COALESCE(?, cutoff_days),
                contact_name  = COALESCE(?, contact_name),
                contact_email = COALESCE(?, contact_email),
                updated_at    = datetime('now')
            WHERE id=?
        ''', (contracted_rate, attrition_pct, cutoff_days,
              hotel_contact, hotel_contact_email, selected_hotel['id']))

    # ── Update or create pickup_config ────────────────────────────────────────
    pickup_updated = False
    if block and rfp['booking_id']:
        existing_pc = db.execute(
            'SELECT id FROM pickup_config WHERE CAST(booking_id AS INTEGER)=CAST(? AS INTEGER)',
            (rfp['booking_id'],)
        ).fetchone()
        hotel_name = selected_hotel['hotel_name'] if selected_hotel else (rfp['event_name'] or '')
        org        = rfp['client_org'] or ''
        event_name = rfp['event_name'] or ''
        cutoff_iso = cutoff_str or None

        if existing_pc:
            db.execute('''
                UPDATE pickup_config SET
                    contracted_block     = ?,
                    contracted_rate      = COALESCE(?, contracted_rate),
                    cutoff_date          = COALESCE(?, cutoff_date),
                    attrition_pct        = COALESCE(?, attrition_pct),
                    block_is_estimated   = 0,
                    contract_filename    = COALESCE(?, contract_filename),
                    contract_data        = COALESCE(?, contract_data),
                    hotel_contact        = COALESCE(?, hotel_contact),
                    hotel_contact_email  = COALESCE(?, hotel_contact_email),
                    hotel_contact2       = COALESCE(?, hotel_contact2),
                    hotel_contact2_email = COALESCE(?, hotel_contact2_email),
                    group_contact        = COALESCE(?, group_contact),
                    group_contact_email  = COALESCE(?, group_contact_email)
                WHERE id=?
            ''', (json.dumps(block), contracted_rate, cutoff_iso, attrition_pct,
                  contract_filename or None, file_blob,
                  hotel_contact, hotel_contact_email,
                  hotel_contact2, hotel_contact2_email,
                  group_contact, group_contact_email,
                  existing_pc['id']))
            pickup_updated = True
        else:
            from datetime import timedelta as _td
            cutoff_default = (
                cutoff_iso if cutoff_iso else
                ((_date.fromisoformat(start_date) - _td(days=30)).isoformat()
                 if start_date else None)
            )
            db.execute('''
                INSERT INTO pickup_config
                    (booking_id, organization, event_name, hotel,
                     contracted_block, contracted_rate, cutoff_date,
                     attrition_pct, status, block_is_estimated,
                     contract_filename, contract_data,
                     hotel_contact, hotel_contact_email,
                     hotel_contact2, hotel_contact2_email,
                     group_contact, group_contact_email,
                     hotel_contacts, cc_emails, force_current, force_past)
                VALUES (?,?,?,?,?,?,?,?,'active',0,?,?,?,?,?,?,?,?,'[]','[]',0,0)
            ''', (str(rfp['booking_id']), org, event_name, hotel_name,
                  json.dumps(block), contracted_rate, cutoff_default,
                  attrition_pct or 0.80,
                  contract_filename or None, file_blob,
                  hotel_contact, hotel_contact_email,
                  hotel_contact2, hotel_contact2_email,
                  group_contact, group_contact_email))
            pickup_updated = True

    # Cascade rich extraction → cost_savings_report (via booking link)
    cs_flash = None
    if file_blob and rfp['booking_id']:
        cs_res = cascade_contract_cost_savings(
            db, file_blob, contract_filename,
            booking_id=rfp['booking_id'],
            pickup_config_id=None,
        )
        if cs_res and not cs_res.get('error'):
            cs_flash = 'Cost Savings reports updated with contract values.'

    db.commit()
    msg = 'Contract saved — RFP and hotel updated.'
    if pickup_updated:
        msg += ' Pickup tracking updated with real block data.'
    flash(msg, 'success')
    if cs_flash:
        flash(cs_flash, 'success')
    return redirect(url_for('rfp_detail', rid=rid))


@app.route('/rfp/<int:rid>/contract/download')
def rfp_contract_download(rid):
    import io as _io
    db = get_db()
    rfp = db.execute('SELECT contract_filename, contract_data FROM rfp WHERE id=?', (rid,)).fetchone()
    if not rfp or not rfp['contract_data']:
        flash('Contract not found.', 'error')
        return redirect(url_for('rfp_detail', rid=rid))
    return send_file(_io.BytesIO(rfp['contract_data']),
                     download_name=rfp['contract_filename'] or 'contract.pdf',
                     as_attachment=True)


@app.route('/rfp/<int:rid>/upload-rfp', methods=['POST'])
def rfp_upload_document(rid):
    f = request.files.get('rfp_file')
    if not f or not f.filename:
        flash('No file selected.', 'error')
        return redirect(url_for('rfp_detail', rid=rid))
    db = get_db()
    db.execute("UPDATE rfp SET rfp_filename=?, rfp_data=?, updated_at=datetime('now') WHERE id=?",
               (f.filename, f.read(), rid))
    db.commit()
    flash(f'RFP document "{f.filename}" uploaded.', 'success')
    return redirect(url_for('rfp_detail', rid=rid))


@app.route('/rfp/<int:rid>/rfp/download')
def rfp_document_download(rid):
    import io as _io
    db = get_db()
    rfp = db.execute('SELECT * FROM rfp WHERE id=?', (rid,)).fetchone()
    if not rfp or not rfp['rfp_data']:
        flash('RFP document not found.', 'error')
        return redirect(url_for('rfp_detail', rid=rid))
    return send_file(_io.BytesIO(rfp['rfp_data']),
                     download_name=rfp['rfp_filename'] or 'rfp.pdf',
                     as_attachment=True)


@app.route('/rfp/<int:rid>/note/add', methods=['POST'])
def rfp_note_add(rid):
    db = get_db()
    note_text = request.form.get('note_text', '').strip()
    if note_text:
        db.execute(
            "INSERT INTO rfp_note (rfp_id, note_date, note_type, note_text) VALUES (?,date('now'),?,?)",
            (rid, request.form.get('note_type', 'internal'), note_text)
        )
        db.commit()
    return redirect(url_for('rfp_detail', rid=rid))


@app.route('/rfp/<int:rid>/note/<int:nid>/delete', methods=['POST'])
def rfp_note_delete(rid, nid):
    db = get_db()
    db.execute('DELETE FROM rfp_note WHERE id=? AND rfp_id=?', (nid, rid))
    db.commit()
    return redirect(url_for('rfp_detail', rid=rid))


# ── AI Assistant ──────────────────────────────────────────────────────────────

try:
    import anthropic as _anthropic
    from config import ANTHROPIC_API_KEY as _ANTHROPIC_KEY
    _ANTHROPIC_AVAILABLE = bool(_ANTHROPIC_KEY and _ANTHROPIC_KEY.strip())
except Exception:
    _anthropic = None
    _ANTHROPIC_KEY = ''
    _ANTHROPIC_AVAILABLE = False

_ASSISTANT_SYSTEM = """You are Kristin's personal business assistant for CPAinc, a conference planning and booking company.
You have access to the booking database (SQLite).

Your tools let you:
- query_database: Run read-only SQL against the bookings database

Key tables:
- ReportPipeline: BookingId, BookingName, EventName, AccountName, Customer (hotel), BookingAssociate,
  StartDate, EndDate, BookingStatus, Revenue, USDRevenue, CommissionPercent, USDCommissionableAmount,
  PeakRooms, TotalRoomNights, RoomRate, ContractedAmount, City, State, Country
- ChkRegNote: ChkRegID, BookingID, FinalPayment, Check_, DateOnCheck, DepositDate, EntryDate, Cancelled
- Pickup: rowid, BookingID, Brand, ActualPickup, TotalRevenue, EntryDate
- pickup_config: id, booking_id, organization, event_name, hotel, contracted_block, contracted_rate, status, cutoff_date

Today's date is {today}.

Be concise and helpful. Format numbers as currency where appropriate.
"""

_ASSISTANT_TOOLS = [
    {
        'name': 'query_database',
        'description': 'Run a read-only SQL query against the CPAinc SQLite database.',
        'input_schema': {
            'type': 'object',
            'properties': {
                'sql': {'type': 'string', 'description': 'A valid SQLite SELECT statement'},
            },
            'required': ['sql'],
        },
    },
]


def _run_assistant_tool(tool_name, tool_input):
    try:
        if tool_name == 'query_database':
            sql = tool_input.get('sql', '')
            if any(kw in sql.upper() for kw in ('INSERT', 'UPDATE', 'DELETE', 'DROP', 'ALTER', 'CREATE')):
                return 'Error: Only SELECT queries are allowed.'
            db = get_db()
            rows = db.execute(sql).fetchall()
            if not rows:
                return 'No results.'
            cols = rows[0].keys()
            lines = ['\t'.join(str(c) for c in cols)]
            for r in rows[:100]:
                lines.append('\t'.join(str(r[c]) if r[c] is not None else '' for c in cols))
            return '\n'.join(lines)
    except Exception as e:
        return f'Error: {e}'


@app.route('/assistant', methods=['GET', 'POST'])
def assistant():
    if not _ANTHROPIC_AVAILABLE:
        flash('AI Assistant is not configured (no API key).', 'error')
        return redirect(url_for('pipeline'))
    if request.method == 'GET':
        return render_template('assistant.html')

    data = request.get_json()
    messages = data.get('messages', [])

    client = _anthropic.Anthropic(api_key=_ANTHROPIC_KEY)
    system_prompt = _ASSISTANT_SYSTEM.format(today=datetime.now().strftime('%B %d, %Y'))
    api_messages = [{'role': m['role'], 'content': m['content']} for m in messages]

    max_iterations = 8
    for _ in range(max_iterations):
        response = client.messages.create(
            model='claude-sonnet-4-6',
            max_tokens=4096,
            system=system_prompt,
            tools=_ASSISTANT_TOOLS,
            messages=api_messages,
        )
        assistant_content = response.content
        api_messages.append({'role': 'assistant', 'content': assistant_content})
        if response.stop_reason == 'tool_use':
            tool_results = []
            for block in assistant_content:
                if block.type == 'tool_use':
                    result = _run_assistant_tool(block.name, block.input)
                    tool_results.append({'type': 'tool_result', 'tool_use_id': block.id, 'content': result})
            api_messages.append({'role': 'user', 'content': tool_results})
        else:
            break

    final_text = ''
    for block in response.content:
        if hasattr(block, 'text'):
            final_text += block.text

    return jsonify({'reply': final_text})


# ── Outlook Integration ───────────────────────────────────────────────────────

def _outlook_redirect_uri():
    return request.host_url.rstrip('/') + '/outlook/callback'


@app.route('/outlook/status')
def outlook_status():
    if not _OUTLOOK_AVAILABLE:
        flash('Outlook connector not available.', 'error')
        return redirect(url_for('pipeline'))
    connected = _oc.is_connected()
    user_info = None
    if connected:
        try:
            user_info = _oc.get_user_info()
        except Exception:
            pass
    return render_template('outlook_status.html', connected=connected, user_info=user_info)


@app.route('/outlook/auth')
def outlook_auth():
    if not _OUTLOOK_AVAILABLE:
        return redirect(url_for('pipeline'))
    import urllib.parse
    try:
        from config import MS_CLIENT_ID, MS_TENANT_ID
        scopes = 'Mail.ReadWrite Mail.Send Calendars.ReadWrite Contacts.ReadWrite User.Read'
        params = {
            'client_id': MS_CLIENT_ID,
            'response_type': 'code',
            'redirect_uri': _outlook_redirect_uri(),
            'scope': scopes,
            'response_mode': 'query',
        }
        auth_url = (f'https://login.microsoftonline.com/{MS_TENANT_ID}/oauth2/v2.0/authorize?'
                    + urllib.parse.urlencode(params))
        return redirect(auth_url)
    except Exception as e:
        flash(f'Could not build login URL: {e}', 'error')
        return redirect(url_for('outlook_status'))


@app.route('/outlook/callback')
def outlook_callback():
    if not _OUTLOOK_AVAILABLE:
        return redirect(url_for('pipeline'))
    code = request.args.get('code')
    error = request.args.get('error_description') or request.args.get('error')
    if error:
        flash(f'Microsoft login failed: {error}', 'error')
        return redirect(url_for('outlook_status'))
    if not code:
        flash('No authorization code received.', 'error')
        return redirect(url_for('outlook_status'))
    ok, err = _oc.exchange_code_for_token(code, _outlook_redirect_uri())
    if ok:
        flash('Connected to Microsoft 365 successfully!', 'success')
    else:
        flash(f'Token exchange failed: {err}', 'error')
    return redirect(url_for('outlook_status'))


@app.route('/outlook/disconnect', methods=['POST'])
def outlook_disconnect():
    if _OUTLOOK_AVAILABLE:
        _oc.disconnect()
    flash('Disconnected from Microsoft 365.', 'success')
    return redirect(url_for('outlook_status'))


@app.route('/outlook/inbox')
def outlook_inbox():
    if not _OUTLOOK_AVAILABLE or not _oc.is_connected():
        flash('Connect to Microsoft 365 first.', 'error')
        return redirect(url_for('outlook_status'))
    search = request.args.get('search', '')
    try:
        emails = _oc.get_emails(count=50, search=search if search else None)
    except Exception as e:
        flash(f'Could not load inbox: {e}', 'error')
        emails = []
    return render_template('outlook_inbox.html', emails=emails, search=search)


@app.route('/outlook/calendar')
def outlook_calendar():
    if not _OUTLOOK_AVAILABLE or not _oc.is_connected():
        flash('Connect to Microsoft 365 first.', 'error')
        return redirect(url_for('outlook_status'))
    try:
        events = _oc.get_calendar_events(days_ahead=60, days_back=7)
    except Exception as e:
        flash(f'Could not load calendar: {e}', 'error')
        events = []
    return render_template('outlook_calendar.html', events=events)


# ── Per-user Microsoft OAuth ──────────────────────────────────────────────────

def _ms_redirect_uri():
    base = request.host_url.rstrip('/')
    # Railway is always https — fix scheme if Flask reports http
    if 'railway.app' in base:
        base = base.replace('http://', 'https://')
    return base + '/auth/microsoft/callback'


@app.route('/auth/microsoft')
def auth_microsoft():
    user = get_current_user()
    if not user:
        return redirect(url_for('login'))
    import urllib.parse, secrets as _secrets
    from config import MS_CLIENT_ID, MS_TENANT_ID
    state = _secrets.token_urlsafe(16)
    session['ms_oauth_state'] = state
    params = {
        'client_id':     MS_CLIENT_ID,
        'response_type': 'code',
        'redirect_uri':  _ms_redirect_uri(),
        'scope':         'Mail.ReadWrite User.Read offline_access',
        'response_mode': 'query',
        'state':         state,
        'login_hint':    user['email'],
    }
    url = (f'https://login.microsoftonline.com/{MS_TENANT_ID}/oauth2/v2.0/authorize?'
           + urllib.parse.urlencode(params))
    return redirect(url)


@app.route('/auth/microsoft/callback')
def auth_microsoft_callback():
    user = get_current_user()
    if not user:
        return redirect(url_for('login'))

    error = request.args.get('error_description') or request.args.get('error')
    if error:
        flash(f'Microsoft login failed: {error}', 'error')
        return redirect(url_for('profile'))

    if request.args.get('state') != session.pop('ms_oauth_state', None):
        flash('Invalid state — please try again.', 'error')
        return redirect(url_for('profile'))

    code = request.args.get('code')
    if not code:
        flash('No authorisation code received from Microsoft.', 'error')
        return redirect(url_for('profile'))

    import time, requests as _req
    from config import MS_CLIENT_ID, MS_TENANT_ID, MS_CLIENT_SECRET

    resp = _req.post(
        f'https://login.microsoftonline.com/{MS_TENANT_ID}/oauth2/v2.0/token',
        data={
            'grant_type':    'authorization_code',
            'client_id':     MS_CLIENT_ID,
            'client_secret': MS_CLIENT_SECRET,
            'code':          code,
            'redirect_uri':  _ms_redirect_uri(),
            'scope':         'Mail.ReadWrite User.Read offline_access',
        }
    )
    if resp.status_code != 200:
        flash(f'Token exchange failed: {resp.text[:200]}', 'error')
        return redirect(url_for('profile'))

    tok           = resp.json()
    access_token  = tok['access_token']
    refresh_token = tok.get('refresh_token', '')
    expires_at    = time.time() + tok.get('expires_in', 3600) - 60

    me = _req.get('https://graph.microsoft.com/v1.0/me',
                  headers={'Authorization': f'Bearer {access_token}'},
                  params={'$select': 'mail,userPrincipalName'})
    ms_email = ''
    if me.status_code == 200:
        d = me.json()
        ms_email = d.get('mail') or d.get('userPrincipalName', '')

    db = get_db()
    db.execute('''
        INSERT INTO UserMicrosoftTokens (user_id, access_token, refresh_token, expires_at, ms_user_email)
        VALUES (?, ?, ?, ?, ?)
        ON CONFLICT(user_id) DO UPDATE SET
            access_token  = excluded.access_token,
            refresh_token = excluded.refresh_token,
            expires_at    = excluded.expires_at,
            ms_user_email = excluded.ms_user_email,
            connected_at  = datetime('now')
    ''', (user['id'], access_token, refresh_token, expires_at, ms_email))
    db.commit()

    flash(f'Microsoft account connected ({ms_email}).', 'success')
    return redirect(url_for('profile'))


@app.route('/auth/microsoft/disconnect', methods=['POST'])
def auth_microsoft_disconnect():
    user = get_current_user()
    if not user:
        return redirect(url_for('login'))
    db = get_db()
    db.execute('DELETE FROM UserMicrosoftTokens WHERE user_id = ?', (user['id'],))
    db.commit()
    flash('Microsoft account disconnected.', 'success')
    return redirect(url_for('profile'))


@app.route('/profile')
def profile():
    user = get_current_user()
    if not user:
        return redirect(url_for('login'))
    db = get_db()
    ms_row = db.execute(
        'SELECT ms_user_email, connected_at FROM UserMicrosoftTokens WHERE user_id = ?',
        (user['id'],)
    ).fetchone()
    return render_template('profile.html', user=user, ms_token=ms_row)



@app.route('/admin/change-log')
def admin_change_log():
    user = get_current_user()
    if not user or user['role'] != 'admin':
        flash('Admin access required.', 'error')
        return redirect(url_for('pipeline'))
    db = get_db()
    filter_cid  = request.args.get('cid', '').strip()
    filter_user = request.args.get('user', '').strip()
    filter_date = request.args.get('date', '').strip()
    page     = max(1, int(request.args.get('page', 1)))
    per_page = 100
    where, params = [], []
    if filter_cid:
        where.append('cl.config_id = ?'); params.append(filter_cid)
    if filter_user:
        where.append('cl.username = ?'); params.append(filter_user)
    if filter_date:
        where.append("DATE(cl.timestamp) = ?"); params.append(filter_date)
    where_sql = ('WHERE ' + ' AND '.join(where)) if where else ''
    total = db.execute(f'SELECT COUNT(*) FROM pickup_change_log cl {where_sql}', params).fetchone()[0]
    logs  = db.execute(
        f'''SELECT cl.*, pc.event_name, pc.hotel
            FROM pickup_change_log cl
            LEFT JOIN pickup_config pc ON pc.id = cl.config_id
            {where_sql}
            ORDER BY cl.id DESC LIMIT ? OFFSET ?''',
        params + [per_page, (page-1)*per_page]
    ).fetchall()
    all_users = [r[0] for r in db.execute(
        "SELECT DISTINCT username FROM pickup_change_log WHERE username IS NOT NULL ORDER BY username"
    ).fetchall()]
    total_pages = max(1, (total + per_page - 1) // per_page)
    return render_template('change_log.html', logs=logs, total=total, page=page,
                           total_pages=total_pages, per_page=per_page,
                           filter_cid=filter_cid, filter_user=filter_user,
                           filter_date=filter_date, all_users=all_users)


@app.route('/admin/run-daily-tasks')
def admin_run_daily_tasks():
    """Called by an external cron service daily. Protected by DAILY_TASK_KEY.
    Sends alert emails for overdue items and a DB backup every Sunday."""
    key = request.args.get('key', '')
    if DAILY_TASK_KEY and key != DAILY_TASK_KEY:
        return 'Forbidden', 403
    if not MAIL_TO:
        return jsonify({'ok': False, 'error': 'MAIL_TO not configured'}), 500

    db      = get_db()
    today   = datetime.today().strftime('%Y-%m-%d')
    log     = []

    # ── 1. Build alert summary ────────────────────────────────────────────────
    configs = db.execute(
        "SELECT * FROM pickup_config WHERE status='active'"
    ).fetchall()
    has_hhr = {str(r[0]) for r in db.execute(
        "SELECT DISTINCT booking_id FROM housing_history_files WHERE booking_id IS NOT NULL"
    ).fetchall()}
    has_history = {r[0] for r in db.execute(
        "SELECT DISTINCT config_id FROM pickup_weekly"
    ).fetchall()}
    latest_pickup = {r[0]: r[1] for r in db.execute(
        "SELECT config_id, MAX(report_date) FROM pickup_weekly GROUP BY config_id"
    ).fetchall()}

    alerts = {'overdue': [], 'cutoff_soon': [], 'ended_no_hhr': []}

    for cfg in configs:
        cid   = cfg['id']
        block = {}
        try:
            block = json.loads(cfg['contracted_block'] or '{}')
        except Exception:
            pass
        dates       = sorted(block.keys())
        event_start = dates[0]  if dates else None
        event_end   = dates[-1] if dates else None
        cutoff      = (cfg['cutoff_date'] or '').strip()
        is_started  = event_start and event_start <= today
        is_ended    = event_end   and event_end   <  today
        is_current  = is_started and not is_ended

        if is_current:
            last_rpt = latest_pickup.get(cid)
            if not last_rpt:
                alerts['overdue'].append(f"  • {cfg['event_name']} @ {cfg['hotel']} — no reports yet")
            else:
                days_since = (datetime.strptime(today, '%Y-%m-%d') -
                              datetime.strptime(last_rpt, '%Y-%m-%d')).days
                if days_since >= 7:
                    alerts['overdue'].append(
                        f"  • {cfg['event_name']} @ {cfg['hotel']} — last report {days_since} days ago ({last_rpt})"
                    )
            if cutoff:
                try:
                    days_to = (datetime.strptime(cutoff, '%Y-%m-%d') -
                               datetime.strptime(today, '%Y-%m-%d')).days
                    if 0 <= days_to <= 7:
                        alerts['cutoff_soon'].append(
                            f"  • {cfg['event_name']} @ {cfg['hotel']} — cutoff in {days_to} day(s) ({cutoff})"
                        )
                except Exception:
                    pass

        if is_ended and str(cfg['booking_id'] or '') not in has_hhr:
            alerts['ended_no_hhr'].append(
                f"  • {cfg['event_name']} @ {cfg['hotel']} — ended {event_end}"
            )

    # ── 2. Build email body ───────────────────────────────────────────────────
    sections = []
    if alerts['overdue']:
        sections.append('<h3 style="color:#c0392b">⚠️ Overdue Pickup Reports</h3><pre style="background:#fdf2f2;padding:12px">'
                        + '\n'.join(alerts['overdue']) + '</pre>')
    if alerts['cutoff_soon']:
        sections.append('<h3 style="color:#e67e22">📅 Cutoff Approaching (within 7 days)</h3><pre style="background:#fef9e7;padding:12px">'
                        + '\n'.join(alerts['cutoff_soon']) + '</pre>')
    if alerts['ended_no_hhr']:
        sections.append('<h3 style="color:#8e44ad">📂 Events Ended — No HHR Uploaded</h3><pre style="background:#f5eef8;padding:12px">'
                        + '\n'.join(alerts['ended_no_hhr']) + '</pre>')

    sent_alert = False
    if sections:
        body = (f'<p>CPAinc daily alert — {today}</p>'
                + ''.join(sections)
                + '<p style="color:#888;font-size:.85em">Sent automatically by CPAinc. '
                  'View the <a href="https://cpainc.up.railway.app/status-board">Status Board</a>.</p>')
        ok, err = send_email(MAIL_TO, f'CPAinc Alerts — {today}', body)
        sent_alert = ok
        log.append(f"Alert email: {'sent' if ok else 'FAILED — ' + str(err)}")
    else:
        log.append('No alerts to send today.')

    # ── 3. Sunday DB backup ───────────────────────────────────────────────────
    import sqlite3 as _sq3, tempfile as _tmp
    sent_backup = False
    if datetime.today().weekday() == 6:   # 6 = Sunday
        try:
            with _tmp.NamedTemporaryFile(suffix='.sqlite', delete=False) as tf:
                tmp_path = tf.name
            src = _sq3.connect(DATABASE)
            dst = _sq3.connect(tmp_path)
            src.backup(dst); src.close(); dst.close()
            with open(tmp_path, 'rb') as f:
                db_bytes = f.read()
            os.unlink(tmp_path)
            backup_name = f'CPAinc_backup_{today}.sqlite'
            ok, err = send_email(
                MAIL_TO,
                f'CPAinc Weekly DB Backup — {today}',
                f'<p>Automated weekly database backup attached ({len(db_bytes)//1024} KB).</p>',
                attachments=[(backup_name, db_bytes)]
            )
            sent_backup = ok
            log.append(f"DB backup email: {'sent' if ok else 'FAILED — ' + str(err)}")
        except Exception as e:
            log.append(f'DB backup error: {e}')
    else:
        log.append('DB backup skipped (not Sunday).')

    return jsonify({'ok': True, 'date': today, 'log': log,
                    'alerts_sent': sent_alert, 'backup_sent': sent_backup})


@app.route('/admin/test-email')
def admin_test_email():
    """Send a test email to verify MAIL_* settings are correct."""
    user = get_current_user()
    if not user or user['role'] != 'admin':
        return 'Forbidden', 403
    to = request.args.get('to') or MAIL_TO
    if not to:
        return 'No recipient — set MAIL_TO env var or pass ?to=email@example.com', 400
    ok, err = send_email(to, 'CPAinc — Test Email',
                         '<p>Your CPAinc email alerts are configured correctly!</p>')
    if ok:
        return f'Test email sent to {to} ✓'
    return f'Failed: {err}', 500


# ════════════════════════════════════════════════════════════════════════════
# Hotel Planner Points — module
# ════════════════════════════════════════════════════════════════════════════

def _get_default_recipient_user_id(db):
    """Return the user_id of the default Hotel Points recipient (Kristin)."""
    row = db.execute("SELECT value FROM Settings WHERE key='hotel_points_default_recipient_user_id'").fetchone()
    if row and row['value']:
        try:
            return int(row['value'])
        except Exception:
            pass
    # Fallback: try to find Kristin
    krow = db.execute(
        "SELECT id FROM Users WHERE LOWER(name) LIKE '%kristin%house%' "
        "OR LOWER(username)='kristin' LIMIT 1"
    ).fetchone()
    return krow['id'] if krow else None


def _get_user_loyalty_profile(db, user_id):
    """Fetch user_loyalty_profile row, falling back to Users.name/email if blank."""
    if not user_id:
        return None
    row = db.execute("SELECT * FROM user_loyalty_profile WHERE user_id=?",
                     (user_id,)).fetchone()
    if not row:
        # Auto-create from Users defaults so the row always exists
        u = db.execute("SELECT id, name, email FROM Users WHERE id=?",
                       (user_id,)).fetchone()
        if not u:
            return None
        db.execute('''INSERT INTO user_loyalty_profile (user_id, full_name, email)
                      VALUES (?, ?, ?)''',
                   (u['id'], u['name'] or '', u['email'] or ''))
        db.commit()
        row = db.execute("SELECT * FROM user_loyalty_profile WHERE user_id=?",
                         (user_id,)).fetchone()
    return row


_CHAIN_TO_NUMBER_COL = {
    'Marriott': 'marriott_number',
    'Hyatt':    'hyatt_number',
    'Hilton':   'hilton_number',
    'IHG':      'ihg_number',
    'Omni':     'omni_number',
    'Choice':   'choice_number',
    'Sonesta':  'sonesta_number',
}


def _get_user_profile(db):
    """Return dict of user profile settings for points form filling.

    In CPAinc, the primary recipient is whichever user is configured as the
    default Hotel Points recipient (Kristin House by default). All forms fill
    that user's name/email/phone into the primary block.
    """
    out = {'user_full_name': '', 'user_email': '', 'user_phone': ''}
    uid = _get_default_recipient_user_id(db)
    if uid:
        prof = _get_user_loyalty_profile(db, uid)
        if prof:
            out['user_full_name'] = prof['full_name'] or ''
            out['user_email']     = prof['email']     or ''
            out['user_phone']     = prof['phone']     or ''
    # Legacy fallback to Settings keys (covers fresh installs before profile set)
    for k in out:
        if not out[k]:
            r = db.execute('SELECT value FROM Settings WHERE key=?', (k,)).fetchone()
            if r and r['value']:
                out[k] = r['value']
    return out


def _get_program(db, chain_name=None, program_id=None):
    if program_id:
        return db.execute('SELECT * FROM hotel_points_program WHERE id=?',
                          (program_id,)).fetchone()
    if chain_name:
        return db.execute('SELECT * FROM hotel_points_program WHERE chain_name=?',
                          (chain_name,)).fetchone()
    return None


def _build_field_values_for_request(db, request_row):
    """Helper: assemble field_values from a request row.

    Primary recipient = default recipient (Kristin by default). The chain-
    specific loyalty number comes from her user_loyalty_profile, with a
    fallback to the program's member_number (legacy behavior).
    """
    from points_utils import build_field_values
    program = _get_program(db, program_id=request_row['program_id'])
    pickup_cfg = None
    if request_row['pickup_config_id']:
        pickup_cfg = db.execute('SELECT * FROM pickup_config WHERE id=?',
                                (request_row['pickup_config_id'],)).fetchone()
    booking = None
    if request_row['booking_id']:
        booking = db.execute('SELECT * FROM ReportPipeline WHERE BookingId=?',
                             (request_row['booking_id'],)).fetchone()
    user_profile = _get_user_profile(db)

    # Override the program's member_number with the default recipient's own
    # per-chain number if set. The program's member_number remains as a
    # fallback for chains the recipient hasn't filled in yet.
    program_d = dict(program) if program else None
    if program_d:
        uid = _get_default_recipient_user_id(db)
        prof = _get_user_loyalty_profile(db, uid) if uid else None
        chain = program_d.get('chain_name')
        col = _CHAIN_TO_NUMBER_COL.get(chain)
        if prof and col and prof[col]:
            program_d['member_number'] = prof[col]

    pickup_cfg_d = dict(pickup_cfg) if pickup_cfg else None
    booking_d    = dict(booking)    if booking    else None
    return build_field_values(user_profile, program_d,
                              dict(request_row), pickup_cfg_d, booking_d), program, pickup_cfg, booking


@app.route('/points')
def points_dashboard():
    """List all hotel points requests across chains."""
    db = get_db()
    today = datetime.today().strftime('%Y-%m-%d')

    requests_rows = db.execute('''
        SELECT r.*, p.chain_name, p.member_number, p.submission_type,
               pc.event_name AS pickup_event_name, pc.hotel AS pickup_hotel,
               pc.organization AS pickup_org,
               pc.event_start AS pickup_event_start,
               pc.event_end   AS pickup_event_end
        FROM hotel_points_request r
        LEFT JOIN hotel_points_program p ON p.id = r.program_id
        LEFT JOIN pickup_config pc       ON pc.id = r.pickup_config_id
        ORDER BY COALESCE(r.points_received_date, r.form_sent_date, r.created_at) DESC
    ''').fetchall()

    kpi = {
        'total':     len(requests_rows),
        'pending':   sum(1 for r in requests_rows if r['status'] == 'pending'),
        'submitted': sum(1 for r in requests_rows if r['status'] == 'submitted'),
        'received':  sum(1 for r in requests_rows if r['status'] == 'received'),
        'overdue':   0,
        'points_ytd': 0,
    }
    current_year = datetime.today().strftime('%Y')
    for r in requests_rows:
        if r['points_awarded'] and r['points_received_date'] and \
           str(r['points_received_date']).startswith(current_year):
            kpi['points_ytd'] += int(r['points_awarded'] or 0)
        ev_end = r['pickup_event_end']
        if ev_end and r['status'] not in ('received', 'cancelled', 'disallowed'):
            try:
                d = datetime.strptime(str(ev_end)[:10], '%Y-%m-%d')
                if (datetime.today() - d).days > 60:
                    kpi['overdue'] += 1
            except Exception:
                pass

    by_chain = {}
    for r in requests_rows:
        cn = r['chain_name'] or 'Other'
        if cn not in by_chain:
            by_chain[cn] = {'count': 0, 'points': 0, 'received': 0}
        by_chain[cn]['count'] += 1
        if r['status'] == 'received':
            by_chain[cn]['received'] += 1
        if r['points_awarded']:
            by_chain[cn]['points'] += int(r['points_awarded'] or 0)

    chain_filter  = request.args.get('chain', '').strip()
    status_filter = request.args.get('status', '').strip()
    search_q      = request.args.get('q', '').strip()
    filtered = requests_rows
    if chain_filter:
        filtered = [r for r in filtered if (r['chain_name'] or '') == chain_filter]
    if status_filter:
        filtered = [r for r in filtered if (r['status'] or '') == status_filter]
    if search_q:
        q_lower = search_q.lower()
        def _hay(r):
            return ' '.join(str(r[k] or '') for k in (
                'pickup_event_name', 'pickup_hotel', 'pickup_org',
                'booking_id', 'chain_name', 'sent_to_name',
                'sent_to_email', 'cvent_rfp_code', 'notes',
                'rewards_form_link',
            )).lower()
        filtered = [r for r in filtered if q_lower in _hay(r)]

    # Optional sort: ?sort=<key>  or  ?sort=-<key>  (desc)
    # Supported keys: booking_id, event_start, form_sent_date, points_received_date
    sort_param = (request.args.get('sort', '') or '').strip()
    sort_dir   = 'desc' if sort_param.startswith('-') else 'asc'
    sort_key   = sort_param.lstrip('-')
    SORT_FIELD = {
        'booking_id':           ('booking_id',           'int'),
        'event_start':          ('pickup_event_start',   'date'),
        'form_sent_date':       ('form_sent_date',       'date'),
        'points_received_date': ('points_received_date', 'date'),
    }
    if sort_key in SORT_FIELD:
        col, kind = SORT_FIELD[sort_key]
        def _norm_date(v):
            """Return ISO yyyy-mm-dd from ISO, US, or other common formats."""
            s = str(v).strip()
            if len(s) >= 10 and s[4] == '-' and s[7] == '-':
                return s[:10]
            for fmt in ('%m/%d/%Y', '%m/%d/%y', '%d/%m/%Y'):
                try:
                    return datetime.strptime(s[:10] if len(s) >= 10 else s, fmt).strftime('%Y-%m-%d')
                except Exception:
                    pass
            try:
                return datetime.strptime(s.split()[0], '%m/%d/%Y').strftime('%Y-%m-%d')
            except Exception:
                return s
        def _value(r):
            v = r[col] if col in r.keys() else None
            if v is None or v == '':
                return None
            if kind == 'int':
                try:
                    return int(str(v).split('.')[0])
                except Exception:
                    return str(v)
            return _norm_date(v)
        nulls    = [r for r in filtered if _value(r) is None]
        non_null = [r for r in filtered if _value(r) is not None]
        non_null.sort(key=_value, reverse=(sort_dir == 'desc'))
        filtered = non_null + nulls

    return render_template('points_dashboard.html',
                           requests=filtered, kpi=kpi, by_chain=by_chain,
                           chain_filter=chain_filter, status_filter=status_filter,
                           search_q=search_q,
                           sort_key=sort_key, sort_dir=sort_dir,
                           today=today)


@app.route('/points/programs')
def points_programs():
    db = get_db()
    programs = db.execute('SELECT * FROM hotel_points_program ORDER BY chain_name').fetchall()
    return render_template('points_programs.html', programs=programs)


@app.route('/points/programs/<int:pid>/edit', methods=['GET', 'POST'])
def points_program_edit(pid):
    db = get_db()
    program = db.execute('SELECT * FROM hotel_points_program WHERE id=?', (pid,)).fetchone()
    if not program:
        flash('Program not found.', 'error')
        return redirect(url_for('points_programs'))

    if request.method == 'POST':
        member_number = request.form.get('member_number', '').strip() or None
        submission_type = request.form.get('submission_type', 'manual').strip()
        form_url = request.form.get('form_url', '').strip() or None
        submission_window_days = request.form.get('submission_window_days', '90').strip()
        receipt_window_days    = request.form.get('receipt_window_days', '60').strip()
        notes = request.form.get('notes', '').strip() or None
        field_mapping_raw = request.form.get('field_mapping_json', '').strip()
        try:
            json.loads(field_mapping_raw) if field_mapping_raw else {}
            field_mapping_json = field_mapping_raw or '{}'
        except Exception:
            flash('Field mapping JSON is invalid — not saved.', 'error')
            field_mapping_json = program['field_mapping_json']
        active = 1 if request.form.get('active') == 'on' else 0

        f = request.files.get('template_file')
        template_data = program['form_template_data']
        template_filename = program['form_template_filename']
        if f and f.filename:
            if not f.filename.lower().endswith('.docx'):
                flash('Template must be a .docx file.', 'error')
            else:
                template_data = f.read()
                template_filename = f.filename

        db.execute('''UPDATE hotel_points_program
            SET member_number=?, submission_type=?, form_url=?,
                submission_window_days=?, receipt_window_days=?,
                notes=?, field_mapping_json=?, active=?,
                form_template_data=?, form_template_filename=?,
                updated_at=datetime('now')
            WHERE id=?''',
            (member_number, submission_type, form_url,
             int(submission_window_days or 90), int(receipt_window_days or 60),
             notes, field_mapping_json, active,
             template_data, template_filename, pid))
        db.commit()
        flash(f'{program["chain_name"]} program saved.', 'success')
        return redirect(url_for('points_programs'))

    try:
        mapping_pretty = json.dumps(json.loads(program['field_mapping_json'] or '{}'), indent=2)
    except Exception:
        mapping_pretty = program['field_mapping_json'] or '{}'
    return render_template('points_program_edit.html', program=program,
                           mapping_pretty=mapping_pretty)


@app.route('/points/programs/<int:pid>/template')
def points_program_template_download(pid):
    db = get_db()
    row = db.execute('SELECT chain_name, form_template_filename, form_template_data '
                     'FROM hotel_points_program WHERE id=?', (pid,)).fetchone()
    if not row or not row['form_template_data']:
        flash('No template on file for this program.', 'error')
        return redirect(url_for('points_program_edit', pid=pid))
    return send_file(io.BytesIO(row['form_template_data']),
                     as_attachment=True,
                     download_name=row['form_template_filename'] or f'{row["chain_name"]}_template.docx')


@app.route('/pickup/<int:cid>/points/generate', methods=['GET', 'POST'])
def points_generate(cid):
    from points_utils import detect_chain
    db = get_db()
    config = db.execute('SELECT * FROM pickup_config WHERE id=?', (cid,)).fetchone()
    if not config:
        flash('Pickup event not found.', 'error')
        return redirect(url_for('pickup_dashboard'))

    chain = detect_chain(config['hotel'])
    if not chain:
        flash(f'Could not detect a hotel chain from "{config["hotel"]}". '
              f'Add the chain in Hotel Points → Programs and try again.', 'error')
        return redirect(url_for('pickup_event', cid=cid))

    program = _get_program(db, chain_name=chain)
    if not program or not program['active']:
        flash(f'No active program for {chain}.', 'error')
        return redirect(url_for('pickup_event', cid=cid))

    existing = db.execute(
        'SELECT * FROM hotel_points_request WHERE pickup_config_id=? AND program_id=?',
        (cid, program['id'])).fetchone()
    if existing:
        rid = existing['id']
    else:
        cur = db.execute('''INSERT INTO hotel_points_request
            (pickup_config_id, program_id, booking_id, form_generated_at,
             sent_to_name, sent_to_email, status)
            VALUES (?, ?, ?, datetime('now'), ?, ?, 'pending')''',
            (cid, program['id'], config['booking_id'],
             config['hotel_contact'], config['hotel_contact_email']))
        rid = cur.lastrowid
        db.commit()
    return redirect(url_for('points_request_detail', rid=rid))


@app.route('/points/<int:rid>')
def points_request_detail(rid):
    db = get_db()
    req = db.execute('SELECT * FROM hotel_points_request WHERE id=?', (rid,)).fetchone()
    if not req:
        flash('Points request not found.', 'error')
        return redirect(url_for('points_dashboard'))
    field_values, program, pickup_cfg, booking = _build_field_values_for_request(db, req)
    try:
        mapping = json.loads(program['field_mapping_json'] or '{}') if program else {}
    except Exception:
        mapping = {}
    return render_template('points_request_detail.html',
                           req=req, program=program,
                           pickup_cfg=pickup_cfg, booking=booking,
                           field_values=field_values, mapping=mapping)


@app.route('/points/<int:rid>/edit', methods=['POST'])
def points_request_edit(rid):
    from points_utils import detect_chain as _detect_chain_edit
    db = get_db()
    fields = {}
    for key in ('form_sent_date', 'sent_to_name', 'sent_to_email',
                'points_received_date', 'points_awarded', 'status',
                'cvent_rfp_code', 'contract_signature_date',
                'incentive_type', 'award_timing',
                'second_recipient_name', 'second_recipient_email',
                'second_recipient_phone', 'second_recipient_number',
                'rewards_form_link', 'notes',
                'booking_id'):
        v = request.form.get(key, '').strip()
        fields[key] = v or None
    if fields['points_awarded']:
        try:
            fields['points_awarded'] = int(str(fields['points_awarded']).replace(',', ''))
        except Exception:
            fields['points_awarded'] = None
    if fields['points_received_date'] and fields['status'] in ('pending', 'submitted', None):
        fields['status'] = 'received'
    elif fields['form_sent_date'] and fields['status'] in ('pending', None):
        fields['status'] = 'submitted'

    # If booking_id changed: re-link pickup_config and re-detect chain
    extra_msg = ''
    old_bid = db.execute('SELECT booking_id FROM hotel_points_request WHERE id=?',
                         (rid,)).fetchone()
    if old_bid and fields.get('booking_id') and str(old_bid['booking_id']) != str(fields['booking_id']):
        new_bid = fields['booking_id']
        pc = db.execute('SELECT id, hotel FROM pickup_config WHERE booking_id=? '
                        'ORDER BY id DESC LIMIT 1', (new_bid,)).fetchone()
        hotel = None
        if pc:
            fields['pickup_config_id'] = pc['id']
            hotel = pc['hotel']
        else:
            fields['pickup_config_id'] = None
            pip = db.execute('SELECT Customer FROM ReportPipeline WHERE '
                             'CAST(BookingId AS INTEGER)=CAST(? AS INTEGER) LIMIT 1',
                             (new_bid,)).fetchone()
            if pip: hotel = pip['Customer']
        if hotel:
            new_chain = _detect_chain_edit(hotel)
            if new_chain:
                new_prog = db.execute('SELECT id FROM hotel_points_program '
                                      'WHERE chain_name=?', (new_chain,)).fetchone()
                if new_prog:
                    fields['program_id'] = new_prog['id']
                    extra_msg = f' Re-classified as {new_chain} based on hotel "{hotel}".'

    sets = ', '.join(f'{k}=?' for k in fields.keys())
    params = list(fields.values()) + [rid]
    db.execute(f'UPDATE hotel_points_request SET {sets}, updated_at=datetime("now") WHERE id=?',
               params)
    db.commit()
    flash('Points request updated.' + extra_msg, 'success')
    return redirect(url_for('points_request_detail', rid=rid))


@app.route('/points/<int:rid>/download')
def points_request_download(rid):
    from points_utils import fill_docx_template
    db = get_db()
    req = db.execute('SELECT * FROM hotel_points_request WHERE id=?', (rid,)).fetchone()
    if not req:
        flash('Points request not found.', 'error')
        return redirect(url_for('points_dashboard'))
    program = _get_program(db, program_id=req['program_id'])
    if not program or not program['form_template_data']:
        flash(f'No .docx template uploaded for {program["chain_name"] if program else "this chain"}. '
              f'Upload one in Hotel Points → Programs.', 'error')
        return redirect(url_for('points_request_detail', rid=rid))

    field_values, _, _, _ = _build_field_values_for_request(db, req)
    try:
        mapping = json.loads(program['field_mapping_json'] or '{}')
    except Exception:
        mapping = {}

    filled = fill_docx_template(program['form_template_data'], field_values, mapping)
    fname = f'{(field_values.get("event_name") or "Event").replace("/", "-")[:60]} - {program["chain_name"]} Points Request.docx'
    db.execute('''UPDATE hotel_points_request
        SET generated_doc_data=?, generated_doc_filename=?,
            form_generated_at=datetime('now'), updated_at=datetime('now')
        WHERE id=?''', (filled, fname, rid))
    db.commit()
    return send_file(io.BytesIO(filled), as_attachment=True,
                     download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document')


@app.route('/points/<int:rid>/mailto')
def points_request_mailto(rid):
    from points_utils import build_mailto
    db = get_db()
    req = db.execute('SELECT * FROM hotel_points_request WHERE id=?', (rid,)).fetchone()
    if not req:
        flash('Points request not found.', 'error')
        return redirect(url_for('points_dashboard'))
    field_values, program, _, _ = _build_field_values_for_request(db, req)
    profile = _get_user_profile(db)

    chain = program['chain_name'] if program else 'Hotel'
    event = field_values.get('event_name') or 'event'
    subject = f'{event} — {chain} Planner Points Request'
    body_lines = [
        f"Hello {req['sent_to_name'] or field_values.get('hotel_contact_name') or ''},",
        '',
        f'Attached is the {chain} planner points request form for {event}.',
        '',
        'Event details:',
        f"  Hotel:       {field_values.get('hotel', '')}",
        f"  Event Dates: {field_values.get('meeting_dates_formatted') or (field_values.get('event_start','') + ' to ' + field_values.get('event_end',''))}",
    ]
    if field_values.get('cvent_rfp_code'):
        body_lines.append(f"  Cvent RFP:   {field_values['cvent_rfp_code']}")
    if field_values.get('total_contracted_rooms'):
        body_lines.append(f"  Room Block:  {field_values['total_contracted_rooms']} total room nights")
    body_lines += [
        '',
        'Please process the attached form for the planner points credit.',
        '',
        'Thank you,',
        profile.get('user_full_name') or '',
        profile.get('user_email') or '',
        profile.get('user_phone') or '',
        'ConferenceDirect',
    ]
    url = build_mailto(req['sent_to_email'] or '', subject, '\n'.join(body_lines))
    return redirect(url)


@app.route('/points/<int:rid>/resend', methods=['POST'])
def points_request_resend(rid):
    db = get_db()
    req = db.execute('SELECT * FROM hotel_points_request WHERE id=?', (rid,)).fetchone()
    if not req:
        flash('Points request not found.', 'error')
        return redirect(url_for('points_dashboard'))
    db.execute('''UPDATE hotel_points_request
        SET resend_count=COALESCE(resend_count,0)+1,
            last_resend_date=date('now'),
            updated_at=datetime('now')
        WHERE id=?''', (rid,))
    db.commit()
    flash('Resend logged — generate and re-attach the form, then send.', 'success')
    return redirect(url_for('points_request_detail', rid=rid))


@app.route('/points/<int:rid>/delete', methods=['POST'])
def points_request_delete(rid):
    db = get_db()
    db.execute('UPDATE hotel_points_request SET status="cancelled", '
               'updated_at=datetime("now") WHERE id=?', (rid,))
    db.commit()
    flash('Points request marked cancelled.', 'success')
    return redirect(url_for('points_dashboard'))


@app.route('/points/<int:rid>/weblink-helper')
def points_weblink_helper(rid):
    db = get_db()
    req = db.execute('SELECT * FROM hotel_points_request WHERE id=?', (rid,)).fetchone()
    if not req:
        flash('Points request not found.', 'error')
        return redirect(url_for('points_dashboard'))
    field_values, program, _, _ = _build_field_values_for_request(db, req)
    try:
        mapping = json.loads(program['field_mapping_json'] or '{}') if program else {}
    except Exception:
        mapping = {}
    return render_template('points_weblink_helper.html',
                           req=req, program=program,
                           field_values=field_values, mapping=mapping)


@app.route('/points/import', methods=['GET', 'POST'])
def points_import():
    from points_utils import import_tracking_xlsx
    db = get_db()
    programs = db.execute(
        'SELECT id, chain_name FROM hotel_points_program ORDER BY chain_name'
    ).fetchall()
    program_by_chain = {p['chain_name']: p['id'] for p in programs}

    if request.method == 'POST':
        action = request.form.get('action', 'preview')

        if action == 'preview':
            chain_name = request.form.get('chain_name', '').strip()
            f = request.files.get('file')
            if not f or not f.filename or not chain_name:
                flash('Choose a chain and an .xlsx file.', 'error')
                return redirect(url_for('points_import'))
            if chain_name not in program_by_chain:
                flash(f'Unknown chain "{chain_name}".', 'error')
                return redirect(url_for('points_import'))
            try:
                rows = import_tracking_xlsx(f.read(), chain_name)
            except Exception as e:
                flash(f'Could not parse file: {e}', 'error')
                return redirect(url_for('points_import'))

            preview = []
            for row in rows:
                match = db.execute('''
                    SELECT id, booking_id, event_name, hotel, event_start, event_end
                    FROM pickup_config
                    WHERE LOWER(TRIM(event_name)) = LOWER(TRIM(?))
                    ORDER BY id DESC LIMIT 1
                ''', (row['event_name'],)).fetchone()
                if not match and row['start_date']:
                    match = db.execute('''
                        SELECT id, booking_id, event_name, hotel, event_start, event_end
                        FROM pickup_config
                        WHERE LOWER(event_name) LIKE LOWER(?)
                          AND event_start = ?
                        LIMIT 1
                    ''', ('%' + row['event_name'][:30] + '%', row['start_date'])).fetchone()
                dup = db.execute('''
                    SELECT id FROM hotel_points_request
                    WHERE program_id=? AND
                          (pickup_config_id=? OR
                           (pickup_config_id IS NULL AND notes LIKE ?))
                    LIMIT 1
                ''', (program_by_chain[chain_name],
                      match['id'] if match else -1,
                      '%' + row['event_name'][:30] + '%')).fetchone()
                preview.append({
                    'row': row,
                    'pickup_config_id': match['id'] if match else None,
                    'pickup_booking_id': match['booking_id'] if match else None,
                    'duplicate': bool(dup),
                })

            return render_template('points_import.html',
                                   programs=programs, preview=preview,
                                   chain_name=chain_name,
                                   matched=sum(1 for p in preview if p['pickup_config_id'] and not p['duplicate']),
                                   unmatched=sum(1 for p in preview if not p['pickup_config_id'] and not p['duplicate']),
                                   skipped=sum(1 for p in preview if p['duplicate']))

        elif action == 'commit':
            chain_name = request.form.get('chain_name', '').strip()
            payload = request.form.get('payload_json', '')
            try:
                rows = json.loads(payload)
            except Exception:
                flash('Invalid payload — re-run preview.', 'error')
                return redirect(url_for('points_import'))
            program_id = program_by_chain.get(chain_name)
            if not program_id:
                flash('Unknown chain.', 'error')
                return redirect(url_for('points_import'))

            from points_utils import detect_chain as _detect_chain_imp
            inserted = 0
            for item in rows:
                if item.get('duplicate'):
                    continue
                row = item['row']
                hotel_name = row.get('hotel', '')
                detected   = _detect_chain_imp(hotel_name)
                row_program_id = (program_by_chain.get(detected, program_id)
                                  if detected else program_id)
                db.execute('''INSERT INTO hotel_points_request
                    (pickup_config_id, program_id, booking_id,
                     form_sent_date, points_received_date, points_awarded,
                     status, rewards_form_link, notes)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                    (item.get('pickup_config_id'), row_program_id,
                     item.get('pickup_booking_id'),
                     row.get('form_sent_date'),
                     row.get('points_received_date'),
                     row.get('points_awarded'),
                     row.get('status', 'pending'),
                     row.get('rewards_form_link'),
                     (f"[Imported from {row.get('sheet','')} | hotel: {row.get('hotel','')}] " +
                      (row.get('notes') or ''))[:2000]))
                inserted += 1
            db.commit()
            flash(f'Imported {inserted} records from {chain_name} workbook '
                  f'(each tagged by detected hotel brand).', 'success')
            return redirect(url_for('points_dashboard'))

    return render_template('points_import.html', programs=programs,
                           preview=None, chain_name=None)


@app.route('/points/rechain', methods=['GET', 'POST'])
def points_rechain():
    """Re-classify points requests by detecting chain from hotel name."""
    import re as _re_rc
    from points_utils import detect_chain as _detect_chain_rc
    db = get_db()
    programs = {p['chain_name']: p['id'] for p in db.execute(
        'SELECT id, chain_name FROM hotel_points_program').fetchall()}

    rows = db.execute('''
        SELECT r.id, r.program_id, r.booking_id, r.notes,
               r.pickup_config_id, r.status, r.points_awarded,
               p.chain_name AS current_chain,
               pc.hotel     AS pc_hotel,
               pc.event_name AS pickup_event_name,
               pip.Customer  AS pip_customer,
               pip.EventName AS pip_event_name
        FROM hotel_points_request r
        LEFT JOIN hotel_points_program p ON p.id = r.program_id
        LEFT JOIN pickup_config pc       ON pc.id = r.pickup_config_id
        LEFT JOIN ReportPipeline pip     ON CAST(pip.BookingId AS INTEGER)
                                          = CAST(r.booking_id AS INTEGER)
    ''').fetchall()

    def _hotel_of(r):
        # Priority: user-maintained pickup_config.hotel > booking's Customer
        # (the actual hotel on the booking record) > workbook value in notes
        # (which may be stale or contain a typo).
        h = r['pc_hotel']
        if not h: h = r['pip_customer']
        if not h and r['notes']:
            m = _re_rc.search(r'\bhotel:\s*([^|\]]+?)\s*\]', r['notes'])
            if m: h = m.group(1).strip()
        return h or ''

    mismatches, unknown = [], []
    for r in rows:
        hotel = _hotel_of(r)
        detected = _detect_chain_rc(hotel) if hotel else None
        if not detected:
            unknown.append({'id': r['id'], 'current_chain': r['current_chain'],
                            'hotel': hotel,
                            'event': r['pickup_event_name'] or r['pip_event_name'] or '',
                            'status': r['status']})
        elif detected != r['current_chain']:
            mismatches.append({'id': r['id'], 'current_chain': r['current_chain'],
                               'detected_chain': detected,
                               'hotel': hotel,
                               'event': r['pickup_event_name'] or r['pip_event_name'] or '',
                               'status': r['status'],
                               'points_awarded': r['points_awarded']})

    if request.method == 'POST' and request.form.get('action') == 'apply':
        applied = 0
        for m in mismatches:
            target = programs.get(m['detected_chain'])
            if not target:
                continue
            db.execute('UPDATE hotel_points_request '
                       'SET program_id=?, updated_at=datetime("now") WHERE id=?',
                       (target, m['id']))
            applied += 1
        db.commit()
        flash(f'Re-classified {applied} request{"s" if applied != 1 else ""}.', 'success')
        return redirect(url_for('points_rechain'))

    return render_template('points_rechain.html',
                           mismatches=mismatches, unknown=unknown,
                           total_requests=len(rows))


def _points_report_rows(db, chain, start_iso, end_iso):
    """Per-chain report rows whose meeting dates overlap [start, end]."""
    sql = '''
        SELECT
            r.id, r.booking_id, r.form_sent_date, r.points_received_date,
            r.points_awarded, r.status, r.notes,
            COALESCE(pc.hotel,      pip.Customer)                  AS hotel,
            COALESCE(pc.event_name, pip.EventName)                 AS event,
            COALESCE(pc.event_start, substr(pip.StartDate,1,10))   AS event_start,
            COALESCE(pc.event_end,   substr(pip.EndDate,  1,10))   AS event_end
        FROM hotel_points_request r
        LEFT JOIN hotel_points_program p ON p.id = r.program_id
        LEFT JOIN pickup_config pc       ON pc.id = r.pickup_config_id
        LEFT JOIN ReportPipeline pip
            ON CAST(pip.BookingId AS INTEGER) = CAST(r.booking_id AS INTEGER)
        WHERE p.chain_name = ?
    '''
    params = [chain]
    if start_iso:
        sql += ' AND (COALESCE(pc.event_end, substr(pip.EndDate,1,10)) >= ? OR ' \
               '     COALESCE(pc.event_start, substr(pip.StartDate,1,10)) >= ?)'
        params += [start_iso, start_iso]
    if end_iso:
        sql += ' AND (COALESCE(pc.event_start, substr(pip.StartDate,1,10)) <= ? OR ' \
               '     COALESCE(pc.event_end, substr(pip.EndDate,1,10)) <= ?)'
        params += [end_iso, end_iso]
    sql += ' ORDER BY event_start ASC, event ASC'
    return db.execute(sql, params).fetchall()


@app.route('/points/report', methods=['GET'])
def points_report():
    db = get_db()
    chain = (request.args.get('chain') or 'Marriott').strip()
    start = (request.args.get('start') or '').strip()
    end   = (request.args.get('end')   or '').strip()
    fmt   = (request.args.get('format') or '').strip().lower()

    programs = db.execute(
        'SELECT chain_name FROM hotel_points_program WHERE active=1 ORDER BY chain_name'
    ).fetchall()

    rows = _points_report_rows(db, chain, start, end) if chain else []
    total_points = sum(int(r['points_awarded'] or 0) for r in rows)

    if fmt == 'xlsx':
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        def _us_date(s):
            if not s: return ''
            s = str(s)[:10]
            try:
                d = datetime.strptime(s, '%Y-%m-%d')
                try:    return d.strftime('%-m/%-d/%Y')
                except ValueError: return d.strftime('%m/%d/%Y')
            except Exception:
                return s

        wb = Workbook()
        ws = wb.active
        ws.title = f'{chain} Points'
        headers = ['Hotel', 'Event', 'Booking ID', 'Event Dates', 'Sent', 'Received', 'Points']
        ws.append(headers)
        hdr_font = Font(bold=True, color='FFFFFF')
        hdr_fill = PatternFill('solid', fgColor='1A3A5C')
        thin = Side(border_style='thin', color='888888')
        bord = Border(left=thin, right=thin, top=thin, bottom=thin)
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = bord
        for r in rows:
            s, e = _us_date(r['event_start']), _us_date(r['event_end'])
            ev_dates = f'{s} – {e}' if s and e else (s or e)
            ws.append([
                r['hotel'] or '',
                r['event'] or '',
                str(r['booking_id'] or ''),
                ev_dates,
                _us_date(r['form_sent_date']),
                _us_date(r['points_received_date']),
                int(r['points_awarded']) if r['points_awarded'] else '',
            ])
        if rows:
            last_row = ws.max_row + 1
            ws.cell(row=last_row, column=6, value='TOTAL').font = Font(bold=True)
            tcell = ws.cell(row=last_row, column=7, value=total_points)
            tcell.font = Font(bold=True)
            tcell.number_format = '#,##0'
        for r_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=r_idx, column=7)
            if isinstance(cell.value, int):
                cell.number_format = '#,##0'
        widths = [36, 50, 12, 24, 12, 12, 12]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = w
        ws.freeze_panes = 'A2'

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        rng = f'{start or "all"}_to_{end or "all"}'
        fname = f'{chain}_Points_{rng}.xlsx'.replace(' ', '_')
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    return render_template('points_report.html',
                           programs=programs,
                           chain=chain, start=start, end=end,
                           rows=rows, total_points=total_points)


# ═════════════════════════════════════════════════════════════════════════════
# COST SAVINGS ANALYSIS
# ═════════════════════════════════════════════════════════════════════════════

COST_SAVINGS_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'assets', 'cost_savings_template.xlsx')


def _cs_get_report(db, report_id):
    return db.execute('SELECT * FROM cost_savings_report WHERE id=?', (report_id,)).fetchone()


def _cs_get_items(db, report_id):
    return db.execute(
        'SELECT * FROM cost_savings_item WHERE report_id=? ORDER BY sort_order, id',
        (report_id,)
    ).fetchall()


def _cs_seed_items(db, report_id):
    for i, item in enumerate(COST_SAVINGS_SEED_ITEMS):
        db.execute('''INSERT INTO cost_savings_item
            (report_id, sort_order, item_name, calc_type,
             standard_price, negotiated_price, quantity, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
            (report_id, i, item['item_name'], item['calc_type'],
             item.get('standard_price'), item.get('negotiated_price'),
             item.get('quantity'), item.get('notes')))


def _cs_recompute_and_save(db, report_id):
    report = _cs_get_report(db, report_id)
    items = _cs_get_items(db, report_id)
    if not report:
        return
    totals = _compute_report_totals(report, items)
    for calc in totals['item_calcs']:
        db.execute('UPDATE cost_savings_item SET cost_savings=? WHERE id=?',
                   (calc['savings'], calc['id']))
    db.execute("UPDATE cost_savings_report SET updated_at=datetime('now') WHERE id=?",
               (report_id,))
    db.commit()


def _cs_apply_extraction(db, report_id, side, data):
    if side == 'proposal':
        sets, params = [], []
        for col, key in [
            ('gr_rack_rate',           'rack_rate'),
            ('sr_group_rate',          'group_rate'),
            ('mr_initial',             'meeting_room_rental'),
            ('fb_initial',             'f_and_b_minimum'),
            ('comp_industry_standard', 'comp_industry_standard'),
        ]:
            v = data.get(key)
            if v is not None:
                sets.append(f'{col}=?'); params.append(v)
        if data.get('hotel_brand'):
            sets.append('hotel_brand=?'); params.append(data['hotel_brand'])
        sets.append("proposal_extracted_at=datetime('now')")
        sets.append("updated_at=datetime('now')")
        params.append(report_id)
        db.execute(f'UPDATE cost_savings_report SET {", ".join(sets)} WHERE id=?', params)
    else:
        sets, params = [], []
        for col, key in [
            ('gr_contracted_rate',     'contracted_rate'),
            ('sr_contracted_rate',     'staff_contracted_rate'),
            ('gr_total_nights',        'total_room_nights'),
            ('sr_total_nights',        'staff_total_nights'),
            ('mr_negotiated',          'meeting_room_rental'),
            ('fb_negotiated',          'f_and_b_minimum'),
            ('comp_negotiated_policy', 'comp_negotiated_policy'),
        ]:
            v = data.get(key)
            if v is not None:
                sets.append(f'{col}=?'); params.append(v)
        sets.append("contract_extracted_at=datetime('now')")
        sets.append("updated_at=datetime('now')")
        params.append(report_id)
        db.execute(f'UPDATE cost_savings_report SET {", ".join(sets)} WHERE id=?', params)
    db.commit()


def _cs_extract_from_crf_columns(rh):
    data = {}
    if rh['proposed_rate'] is not None:
        data['rack_rate'] = rh['proposed_rate']
        data['group_rate'] = rh['proposed_rate']
    if rh['f_and_b_minimum'] is not None:
        data['f_and_b_minimum'] = rh['f_and_b_minimum']
    if rh['meeting_room_rental'] is not None:
        data['meeting_room_rental'] = rh['meeting_room_rental']
    if rh['attrition_pct'] is not None:
        data['proposal_attrition_pct'] = rh['attrition_pct']
    brand_val = rh['brand'] or _detect_brand_from_hotel(rh['hotel_name'])
    if brand_val:
        bl = brand_val.lower()
        for canonical in ('Hyatt', 'Hilton', 'Marriott', 'IHG'):
            if canonical.lower() in bl:
                brand_val = canonical
                break
        data['hotel_brand'] = brand_val
    try:
        crf_qa = json.loads(rh['crf_row_data'] or '{}')
        for q, a in crf_qa.items():
            ql = q.lower()
            if 'gross room rate' in ql and a:
                import re as _re
                m = _re.search(r'\$?\s*([\d,]+(?:\.\d+)?)', str(a))
                if m:
                    try:
                        data['rack_rate'] = float(m.group(1).replace(',', ''))
                    except Exception:
                        pass
    except Exception:
        pass
    return data


def _cs_apply_attrition_to_seed(db, report_id, attrition_pct):
    if attrition_pct is None:
        return
    db.execute('''UPDATE cost_savings_item SET quantity=?
                  WHERE id = (SELECT id FROM cost_savings_item
                              WHERE report_id=? AND calc_type='attrition'
                              ORDER BY sort_order LIMIT 1)''',
               (attrition_pct, report_id))


def cascade_contract_cost_savings(db, file_bytes, filename, booking_id=None,
                                  pickup_config_id=None):
    """Run rich cost-savings extraction on the contract file, cache JSON in
    pickup_config, and cascade contract-side values onto any linked reports."""
    from pickup_utils import parse_contract_for_cost_savings
    if not file_bytes:
        return None
    try:
        result = parse_contract_for_cost_savings(file_bytes, filename or 'contract.pdf')
    except Exception as e:
        return {'error': f'cost-savings extraction failed: {e}'}
    if result.get('error'):
        return result
    payload = json.dumps({k: v for k, v in result.items() if k != 'raw_text'})
    if pickup_config_id:
        db.execute('UPDATE pickup_config SET contract_extracted_data=? WHERE id=?',
                   (payload, pickup_config_id))
    elif booking_id is not None:
        db.execute('''UPDATE pickup_config SET contract_extracted_data=?
                      WHERE CAST(booking_id AS INTEGER)=CAST(? AS INTEGER)''',
                   (payload, booking_id))
    if booking_id is not None:
        rows = db.execute('''
            SELECT csr.id FROM cost_savings_report csr
            JOIN rfp_hotel rh ON rh.id = csr.rfp_hotel_id
            JOIN rfp r        ON r.id  = rh.rfp_id
            WHERE CAST(r.booking_id AS INTEGER) = CAST(? AS INTEGER)
        ''', (booking_id,)).fetchall()
        for row in rows:
            _cs_apply_extraction(db, row['id'], 'contract', result)
            _cs_apply_attrition_to_seed(db, row['id'], result.get('attrition_pct'))
            _cs_recompute_and_save(db, row['id'])
    db.commit()
    return result


def _cs_seed_from_selected_hotel(db, rid, hid):
    """When a hotel is selected, create/refresh a Cost Savings Report pre-filled
    from CRF / structured rfp_hotel columns."""
    rfp = db.execute('SELECT * FROM rfp WHERE id=?', (rid,)).fetchone()
    hotel = db.execute('SELECT * FROM rfp_hotel WHERE id=? AND rfp_id=?',
                       (hid, rid)).fetchone()
    if not (rfp and hotel):
        return None
    existing = db.execute(
        'SELECT id FROM cost_savings_report WHERE rfp_hotel_id=? ORDER BY id DESC LIMIT 1',
        (hid,)
    ).fetchone()
    if existing:
        report_id = existing['id']
    else:
        brand = _detect_brand_from_hotel(hotel['hotel_name']) or 'Preferred'
        meeting_dates = ''
        if rfp['start_date'] and rfp['end_date']:
            meeting_dates = f"{rfp['start_date']} to {rfp['end_date']}"
        cur = db.execute('''
            INSERT INTO cost_savings_report
                (rfp_hotel_id, meeting_name, hotel_name, meeting_dates,
                 lead_date_requested, booked_date,
                 comp_industry_standard, hotel_brand, status)
            VALUES (?, ?, ?, ?, ?, ?, 50, ?, 'draft')
        ''', (
            hid,
            rfp['event_name'] or rfp['rfp_name'] or '',
            hotel['hotel_name'] or '',
            meeting_dates,
            rfp['created_at'][:10] if rfp['created_at'] else '',
            '',
            brand,
        ))
        report_id = cur.lastrowid
        _cs_seed_items(db, report_id)
        db.commit()
    data = _cs_extract_from_crf_columns(hotel)
    if data:
        _cs_apply_extraction(db, report_id, 'proposal', data)
        _cs_apply_attrition_to_seed(db, report_id, data.get('proposal_attrition_pct'))
        _cs_recompute_and_save(db, report_id)
    return report_id


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route('/cost-savings')
def cost_savings_dashboard():
    db = get_db()
    q = (request.args.get('q') or '').strip()
    sql = '''
        SELECT csr.*, rh.rfp_id AS rfp_id, r.rfp_code AS rfp_code, r.event_name AS rfp_event_name
        FROM cost_savings_report csr
        JOIN rfp_hotel rh ON rh.id = csr.rfp_hotel_id
        JOIN rfp r        ON r.id  = rh.rfp_id
    '''
    params = []
    if q:
        sql += ' WHERE csr.meeting_name LIKE ? OR csr.hotel_name LIKE ? OR r.rfp_code LIKE ?'
        params = [f'%{q}%', f'%{q}%', f'%{q}%']
    sql += ' ORDER BY csr.updated_at DESC'
    reports = db.execute(sql, params).fetchall()
    enriched = []
    total_savings = 0
    total_hours = 0
    for r in reports:
        items = _cs_get_items(db, r['id'])
        totals = _compute_report_totals(r, items)
        enriched.append({'r': r, 'grand_total': totals['grand_total']})
        total_savings += totals['grand_total']
        total_hours += (r['hours_worked'] or 0)
    stat_count = len(enriched)
    avg = (total_savings / stat_count) if stat_count else 0
    return render_template('cost_savings_dashboard.html',
                           reports=enriched, q=q,
                           stat_count=stat_count,
                           total_savings=total_savings,
                           total_hours=total_hours,
                           avg_savings=avg)


@app.route('/rfp/<int:rid>/cost-savings')
def cost_savings_rfp(rid):
    db = get_db()
    rfp = db.execute('SELECT * FROM rfp WHERE id=?', (rid,)).fetchone()
    if not rfp:
        flash('RFP not found.', 'error')
        return redirect(url_for('rfp_dashboard'))
    hotels = db.execute('SELECT * FROM rfp_hotel WHERE rfp_id=? ORDER BY id', (rid,)).fetchall()
    rows = []
    for h in hotels:
        rep = db.execute(
            'SELECT * FROM cost_savings_report WHERE rfp_hotel_id=? ORDER BY id DESC LIMIT 1',
            (h['id'],)
        ).fetchone()
        grand = None
        if rep:
            items = _cs_get_items(db, rep['id'])
            grand = _compute_report_totals(rep, items)['grand_total']
        rows.append({'hotel': h, 'report': rep, 'grand_total': grand})
    return render_template('cost_savings_rfp.html', rfp=rfp, rows=rows)


@app.route('/rfp/<int:rid>/hotel/<int:hid>/cost-savings/create', methods=['POST'])
def cost_savings_create(rid, hid):
    db = get_db()
    rfp = db.execute('SELECT * FROM rfp WHERE id=?', (rid,)).fetchone()
    hotel = db.execute('SELECT * FROM rfp_hotel WHERE id=? AND rfp_id=?',
                       (hid, rid)).fetchone()
    if not (rfp and hotel):
        flash('RFP or hotel not found.', 'error')
        return redirect(url_for('rfp_dashboard'))
    existing = db.execute(
        'SELECT id FROM cost_savings_report WHERE rfp_hotel_id=? ORDER BY id DESC LIMIT 1',
        (hid,)
    ).fetchone()
    if existing:
        return redirect(url_for('cost_savings_edit', report_id=existing['id']))
    brand = _detect_brand_from_hotel(hotel['hotel_name']) or 'Preferred'
    meeting_dates = ''
    if rfp['start_date'] and rfp['end_date']:
        meeting_dates = f"{rfp['start_date']} to {rfp['end_date']}"
    cur = db.execute('''
        INSERT INTO cost_savings_report
            (rfp_hotel_id, meeting_name, hotel_name, meeting_dates,
             lead_date_requested, booked_date,
             gr_contracted_rate, sr_contracted_rate,
             comp_industry_standard, hotel_brand,
             mr_negotiated, fb_negotiated, status)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, 50, ?, ?, ?, 'draft')
    ''', (hid,
          rfp['event_name'] or rfp['rfp_name'] or '',
          hotel['hotel_name'] or '',
          meeting_dates,
          rfp['created_at'][:10] if rfp['created_at'] else '',
          '',
          hotel['proposed_rate'], hotel['proposed_rate'],
          brand,
          hotel['meeting_room_rental'], hotel['f_and_b_minimum']))
    report_id = cur.lastrowid
    _cs_seed_items(db, report_id)
    db.commit()
    _cs_recompute_and_save(db, report_id)
    flash('Cost Savings Report created.', 'success')
    return redirect(url_for('cost_savings_edit', report_id=report_id))


@app.route('/cost-savings/<int:report_id>')
def cost_savings_edit(report_id):
    db = get_db()
    report = _cs_get_report(db, report_id)
    if not report:
        flash('Report not found.', 'error')
        return redirect(url_for('cost_savings_dashboard'))
    items = _cs_get_items(db, report_id)
    rh = db.execute('SELECT * FROM rfp_hotel WHERE id=?',
                    (report['rfp_hotel_id'],)).fetchone()
    rfp = db.execute('SELECT * FROM rfp WHERE id=?', (rh['rfp_id'],)).fetchone() if rh else None
    totals = _compute_report_totals(report, items)
    return render_template('cost_savings_edit.html',
                           report=report, items=items,
                           rfp=rfp, rfp_hotel=rh,
                           totals=totals,
                           brands=list(PLANNER_POINTS_BY_BRAND.keys()))


@app.route('/cost-savings/<int:report_id>/save', methods=['POST'])
def cost_savings_save(report_id):
    db = get_db()
    report = _cs_get_report(db, report_id)
    if not report:
        flash('Report not found.', 'error')
        return redirect(url_for('cost_savings_dashboard'))

    def _f(name):
        v = request.form.get(name, '').strip().replace(',', '').replace('$', '')
        try:
            return float(v) if v else None
        except Exception:
            return None

    def _i(name):
        v = request.form.get(name, '').strip()
        try:
            return int(float(v)) if v else None
        except Exception:
            return None

    db.execute('''UPDATE cost_savings_report SET
        meeting_name=?, hotel_name=?, meeting_dates=?,
        lead_date_requested=?, booked_date=?, hours_worked=?,
        gr_rack_rate=?, gr_contracted_rate=?, gr_total_nights=?, gr_notes=?,
        sr_group_rate=?, sr_contracted_rate=?, sr_total_nights=?,
        comp_industry_standard=?, comp_negotiated_policy=?,
        mr_initial=?, mr_negotiated=?, mr_notes=?,
        fb_initial=?, fb_negotiated=?, fb_notes=?,
        hotel_brand=?, status=?, updated_at=datetime('now')
        WHERE id=?''', (
        request.form.get('meeting_name', '').strip(),
        request.form.get('hotel_name', '').strip(),
        request.form.get('meeting_dates', '').strip(),
        request.form.get('lead_date_requested', '').strip(),
        request.form.get('booked_date', '').strip(),
        _f('hours_worked') or 0,
        _f('gr_rack_rate'), _f('gr_contracted_rate'), _i('gr_total_nights'),
        request.form.get('gr_notes', '').strip() or None,
        _f('sr_group_rate'), _f('sr_contracted_rate'), _i('sr_total_nights'),
        _i('comp_industry_standard') or 50, _i('comp_negotiated_policy'),
        _f('mr_initial'), _f('mr_negotiated'),
        request.form.get('mr_notes', '').strip() or None,
        _f('fb_initial'), _f('fb_negotiated'),
        request.form.get('fb_notes', '').strip() or None,
        request.form.get('hotel_brand', 'Preferred').strip(),
        request.form.get('status', 'draft').strip(),
        report_id,
    ))
    items = _cs_get_items(db, report_id)
    for idx, it in enumerate(items):
        prefix = f'item_{it["id"]}_'
        db.execute('''UPDATE cost_savings_item SET
            sort_order=?, item_name=?, calc_type=?,
            standard_price=?, negotiated_price=?, quantity=?, notes=?
            WHERE id=?''', (
            _i(prefix + 'sort_order') if request.form.get(prefix + 'sort_order') else idx,
            request.form.get(prefix + 'item_name', it['item_name']).strip(),
            request.form.get(prefix + 'calc_type', it['calc_type']).strip(),
            _f(prefix + 'standard_price'),
            _f(prefix + 'negotiated_price'),
            _f(prefix + 'quantity'),
            request.form.get(prefix + 'notes', '').strip() or None,
            it['id'],
        ))
    db.commit()
    _cs_recompute_and_save(db, report_id)
    flash('Saved.', 'success')
    return redirect(url_for('cost_savings_edit', report_id=report_id))


@app.route('/cost-savings/<int:report_id>/extract-proposal', methods=['POST'])
def cost_savings_extract_proposal(report_id):
    from pickup_utils import parse_proposal_for_cost_savings
    db = get_db()
    report = _cs_get_report(db, report_id)
    if not report:
        flash('Report not found.', 'error')
        return redirect(url_for('cost_savings_dashboard'))
    rh = db.execute('SELECT * FROM rfp_hotel WHERE id=?',
                    (report['rfp_hotel_id'],)).fetchone()
    if rh and rh['proposal_data']:
        result = parse_proposal_for_cost_savings(bytes(rh['proposal_data']),
                                                 rh['proposal_filename'] or 'proposal.pdf')
        if result.get('error'):
            flash(f"AI extraction problem: {result['error']}", 'error')
            return redirect(url_for('cost_savings_edit', report_id=report_id))
        _cs_apply_extraction(db, report_id, 'proposal', result)
        _cs_apply_attrition_to_seed(db, report_id, result.get('attrition_pct'))
        _cs_recompute_and_save(db, report_id)
        flash('Proposal values extracted from uploaded PDF — review and Save.', 'success')
        return redirect(url_for('cost_savings_edit', report_id=report_id))
    if rh and (rh['proposed_rate'] is not None or rh['crf_row_data']):
        data = _cs_extract_from_crf_columns(rh)
        _cs_apply_extraction(db, report_id, 'proposal', data)
        _cs_apply_attrition_to_seed(db, report_id, data.get('proposal_attrition_pct'))
        _cs_recompute_and_save(db, report_id)
        flash('Proposal values pulled from the CRF — review and Save.', 'success')
        return redirect(url_for('cost_savings_edit', report_id=report_id))
    flash('No proposal data found on this hotel — upload a proposal PDF or import a CRF first.', 'error')
    return redirect(url_for('cost_savings_edit', report_id=report_id))


@app.route('/cost-savings/<int:report_id>/extract-contract', methods=['POST'])
def cost_savings_extract_contract(report_id):
    from pickup_utils import parse_contract_for_cost_savings
    db = get_db()
    report = _cs_get_report(db, report_id)
    if not report:
        flash('Report not found.', 'error')
        return redirect(url_for('cost_savings_dashboard'))
    rh = db.execute('SELECT * FROM rfp_hotel WHERE id=?',
                    (report['rfp_hotel_id'],)).fetchone()
    rfp = db.execute('SELECT * FROM rfp WHERE id=?', (rh['rfp_id'],)).fetchone() if rh else None
    if rfp and rfp['booking_id']:
        cached = db.execute(
            '''SELECT contract_extracted_data FROM pickup_config
               WHERE CAST(booking_id AS INTEGER)=CAST(? AS INTEGER)
                 AND contract_extracted_data IS NOT NULL LIMIT 1''',
            (rfp['booking_id'],)
        ).fetchone()
        if cached and cached['contract_extracted_data']:
            try:
                data = json.loads(cached['contract_extracted_data'])
                _cs_apply_extraction(db, report_id, 'contract', data)
                _cs_apply_attrition_to_seed(db, report_id, data.get('attrition_pct'))
                _cs_recompute_and_save(db, report_id)
                flash('Contract values pulled from cached extraction — review and Save.', 'success')
                return redirect(url_for('cost_savings_edit', report_id=report_id))
            except Exception:
                pass
    file_bytes, filename = None, None
    if rfp and rfp['contract_data']:
        file_bytes = bytes(rfp['contract_data'])
        filename = rfp['contract_filename'] or 'contract.pdf'
    elif rfp and rfp['booking_id']:
        bc = db.execute('''SELECT filename, file_data FROM booking_contract
                           WHERE booking_id=? ORDER BY id DESC LIMIT 1''',
                        (rfp['booking_id'],)).fetchone()
        if bc and bc['file_data']:
            file_bytes = bytes(bc['file_data'])
            filename = bc['filename']
    if not file_bytes:
        flash('No contract PDF found on this RFP — upload one first.', 'error')
        return redirect(url_for('cost_savings_edit', report_id=report_id))
    result = parse_contract_for_cost_savings(file_bytes, filename)
    if result.get('error'):
        flash(f"Extraction problem: {result['error']}", 'error')
    else:
        _cs_apply_extraction(db, report_id, 'contract', result)
        _cs_apply_attrition_to_seed(db, report_id, result.get('attrition_pct'))
        _cs_recompute_and_save(db, report_id)
        if rfp and rfp['booking_id']:
            payload = json.dumps({k: v for k, v in result.items() if k != 'raw_text'})
            db.execute('''UPDATE pickup_config SET contract_extracted_data=?
                          WHERE CAST(booking_id AS INTEGER)=CAST(? AS INTEGER)''',
                       (payload, rfp['booking_id']))
            db.commit()
        flash('Contract values extracted — review and Save.', 'success')
    return redirect(url_for('cost_savings_edit', report_id=report_id))


@app.route('/cost-savings/<int:report_id>/item/add', methods=['POST'])
def cost_savings_item_add(report_id):
    db = get_db()
    report = _cs_get_report(db, report_id)
    if not report:
        return redirect(url_for('cost_savings_dashboard'))
    max_order = db.execute(
        'SELECT COALESCE(MAX(sort_order), -1) FROM cost_savings_item WHERE report_id=?',
        (report_id,)
    ).fetchone()[0]
    db.execute('''INSERT INTO cost_savings_item
        (report_id, sort_order, item_name, calc_type) VALUES (?, ?, ?, ?)''',
        (report_id, max_order + 1, 'New Item', 'simple'))
    db.commit()
    return redirect(url_for('cost_savings_edit', report_id=report_id))


@app.route('/cost-savings/<int:report_id>/item/<int:iid>/delete', methods=['POST'])
def cost_savings_item_delete(report_id, iid):
    db = get_db()
    db.execute('DELETE FROM cost_savings_item WHERE id=? AND report_id=?',
               (iid, report_id))
    db.commit()
    _cs_recompute_and_save(db, report_id)
    return redirect(url_for('cost_savings_edit', report_id=report_id))


@app.route('/cost-savings/<int:report_id>/delete', methods=['POST'])
def cost_savings_delete(report_id):
    db = get_db()
    db.execute('DELETE FROM cost_savings_item WHERE report_id=?', (report_id,))
    db.execute('DELETE FROM cost_savings_report WHERE id=?', (report_id,))
    db.commit()
    flash('Report deleted.', 'success')
    return redirect(url_for('cost_savings_dashboard'))


def _sanitize_sheet_name(name):
    bad = '\\/?*[]:'
    out = ''.join((c if c not in bad else ' ') for c in (name or 'Sheet'))
    return out[:31] or 'Sheet'


def _cs_populate_template_sheet(ws, report, items):
    from openpyxl.styles import Font
    ws['A1'] = report['meeting_name'] or ''
    ws['A2'] = report['hotel_name'] or ''
    ws['A3'] = report['meeting_dates'] or ''
    ws['A4'] = f"Original Lead Date Request: {report['lead_date_requested'] or ''}"
    ws['A5'] = f"Booked: {report['booked_date'] or ''}"
    ws['B8']  = report['gr_rack_rate']
    ws['B9']  = report['gr_contracted_rate']
    ws['B10'] = '=B8-B9'
    ws['B11'] = report['gr_total_nights']
    ws['B12'] = '=B11*B10'
    if report['gr_notes']:
        ws['C7'] = 'Notes'
        ws['C8'] = report['gr_notes']
    ws['B15'] = report['sr_group_rate']
    ws['B16'] = report['sr_contracted_rate']
    ws['B17'] = '=B15-B16'
    ws['B18'] = report['sr_total_nights']
    ws['B19'] = '=B17*B18'
    ws['B22'] = report['comp_industry_standard'] or 50
    ws['B23'] = report['comp_negotiated_policy']
    if report['comp_negotiated_policy']:
        ws['B24'] = '=(B11/B23)*B9'
    else:
        ws['B24'] = 0
    ws['B27'] = report['mr_initial']
    ws['C27'] = report['mr_negotiated']
    ws['D27'] = '=B27-C27'
    if report['mr_notes']: ws['E27'] = report['mr_notes']
    ws['B28'] = report['fb_initial']
    ws['C28'] = report['fb_negotiated']
    ws['D28'] = '=B28-C28'
    if report['fb_notes']: ws['E28'] = report['fb_notes']
    brand = report['hotel_brand'] or 'Preferred'
    brand_cell = {'Hyatt': 'K7', 'Hilton': 'K8', 'Marriott': 'K9',
                  'IHG': 'K10', 'Preferred': 'K11'}.get(brand, 'K11')
    row = 31
    item_rows = []
    for it in items:
        ws.cell(row=row, column=1, value=it['item_name'])
        ct = it['calc_type']
        if ct == 'simple':
            if it['standard_price'] is not None:
                ws.cell(row=row, column=2, value=it['standard_price'])
            if it['negotiated_price'] is not None:
                ws.cell(row=row, column=3, value=it['negotiated_price'])
            if it['quantity'] is not None:
                ws.cell(row=row, column=4, value=it['quantity'])
            ws.cell(row=row, column=5, value=f'=(B{row}-C{row})*D{row}')
        elif ct == 'attrition':
            if it['quantity'] is not None:
                ws.cell(row=row, column=3, value=it['quantity'])
            ws.cell(row=row, column=5, value=f'=B9*B11*MAX(0.9-C{row},0)')
        elif ct == 'points':
            ws.cell(row=row, column=2, value=f'={brand_cell}')
            ws.cell(row=row, column=3, value=it['negotiated_price'] or 0)
            ws.cell(row=row, column=4, value=f'=(B11*B9)+C27')
            ws.cell(row=row, column=5, value=f'=(B{row}-C{row})*D{row}')
        if it['notes']:
            ws.cell(row=row, column=6, value=it['notes'])
        item_rows.append(row)
        row += 1
    total_items_row = row + 1
    ws.cell(row=total_items_row, column=1, value='TOTAL').font = Font(bold=True)
    if item_rows:
        ws.cell(row=total_items_row, column=5,
                value=f'=SUM(E{item_rows[0]}:E{item_rows[-1]})')
    else:
        ws.cell(row=total_items_row, column=5, value=0)
    hours_row = total_items_row + 2
    ws.cell(row=hours_row, column=1, value='Hours Worked:')
    ws.cell(row=hours_row, column=2, value=report['hours_worked'] or 0)
    grand_total_row = hours_row + 2
    ws.cell(row=grand_total_row, column=1, value='TOTAL COST SAVINGS:  ').font = Font(bold=True)
    ws.cell(row=grand_total_row, column=2,
            value=f'=B12+B19+B24+D27+D28+E{total_items_row}').font = Font(bold=True)
    money_cells = ['B8','B9','B10','B12','B15','B16','B17','B19','B24',
                   'B27','C27','D27','B28','C28','D28', f'B{grand_total_row}']
    for addr in money_cells:
        ws[addr].number_format = '$#,##0.00'
    for r in item_rows:
        for col in ('B', 'C', 'E'):
            ws[f'{col}{r}'].number_format = '$#,##0.00'
    ws[f'E{total_items_row}'].number_format = '$#,##0.00'
    return {'grand_total_row': grand_total_row, 'hours_row': hours_row}


def _cs_build_workbook_for_report(report, items):
    from openpyxl import load_workbook
    wb = load_workbook(COST_SAVINGS_TEMPLATE_PATH)
    ws = wb.active
    ws.title = _sanitize_sheet_name(report['meeting_name'] or 'Cost Savings')
    _cs_populate_template_sheet(ws, report, items)
    return wb


@app.route('/cost-savings/<int:report_id>/export')
def cost_savings_export(report_id):
    db = get_db()
    report = _cs_get_report(db, report_id)
    if not report:
        flash('Report not found.', 'error')
        return redirect(url_for('cost_savings_dashboard'))
    items = _cs_get_items(db, report_id)
    wb = _cs_build_workbook_for_report(report, items)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    safe = ''.join(c if c.isalnum() or c in '-_' else '_'
                   for c in (report['meeting_name'] or f'report_{report_id}'))
    filename = f"{safe}_Cost_Savings.xlsx"
    return send_file(output, download_name=filename, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/rfp/<int:rid>/cost-savings/export-all')
def cost_savings_export_all(rid):
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from copy import copy
    db = get_db()
    rfp = db.execute('SELECT * FROM rfp WHERE id=?', (rid,)).fetchone()
    if not rfp:
        flash('RFP not found.', 'error')
        return redirect(url_for('rfp_dashboard'))
    rows = db.execute('''
        SELECT csr.* FROM cost_savings_report csr
        JOIN rfp_hotel rh ON rh.id = csr.rfp_hotel_id
        WHERE rh.rfp_id = ? ORDER BY csr.id
    ''', (rid,)).fetchall()
    if not rows:
        flash('No cost savings reports under this RFP yet.', 'error')
        return redirect(url_for('cost_savings_rfp', rid=rid))
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = 'Summary'
    summary_ws['A1'] = f"{rfp['rfp_name'] or rfp['rfp_code'] or 'RFP'} Cost Savings Summary"
    summary_ws['A1'].font = Font(bold=True, size=14)
    headers = ['Meeting Name', 'Hotel', 'Meeting Dates', 'Total Cost Savings', 'Hours Worked']
    for i, h in enumerate(headers, 1):
        c = summary_ws.cell(row=4, column=i, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', start_color='1A3A5C')
        c.alignment = Alignment(horizontal='center')
    summary_data_rows = []
    for report in rows:
        tpl = load_workbook(COST_SAVINGS_TEMPLATE_PATH).active
        sheet_name = _sanitize_sheet_name(report['meeting_name'] or f'Report {report["id"]}')
        base = sheet_name; n = 2
        while sheet_name in wb.sheetnames:
            sheet_name = (base[:28] + f' {n}')[:31]; n += 1
        new_ws = wb.create_sheet(sheet_name)
        for row in tpl.iter_rows():
            for cell in row:
                tgt = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    tgt.font = copy(cell.font); tgt.fill = copy(cell.fill)
                    tgt.border = copy(cell.border); tgt.alignment = copy(cell.alignment)
                    tgt.number_format = cell.number_format; tgt.protection = copy(cell.protection)
        for col_letter, dim in tpl.column_dimensions.items():
            new_ws.column_dimensions[col_letter].width = dim.width
        items = _cs_get_items(db, report['id'])
        meta = _cs_populate_template_sheet(new_ws, report, items)
        summary_data_rows.append((report, sheet_name, meta))
    for i, (report, sheet_name, meta) in enumerate(summary_data_rows):
        srow = 5 + i
        sn = sheet_name.replace("'", "''")
        summary_ws.cell(row=srow, column=1, value=f"='{sn}'!A1")
        summary_ws.cell(row=srow, column=2, value=f"='{sn}'!A2")
        summary_ws.cell(row=srow, column=3, value=f"='{sn}'!A3")
        summary_ws.cell(row=srow, column=4, value=f"='{sn}'!B{meta['grand_total_row']}")
        summary_ws.cell(row=srow, column=4).number_format = '$#,##0.00'
        summary_ws.cell(row=srow, column=5, value=f"='{sn}'!B{meta['hours_row']}")
    tot_row = 5 + len(summary_data_rows) + 1
    summary_ws.cell(row=tot_row, column=3, value='Totals').font = Font(bold=True)
    summary_ws.cell(row=tot_row, column=4,
                    value=f'=SUM(D5:D{4 + len(summary_data_rows)})').font = Font(bold=True)
    summary_ws.cell(row=tot_row, column=4).number_format = '$#,##0.00'
    summary_ws.cell(row=tot_row, column=5,
                    value=f'=SUM(E5:E{4 + len(summary_data_rows)})').font = Font(bold=True)
    widths = [36, 36, 22, 18, 12]
    for i, w in enumerate(widths, 1):
        summary_ws.column_dimensions[summary_ws.cell(1, i).column_letter].width = w
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    code = (rfp['rfp_code'] or rfp['rfp_name'] or f'RFP_{rid}').strip().replace(' ', '_')
    filename = f"{code}_Cost_Savings_Report.xlsx"
    return send_file(output, download_name=filename, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


if __name__ == '__main__':
    app.run(debug=True, port=5051, use_reloader=False)
